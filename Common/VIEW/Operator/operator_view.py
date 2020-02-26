from datetime import date, datetime
from django.conf import settings
from django.contrib.auth import authenticate
from django.contrib.auth.hashers import make_password
from django.shortcuts import render, redirect
import requests
from django.contrib.auth import authenticate, login as auth_login, logout
import json
from django_global_request.middleware import get_request
from django.http import HttpResponseRedirect, HttpResponse
from django.contrib import messages
from openpyxl import Workbook

from Common.models import Operator_Login_Access_Token


def operator_homepage(request):
    if 'operator_login_type' in request.session:
        user_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        payload = {'operator_id':request.user.id}
        url = settings.API_BASE_URL + "operator_dashboard"
        data = getDataFromAPI(user_type, access_token, url, payload)
        print(data)
        dataDashboard = data['Dashboard']
        return render(request,'Operator/operator_home_page.html',{'user': request.user,'dataDashboard':dataDashboard})
    else:
        return HttpResponseRedirect("/operator/login")


def user_profile(request):
    return render(request, 'Operator/user_profile.html', {'user': request.user})


def operator_login_action(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('email', '')
        password = make_password(request.POST.get('password', ''))
        user = authenticate(username=username, post_password=password, login_type="7")
        print(user)
        if user is not None:
            if user:
                request.session.set_expiry(7200)  # sets the exp. value of the session
                request.session.set_expiry(7200)  # sets the exp. value of the session
                user_type_login = request.session['operator_login_type']
                access_token_login = request.session['operator_access_token']
                auth_login(request, user, backend='Common.backends.CustomCompanyUserAuth')  # the user is now logged in
                request.session['operator_access_token'] = access_token_login
                request.session['operator_login_type'] = user_type_login
                request.session.set_expiry(7200)  # sets the exp. value of the session
                messages.success(request, 'Login Successfully..!')
                return redirect("/operator/operator_home")
        else:
            messages.error(request, 'Invalid Email Or Password..!')
            return render(request,'Operator/corporate_operator_login.html',context)
    else:
        return render(request,'Operator/corporate_operator_login.html',context)


def operator_logout_action(request):
    if 'operator_login_type' in request.session:
        request = get_request()
        access_token = request.session['operator_access_token']
        user = Operator_Login_Access_Token.objects.get(access_token=access_token)
        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        del request.session['operator_login_type']
        del request.session['operator_access_token']
        return redirect("/operator/login")
    else:
        return redirect("/operator/login")


def operator_contacts(request, id):
    if 'operator_login_type' in request.session:
        if request.method == 'POST':
            user_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']

            operator_id = request.POST.get('operator_id', '')
            operator_address = request.POST.get('operator_address', '')
            contact_name = request.POST.get('contact_name', '')
            contact_email = request.POST.get('contact_email', '')
            contact_no = request.POST.get('contact_no', '')

            contact_id = request.POST.get('contact_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')
            url = ""
            payload = {}

            if contact_id:
                if delete_id == '1':
                    print("in delete")
                    url = settings.API_BASE_URL + "delete_operator_contact"
                    payload = {'contact_id':contact_id,'user_id':user_id,'user_type':user_type}
                    operation_message= "Operator Contact Deleted Successfully..!"
                else:
                    print("in edit")
                    url = settings.API_BASE_URL + "update_operator_contact"
                    payload = {'contact_id':contact_id,'operator_id':operator_id,'operator_address':operator_address,'contact_name':contact_name,'contact_email':contact_email,
                               'contact_no':contact_no,'user_id':user_id,'user_type':user_type}
                    operation_message = "Operator Contact Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operator_contact"
                payload = {'operator_id':operator_id,'operator_address':operator_address,'contact_name':contact_name,'contact_email':contact_email,
                               'contact_no':contact_no,'user_id':user_id,'user_type':user_type}
                operation_message = "Operator Contact Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/operator/operator-contacts/"+str(id), {'message': "Added Successfully"})
            else:
                messages.error(request, taxi['message'])
                return HttpResponseRedirect("/operator/operator-contacts/"+str(id), {'message': "Added Successfully"})
        else:
            login_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']

            url = settings.API_BASE_URL+"operator_contacts"
            payload = {'operator_id':id}
            operator_contacts = getDataFromAPI(login_type, access_token, url, payload)

            if operator_contacts['success'] == 1:
                operator_contacts = operator_contacts['OperatorContacts']
                return render(request,"Operator/operator_contacts.html",{'operator_contacts':operator_contacts,'operator_id':id})
            else:
                return render(request,"Operator/operator_contacts.html",{'operator_contacts':{}})
    else:
        return HttpResponseRedirect("/operator/login")



def operator_banks(request,id):
    if 'operator_login_type' in request.session:
        if request.method == 'POST':
            user_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']

            operator_id = request.POST.get('operator_id', '')
            beneficiary_name = request.POST.get('beneficiary_name', '')
            beneficiary_account_no = request.POST.get('beneficiary_account_no', '')
            bank_name = request.POST.get('bank_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')

            bank_id = request.POST.get('bank_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')

            url = ""
            payload = {}

            if bank_id:
                if delete_id == '1':
                    print("in delete")
                    url = settings.API_BASE_URL + "delete_operator_bank"
                    payload = {'bank_id': bank_id, 'user_id': user_id, 'user_type': user_type}
                    operation_message = "Operator Bank Deleted Successfully..!"
                else:
                    print("in edit")
                    url = settings.API_BASE_URL + "update_operator_bank"
                    payload = {'bank_id': bank_id, 'operator_id': operator_id,'beneficiary_name':beneficiary_name,
                               'beneficiary_account_no':beneficiary_account_no,'bank_name':bank_name,'ifsc_code':ifsc_code,
                               'user_id': user_id, 'user_type': user_type}
                    operation_message = "Operator Bank Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operator_bank"
                payload = {'operator_id': operator_id,'beneficiary_name':beneficiary_name,'beneficiary_account_no':beneficiary_account_no,'bank_name':bank_name,'ifsc_code':ifsc_code,
                           'user_id': user_id, 'user_type': user_type}
                operation_message = "Operator Bank Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)
            print(taxi)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/operator/operator-banks/"+str(id), {'message': "Added Successfully"})
            else :
                messages.error(request,taxi['message'])
                return HttpResponseRedirect("/operator/operator-banks/"+str(id), {'message': "Operation Fails"})

        else:
            login_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']

            url = settings.API_BASE_URL + "operator_banks"
            payload = {'operator_id': id}
            operator_banks = getDataFromAPI(login_type, access_token, url, payload)
            if operator_banks['success'] == 1:
                operator_banks = operator_banks['OperatorBanks']
                return render(request, "Operator/operator_banks.html", {'operator_banks': operator_banks,'operator_id':id})
            else:
                return render(request, "Operator/operator_banks.html", {'operator_contacts': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def operator_rates(request,id):
    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "operator_rates"
        payload = {'user_id':id}
        op_rates = getDataFromAPI(login_type, access_token, url, payload)
        if op_rates['success'] == 1:
            op_rates = op_rates['Rates']
            return render(request, "Operator/operator_rates.html", {'op_rates': op_rates})
        else:
            return render(request, "Operator/operator_rates.html", {'op_rates': {}})
    else:
        return HttpResponseRedirect("/operator/login")

    
def operator_drivers(request,id):
    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "operator_drivers"
        payload = {'user_id': request.user.id}
        op_drivers = getDataFromAPI(login_type, access_token, url, payload)
        if op_drivers['success'] == 1:
            drivers = op_drivers['Drivers']
            return render(request, "Operator/operator_drivers.html", {'op_drivers': drivers})
        else:
            return render(request, "Operator/operator_drivers.html", {'op_drivers': {}})
    else:
        return HttpResponseRedirect("/operator/login")



def add_operator_driver(request,id):
    if request.method == 'POST':
        if 'operator_login_type' in request.session:
            login_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']
            user_id = request.POST.get('cotrav_agent_id', '')

            operator_id = request.POST.get('operator_id', '')
            driver_name = request.POST.get('driver_name', '')
            driver_contact = request.POST.get('driver_contact', '')
            driver_email = request.POST.get('driver_email', '')
            licence_no = request.POST.get('licence_no', '')
            fcm_regid = request.POST.get('fcm_regid', '')
            taxi_id = request.POST.get('taxi_id', '')

            driver_id = request.POST.get('driver_id', '')
            if driver_id:
                password = ''
            else:
                password = "taxi123"
                driver_id = 0

            payload = {'operator_id':operator_id,'driver_name':driver_name,'driver_contact':driver_contact,'driver_email':driver_email,'licence_no':licence_no,
                          'fcm_regid':fcm_regid,'taxi_id':taxi_id,'password':password,'driver_id':driver_id,'user_id':user_id,'login_type':login_type}

            url = ""

            if driver_id:
                url = settings.API_BASE_URL + "update_operator_driver"
                operation_message = 'Operator Driver Updated Successfully..!'
            else:
                url = settings.API_BASE_URL + "add_operator_driver"
                operation_message = 'Operator Driver Added Successfully..!'

            op_drivers = getDataFromAPI(login_type, access_token, url, payload)

            if op_drivers['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/operator/operator-drivers/0", {'message': "Operation Successfully"})
            else:
                messages.error(request, op_drivers['message'])
                return HttpResponseRedirect("/operator/operator-drivers/0", {'message': "Operation Fails"})
        else:
            return HttpResponseRedirect("/operator/login")

    else:
        request = get_request()
        if 'operator_login_type' in request.session:
            if id:
                if 'operator_login_type' in request.session:
                    login_type = request.session['operator_login_type']
                    access_token = request.session['operator_access_token']
                    payload = {'driver_id': id}
                    url = settings.API_BASE_URL + "view_operator_driver"
                    op_drivers = getDataFromAPI(login_type, access_token, url, payload)
                    op_drivers = op_drivers['Drivers']

                    url = settings.API_BASE_URL + "operators"
                    operators = getDataFromAPI(login_type, access_token, url, payload)
                    operators = operators['Operators']

                    url_taxi = settings.API_BASE_URL + "taxis"
                    taxi = getDataFromAPI(login_type, access_token, url_taxi, payload)
                    taxi = taxi['Taxis']

                    return render(request, 'Operator/add_operator_driver.html', {'operator_drivers': op_drivers,'operators':operators,'taxies':taxi})
                else:
                    return HttpResponseRedirect("/operator/login")
            else:
                if 'operator_login_type' in request.session:
                    login_type = request.session['operator_login_type']
                    access_token = request.session['operator_access_token']

                    payload = {'operator_id': id}
                    url = settings.API_BASE_URL + "operators"
                    operators = getDataFromAPI(login_type, access_token, url, payload)
                    operators = operators['Operators']

                    url_taxi = settings.API_BASE_URL + "taxis"
                    taxi = getDataFromAPI(login_type, access_token, url_taxi, payload)
                    taxi = taxi['Taxis']

                    return render(request, 'Operator/add_operator_driver.html', {'operators':operators,'taxies':taxi})
                else:
                    return HttpResponseRedirect("/operator/login")
        else:
            return HttpResponseRedirect("/operator/login")


def delete_operator_driver(request,id):
    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        driver_id = request.POST.get('driver_id', '')
        user_id = request.POST.get('user_id', '')

        url = settings.API_BASE_URL+"delete_operator_driver"
        payload = {'driver_id': driver_id,'user_id':user_id,'user_type':login_type}
        operators = getDataFromAPI(login_type, access_token, url, payload)

        if operators['success'] == 1:
            messages.success(request, 'Operator Driver Deleted Successfully..!')
            return HttpResponseRedirect("/operator/operator-drivers/0", {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Delete Operator Driver..!')
            return HttpResponseRedirect("/operator/operator-drivers/0", {'message': "Operation Fails"})
    else:
        return HttpResponseRedirect("/operator/login")


############################################# TAXI #########################################

def taxi_bookings(request,id):

    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "operator_taxi_bookings"
        payload = {'operator_id': user_id, 'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/taxi_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Operator/taxi_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def view_taxi_booking(request,id):

    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/view_taxi_booking.html",{'bookings': booking})
        else:
            return render(request, "Operator/view_taxi_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def reject_taxi_booking(request,id):
    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "operator_reject_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Taxi Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/operator/login')


def assign_taxi_booking(request,id):
    if 'operator_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']
            current_url = request.POST.get('current_url', '')

            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')

            vendor_booking_id = request.POST.get('vendor_booking_id', '')
            operator_id = request.POST.get('operator_id', '')
            driver_contact = request.POST.get('driver_contact', '')
            driver_id = request.POST.get('driver_id', '')
            taxi_id = request.POST.get('taxi_id', '')
            taxi_types = request.POST.get('taxi_types', '')
            taxi_model = request.POST.get('taxi_model', '')
            tour_type = request.POST.get('tour_typ_save', '')

            taxi_type_id =0
            taxi_model_id= 0
            taxi_act_id =0
            oper_id= 0
            driver_id_id=0
            taxi_types_url = {'operator_id':operator_id, 'type_name':taxi_types,'user_type':login_type,'user_id':user_id}
            url_add = settings.API_BASE_URL + "add_taxi_type"
            country_id = getDataFromAPI(login_type, access_token, url_add, taxi_types_url)
            for conty_id in country_id['id']:
                taxi_type_id = conty_id['id']

            taxi_model_data = {'brand_name': "NA", 'model_name': taxi_model,'taxitype_id':taxi_type_id,'no_of_seats':4,'user_type':login_type,'user_id':user_id}
            url_add_model = settings.API_BASE_URL + "add_taxi_model"
            country_id = getDataFromAPI(login_type, access_token, url_add_model, taxi_model_data)
            for conty_id in country_id['id']:
                taxi_model_id = conty_id['id']

            taxi_model_data = {'model_id': taxi_model_id, 'taxi_reg_no': taxi_id,'make_year':0,'garage_location':"0", 'garage_distance':0,'user_type':login_type,'user_id':user_id}
            print(taxi_model_data)
            url_add_model = settings.API_BASE_URL + "add_taxi"
            country_id = getDataFromAPI(login_type, access_token, url_add_model, taxi_model_data)
            for conty_id in country_id['id']:
                taxi_act_id = conty_id['id']

            taxi_model_data = {'operator_name':operator_id, 'operator_email':"NA",'operator_contact':"NA",'type':tour_type, 'user_type': login_type,
                               'user_id': user_id,'username':"NA", 'password':"NA",'is_service_tax_applicable':0}

            url_add_model = settings.API_BASE_URL + "add_operator"
            country_id = getDataFromAPI(login_type, access_token, url_add_model, taxi_model_data)
            for conty_id in country_id['id']:
                oper_id = conty_id['id']

            taxi_model_data = {'operator_id':oper_id, 'driver_name':driver_id,'driver_contact':driver_contact, 'user_type': login_type,
                               'user_id': user_id,'fcm_regid':0,'password':"NA"}

            url_add_model = settings.API_BASE_URL + "add_operator_driver"
            country_id = getDataFromAPI(login_type, access_token, url_add_model, taxi_model_data)
            for conty_id in country_id['id']:
                driver_id_id = conty_id['id']

            is_client_sms = request.POST.get('is_client_sms', '')
            is_client_email = request.POST.get('is_client_email', '')
            is_driver_sms = request.POST.get('is_driver_sms', '')

            url = settings.API_BASE_URL + "assign_taxi_booking"
            payload = {'vendor_booking_id':vendor_booking_id,'operator_id':oper_id,'driver_id':driver_id_id,'is_client_sms':is_client_sms,
                       'is_client_email':is_client_email,'is_driver_sms':is_driver_sms,
                       'taxi_id':taxi_act_id,'booking_id': booking_id,'user_id':user_id,'user_type':login_type}
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Taxi Booking Assigned..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Assign Taxi Booking..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Fails"})
        else:
            login_type = request.session['operator_login_type']
            access_token = request.session['operator_access_token']
            payload = {'booking_id': id}
            opr_url = settings.API_BASE_URL + "operators"
            operators = getDataFromAPI(login_type, access_token, opr_url, payload)
            operators = operators['Operators']

            drivers_url = settings.API_BASE_URL + "operator_drivers"
            operator_drivers = getDataFromAPI(login_type, access_token, drivers_url, payload)
            operator_drivers = operator_drivers['Drivers']

            url_taxi = settings.API_BASE_URL + "taxi_models"
            taxis = getDataFromAPI(login_type, access_token, url_taxi, payload)
            models = taxis['Models']

            url_taxi_types = settings.API_BASE_URL + "taxi_types"
            url_taxi_types = getDataFromAPI(login_type, access_token, url_taxi_types, payload)
            taxi_types = url_taxi_types['taxi_types']

            url_taxis = settings.API_BASE_URL + "taxis"
            url_taxis = getDataFromAPI(login_type, access_token, url_taxis, payload)
            taxis = url_taxis['Taxis']

            url = settings.API_BASE_URL + "view_taxi_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)
            booking = booking['Bookings']

            return render(request, 'Operator/assign_taxi_booking.html',
                          {'bookings': booking, 'operators': operators, 'operator_drivers': operator_drivers,
                           'models': models,'taxi_types':taxi_types,'taxis':taxis})

    else:
        return HttpResponseRedirect("/operator/login")


############################################# BUS #########################################


def bus_bookings(request,id):

    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "operator_bus_bookings"
        payload = {'operator_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/bus_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Operator/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def view_bus_booking(request,id):

    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Operator/view_bus_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


############################################# TRAIN #########################################

def train_bookings(request,id):
    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "operator_train_bookings"
        payload = {'operator_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/train_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Operator/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def view_train_booking(request,id):
    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Operator/view_train_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")





############################################# HOTELS #########################################


def hotel_bookings(request,id):
    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "operator_hotel_bookings"
        payload = {'operator_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/hotel_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Operator/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def view_hotel_booking(request,id):
    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Operator/view_hotel_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")




############################################# TRAIN #########################################


def flight_bookings(request,id):
    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "operator_flight_bookings"
        payload = {'operator_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/flight_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Operator/flight_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")


def view_flight_booking(request,id):
    if 'operator_login_type' in request.session:
        request = get_request()
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Operator/view_flight_booking.html",{'bookings': booking})
        else:
            return render(request, "Operator/view_flight_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/operator/login")




####################### Download MIS ##################################

def dateonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError:
        return ''


def download_taxi_bookings(request):
    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

        # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [

        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',

        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'Passengers',

        'Booking Date',
        'Booking Time',


        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',

        'Package Name',
        'Tour Type',
        'Vehicle Type',

        'Driver Name',
        'Driver Contact	',
        'Taxi Reg No.',
        'No. Of Seats',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        spoc_status =''
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by= ''
        spoc_canceled_date =''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

            if bk['spoc_status'] == 1:
                spoc_status = "In-Active"
            else:
                spoc_status = "Active"

            if (act['action'] == 3 and act['user_type'] == 6):
                print('canceled')
                spoc_canceled_by = act['employee_name']
                spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [
            bk['reference_no'],
            bk['assessment_city_id'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],

            bk['zone_name'],
            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),

            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),
            dateonly(bk['pickup_datetime']),
            timeonly(bk['pickup_datetime']),

            bk['package_name'],
            bk['tour_type'],
            bk['taxi_type_name'],

            bk['driver_name'],
            bk['driver_contact'],
            'Taxi Reg No.',
            bk['no_of_seats'],

            bk['client_status'],

            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Bus Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Reason For Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Booking Date',
        'Booking Time',

        'Passanger Name',
        'Pickup City',
        'Drop City',
        'Journey Date',
        'Journey Time',
        'Current Booking Status',
        'Bus Type Allocated',
        'PNR Number',
        'Ticket Price',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by=''
        spoc_canceled_date=''
        spoc_status=''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']


                # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            passanger_list,
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),
            bk['cotrav_status'],
            bk['assign_bus_type_id'],
            bk['pnr_no'],
            bk['client_status'],
            bk['cotrav_status'],

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "Pickup City",
        "Drop City",
        "Booking Date",
        "Booking Time",
        "Journey Date",
        "Journey Time",
        "Passengers",
        "Coach Type Allocated",
        "Quota Used",
        'No of seats',
        'Operator name',
        'Operator contact',
        'Train name',
        'Ticket no',
        'pnr no',
        'Assign bus type id',
        'Seat no',
        'Portal used',
        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status=''
        spoc_canceled_by=''
        spoc_canceled_date=''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),
            passanger_list,
            bk['zone_name'],
            bk['train_type_priority_1'],
            bk['seat_no'],
            bk['no_of_seats'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['train_name'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['seat_no'],
            bk['portal_used'],
            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "From City",
        "To City",
        "Booking Date",
        "Booking Time",
        "Departure Date",
        "Departure Time",

        "Return Date",
        "Booking Status",

        "Passengers",

        'First Flight Name',
        'First Flight No',
        'First Flight PNR No.',
        'First Flight From',
        'First Flight To',
        'First Flight Departure Date',
        'First Flight Departure time',
        'First Flight Arrival Date',
        'First Flight Arrival time',

        'Second Flight Name',
        'Second Flight No',
        'Second Flight PNR No.',
        'Second Flight From',
        'Second Flight To',
        'Second Flight Departure Date',
        'Second Flight Departure Time',
        'Second Flight Arrival Datetime',
        'Second Flight Arrival Time',

        'Third Flight Name',
        'Third Flight No',
        'Third Flight PNR No.',
        'Third Flight From',
        'Third Flight To',
        'Third Flight Departure Date',
        'Third Flight Departure Time',
        'Third Flight Arrival Date',
        'Third Flight Arrival Time',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        spoc_status= ''
        spoc_canceled_by=''
        spoc_canceled_date=''


        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            "Booking Remarks",
            bk['from_location'],
            bk['to_location'],

            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),

            "",
            bk['cotrav_status'],

            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],

            passanger_list,

            flight_name1,
            flight_no1,
            pnr_no1,
            from_city1,
            to_city1,
            dateonly(departure_datetime1),
            timeonly(departure_datetime1),
            dateonly(arrival_datetime1),
            timeonly(arrival_datetime1),

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            dateonly(departure_datetime2),
            timeonly(departure_datetime2),
            dateonly(arrival_datetime2),
            timeonly(arrival_datetime2),

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            dateonly(departure_datetime3),
            timeonly(departure_datetime3),
            dateonly(arrival_datetime3),
            timeonly(arrival_datetime3),

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'operator_login_type' in request.session:
        login_type = request.session['operator_login_type']
        access_token = request.session['operator_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']



    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'No. of Nights',
        'Room Type',
        'Room Occupancy',
        'Per Night Price',
        'Total Room Price',
        'Tax On Room Cancellation',
        'Is Auto Approved',
        'Bill ID',
        'Is TBA Booking',
        'Is Offline Booking',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''


        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            'Billing Entity',
            'Travel request Code',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],

            bk['room_type_name'],
            bk['hotel_type_name'],
            bk['bucket_price_1'],
            bk['bucket_price_1'],
            '',
            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],


        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response




def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    r.close()
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
