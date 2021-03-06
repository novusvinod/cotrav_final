from django.conf import settings
from django.shortcuts import render, redirect
from django.http import HttpResponseRedirect, HttpResponse
import requests
import json
import string
import random
from datetime import date, datetime
from django_global_request.middleware import get_request
from django.contrib.auth.hashers import make_password
from django.contrib import messages
from openpyxl import Workbook

from Common.models import Corporate_Login_Access_Token

def logout_action(request):
    if 'admin_login_type' in request.session:
        request = get_request()
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user = Corporate_Login_Access_Token.objects.get(access_token=access_token)
        del request.session['admin_login_type']
        del request.session['admin_access_token']

        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        #logout(request)  # the user is now LogOut
        return redirect("/login")
    else:
        return redirect("/login")


def homepage(request):

    if 'admin_login_type' in request.session:

        user_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        print(request.user)
        print("Admin Home Page ")
        print(user_type)
        print(access_token)
        payload = {'admin_id':request.user.id, 'corporate_id':request.user.corporate_id}
        url = settings.API_BASE_URL + "admin_dashboard"

        data = getDataFromAPI(user_type, access_token, url, payload)
        dataDashboard = data['Dashboard']
        print(dataDashboard)
        return render(request, 'Company/Admin/home_page.html', {'user': request.user,'dataDashboard':dataDashboard})
    else:
        print("i m from home els part")
        return HttpResponseRedirect("/login")

def user_profile(request):
    return render(request, 'Company/Admin/user_profile.html', {'user': request.user})


def company_admins(request, id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            admins = company['Admins']
            return render(request, "Company/Admin/company_admins.html", {'admins': admins})
        else:
            return render(request, "Company/Admin/company_admins.html", {'admins': {}})
    else:
        return HttpResponseRedirect("/login")


def assessment_cities(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id;

        if request.method == 'POST':
            corporate_id = request.POST.get('corporate_id', '')
            city_name = request.POST.get('city_name', '')
            current_url = request.POST.get('current_url', '')
            city_id = request.POST.get('city_id', '')

            payload = {'corporate_id':corporate_id,'city_name':city_name,'city_id':city_id,'login_type':login_type,'user_id':user_id}

            if city_id:
                url = settings.API_BASE_URL + "update_assessment_cities"
                opr_msg = "Assessment City Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_assessment_cities"
                opr_msg = "Assessment City Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, company['error'])
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "assessment_cities"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            if company['success'] == 1:
                cities = company['Cities']
                companies = company1['Corporates']
                return render(request, "Company/Admin/assessment_cities.html", {'cities': cities,'companies':companies})
            else:
                return render(request, "Company/Admin/assessment_cities.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/login")


def assessment_codes(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id;
        if request.method == 'POST':
            current_url = request.POST.get('current_url', '')
            corporate_id = request.POST.get('corporate_id', '')
            assessment_code = request.POST.get('assessment_code', '')
            code_desc = request.POST.get('code_desc', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            service_from = request.POST.get('service_from', '')
            service_to = request.POST.get('service_to', '')

            code_id = request.POST.get('code_id', '')

            payload = {'corporate_id': corporate_id, 'assessment_code': assessment_code, 'code_desc': code_desc,'from_date':from_date,'to_date':to_date,
                       'login_type': login_type, 'user_id': user_id,'service_from':service_from,'service_to':service_to,'code_id':code_id}

            if code_id:
                url = settings.API_BASE_URL + "update_assessment_codes"
                opr_msg = "Assessment Code Updated Successfully"
            else:
                url = settings.API_BASE_URL + "add_assessment_codes"
                opr_msg = "Assessment Code Added Successfully"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, "Operation Fail")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "assessment_codes"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            if company['success'] == 1:
                codes = company['Codes']
                companies = company1['Corporates']
                return render(request, "Company/Admin/assessment_codes.html", {'codes': codes,'companies':companies})
            else:
                return render(request, "Company/Admin/assessment_codes.html", {'codes': {}})
    else:
        return HttpResponseRedirect("/login")


def delete_assessment_codes(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        code_id = request.POST.get('code_id')
        current_url = request.POST.get('current_url', '')

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            payload = {'code_id': code_id, 'user_id': user_id, 'login_type': login_type}
            url = settings.API_BASE_URL + "delete_assessment_codes"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Assessment Code Deleted Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Deleted Successfully"})
            else:
                messages.error(request, 'Failed to Delete Assessment Code..!')
                return HttpResponseRedirect(current_url, {'message': "Fails"})
        else:
            return HttpResponseRedirect("/login")


def delete_assessment_cities(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        city_id = request.POST.get('city_id')
        current_url = request.POST.get('current_url', '')

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            payload = {'city_id': city_id, 'user_id': user_id, 'login_type': login_type}
            url = settings.API_BASE_URL + "delete_assessment_cities"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Assessment City Deleted Successfully')
                return HttpResponseRedirect(current_url, {'message': "Deleted Successfully"})
            else:
                messages.error(request, 'Failed to Delete Assessment City..!')
                return HttpResponseRedirect(current_url, {'message': "Fails"})
        else:
            return HttpResponseRedirect("/login")


def company_billing_entities(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': id}

        company = getDataFromAPI(login_type, access_token, url, payload)
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)

        if company['success'] == 1:
            entities = company['Entitys']
            cities = cities["Cities"]
            return render(request, "Company/Admin/billing_entities.html",
                          {'billing_entities': entities, "cities": cities, })
        else:
            return render(request, "Company/Admin/billing_entities.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_rates(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            company_rates = company['Corporate_Retes']
            return render(request, "Company/Admin/company_rates.html", {'corporate_rates': company_rates})
        else:
            return render(request, "Company/Admin/company_rates.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_groups(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            groups = company['Groups']
            return render(request, "Company/Admin/groups.html", {'groups': groups})
        else:
            return render(request, "Company/Admin/groups.html", {'groups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_subgroups(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            url2 = settings.API_BASE_URL + "groups"
            subgroups = company['Subgroups']
            gr = getDataFromAPI(login_type, access_token, url2, payload)
            groups = gr['Groups']
            return render(request, "Company/Admin/subgroups.html", {'subgroups': subgroups, 'groups': groups})
        else:
            return render(request, "Company/Admin/subgroups.html", {'subgroups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_spocs(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            spocs = company['Spocs']

            url_subgroup = settings.API_BASE_URL + "subgroups"
            company_sub = getDataFromAPI(login_type, access_token, url_subgroup, payload)
            subgroups = company_sub['Subgroups']

            return render(request, "Company/Admin/spocs.html", {'spocs': spocs, 'subgroups':subgroups})
        else:
            return render(request, "Company/Admin/spocs.html", {'spocs': {}})
    else:
        return HttpResponseRedirect("/login")


def company_employees(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            employees = company['Employees']
            return render(request, "Company/Admin/employees.html", {'employees': employees})
        else:
            return render(request, "Company/Admin/employees.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/login")


def add_company_rate(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            url = settings.API_BASE_URL + "company_rates"
            payload = {'corporate_id': id}
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, 'Company Package Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Admin/company-rates/" + id, {'message': "Added Successfully"})
            else:
                messages.error(request, "Similar Package Already Exists..!")
                return HttpResponseRedirect("/Corporate/Admin/company-rates/" + id, {'message': "Operation Fails"})
        else:
            return HttpResponseRedirect("/login")
    else:
        return render(request, "Company/Admin/company_rate_add.html", {'entities': {}})
        pass


def add_company_entity(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            current_url = request.POST.get('current_url', '')
            entity_name = request.POST.get('entity_name', '')
            billing_city_id = request.POST.get('billing_city_id')
            contact_person_name = request.POST.get('contact_person_name', '')
            contact_person_email = request.POST.get('contact_person_email', '')
            contact_person_no = request.POST.get('contact_person_no', '')
            address_line_1 = request.POST.get('address_line_1', '')
            address_line_2 = request.POST.get('address_line_2', '')
            address_line_3 = request.POST.get('address_line_3', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            entity_id = request.POST.get('entity_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token,
                       'entity_name': entity_name, 'billing_city_id': billing_city_id,
                       'contact_person_name': contact_person_name, 'contact_person_email': contact_person_email,
                       'contact_person_no': contact_person_no, 'address_line_1': address_line_1,
                       'address_line_2': address_line_2,
                       'address_line_3': address_line_3, 'gst_id': gst_id, 'pan_no': pan_no, 'entity_id': entity_id,
                       'is_delete': delete_id, }

            if entity_id:
                url = settings.API_BASE_URL + "update_billing_entity"
                operation_message = "Company Entity Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_billing_entity"
                    operation_message = "Company Entity Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_billing_entity"
                operation_message = "Company Entity Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url,{})
            else:
                messages.error(request, 'Billing Entity Already Exists..!')
                return HttpResponseRedirect(current_url,{})
        else:
            return HttpResponseRedirect("/login")


def add_company_group(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = id
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            current_url = request.POST.get('current_url', '')
            group_name = request.POST.get('group_name', '')
            zone_name = request.POST.get('zone_name')

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))
            password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'group_name': group_name, 'zone_name': zone_name,'access_token_auth':access_token_auth,'name':name,
                       'email':email,'cid':cid,'contact_no':contact_no,'is_radio':is_radio,'is_local':is_local,'is_outstation':is_outstation,'is_bus':is_bus,
                       'is_train':is_train,'is_hotel':is_hotel,'is_meal':is_meal,'is_flight':is_flight,'is_water_bottles':is_water_bottles,'is_reverse_logistics':is_reverse_logistics,
                       'password':password,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = settings.API_BASE_URL + "add_group"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Company Group Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, 'Group Name Already Exists..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()
        if 'admin_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            current_url = request.POST.get('current_url', '')
            subgroup_name = request.POST.get('group_name', '')
            group_id = request.POST.get('group_id', '')

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))
            password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'subgroup_name': subgroup_name, 'group_id': group_id,'name':name,
                       'email':email,'cid':cid,'contact_no':contact_no,'is_radio':is_radio,'is_local':is_local,'is_outstation':is_outstation,'is_bus':is_bus,
                       'is_train':is_train,'is_hotel':is_hotel,'is_meal':is_meal,'is_flight':is_flight,'is_water_bottles':is_water_bottles,'is_reverse_logistics':is_reverse_logistics,
                       'password':password,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            #print(payload)
            url = settings.API_BASE_URL + "add_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print("sdsadasd")
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Add Company Subgroup Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, 'Subgroup Name Already Exists..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def update_company_group(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            group_id = request.POST.get('group_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            group_name = request.POST.get('group_name', '')
            zone_name = request.POST.get('zone_name')

            payload = {'group_id': group_id, 'access_token': access_token, 'group_name': group_name, 'zone_name': zone_name,
                       'user_id': user_id, 'login_type': login_type}

            print(payload)
            url = settings.API_BASE_URL + "update_group"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Group Updated Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Update Successfully"})
            else:
                messages.error(request, 'Company Group Update Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Updated"})
        else:
            return HttpResponseRedirect("/login")


def update_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            group_name = request.POST.get('group_name', '')

            payload = {'subgroup_id': subgroup_id, 'access_token': access_token, 'group_name': group_name,
                       'user_id': user_id, 'login_type': login_type}

            print(payload)
            url = settings.API_BASE_URL + "update_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Subgroup Updated Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Update Successfully"})
            else:
                messages.error(request, 'Company Subgroup Update Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Updated"})
        else:
            return HttpResponseRedirect("/login")


def delete_company_group(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            group_id = request.POST.get('group_id')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            access_token_auth = request.session['admin_access_token']
            payload = {'group_id': group_id, 'user_id': user_id, 'login_type': login_type, 'access_token': access_token,
                       'access_token_auth': access_token_auth}
            url = settings.API_BASE_URL + "delete_group"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Group Deleted Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Delete Successfully"})
            else:
                messages.error(request, 'Company Group Deletion Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def delete_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            subgroup_id = request.POST.get('subgroup_id')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = request.session['admin_access_token']

            payload = {'subgroup_id': subgroup_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'access_token_auth': access_token_auth}
            url = settings.API_BASE_URL + "delete_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(payload)
            if company['success'] == 1:
                messages.success(request, 'Company Sub-group Deleted Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Delete Successfully"})
            else:
                messages.error(request, 'Company Sub-group Deletion Failed..!')
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_group_auth(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            group_id = request.POST.get('group_id')
            group_auth_id = request.POST.get('group_auth_id')
            delete_id = request.POST.get('delete_id')

            if group_id:
                group_auth_id = group_auth_id
                password = "taxi123"

            if group_auth_id:
                group_auth_id = group_auth_id
            else:
                group_auth_id = 0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'group_id': group_id, 'delete_id': delete_id, 'password': password, 'group_auth_id': group_auth_id,
                       'access_token_auth': access_token_auth,'is_send_sms':is_send_sms,'is_send_email':is_send_email}

            url = ""
            if group_auth_id:
                url = settings.API_BASE_URL + "update_group_auth"
                oper_msg = "Company Group Authentication Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_group_auth"
                    oper_msg = "Company Group Authentication Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_group_auth"
                oper_msg = "Company Group Authentication Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, oper_msg)
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Authenticator Already Exists. Try with another data..!")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_subgroup_auth(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            subgroup_id = request.POST.get('subgroup_id')
            subgroup_auth_id = request.POST.get('subgroup_auth_id')
            delete_id = request.POST.get('delete_id')

            if subgroup_id:
                subgroup_auth_id = subgroup_auth_id
                password = "taxi123"

            if subgroup_auth_id:
                subgroup_auth_id = subgroup_auth_id
            else:
                subgroup_auth_id = 0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'subgroup_id': subgroup_id, 'delete_id': delete_id, 'password': password,
                       'subgroup_auth_id': subgroup_auth_id, 'access_token_auth': access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = ""
            if subgroup_auth_id:
                url = settings.API_BASE_URL + "update_subgroup_auth"
                operation_msg = "Company SubGroup Authentication Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_subgroup_auth"
                    operation_msg = "Company SubGroup Authentication Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_subgroup_auth"
                operation_msg = "Company SubGroup Authentication Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_msg)
                return HttpResponseRedirect(current_url,{'message': "Added Successfully"})
            else:
                messages.error(request, "Authenticator Already Exists. Try with another data..!")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_company_admins(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            has_billing_access = request.POST.get('has_billing_access', '')
            admin_id = request.POST.get('admin_id')

            delete_id = request.POST.get('delete_id')

            if admin_id:
                password = ''
            else:
                password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'admin_id': admin_id, 'delete_id': delete_id, 'password': password,'has_billing_access':has_billing_access,
                       'access_token_auth': access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = ""
            print(payload)
            if admin_id:
                url = settings.API_BASE_URL + "update_admin"
                opration_msg = "Company Admin Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_admin"
                    opration_msg = "Company Admin Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_admin"
                opration_msg = "Company Admin Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, opration_msg)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, "Admin Already Exists. Try with another data..!")
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")


def add_spocs(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')
            user_cid = request.POST.get('user_cid', '')

            user_name = request.POST.get('user_name', '')
            user_contact = request.POST.get('user_contact', '')
            email = request.POST.get('email', '')
            username = request.POST.get('email', '')
            budget = request.POST.get('budget', '')
            expense = request.POST.get('budget', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            spoc_id = request.POST.get('spoc_id')

            delete_id = request.POST.get('delete_id')

            if spoc_id:
                password = ''
            else:
                password = "taxi123"
                spoc_id =0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'group_id': group_id, 'subgroup_id': subgroup_id, 'user_cid': user_cid, 'user_name': user_name,
                       'user_contact':user_contact,'email':email,'username':username,'budget':budget,'expense':expense,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'spoc_id': spoc_id, 'delete_id': delete_id, 'password': password,'is_send_sms':is_send_sms,'is_send_email':is_send_email}

            url = ""
            print(payload)
            if spoc_id:
                url = settings.API_BASE_URL + "update_spoc"
                success_msg = "Spoc Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_spoc"
                    success_msg = "Spoc De-Activated Successfully..!"
                if delete_id == '2':
                    url = settings.API_BASE_URL + "active_spoc"
                    success_msg = "Spoc Activated Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_spoc"
                success_msg = "Spoc Added Successfully..1"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, success_msg)
                return HttpResponseRedirect("/Corporate/Admin/company-spoc/"+str(corporate_id), {'message': "Added Successfully"})
            else:
                messages.error(request, "Spoc Already Exists. Try with another data..!")
                return HttpResponseRedirect("/Corporate/Admin/company-spoc/"+str(corporate_id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")
    else:
        request = get_request()

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_spoc = settings.API_BASE_URL + "view_spoc"
            payload = {'spoc_id': id}
            r = requests.post(url_spoc, data=payload, headers=headers)
            company_spoc = json.loads(r.text)
            spocs = company_spoc['Spoc']

            url = settings.API_BASE_URL + "groups"
            payload = {'corporate_id': request.user.corporate_id}
            company = getDataFromAPI(login_type, access_token, url, payload)
            groups = company['Groups']

            url_subgroup = settings.API_BASE_URL + "subgroups"
            company_sub = getDataFromAPI(login_type, access_token, url_subgroup, payload)
            subgroups = company_sub['Subgroups']

            if id:
                return render(request, 'Company/Admin/add_spoc.html', {'groups': groups, 'subgroups': subgroups, 'spoc':spocs})
            else:
                return render(request, 'Company/Admin/add_spoc.html', {'groups': groups, 'subgroups': subgroups})
        else:
            return HttpResponseRedirect("/login")


def add_employee(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'admin_login_type' in request.session:
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            spoc_id = request.POST.get('spoc_id', '')
            billing_entity_id = request.POST.get('billing_entity_id', '')
            core_employee_id = request.POST.get('core_employee_id', '')
            employee_cid = request.POST.get('employee_cid', '')

            employee_name = request.POST.get('employee_name', '')
            employee_email = request.POST.get('employee_email', '')
            username = request.POST.get('employee_email', '')
            employee_contact = request.POST.get('employee_contact', '')

            date_of_birth = request.POST.get('date_of_birth','')
            if date_of_birth and date_of_birth != 'None':
                age = calculate_age(date_of_birth)
            else:
                age = 0

            gender = request.POST.get('gender')
            id_proof_type = request.POST.get('id_proof_type')

            id_proof_no = request.POST.get('id_proof_no', '')
            is_active = request.POST.get('is_active', '')
            has_dummy_email = 0
            fcm_regid = request.POST.get('fcm_regid', '')
            is_cxo = request.POST.get('is_cxo', '')
            designation = request.POST.get('designation', '')
            home_city = request.POST.get('home_city', '')
            home_address = request.POST.get('home_address', '')
            reporting_manager = request.POST.get('reporting_manager', '')
            employee_band = request.POST.get('employee_band', '')

            if is_cxo == '1':
                assistant_id = request.POST.get('assistant_id', '')
                if assistant_id == ' ':
                    assistant_id =0
            else:
                assistant_id = 0

            employee_id = request.POST.get('employee_id')

            delete_id = request.POST.get('delete_id')

            if employee_id:
                password = ''
            else:
                password = "taxi123"
                employee_id =0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,'spoc_id':spoc_id,'core_employee_id':core_employee_id,
                       'access_token': access_token,'employee_cid':employee_cid,'employee_name':employee_name,'employee_email':employee_email,
                       'employee_contact':employee_contact,'age':age,'gender':gender,'id_proof_type':id_proof_type,'id_proof_no':id_proof_no,
                       'is_active':is_active,'has_dummy_email':has_dummy_email,'fcm_regid':fcm_regid,'is_cxo':is_cxo,'employee_id': employee_id,
                       'designation':designation,'home_city':home_city,'home_address':home_address,'assistant_id':assistant_id,'date_of_birth':date_of_birth,
                       'delete_id': delete_id, 'password': password,'billing_entity_id':billing_entity_id,'username':username,'reporting_manager':reporting_manager,
                       'employee_band':employee_band}

            url = ""
            print(payload)
            if employee_id:
                url = settings.API_BASE_URL + "update_employee"
                success_msg  = "Employee Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_employee"
                    success_msg  = "Employee Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_employee"
                success_msg = "Employee Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(url)
            print(company)
            if company['success'] == 1:
                messages.success(request, success_msg)
                return HttpResponseRedirect("/Corporate/Admin/company-employees/"+str(corporate_id), {'message': "Added Successfully"})
            else:
                messages.error(request, "Employee Already Exists. Try with another data..!")
                return HttpResponseRedirect("/Corporate/Admin/company-employees/"+str(corporate_id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/login")
    else:
        request = get_request()

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            url_emp = settings.API_BASE_URL + "view_employee"
            payload = {'employee_id': id, 'corporate_id' : request.user.corporate_id}

            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employee']

            url_spoc = settings.API_BASE_URL + "spocs"
            payload_spoc = {'corporate_id': request.user.corporate_id}
            company_spoc = getDataFromAPI(login_type, access_token, url_spoc, payload_spoc)
            spocs = company_spoc['Spocs']

            url_entity = settings.API_BASE_URL + "billing_entities"
            payload_entity = {'corporate_id': request.user.corporate_id}
            company_entity = getDataFromAPI(login_type, access_token, url_entity, payload_entity)
            entitys1 = company_entity['Entitys']

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']

            url_emp = settings.API_BASE_URL + "employees"
            employees1 = getDataFromAPI(login_type, access_token, url_emp, payload)
            employeess = employees1['Employees']

            if id:
                return render(request, 'Company/Admin/add_employee.html', {'employee':employees,'spocs':spocs,'entitys':entitys1,'cities':cities,'employees':employeess})
            else:
                return render(request, 'Company/Admin/add_employee.html', {'spocs':spocs,'entitys':entitys1,'cities':cities,'employees':employeess})
        else:
            return HttpResponseRedirect("/login")


def view_company_group(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_group"
        payload = {'group_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_group_auth"
        grp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(grp_auths)
        if company['success'] == 1:
            groups = company['Groups']
            grp_auths = grp_auths['Groups']
            return render(request, "Company/Admin/view_groups.html", {'group': groups, 'grp_auths': grp_auths})
        else:
            return render(request, "Company/Admin/view_groups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")


def view_company_subgroup(request, id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_subgroup"
        payload = {'subgroup_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_subgroup_auth"
        subgrp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(subgrp_auths)
        if company['success'] == 1:
            subgroups = company['SubGroups']
            subgrp_auths = subgrp_auths['SubGroups']
            return render(request, "Company/Admin/view_subgroups.html",
                          {'subgroup': subgroups, 'subgrp_auths': subgrp_auths})
        else:
            return render(request, "Company/Admin/view_subgroups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")

# #####################   TAXI ###########################################

def taxi_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_taxi_bookings"
        payload = {'booking_type': id,'corporate_id':corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/taxi_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/taxi_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_taxi_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_taxi_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_taxi_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_taxi_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')
        user_id = request.user.id

        url = settings.API_BASE_URL + "admin_accept_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Taxi Booking Failed To Accept..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_taxi_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            messages.error(request, 'Taxi Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Taxi Booking')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_taxi_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            user_id = request.POST.get('user_id', '')

            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            tour_type = request.POST.get('tour_type', '')
            pickup_city = request.POST.get('pickup_city', '')
            pickup_location = request.POST.get('pickup_location', '')
            drop_location = request.POST.get('drop_location', '')
            pickup_datetime = request.POST.get('pickup_datetime', '')
            taxi_type = request.POST.get('taxi_type', '')
            package_id = request.POST.get('package_id', '')
            no_of_days = request.POST.get('no_of_days', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            actual_city_id = request.POST.get('current_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            if tour_type == 1 or tour_type == '1':
                url_add_city = settings.API_BASE_URL + "add_city_name"
                pickup_details = [x.strip() for x in pickup_city.split(',')]
                city_data = {'login_type': login_type, 'access_token': access_token, 'city_name': pickup_details[0], 'state_id': '1'}
                city_id = getDataFromAPI(login_type, access_token, url_add_city, city_data)
                for conty_id in city_id['id']:
                    actual_city_id = conty_id['id']

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'tour_type':tour_type,'pickup_city':actual_city_id,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'pickup_location':pickup_location,'drop_location':drop_location,'pickup_datetime':pickup_datetime+':00','taxi_type':taxi_type,
                       'package_id':package_id,'no_of_days':no_of_days,'reason_booking':reason_booking,'no_of_seats':no_of_seats,
                       'employees':employees,'user_id':user_id,'entity_id':entity_id,'is_sms':1,'is_email':1}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_taxi_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/taxi-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed To Add Taxi Booking')
                return HttpResponseRedirect("/Corporate/Admin/taxi-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_taxi = settings.API_BASE_URL + "taxi_types"
            taxies = getDataFromAPI(login_type, access_token, url_taxi, payload)
            taxies = taxies['taxi_types']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_city1 = settings.API_BASE_URL + "city_by_package"
            cities1 = getDataFromAPI(login_type, access_token, url_city1, payload)
            citiess = cities1['Cities']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_taxi_booking.html', {'employees':employees,'entities':entities,'cities':cities,
                'taxies':taxies,'assessments':ass_code,'citiess':citiess, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_taxi_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

#####################################  BUS  #####################################


def bus_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_bus_bookings"
        payload = {'corporate_id': corporate_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/bus_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_bus_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_bus_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_bus_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Accepted Successfully...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Bus Booking...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_bus_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Bus Booking...!')
            return HttpResponseRedirect(current_url , {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_bus_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            bus_type = request.POST.get('bus_type', '')
            bus_type2 = request.POST.get('bus_type2', '')
            bus_type3 = request.POST.get('bus_type3', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            preferred_board_point = request.POST.get('preferred_board_point', '')
            preferred_drop_point = request.POST.get('preferred_drop_point', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'bus_type':bus_type,'booking_datetime':booking_datetime,'journey_datetime':journey_datetime+':00','journey_datetime_to':journey_datetime_to+':00','entity_id':entity_id,
                       'preferred_bus':preferred_bus,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'is_sms':1,'is_email':1,
                       'preferred_board_point':preferred_board_point, 'preferred_drop_point':preferred_drop_point,'bus_type2':bus_type2,'bus_type3':bus_type3}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_bus_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/bus-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Bus Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/bus-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_bus_type = settings.API_BASE_URL + "bus_types"
            bus_type = getDataFromAPI(login_type, access_token, url_bus_type, payload)
            bus_types = bus_type['Types']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_bus_booking.html', {'bus_types':bus_types,'employees':employees,'cities':cities,
                 'entities':entities,'assessments':ass_code, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_bus_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

#####################################  TRAIN  #####################################


def train_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_train_bookings"
        payload = {'corporate_id':corporate_id,'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/train_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_train_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_train_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_train_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Accepted Successfully...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Train Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_train_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Train Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_train_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            train_type = request.POST.get('train_type', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'train_type':train_type,'booking_datetime':booking_datetime,'journey_datetime':journey_datetime+':00','entity_id':entity_id,
                       'preferred_bus':preferred_bus,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,
                       'is_sms':1,'is_email':1,'journey_datetime_to':journey_datetime_to+':00'}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_train_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/train-bookings/1", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Train Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/train-bookings/1", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_bus_type = settings.API_BASE_URL + "train_types"
            bus_type = getDataFromAPI(login_type, access_token, url_bus_type, payload)
            train_types = bus_type['Types']


            # url_railway_stations = settings.API_BASE_URL + "railway_stations"
            # trains1 = getDataFromAPI(login_type, access_token, url_railway_stations, payload)
            # railway_stations = trains1['Stations']
            railway_stations = ""

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_train_booking.html', {'train_types':train_types,'employees':employees,'cities':cities,
                'entities':entities,'assessments':ass_code,'railway_stations':railway_stations, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_train_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


#####################################  Hotels  #####################################


def hotel_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_hotel_bookings"
        payload = {'corporate_id':corporate_id,'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/hotel_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_hotel_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_hotel_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_hotel_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Hotel Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_hotel_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Reject Hotel Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_hotel_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_city = request.POST.get('from_city')
            city_area = request.POST.get('city_area')
            preferred_hotel_area = request.POST.get('preferred_hotel_area')
            check_in_date = request.POST.get('check_in_date')
            check_out_date = request.POST.get('check_out_date')
            room_type_priority1 = request.POST.get('room_type_priority1')
            room_type_priority2 = request.POST.get('room_type_priority2')
            room_occupancy = request.POST.get('room_occupancy')
            preferred_hotel = request.POST.get('preferred_hotel')
            booking_date = request.POST.get('booking_datetime')
            no_of_nights = request.POST.get('no_of_nights', '')
            assessment_code = request.POST.get('assessment_code')

            assessment_city = request.POST.get('assessment_city')
            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')

            no_of_seats = 1

            employees = []

            for i in range(1,2):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from_city_id': from_city, 'from_area_id': city_area,
                       'preferred_area': preferred_hotel_area, 'checkin_datetime': check_in_date+':00',
                       'checkout_datetime': check_out_date+':00', 'bucket_priority_1': room_type_priority1,
                       'bucket_priority_2': room_type_priority2, 'room_type_id': room_occupancy,
                       'preferred_hotel': preferred_hotel, 'booking_datetime': booking_date,'no_of_nights':no_of_nights,
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city,
                       'billing_entity_id': billing_entity, 'employees': employees,'reason_booking':reason_for_booking,'no_of_seats':no_of_seats,
                       'is_sms':1,'is_email':1}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_hotel_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/hotel-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Hotel Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/hotel-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            # url_city = settings.API_BASE_URL + "cities"
            # cities = getDataFromAPI(login_type, access_token, url_city, payload)
            # cities = cities['Cities']
            cities = ""

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_room_types = settings.API_BASE_URL + "room_types"
            room_types = getDataFromAPI(login_type, access_token, url_room_types, payload)
            room_types = room_types['Types']

            url_hotel_types = settings.API_BASE_URL + "hotel_types"
            hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
            hotel_types = hotel_types['Types']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            if id:
                return render(request, 'Company/Admin/add_hotel_booking.html', {'hotel_types':hotel_types,'room_types':room_types,'employees':employees,
                'cities':cities,'entities':entities,'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs})
            else:
                return render(request, 'Company/Admin/add_hotel_booking.html', {})
        else:
            return HttpResponseRedirect("/login")
#####################################  FLIGHT  #####################################


def flight_bookings(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        corporate_id = request.user.corporate_id

        url = settings.API_BASE_URL + "admin_flight_bookings"
        payload = {'corporate_id':corporate_id,'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/flight_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Admin/flight_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_flight_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/view_flight_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Admin/view_flight_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def accept_flight_booking(request,id):
    request = get_request()
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_accept_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Accepted Successfully...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed To Accept Flight Booking...!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def reject_flight_booking(request,id):
    request = get_request()

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        user_comment = request.POST.get('user_comment', '')

        url = settings.API_BASE_URL + "admin_reject_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id,'user_comment':user_comment}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Reject Flight Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
    else:
        return HttpResponseRedirect("/login")


def add_flight_booking(request,id):
    if request.method == 'POST':
        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')

            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            usage_type = request.POST.get('usage_type', '')
            trip_type = request.POST.get('trip_type', '')
            seat_type = request.POST.get('seat_type', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            departure_date = request.POST.get('departure_date', '')
            preferred_flight = request.POST.get('preferred_flight', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            if entity_id:
                pass
            else:
                entity_id=0

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'usage_type':usage_type,'trip_type':trip_type,'seat_type':seat_type,'from_city':from_city,'to_city':to_city,
                       'booking_datetime':booking_datetime,'departure_datetime':departure_date,'preferred_flight':preferred_flight,'assessment_code':assessment_code,
                       'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'billing_entity_id':entity_id,
                       'is_sms':1,'is_email':1,'assessment_city_id':assessment_city_id}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
            print(booking)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Flight Booking..!')
                return HttpResponseRedirect("/Corporate/Admin/flight-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'admin_login_type' in request.session:
            request = get_request()
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "employees"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            url_spoc = settings.API_BASE_URL + "spocs"
            spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = spoc['Spocs']

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            if id:
                return render(request, 'Company/Admin/add_flight_booking.html', {'employees':employees,'entities':entities,
                            'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access,'spocs':spocs,'airports':airports})
            else:
                return render(request, 'Company/Admin/add_flight_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

####################### Download MIS ##################################

def dateonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError:
        return ''


def download_taxi_bookings(request):
    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

        # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [

        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',

        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'Passengers',

        'Booking Date',
        'Booking Time',

        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        'Approver1 Action',

        'Approver1 Name',

        'Approver1 Date',

        'Approver1 Time',

        'Approver2 Action',

        'Approver2 Name',

        'Approver2 Date',

        'Approver2 Time',

        'Approved Date',

        'Approved Time',

        'Approved By',

        'Canceled Date',

        'Canceled Time',

        'Canceled By',

        'Assigned Date',

        'Assigned Time',

        'Assigned By',

        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',

        'Package Name',
        'Tour Type',
        'Vehicle Type',

        'Driver Name',
        'Driver Contact	',
        'Taxi Reg No.',
        'No. Of Seats',

        'Current Booking Status',

        'Hours Done',

        'Allowed Hours',

        'Extra Hours',

        'Kms Done',

        'Allowed Kms',

        'Extra Kms',

        'Extra Hours Charges',

        'Base Price',

        'Management Fee',
        'Tax on management fee',
        'Sub Total',
        'Cotrav Billing Entity',
        'IGST',
        'CGST',
        'SGST',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',

        'Estimated Amount',
        'Is Auto Approved',
        'Bill ID',
        'Bill Date',
        'Billing Entity',

        'Is Auto Approved',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        spoc_status =''
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by= ''
        spoc_canceled_date =''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

            if bk['spoc_status'] == 1:
                spoc_status = "In-Active"
            else:
                spoc_status = "Active"

            if (act['action'] == 3 and act['user_type'] == 6):
                print('canceled')
                spoc_canceled_by = act['employee_name']
                spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [
            bk['reference_no'],
            bk['assessment_city_id'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],

            bk['zone_name'],
            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),
            dateonly(bk['pickup_datetime']),
            timeonly(bk['pickup_datetime']),

            bk['package_name'],
            bk['tour_type'],
            bk['taxi_type_name'],

            bk['driver_name'],
            bk['driver_contact'],
            'Taxi Reg No.',
            bk['no_of_seats'],

            bk['cotrav_status'],

            bk['ci_hours_done'],

            bk['ci_allowed_hours'],

            bk['ci_extra_hours'],

            bk['ci_kms_done'],

            bk['ci_allowed_kms'],

            bk['ci_extra_kms'],

            bk['ci_extra_hr_charges'],

            bk['base_rate'],

            bk['ci_management_fee'],
            bk['ci_tax_on_management_fee'],
            bk['ci_sub_total'],
            bk['ci_cotrav_billing_entity'],
            bk['ci_igst'],
            bk['ci_cgst'],
            bk['ci_sgst'],
            bk['ci_management_fee_igst'],
            bk['ci_management_fee_cgst'],
            bk['ci_management_fee_sgst'],
            bk['ci_management_fee_igst_rate'],
            bk['ci_management_fee_cgst_rate'],
            bk['ci_management_fee_sgst_rate'],

            '',
            '',
            '',
            '',
            '',
            '',

            bk['client_status'],

            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Bus Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Reason For Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",
        'Approver1 Action',
        'Approver1 Name',
        'Approver1 Date',
        'Approver1 Time',
        'Approver2 Action',
        'Approver2 Name',
        'Approver2 Date',
        'Approver2 Time',
        'Approved Date',
        'Approved Time',
        'Approved By',
        'Canceled Date',
        'Canceled Time',
        'Canceled By',
        'Assigned Date',
        'Assigned Time',
        'Assigned By',
        'Passanger Name',
        'Pickup City',
        'Drop City',
        'Journey Date',
        'Journey Time',
        'Current Booking Status',
        'Bus Type Allocated',
        'PNR Number',
        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',
        'Is Auto Approved',
        'Bill ID',
        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by=''
        spoc_canceled_date=''
        spoc_status=''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']


                # Define the data for each cell in the row
        row = [

            bk['reference_no'],

            bk['assessment_code'],
            bk['assessment_city_id'],

            bk['reason_booking'],

            bk['zone_name'],

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            dateonly(bk['booking_datetime']),

            timeonly(bk['booking_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            passanger_list,

            bk['pickup_location'],

            bk['drop_location'],

            dateonly(bk['pickup_from_datetime']),

            timeonly(bk['pickup_from_datetime']),

            bk['cotrav_status'],

            bk['assign_bus_type_id'],

            bk['pnr_no'],

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            '',

            '',

            bk['client_status'],
            bk['cotrav_status'],

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "Pickup City",
        "Drop City",
        "Booking Date",
        "Booking Time",
        "Journey Date",
        "Journey Time",
        "SPOC Status",
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        "Zone",
        "Coach Type Allocated",
        "Quota Used",

        'No of seats',
        'Operator name',
        'Operator contact',
        'Train name',
        'Ticket no',
        'pnr no',
        'Assign bus type id',
        'Seat no',
        'Portal used',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status=''
        spoc_canceled_by=''
        spoc_canceled_date=''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            bk['zone_name'],
            bk['train_type_priority_1'],
            bk['seat_no'],


            bk['no_of_seats'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['train_name'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['seat_no'],
            bk['portal_used'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "From City",
        "To City",
        "Booking Date",
        "Booking Time",
        "Departure Date",
        "Departure Time",

        "Return Date",
        "Booking Status",

        "SPOC Status",
        "SPOC Cancel By",
        "SPOC Cancel Date",

        "Usage Type",
        "Trip Type",
        "Flight Type",
        "Seat Type",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        'First Flight Name',
        'First Flight No',
        'First Flight PNR No.',
        'First Flight From',
        'First Flight To',
        'First Flight Departure Date',
        'First Flight Departure time',
        'First Flight Arrival Date',
        'First Flight Arrival time',

        'Second Flight Name',
        'Second Flight No',
        'Second Flight PNR No.',
        'Second Flight From',
        'Second Flight To',
        'Second Flight Departure Date',
        'Second Flight Departure Time',
        'Second Flight Arrival Datetime',
        'Second Flight Arrival Time',

        'Third Flight Name',
        'Third Flight No',
        'Third Flight PNR No.',
        'Third Flight From',
        'Third Flight To',
        'Third Flight Departure Date',
        'Third Flight Departure Time',
        'Third Flight Arrival Date',
        'Third Flight Arrival Time',

        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        spoc_status= ''
        spoc_canceled_by=''
        spoc_canceled_date=''


        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            "Booking Remarks",
            bk['from_location'],
            bk['to_location'],

            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),

            "",
            bk['cotrav_status'],

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            flight_name1,
            flight_no1,
            pnr_no1,
            from_city1,
            to_city1,
            dateonly(departure_datetime1),
            timeonly(departure_datetime1),
            dateonly(arrival_datetime1),
            timeonly(arrival_datetime1),

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            dateonly(departure_datetime2),
            timeonly(departure_datetime2),
            dateonly(arrival_datetime2),
            timeonly(arrival_datetime2),

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            dateonly(departure_datetime3),
            timeonly(departure_datetime3),
            dateonly(arrival_datetime3),
            timeonly(arrival_datetime3),

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']



    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Bucket Request Date',
        'Bucket Request Time',
        'Bucket Approved Date',
        'Bucket Approved Time',
        'Preferred Hotel',
        'Assigned Hotel',
        'Assigned Hotel Address',
        'Hotel Contact',
        'Assign Date',
        'Assign Time',
        'TaxiVaxi Status',
        'Employees Name',
        'No. of Persons',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Higher Bucket Requested',
        'Reason for Higher Bucket',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',
        'Current Booking Status',
        'No. of Nights',
        'Room Type',
        'Room Occupancy',
        'Per Night Price',
        'Total Room Price',
        'Tax On Room Cancellation',
        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',
        'Is Auto Approved',
        'Bill ID',
        'Is TBA Booking',
        'Is Offline Booking',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")


        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            '',
            '',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            '',
            '',
            '',
            '',

            bk['preferred_hotel'],
            bk['assign_hotel_id'],
            bk['operator_name'],
            bk['operator_contact'],
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['status_cotrav'],
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),
            bk['status_cotrav'],
            '',
            bk['room_type_name'],
            bk['hotel_type_name'],
            bk['bucket_price_1'],
            bk['bucket_price_1'],
            '',
            bk['ticket_price'],

            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['sub_total'],
            bk['billing_entity_id'],
            bk['igst'],
            bk['cgst'],
            bk['sgst'],
            bk['management_fee_igst'],
            bk['management_fee_cgst'],
            bk['management_fee_sgst'],
            bk['management_fee_igst_rate'],
            bk['management_fee_cgst_rate'],
            bk['management_fee_sgst_rate'],

            '',
            '',
            '',
            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],


        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_billing_entities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Entitys']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=BillingEntitys.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Billing Entitys'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Billing City",
        "Entity Name",
        "Contact Person Name",
        "Contact Person Email",
        "Contact Person Phone No",
        "Address Line 1",
        "Address Line 2",
        "Address Line 3",
        "GST NO",
        "PAN No",
        "Is Active",
    ]

    row_num = 1



    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"
        # Define the data for each cell in the row
        row = [

            row_num-1,
            bk['corporate_name'],
            bk['billing_city'],
            bk['entity_name'],
            bk['contact_person_name'],
            bk['contact_person_email'],
            bk['contact_person_no'],
            bk['address_line_1'],
            bk['address_line_2'],
            bk['address_line_3'],
            bk['gst_id'],
            bk['pan_no'],
            is_deleted,
            
        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_rates(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Corporate_Retes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyRates.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Rates'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City",
        "Taxi Type",
        "Tour Type",
        "KMS",
        "HOURS",
        "KM Rate",
        "HR Rate",
        "Base Rate",
        "Night Rate",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['tour_type'] == 2:
            tour_type = "Local"
        else:
            tour_type = "Outstation"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            bk['taxi_type'],
            tour_type,
            bk['kms'],
            bk['hours'],
            bk['km_rate'],
            bk['hour_rate'],
            bk['base_rate'],
            bk['night_rate'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_assessment_cities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "assessment_cities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Cities']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=AssessmentCities.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Assessment Cities'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_assessment_codes(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "assessment_codes"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Codes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=AssessmentCodes.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Assessment Codes'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Assessment Code",
        "Code Description",
        "From Date",
        "To Date",
        "Service From Date",
        "Service To Date",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_active'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['assessment_code'],
            bk['code_desc'],
            bk['from_date'],
            bk['to_date'],
            bk['service_from'],
            bk['service_to'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_groups(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Groups']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyGroups.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Groups'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Zone Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['zone_name'],

            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_subgroups(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Subgroups']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySubGroups.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company SubGroups'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgoup Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],

            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_admins(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Admins']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyAdmins.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Admins'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Admin Name",
        "Admin Email",
        "Admin Contact No ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['name'],
            bk['email'],
            bk['contact_no'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_spocs(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Spocs']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySpocs.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Spocs'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgroup Name",
        "Spoc Company ID",
        "Spoc Name",
        "Spoc Email",
        "Spoc Contact No ",
        "Login UserName ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['user_cid'],
            bk['user_name'],
            bk['email'],
            bk['user_contact'],
            bk['username'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_employees(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Employees']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyEmployees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Employees'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Spoc Name",
        "Core Employee ID",
        "Employee Company ID",
        "Employee Name",
        "Employee Email",
        "Employee Phone No",
        "Login UserName",
        "Age",
        "Gender",
        "ID Proof Type",
        "ID No",
        "Is CXO",
        "Designation",
        "Home City",
        "Home Address",
        "Assistant ID",
        "Date Of Birth",
        "Billing Entity",
        "Last Login",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_cxo'] == 0:
            is_cxo = "No"
        else:
            is_cxo = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['user_name'],
            bk['core_employee_id'],
            bk['core_employee_id'],
            bk['employee_cid'],
            bk['employee_name'],
            bk['employee_email'],
            bk['employee_contact'],
            bk['username'],
            bk['age'],
            bk['gender'],
            bk['id_proof_type'],
            bk['id_proof_no'],
            is_cxo,
            bk['designation'],
            bk['home_city'],
            bk['home_address'],
            bk['assistant_id'],
            bk['date_of_birth'],
            bk['billing_entity_id'],
            bk['last_login'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def taxi_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_taxi_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_taxi_billing.html", {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_taxi_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def taxi_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)

            vry_url = settings.API_BASE_URL + "admin_verify_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            print("current url")
            print(current_url)
            vry_url = settings.API_BASE_URL + "admin_revise_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def bus_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_bus_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_bus_billing.html", {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_bus_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def train_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_train_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_train_billing.html",
                          {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_train_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def flight_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_flight_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_flight_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_flight_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def hotel_billing(request,id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        url = settings.API_BASE_URL + "admin_hotel_bookings"
        payload = {'booking_type': id, 'corporate_id':request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Admin/admin_hotel_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Admin/admin_hotel_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def bus_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def train_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def hotel_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def flight_billing_verify(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_verify_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "admin_revise_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def accept_bill(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            current_url = request.POST.get('current_url', '')
            user_comment = request.POST.get('user_comment', '')

            url = ""
            operation_message = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "admin_accept_bill"
                operation_message="Bill Accepted successfully..!"

            if reject_id == '1':
                url = settings.API_BASE_URL + "admin_reject_bill"
                operation_message="Bill Rejected successfully..!"

            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, 'Failed to Accept Bill..!')
                return HttpResponseRedirect(current_url, {})
        else:
            return HttpResponseRedirect("/Corporate/Admin/bill/2")
    else:
        return HttpResponseRedirect("/login")

def get_all_generated_bills(request, id):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_all_generated_bills"
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = company['Bill']
        return render(request, 'Company/Admin/bills_geterated.html', {'bills': companies, 'bill_type':id})
    else:
        return HttpResponseRedirect("/login")


def corporate_bank_accounts(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        payload = {'corporate_id': request.user.corporate_id}
        url = settings.API_BASE_URL + "get_corporate_accounts"
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        pos = ''
        companies = ''
        if company['success'] == 1:
            pos = company['Accounts']
        else:
            pos = ''

        return render(request, 'Company/Admin/corporate_accounts.html', {'accounts': pos})
    else:
        return HttpResponseRedirect("/login")


def add_company_accounts(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')

        if 'admin_login_type' in request.session:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            current_url = request.POST.get('current_url', '')
            corporate_id = request.POST.get('corporate_id', '')
            bank_name = request.POST.get('bank_name')
            bank_branch = request.POST.get('bank_branch', '')
            acoount_no = request.POST.get('acoount_no', '')
            acoount_holder_name = request.POST.get('acoount_holder_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')
            micr_code = request.POST.get('micr_code', '')

            account_id = request.POST.get('account_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'bank_name': bank_name, 'bank_branch': bank_branch, 'acoount_no': acoount_no, 'acoount_holder_name': acoount_holder_name,
                       'ifsc_code': ifsc_code,'micr_code':micr_code,  'account_id': account_id, 'is_delete': delete_id, }

            url = ""
            if account_id:
                url = settings.API_BASE_URL + "update_corporate_account"
                operation_message = "Company Account Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_corporate_account"
                    operation_message = "Company Account Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_corporate_account"
                operation_message = "Company Account Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def dashboard_search_admin_api_call(request):
    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']
        serveType = int(request.POST.get('serveType', ''))

        corp_id = request.user.corporate_id

        booking_type = int(request.POST.get('booking_type', ''))

        bookings_from_date = request.POST.get('bookings_from_date', '')

        bookings_to_date = request.POST.get('bookings_to_date', '')

        booking_id = ""
        pickup_location = ""
        pickup_date1 = ""
        pickup_date2 = ""
        pickup_date3 = ""
        pickup_date4 = ""
        spoc_id = ""
        operator_name = ""

        city = ""
        pnr_no = ""
        ass_code = ""

        checkin_date = ""
        voucher_no = ""
        hotel_name = ""

        search_serve_url = ""

        whereClause = "b.corporate_id = '" + str(corp_id) + "' "

        if serveType == 1:
            booking_id = request.POST.get('booking_id', '')
            pickup_location = request.POST.get('pickup_location', '')
            pickup_date1 = request.POST.get('pickup_date1', '')
            spoc_id = request.POST.get('spoc_id', '')
            operator_name = request.POST.get('operator_name', '')
            current_url = request.POST.get('current_url', '')

        if serveType == 2:
            booking_id = request.POST.get('booking_id', '')
            pickup_date2 = request.POST.get('pickup_date2', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')

        if serveType == 3:
            booking_id = request.POST.get('booking_id', '')
            pickup_date3 = request.POST.get('pickup_date3', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')

        if serveType == 4:
            booking_id = request.POST.get('booking_id', '')
            pickup_date4 = request.POST.get('pickup_date4', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/flight_bookings.html'
        if serveType == 5:
            booking_id = request.POST.get('booking_id', '')
            checkin_date = request.POST.get('checkin_date', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            ass_code = request.POST.get('ass_code', '')
            voucher_no = request.POST.get('voucher_no', '')
            hotel_name = request.POST.get('hotel_name', '')
            current_url = request.POST.get('current_url', '')

        if (bookings_from_date and bookings_to_date):
            bookings_from_date = bookings_from_date + ' 00:00:00'
            bookings_to_date = bookings_to_date + ' 00:00:00'

            bookings_from_date_object = datetime.strptime(bookings_from_date, '%d-%m-%Y %H:%M:%S')
            bookings_to_date_object = datetime.strptime(bookings_to_date, '%d-%m-%Y %H:%M:%S')

            bookings_from_date = bookings_from_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            bookings_to_date = bookings_to_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            print(bookings_from_date)

            print(bookings_to_date)

            if serveType == 1:
                whereClause = whereClause + "AND " + "b.booking_date BETWEEN CAST('" + bookings_from_date + "' AS DATE) AND CAST('" + bookings_to_date + "' AS DATE) "
            else:
                whereClause = whereClause + "AND " + "b.booking_datetime BETWEEN CAST('" + bookings_from_date + "' AS DATE) AND CAST('" + bookings_to_date + "' AS DATE) "

        if pickup_location:
            whereClause = whereClause + "AND " + "b.pickup_location LIKE '%" + pickup_location + "%' "

        if pickup_date1:
            pickup_date1 = pickup_date1 + ' 00:00:00'
            pickup_date1_object = datetime.strptime(pickup_date1, '%d-%m-%Y %H:%M:%S')
            pickup_date1 = pickup_date1_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_datetime) = CAST('" + pickup_date1 + "' AS DATE) "

        if pickup_date2:
            pickup_date2 = pickup_date2 + ' 00:00:00'
            pickup_date2_object = datetime.strptime(pickup_date2, '%d-%m-%Y %H:%M:%S')
            pickup_date2 = pickup_date2_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_from_datetime) = CAST('" + pickup_date2 + "' AS DATE) "

        if pickup_date3:
            pickup_date3 = pickup_date3 + ' 00:00:00'
            pickup_date3_object = datetime.strptime(pickup_date3, '%d-%m-%Y %H:%M:%S')
            pickup_date3 = pickup_date3_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_from_datetime) = CAST('" + pickup_date3 + "' AS DATE) "

        if pickup_date4:
            pickup_date4 = pickup_date4 + ' 00:00:00'
            pickup_date4_object = datetime.strptime(pickup_date4, '%d-%m-%Y %H:%M:%S')
            pickup_date4 = pickup_date4_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.departure_datetime) = CAST('" + pickup_date4 + "' AS DATE) "

        if spoc_id:
            whereClause = whereClause + "AND " + "b.spoc_id = '" + spoc_id + "' "

        if operator_name:
            whereClause = whereClause

        if city:
            whereClause = whereClause + "AND " + "b.pickup_location LIKE '%" + city + "%' "

        if pnr_no:
            whereClause = whereClause + "AND " + "b.pnr_no = '" + pnr_no + "' "

        if ass_code:
            whereClause = whereClause + "AND " + "b.assessment_code = '" + ass_code + "' "

        if checkin_date:
            checkin_date = checkin_date + ' 00:00:00'
            checkin_date_object = datetime.strptime(checkin_date, '%d-%m-%Y %H:%M:%S')
            checkin_date = checkin_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.checkin_datetime) = CAST('" + checkin_date + "' AS DATE) "

        if voucher_no:
            whereClause = whereClause + "AND " + "b.voucher_no = '" + voucher_no + "' "

        if hotel_name:
            whereClause = whereClause + "AND " + "ht.name LIKE '%" + hotel_name + "%' "

        if booking_id:
            whereClause = "b.reference_no = '" + booking_id + "' "

        search_serve_url = "Company/Admin/dashboard_search_admin_result.html"

        payload = {'whereClause': whereClause, 'serveType': serveType, 'booking_type': booking_type}
        url = settings.API_BASE_URL + "dashboard_search_admin_bookings"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Search Result..!")
            return render(request, search_serve_url, {'bookings': verify['Result'], 'serveType': serveType})

        else:
            messages.error(request, "Sory for error...!")
            return HttpResponse("error")
    else:
        return HttpResponseRedirect("/login")


def reports_invoice(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0

            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            date_type = request.POST.get('date_type', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')

            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            payload = {'corporate_id': int(corporate_id), 'service_type': service_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}

            print("payload")

            print(payload)

            url = settings.API_BASE_URL + "admin_report_invoice"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            #print(operator)
            if operator['success'] == 1:
                operator = operator['Reports']
            else:
                operator = {}

            return render(request, 'Company/Admin/reports_invoice.html',
                          {'Reports': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Company/Admin/reports_invoice.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/login")



def reports_client_billing(request):
    if 'admin_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0
            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            bill_status = request.POST.get('bill_status', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'bill_status': bill_status, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "admin_report_client_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}

            return render(request, 'Company/Admin/reports_client_billing.html',
                          {'bills': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['admin_login_type']
            access_token = request.session['admin_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Company/Admin/reports_client_billing.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/login")


def download_invoice_reports(request):

    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        if not corporate_id:
            corporate_id = 0
        service_type = int(request.POST.get('service_type', ''))
        date_type = request.POST.get('date_type', '')
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')

        service_text = ""

        if service_type == 1:

            service_text = "taxi"

        elif service_type == 2:

            service_text = "bus"

        elif service_type == 3:

            service_text = "train"

        elif service_type == 4:

            service_text = "flight"

        elif service_type == 5:

            service_text = "hotel"

        else:

            service_text = "all"



        payload = {'corporate_id': int(corporate_id), 'service_type': service_type,'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
        print(payload)
        url = settings.API_BASE_URL + "admin_report_invoice"
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Reports']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report_invoice_'+ service_text +'_'+ from_date +'_'+ to_date + '.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Invoice Reports'

    # Define the titles for columns

    columns = [
            "Sr.No",
            "reference_no",
            "assessment_code",
             "assessment_city_id",
            "pickup_location",
            "drop_location",
            "spoc_id",
            "user_name",
            "user_contact",
            "booking_datetime",
            "boarding_datetime",
            "boarding_point",
            "portal_used",
            "operator_name",
            "operator_contact",
            "ticket_no",
            "pnr_no",
            "assign_bus_type_id",
            "ticket_price",
            "management_fee",
            "tax_on_management_fee",
            "tax_on_management_fee_percentage",
            "sub_total",
            "vi_ticket_price",
            "vender_commission",

            "vender_commission",
            "invoice_status",
    ]



    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1

        if service_type == 1:
            bk['booking_datetime'] = bk['booking_date']
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['portal_used'] = ''
            bk['ticket_no'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''
            bk['ticket_price'] = ''

        if service_type == 3:
            bk['vi_ticket_price'] = ''
            bk['vender_commission'] = ''
            bk['invoice_status'] = ''

        if service_type == 4:
            bk['assessment_city_id'] = ''
            bk['pickup_location'] = bk['from_location']
            bk['drop_location'] = bk['to_location']
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['portal_used'] = ''
            bk['operator_name'] = ''
            bk['operator_contact'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''

        if service_type == 5:
            bk['pickup_location'] = ''
            bk['drop_location'] = ''
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['ticket_no'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['pickup_location'],
            bk['drop_location'],
            bk['spoc_id'],
            bk['user_name'],
            bk['user_contact'],
            bk['booking_datetime'],
            bk['boarding_datetime'],
            bk['boarding_point'],
            bk['portal_used'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['ticket_price'],
            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['tax_on_management_fee_percentage'],
            bk['sub_total'],
            bk['vi_ticket_price'],


            bk['vender_commission'],

            bk['vender_commission'],
            bk['invoice_status'],

        ]





        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_client_bill_reports(request):

    request = get_request()

    booking = ''

    if 'admin_login_type' in request.session:
        login_type = request.session['admin_login_type']
        access_token = request.session['admin_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        if not corporate_id:
            corporate_id = 0
        service_type = int(request.POST.get('service_type', ''))
        bill_status = int(request.POST.get('bill_status', ''))
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')

        service_text = ""

        if service_type == 1:

            service_text = "taxi"

        elif service_type == 2:

            service_text = "bus"

        elif service_type == 3:

            service_text = "train"

        elif service_type == 4:

            service_text = "flight"

        elif service_type == 5:

            service_text = "hotel"

        else:

            service_text = "all"


        if bill_status == 1 :

            bill_text = "Unpaid"

        elif bill_status == 2 :

            bill_text = "partial"

        elif bill_status == 3:

            bill_text = "paid"

        else:
            bill_text = "All"


        payload = {'corporate_id': int(corporate_id), 'service_type': service_type,'bill_status': bill_status, 'from_date': from_date, 'to_date': to_date}
        print(payload)
        url = settings.API_BASE_URL + "admin_report_client_bills"
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bill']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report_client_bills_'+ service_text +'_'+ from_date +'_'+ to_date + '_' + bill_text +'.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Client Bill Reports'

    # Define the titles for columns

    columns = [
        "Sr.No",
        "CorporatemName ",
        "Bill Number",
        "No Of Invoices",
        "Cotrav Billing Entity",
        "Client  Billing Entity ",
        "Billing Type",
        "TDS  Deducted  By Client",
        "System  Calculated TDS",
        "IGST",
        "CGST",
        "SGST",
        "Total Amount",
        "Is Paid",
        "Payment Status",
        "Total GST Paid",
        "Management Fee",
        "Outstanding Pending Payment",
        "Paid Total Amount",
        "Balance Total Amount",
        "Advance Payment",
        "Is Offline",
        "Reimbursement Voucher",
        "ID Taxable Amount",
        "Nontaxable Amount",
        "PO Id",
        "Bill Created Date",
        "Bill Final Date",
        "User Comment",
        "Cotrav Status",
        "Client Status",
    ]



    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_id'],
            bk['bill_number'],
            bk['no_of_invoices'],
            bk['cotrav_billing_entity'],
            bk['client_billing_entity'],
            bk['billing_type'],
            bk['tds_deducted_by_client'],
            bk['system_calculated_tds'],
            bk['igst'],
            bk['cgst'],

            bk['sgst'],
            bk['total_amount'],
            bk['is_paid'],

            bk['payment_status'],
            bk['total_gst_paid'],
            bk['management_fee'],
            bk['outstanding_pending_payment'],

            bk['paid_total_amount'],
            bk['balance_total_amount'],
            bk['advance_payment'],

            bk['is_offline'],
            bk['reimbursement_voucher_id'],
            bk['taxable_amount'],
            bk['nontaxable_amount'],
            bk['po_id'],
            bk['bill_created_date'],
            bk['bill_final_date'],
            bk['user_comment'],
            bk['cotrav_status'],
            bk['client_status'],

        ]



        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response




def calculate_age(born):
    print(born)
    dt_str = datetime.strptime(born, '%d-%m-%Y')
    today = date.today()
    return today.year - dt_str.year - ((today.month, today.day) < (dt_str.month, dt_str.day))


def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
