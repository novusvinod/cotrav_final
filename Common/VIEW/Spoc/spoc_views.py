from datetime import date, datetime
from dateutil.parser import parse
from django.conf import settings
from django.shortcuts import render, redirect
import requests
import json
from time import sleep
from django_global_request.middleware import get_request
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from django.contrib import messages
from openpyxl import Workbook
from Common.models import Corporate_Spoc_Login_Access_Token
from landing.cotrav_messeging import Excelexport
from django.http import HttpResponse
import razorpay
razorpay_client = razorpay.Client(auth=("rzp_test_eipQBGxGd1SmmJ", "r82J3rVV4NEAZKMGxIJVPyGY"))

def logout_action(request):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']

        access_token = request.session['spoc_access_token']
        user = Corporate_Spoc_Login_Access_Token.objects.get(access_token=access_token)
        del request.session['spoc_login_type']
        del request.session['spoc_access_token']

        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        #logout(request)  # the user is now LogOut
        return redirect("/login")
    else:
        return redirect("/login")

def homepage(request):
    print("i ma session Data")
    for key, value in request.session.items():
        print('{} => {}'.format(key, value))
    print("i ma session Data")
    if 'spoc_login_type' in request.session:
        user_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        payload = {'spoc_id': request.user.id, 'corporate_id':request.user.corporate_id}
        url = settings.API_BASE_URL + "spoc_dashboard"
        data = getDataFromAPI(user_type, access_token, url, payload)
        dataDashboard = data['Dashboard']

        url_access = settings.API_BASE_URL + "view_company"
        data = getDataFromAPI(user_type, access_token, url_access, payload)
        access = data['Corporates']

        return render(request, 'Company/Spoc/home_page.html', {'user': request.user,'dataDashboard':dataDashboard, 'corp_access':access})
    else:
        return HttpResponseRedirect("/login")


def user_profile(request):
    return render(request, 'Company/Spoc/user_profile.html', {'user': request.user})


def company_admins(request, id):
    request = get_request()

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            admins = company['Admins']
            return render(request, "Company/Spoc/company_admins.html", {'admins': admins})
        else:
            return render(request, "Company/Spoc/company_admins.html", {'admins': {}})
    else:
        return HttpResponseRedirect("/login")


def company_billing_entities(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        if company['success'] == 1:
            entities = company['Entitys']
            cities = cities["Cities"]
            return render(request, "Company/Spoc/billing_entities.html",
                          {'billing_entities': entities, "cities": cities, })
        else:
            return render(request, "Company/Spoc/billing_entities.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_rates(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            company_rates = company['Corporate_Retes']
            return render(request, "Company/Spoc/company_rates.html", {'corporate_rates': company_rates})
        else:
            return render(request, "Company/Spoc/company_rates.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/login")


def company_groups(request, id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            groups = company['Groups']
            return render(request, "Company/Spoc/groups.html", {'groups': groups})
        else:
            return render(request, "Company/Spoc/groups.html", {'groups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_subgroups(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            url2 = settings.API_BASE_URL + "groups"
            subgroups = company['Subgroups']
            gr = getDataFromAPI(login_type, access_token, url2, payload)
            groups = gr['Groups']
            return render(request, "Company/Spoc/subgroups.html", {'subgroups': subgroups, 'groups': groups})
        else:
            return render(request, "Company/Spoc/subgroups.html", {'subgroups': {}})
    else:
        return HttpResponseRedirect("/login")


def company_spocs(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            spocs = company['Spocs']
            return render(request, "Company/Spoc/spocs.html", {'spocs': spocs})
        else:
            return render(request, "Company/Spoc/spocs.html", {'spocs': {}})
    else:
        return HttpResponseRedirect("/login")


def company_employees(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            employees = company['Employees']
            return render(request, "Company/Spoc/employees.html", {'employees': employees})
        else:
            return render(request, "Company/Spoc/employees.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/login")


def view_company_group(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_group"
        payload = {'group_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_group_auth"
        grp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(grp_auths)
        if company['success'] == 1:
            groups = company['Groups']
            grp_auths = grp_auths['Groups']
            return render(request, "Company/Spoc/view_groups.html", {'group': groups, 'grp_auths': grp_auths})
        else:
            return render(request, "Company/Spoc/view_groups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")


def view_company_subgroup(request, id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_subgroup"
        payload = {'subgroup_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        subgrp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(subgrp_auths)
        if company['success'] == 1:
            subgroups = company['SubGroups']
            subgrp_auths = subgrp_auths['SubGroups']
            return render(request, "Company/Spoc/view_subgroups.html",
                          {'subgroup': subgroups, 'subgrp_auths': subgrp_auths})
        else:
            return render(request, "Company/Spoc/view_subgroups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/login")


###################################### TAXI #############################################

def taxi_bookings(request,id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "spoc_taxi_bookings"
        payload = {'spoc_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/taxi_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Spoc/taxi_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_taxi_booking(request,id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/view_taxi_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Spoc/view_taxi_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_taxi_booking(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')

        url = settings.API_BASE_URL + "spoc_reject_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Reject Taxi Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_taxi_booking(request,id):
    if request.method == 'POST':
        if 'spoc_login_type' in request.session:
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            user_id = request.POST.get('user_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            tour_type = request.POST.get('tour_type', '')
            pickup_city = request.POST.get('pickup_city', '')
            pickup_location = request.POST.get('pickup_location', '')
            drop_location = request.POST.get('drop_location', '')
            pickup_datetime = request.POST.get('pickup_datetime', '')
            taxi_type = request.POST.get('taxi_type', '')
            package_id = request.POST.get('package_id', '')
            no_of_days = request.POST.get('no_of_days', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            actual_city_id = request.POST.get('current_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'tour_type':tour_type,'pickup_city':actual_city_id,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'pickup_location':pickup_location,'drop_location':drop_location,'pickup_datetime':pickup_datetime,'taxi_type':taxi_type,
                       'package_id':package_id,'no_of_days':no_of_days,'reason_booking':reason_booking,'no_of_seats':no_of_seats,
                       'employees':employees,'user_id':user_id,'entity_id':entity_id,'is_sms':1,'is_email':1}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_taxi_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, 'Taxi Booking Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Spoc/taxi-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'spoc_login_type' in request.session:
            request = get_request()
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "spoc_employee"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_taxi = settings.API_BASE_URL + "taxi_types"
            taxies = getDataFromAPI(login_type, access_token, url_taxi, payload)
            taxies = taxies['taxi_types']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            # url_city1 = settings.API_BASE_URL + "cities"
            # cities1 = getDataFromAPI(login_type, access_token, url_city1, payload)
            # citiess = cities1['Cities']
            citiess = ""

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Spoc/add_taxi_booking.html', {'employees':employees,'entities':entities,'cities':cities,
                                                                              'taxies':taxies,'assessments':ass_code,'citiess':citiess, 'corp_access':access})
            else:
                return render(request, 'Company/Spoc/add_taxi_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


############################################# BUS #########################################


def bus_bookings(request,id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "spoc_bus_bookings"
        payload = {'spoc_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/bus_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Spoc/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_bus_booking(request,id):

    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Spoc/view_bus_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_bus_booking(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')

        url = settings.API_BASE_URL + "spoc_reject_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Reject Bus Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_bus_booking(request,id):
    if request.method == 'POST':
        if 'spoc_login_type' in request.session:
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']

            user_id = request.POST.get('spoc_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            bus_type = request.POST.get('bus_type', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'bus_type':bus_type,'booking_datetime':booking_datetime,'journey_datetime':journey_datetime,'entity_id':entity_id,
                       'preferred_bus':preferred_bus,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,
                       'is_sms':1,'is_email':1,'journey_datetime_to':journey_datetime_to}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_bus_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, 'Bus Booking Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Spoc/bus-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'spoc_login_type' in request.session:
            request = get_request()
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "spoc_employee"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_bus_type = settings.API_BASE_URL + "bus_types"
            bus_type = getDataFromAPI(login_type, access_token, url_bus_type, payload)
            bus_types = bus_type['Types']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Spoc/add_bus_booking.html', {'bus_types':bus_types,'employees':employees,'cities':cities,
                                                                             'entities':entities,'assessments':ass_code, 'corp_access':access})
            else:
                return render(request, 'Company/Spoc/add_bus_booking.html', {})
        else:
            return HttpResponseRedirect("/login")



############################################# TRAIN #########################################

def train_bookings(request,id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "spoc_train_bookings"
        payload = {'spoc_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/train_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Spoc/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_train_booking(request,id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Spoc/view_train_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_train_booking(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')

        url = settings.API_BASE_URL + "spoc_reject_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Reject Train Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_train_booking(request,id):
    if request.method == 'POST':
        if 'spoc_login_type' in request.session:
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']

            user_id = request.POST.get('spoc_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            train_type = request.POST.get('train_type', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,
                       'train_type':train_type,'booking_datetime':booking_datetime,'journey_datetime':journey_datetime,'entity_id':entity_id,
                       'preferred_bus':preferred_bus,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,
                       'is_sms':1,'is_email':1,'journey_datetime_to':journey_datetime_to}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_train_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, 'Train Booking Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Spoc/train-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'spoc_login_type' in request.session:
            request = get_request()
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "spoc_employee"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_city = settings.API_BASE_URL + "get_assessment_city"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['AssCity']

            url_bus_type = settings.API_BASE_URL + "train_types"
            bus_type = getDataFromAPI(login_type, access_token, url_bus_type, payload)
            train_types = bus_type['Types']

            # url_railway_stations = settings.API_BASE_URL + "railway_stations"
            # trains1 = getDataFromAPI(login_type, access_token, url_railway_stations, payload)
            # railway_stations = trains1['Stations']
            railway_stations = ""

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Spoc/add_train_booking.html', {'train_types':train_types,'employees':employees,'cities':cities,
                'entities':entities,'assessments':ass_code,'railway_stations':railway_stations, 'corp_access':access})
            else:
                return render(request, 'Company/Spoc/add_train_booking.html', {})
        else:
            return HttpResponseRedirect("/login")

############################################# HOTELS #########################################


def hotel_bookings(request,id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "spoc_hotel_bookings"
        payload = {'spoc_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/hotel_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Spoc/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_hotel_booking(request,id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Spoc/view_hotel_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_hotel_booking(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')

        url = settings.API_BASE_URL + "spoc_reject_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Reject Hotel Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_hotel_booking(request,id):
    if request.method == 'POST':
        if 'spoc_login_type' in request.session:
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']

            user_id = request.POST.get('spoc_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            from_city = request.POST.get('from_city')
            city_area = request.POST.get('city_area')
            preferred_hotel_area = request.POST.get('preferred_hotel_area')
            check_in_date = request.POST.get('check_in_date')
            check_out_date = request.POST.get('check_out_date')
            room_type_priority1 = request.POST.get('room_type_priority1')
            room_type_priority2 = request.POST.get('room_type_priority2')
            room_occupancy = request.POST.get('room_occupancy')
            preferred_hotel = request.POST.get('preferred_hotel')
            booking_date = request.POST.get('booking_datetime')

            assessment_code = request.POST.get('assessment_code')

            assessment_city = request.POST.get('assessment_city')
            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')

            no_of_seats = 1

            employees = []

            for i in range(1,2):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from_city_id': from_city, 'from_area_id': city_area,
                       'preferred_area': preferred_hotel_area, 'checkin_datetime': check_in_date,
                       'checkout_datetime': check_out_date, 'bucket_priority_1': room_type_priority1,
                       'bucket_priority_2': room_type_priority2, 'room_type_id': room_occupancy,
                       'preferred_hotel': preferred_hotel, 'booking_datetime': booking_date,
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city,
                       'billing_entity_id': billing_entity, 'employees': employees,'reason_booking':reason_for_booking,'no_of_seats':no_of_seats,
                       'is_sms':1,'is_email':1}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_hotel_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, 'Hotel Booking Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Spoc/hotel-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'spoc_login_type' in request.session:
            request = get_request()
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "spoc_employee"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            # url_city = settings.API_BASE_URL + "cities"
            # cities = getDataFromAPI(login_type, access_token, url_city, payload)
            # cities = cities['Cities']
            cities = ""

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_room_types = settings.API_BASE_URL + "room_types"
            room_types = getDataFromAPI(login_type, access_token, url_room_types, payload)
            room_types = room_types['Types']

            url_hotel_types = settings.API_BASE_URL + "hotel_types"
            hotel_types = getDataFromAPI(login_type, access_token, url_hotel_types, payload)
            hotel_types = hotel_types['Types']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Spoc/add_hotel_booking.html', {'hotel_types':hotel_types,'room_types':room_types,'employees':employees,
                'cities':cities,'entities':entities,'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access})
            else:
                return render(request, 'Company/Spoc/add_hotel_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


############################################# TRAIN #########################################


def flight_bookings(request,id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "spoc_flight_bookings"
        payload = {'spoc_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/flight_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Spoc/flight_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def view_flight_booking(request,id):
    if 'spoc_login_type' in request.session:
        request = get_request()
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/view_flight_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Spoc/view_flight_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/login")


def reject_flight_booking(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        user_id = request.user.id
        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')

        url = settings.API_BASE_URL + "spoc_reject_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Reject flight Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def add_flight_booking(request,id):
    if request.method == 'POST':
        if 'spoc_login_type' in request.session:
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']

            user_id = request.POST.get('spoc_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            usage_type = request.POST.get('usage_type', '')
            trip_type = request.POST.get('trip_type', '')
            seat_type = request.POST.get('seat_type', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            departure_date = request.POST.get('departure_date', '')
            preferred_flight = request.POST.get('preferred_flight', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')
            entity_id = request.POST.get('entity_id', '')
            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            if entity_id:
                pass
            else:
                entity_id=0

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'usage_type':usage_type,'trip_type':trip_type,'seat_type':seat_type,'from_city':from_city,'to_city':to_city,
                       'booking_datetime':booking_datetime,'departure_datetime':departure_date,'preferred_flight':preferred_flight,'assessment_code':assessment_code,
                       'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'billing_entity_id':entity_id,
                       'is_sms':1,'is_email':1,'assessment_city_id':assessment_city_id}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, 'Flight Booking Added Successfully..!')
                return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/login")
    else:
        if 'spoc_login_type' in request.session:
            request = get_request()
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}

            url_emp = settings.API_BASE_URL + "spoc_employee"
            payload = {'corporate_id': id,'spoc_id':request.user.id}
            r = requests.post(url_emp, data=payload, headers=headers)
            company_emp = json.loads(r.text)
            employees = company_emp['Employees']

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']

            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities_ass = cities_ass['AssCity']

            url_access = settings.API_BASE_URL + "view_company"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            access = data['Corporates']

            if id:
                return render(request, 'Company/Spoc/add_flight_booking.html', {'employees':employees,'cities':cities,'entities':entities,
                            'assessments':ass_code,'cities_ass':cities_ass, 'corp_access':access})
            else:
                return render(request, 'Company/Spoc/add_flight_booking.html', {})
        else:
            return HttpResponseRedirect("/login")


def add_flight_booking_self(request,id):
    if request.method == 'POST':
        if 'spoc_login_type' in request.session:
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']

            user_id = request.POST.get('spoc_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')

            trip_type = request.POST.get('trip_type', '')
            return_date = request.POST.get('return_date', '')
            fl_class = request.POST.get('fl_class', '')
            no_of_seats = request.POST.get('no_of_seats', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            departure_date = request.POST.get('departure_date', '')

            booking_data = {'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from_city':from_city,'to_city':to_city,
                       'departure_datetime':departure_date,'return_date':return_date,'trip_type':trip_type,'fl_class':fl_class,'no_of_seats':no_of_seats}

            payload = {'auth_token':"",'session_id':access_token,'from_city':from_city,'to_city':to_city,'departure_date':departure_date,
                       'fl_class':fl_class,'return_date':return_date,'trip_type':trip_type,'no_of_seats':no_of_seats,}
            print(payload)
            url_flt = settings.API_BASE_URL + "get_flight_search"
            try:
                flightdata = getDataFromAPI(login_type, access_token, url_flt, payload)
            except Exception as e:
                url_access = settings.API_BASE_URL + "get_airports"
                data11 = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data11['Airports']
                messages.success(request, 'No Flight Found Please Try Another Flight.!')
                return render(request, 'Company/Spoc/add_flight_booking_self.html',
                              {'booking_datas': booking_data, 'flights': '', 'airports': airports})

            #print(flightdata['Data'])
            if 'FLIGHT' in flightdata['Data'] or 'FLIGHTOW' in flightdata['Data']:
                flight = flightdata['Data']
                #print(flight)
                url_access = settings.API_BASE_URL + "get_airports"
                data11 = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data11['Airports']
                flight1 = ""
                flight2 = ""
                unique_flights1 = ""
                unique_flights2 = ""
                print("trip type")
                print(trip_type)
                if trip_type == '2':
                    flight1 = flight['FLIGHTOW']
                    flight2 = flight['FLIGHTRT']
                    uniq_flights = ''
                    uniq_fl = []
                    uniq_code = []
                    uniq_img = []
                    uniq_flights2 = ''
                    uniq_fl2 = []
                    uniq_code2 = []
                    uniq_img2 = []
                    for fl_name in flight1:
                        if fl_name['F_NAME'] not in uniq_flights:
                            uniq_fl.append(fl_name['F_NAME'])
                            uniq_code.append(fl_name['F_CODE'])
                            uniq_img.append(fl_name['F_LOGO'])
                    for fl_name2 in flight2:
                        if fl_name2['F_NAME'] not in uniq_flights2:
                            uniq_fl2.append(fl_name2['F_NAME'])
                            uniq_code2.append(fl_name2['F_CODE'])
                            uniq_img2.append(fl_name2['F_LOGO'])

                else:
                    flight1 = flight['FLIGHT']
                    uniq_fl = []
                    uniq_code = []
                    uniq_img = []
                    uniq_flights = ''
                    for fl_name in flight1:
                        if fl_name['F_NAME'] not in uniq_flights:
                            uniq_fl.append(fl_name['F_NAME'])
                            uniq_code.append(fl_name['F_CODE'])
                            uniq_img.append(fl_name['F_LOGO'])

                    uniq_flights = set(zip(uniq_fl,uniq_code,uniq_img))
                    print("in trip 1")

                return render(request, 'Company/Spoc/add_flight_booking_serarch_result.html', {'booking_datas': booking_data,'params':flight['PARAM'], 'flights': flight1, 'flights2': flight2,
                'airports':airports,'no_of_seats':no_of_seats, 'uniq_flights':uniq_flights,'unique_flights2':unique_flights2, 'Deals':flight['Deal']})
            else:
                url_access = settings.API_BASE_URL + "get_airports"
                data11 = getDataFromAPI(login_type, access_token, url_access, payload)
                airports = data11['Airports']

                messages.success(request, 'No Flight Found Please Try Another Flight.!')
                return render(request, 'Company/Spoc/add_flight_booking_self.html',{'booking_datas': booking_data, 'flights': '', 'airports':airports})

        else:
            return HttpResponseRedirect("/login")
    else:
        if 'spoc_login_type' in request.session:
            request = get_request()
            login_type = request.session['spoc_login_type']
            access_token = request.session['spoc_access_token']
            payload = {'corporate_id': id,'spoc_id':request.user.id}

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            if id:
                return render(request, 'Company/Spoc/add_flight_booking_self.html', {'airports':airports})
            else:
                return render(request, 'Company/Spoc/add_flight_booking_self.html', {})
        else:
            return HttpResponseRedirect("/login")


def add_flight_booking_self_conformation(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        UID = request.POST.get('UID', '')
        ID = request.POST.get('ID', '')
        TID = request.POST.get('TID', '')

        UID2 = request.POST.get('UID2', '')
        ID2 = request.POST.get('ID2', '')
        TID2 = request.POST.get('TID2', '')

        src = request.POST.get('src', '')
        des = request.POST.get('des', '')
        ret_date = request.POST.get('ret_date', '')
        adt = request.POST.get('adt', '')
        chd = request.POST.get('chd', '')
        inf = request.POST.get('inf', '')
        L_OW = request.POST.get('L_OW', '')
        L_RT = request.POST.get('L_RT', '')
        H_OW = request.POST.get('H_OW', '')
        H_RT = request.POST.get('H_RT', '')
        T_TIME = request.POST.get('T_TIME', '')
        dep_date = request.POST.get('dep_date', '')
        trip_string = request.POST.get('trip_string', '')
        booking_datass = request.POST.get('booking_data', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        flight_class_is_international = request.POST.get('flight_class_is_international', '')
        flight1 = ""
        flight2 = ""
        booking_data = {'UID':UID,'ID':ID,'TID':TID,'UID2':UID2,'ID2':ID2,'TID2':TID2,'src':src,'des':des,'dep_date':dep_date,'ret_date':ret_date,
            'adt':adt,'chd':chd,'inf':inf,'L_OW':L_OW,'H_OW':H_OW,'T_TIME':T_TIME,'trip_string':trip_string,'flight_class_is_international':flight_class_is_international}

        url_tokn = settings.API_BASE_URL + "get_flight_fare_search"
        data = getDataFromAPI(login_type, access_token, url_tokn, booking_data)
        print("DATA TYPE")
        print(type(data))
        if data['success'] == 1:
            api_response = data['Data']
            print("SEARCH PAI RESPONSE")

            payload = {'corporate_id': request.user.corporate_id, 'spoc_id': request.user.id}
            url_enty = settings.API_BASE_URL + "billing_entities"
            entys = getDataFromAPI(login_type, access_token, url_enty, payload)
            entities = entys['Entitys']

            url_ass_code = settings.API_BASE_URL + "get_assessment_code"
            ass_code = getDataFromAPI(login_type, access_token, url_ass_code, payload)
            ass_code = ass_code['AssCodes']

            url_cities_ass = settings.API_BASE_URL + "get_assessment_city"
            cities_ass = getDataFromAPI(login_type, access_token, url_cities_ass, payload)
            cities = cities_ass['AssCity']

            url_emp = settings.API_BASE_URL + "employees"
            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employees']

            url_nat = settings.API_BASE_URL + "get_nationality"
            nationality = getDataFromAPI(login_type, access_token, url_nat, payload)
            nationalities = nationality['Nationality']

            return render(request, 'Company/Spoc/add_flight_booking_conformation.html', {'booking_datas': booking_data, 'flights': api_response, 'UID2': UID2, 'employees': employees, 'cities_ass': cities,
                'entities': entities, 'assessments': ass_code, 'no_of_seats':no_of_seats, 'flight_class_is_international':flight_class_is_international,'nationalities':nationalities})
        else:
            return render(request, 'Company/Spoc/add_flight_booking_conformation.html', {'booking_datas': booking_data, 'flights': ''})

    else:
        return HttpResponseRedirect("/login")


def add_flight_booking_self_final(request, id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        flightdata = request.POST.get('flightdata', '')
        UID2 = request.POST.get('UID2', '')

        user_id = request.POST.get('spoc_id', '')
        corporate_id = request.POST.get('corporate_id', '')
        spoc_id = request.POST.get('spoc_id', '')
        group_id = request.POST.get('group_id', '')
        subgroup_id = request.POST.get('subgroup_id', '')

        usage_type = request.POST.get('usage_type', '')
        trip_type = request.POST.get('trip_type', '')
        seat_type = request.POST.get('seat_type', '')
        from_city = request.POST.get('from_city', '')
        to_city = request.POST.get('to_city', '')
        booking_datetime = request.POST.get('booking_datetime', '')
        departure_date = request.POST.get('departure_date', '')
        departure_date = datetime.strptime(departure_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        preferred_flight = request.POST.get('preferred_flight', '')
        assessment_code = request.POST.get('assessment_code', '')
        assessment_city_id = request.POST.get('assessment_city_id', '')
        entity_id = request.POST.get('entity_id', '')
        reason_booking = request.POST.get('reason_booking', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        no_of_emp = int(no_of_seats) + 1
        employee_name_1 = request.POST.get('employee_name_1', '')
        flight_class_is_international = request.POST.get('flight_class_is_international', '')
        emp_info_international = []
        emp_data = {}
        if flight_class_is_international:
            for i in range(1, no_of_emp):
                emp_data['emp_id'] =  request.POST.get('employee_id_pass_' + str(i), '')
                emp_data['emp_title'] =  request.POST.get('employee_title_' + str(i), '')
                emp_data['emp_fname'] =  request.POST.get('employee_fname_' + str(i), '')
                emp_data['emp_lname'] =  request.POST.get('employee_lname_' + str(i), '')
                emp_data['emp_dob'] =  request.POST.get('employee_dob_' + str(i), '')
                emp_data['emp_passport_no'] =  request.POST.get('employee_passport_' + str(i), '')
                emp_data['emp_passport_exp'] =  request.POST.get('employee_pass_exp_' + str(i), '')
                emp_data['emp_nationality'] =  request.POST.get('employee_nationality_' + str(i), '')
                emp_info_international.append(emp_data)
        else:
            for i in range(1, no_of_emp):
                emp_data['emp_title'] =  request.POST.get('employee_ttl_' + str(i), '')
                emp_data['emp_fname'] =  request.POST.get('employee_ffname_' + str(i), '')
                emp_data['emp_lname'] =  request.POST.get('employee_llname_' + str(i), '')
                emp_data['emp_dob'] =  request.POST.get('employee_edob_' + str(i), '')
                emp_info_international.append(emp_data)

        if entity_id:
            pass
        else:
            entity_id = 0

        employees = []
        employees_name = []

        for i in range(1, no_of_emp):
            employees.append(request.POST.get('employee_id_' + str(i), ''))
            employees_name.append(request.POST.get('employee_name_' + str(i), ''))
            print(employees)

        payload11 = {'flightdata': flightdata, 'employee_name_1': employees_name,'UID2':UID2,'flight_class_is_international':flight_class_is_international,'emp_info_international':str(emp_info_international)}
        print("adsad")
        print(payload11)
        url_save = settings.API_BASE_URL + "save_flight_booking"
        booking1 = getDataFromAPI(login_type, access_token, url_save, payload11)
        print("API BOOK")
        print(booking1)
        if not 'RESULT' in booking1['Data']:
            messages.error(request, 'FLIGHT/FARE NOT AVAILABLE')
            return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/30", {'message': "Operation Successfully"})

        if 'BOOKINGID' in booking1['Data']['RESULT'][0]:
            vendor_booking = booking1['Data']['RESULT'][0]['BOOKINGID']
        else:
            vendor_booking = ""
        payload = {'user_id': user_id, 'user_type': login_type, 'corporate_id': corporate_id, 'spoc_id': spoc_id,
                   'group_id': group_id,
                   'subgroup_id': subgroup_id, 'usage_type': usage_type, 'trip_type': trip_type, 'seat_type': seat_type,
                   'from_city': from_city, 'to_city': to_city,
                   'booking_datetime': booking_datetime, 'departure_datetime': departure_date,
                   'preferred_flight': preferred_flight, 'assessment_code': assessment_code,
                   'reason_booking': reason_booking, 'no_of_seats': no_of_seats, 'employees': employees,
                   'billing_entity_id': entity_id,
                   'is_sms': 1, 'is_email': 1, 'assessment_city_id': assessment_city_id, 'flightdata': flightdata,'UID2':UID2,
                   'employee_name_1': employee_name_1,'vendor_booking':vendor_booking}
        #print(payload)
        if vendor_booking:
            url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            print("MYBOOOK")
            #print(booking)
            if booking['success'] == 1:
                last_booking_id = booking['last_booking_id']

                for i in range(3):
                    url_save = settings.API_BASE_URL + "get_flight_pnr_details"
                    pnr_no = {'pnr': vendor_booking}
                    booking1 = getDataFromAPI(login_type, access_token, url_save, pnr_no)
                    print(booking1)
                    if 'ERROR' in booking1['Data']:
                        sleep(5)
                    else:
                        if UID2:
                            if booking1['Data']['PAXOW'][0]['apnr']:
                                print("GENP PNPRPPRPRPRPPRPRR ......................")
                                print(booking1['Data']['PAXOW'][0]['apnr'])
                                print("GENP PNPRPPRPRPRPPRPRR ......................")
                                pass
                            else:
                                sleep(5)
                        else:
                            print("GENP PNPRPPRPRPRPPRPRR ......................")
                            print(booking1['Data']['PAX'][0]['apnr'])
                            print("GENP PNPRPPRPRPRPPRPRR ......................")
                            if booking1['Data']['PAX'][0]['apnr']:
                                pass
                            else:
                                sleep(5)

                if not 'FLIGHT' in booking1['Data'] or 'FLIGHTOW' in booking1['Data'] :
                    messages.error(request, 'Booking successful, your CoTrav booking id is - ' +str(last_booking_id)+ ', but pending for PNR status, please check the status under Pending for PNR tab')
                    return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/30", {'message': "Operation Successfully"})
                else:
                    if UID2:
                        if not 'apnr' in booking1['Data']['PAXOW'][0] or len(booking1['Data']['PAXOW'][0]['apnr']) == 0:
                            messages.error(request, 'Booking successful, your CoTrav booking id is - ' + str(last_booking_id) + ', but pending for PNR status, please check the status under Pending for PNR tab')
                            return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/30",{'message': "Operation Successfully"})
                        else:
                            pass
                    else:
                        if not 'apnr' in booking1['Data']['PAX'][0] or len(booking1['Data']['PAX'][0]['apnr']) == 0:
                            messages.error(request, 'Booking successful, your CoTrav booking id is - ' +str(last_booking_id)+ ', but pending for PNR status, please check the status under Pending for PNR tab')
                            return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/2",{'message': "Operation Successfully"})
                        else:
                            pass

                    ticket_number =[]
                    pnr_no =[]
                    flight_no =[]
                    flight_name =[]
                    arrival_time =[]
                    departure_time =[]
                    flight_to =[]
                    flight_from =[]
                    is_return_flight =[]

                    if UID2:
                        no_of_stops = booking1['Data']['FLIGHTOW'][0]['STOP']
                        flight_type = seat_type
                        fare_type = booking1['Data']['FLIGHTOW'][0]['FARE_TYPE']
                        meal_is_include = ''
                        no_of_passanger = booking1['Data']['PARAM'][0]['adt']
                        employee_booking_id = employees
                        ticket_price = booking1['Data']['FLIGHTOW'][0]['AMOUNT']

                        if booking1['Data']['CON_FLIGHTOW']:
                            for flightt in booking1['Data']['CON_FLIGHTOW']:
                                ticket_number.append(booking1['Data']['FLIGHTOW'][0]['PCC'])
                                pnr_no.append(booking1['Data']['PAXOW'][0]['apnr'])
                                flight_no.append(flightt['FLIGHT_NO'])
                                flight_name.append(flightt['FLIGHT_NAME'])
                                arrival_time1 = datetime.strptime(flightt['ARRV_DATE'] + " " + flightt['ARRV_TIME']+":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                                arrival_time.append(arrival_time1)
                                arrival_time2 = datetime.strptime(flightt['DEP_DATE'] + " " + flightt['DEP_TIME']+":00",  "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                                departure_time.append(arrival_time2)
                                flight_to.append(flightt['DES_NAME'])
                                flight_from.append(flightt['ORG_NAME'])
                                is_return_flight.append('0')
                        else:
                            ticket_number.append(booking1['Data']['FLIGHTOW'][0]['PCC'])
                            pnr_no.append(booking1['Data']['PAXOW'][0]['apnr'])
                            flight_no.append(booking1['Data']['FLIGHTOW'][0]['FLIGHT_NO'])
                            flight_name.append(booking1['Data']['FLIGHTOW'][0]['FLIGHT_NAME'])
                            arrival_time1 = datetime.strptime(booking1['Data']['FLIGHTOW'][0]['ARRV_DATE'] + " " + booking1['Data']['FLIGHTOW'][0]['ARRV_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                            arrival_time.append(arrival_time1)
                            arrival_time2 = datetime.strptime(booking1['Data']['FLIGHTOW'][0]['DEP_DATE'] + " " + booking1['Data']['FLIGHTOW'][0]['DEP_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                            departure_time.append(arrival_time2)
                            flight_to.append(booking1['Data']['FLIGHTOW'][0]['DES_NAME'])
                            flight_from.append(booking1['Data']['FLIGHTOW'][0]['ORG_NAME'])
                            is_return_flight.append('0')


                        if booking1['Data']['CON_FLIGHTRT']:
                            for flightt in booking1['Data']['CON_FLIGHTRT']:
                                ticket_number.append(booking1['Data']['FLIGHTRT'][0]['PCC'])
                                pnr_no.append(booking1['Data']['PAXRT'][0]['apnr'])
                                flight_no.append(flightt['FLIGHT_NO'])
                                flight_name.append(flightt['FLIGHT_NAME'])
                                arrival_time1 = datetime.strptime(flightt['ARRV_DATE'] + " " + flightt['ARRV_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                                arrival_time.append(arrival_time1)
                                arrival_time2 = datetime.strptime( flightt['DEP_DATE'] + " " + flightt['DEP_TIME'] + ":00","%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                                departure_time.append(arrival_time2)
                                flight_to.append(flightt['DES_NAME'])
                                flight_from.append(flightt['ORG_NAME'])
                                is_return_flight.append('1')
                        else:
                            ticket_number.append(booking1['Data']['FLIGHTRT'][0]['PCC'])
                            pnr_no.append(booking1['Data']['PAXRT'][0]['apnr'])
                            flight_no.append(booking1['Data']['FLIGHTRT'][0]['FLIGHT_NO'])
                            flight_name.append(booking1['Data']['FLIGHTRT'][0]['FLIGHT_NAME'])
                            arrival_time1 = datetime.strptime(booking1['Data']['FLIGHTRT'][0]['ARRV_DATE'] + " " + booking1['Data']['FLIGHTRT'][0]['ARRV_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                            arrival_time.append(arrival_time1)
                            arrival_time2 = datetime.strptime(booking1['Data']['FLIGHTRT'][0]['DEP_DATE'] + " " + booking1['Data']['FLIGHTRT'][0]['DEP_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                            departure_time.append(arrival_time2)
                            flight_to.append(booking1['Data']['FLIGHTRT'][0]['DES_NAME'])
                            flight_from.append(booking1['Data']['FLIGHTRT'][0]['ORG_NAME'])
                            is_return_flight.append('1')

                            print("INNNNNNNNNNNN ELSEEEEEEEEEEEEEEEE")


                    else:
                        no_of_stops = booking1['Data']['FLIGHT'][0]['STOP']
                        flight_type = seat_type
                        fare_type = booking1['Data']['FLIGHT'][0]['FARE_TYPE']
                        meal_is_include = ''
                        no_of_passanger = booking1['Data']['FLIGHT'][0]['SEAT']
                        employee_booking_id = employees
                        ticket_price = booking1['Data']['FLIGHT'][0]['AMOUNT']

                        if booking1['Data']['CON_FLIGHT']:
                            for flightt in booking1['Data']['CON_FLIGHT']:
                                ticket_number.append(booking1['Data']['FLIGHT'][0]['PCC'])
                                pnr_no.append(booking1['Data']['PAX'][0]['apnr'])
                                flight_no.append(flightt['FLIGHT_NO'])
                                flight_name.append(flightt['FLIGHT_NAME'])
                                arrival_time1 = datetime.strptime(flightt['ARRV_DATE'] + " " + flightt['ARRV_TIME']+":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                                arrival_time.append(arrival_time1)
                                arrival_time2 = datetime.strptime(flightt['DEP_DATE'] + " " + flightt['DEP_TIME']+":00",  "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                                departure_time.append(arrival_time2)
                                flight_to.append(flightt['DES_NAME'])
                                flight_from.append(flightt['ORG_NAME'])
                                is_return_flight.append('0')
                            print("INNNNNNNNNNNN IFFFFFFFFFFFFF")
                        else:
                            ticket_number.append(booking1['Data']['FLIGHT'][0]['PCC'])
                            pnr_no.append(booking1['Data']['PAX'][0]['apnr'])
                            flight_no.append(booking1['Data']['FLIGHT'][0]['FLIGHT_NO'])
                            flight_name.append(booking1['Data']['FLIGHT'][0]['FLIGHT_NAME'])
                            arrival_time1 = datetime.strptime(booking1['Data']['FLIGHT'][0]['ARRV_DATE'] + " " + booking1['Data']['FLIGHT'][0]['ARRV_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                            arrival_time.append(arrival_time1)
                            arrival_time2 = datetime.strptime(booking1['Data']['FLIGHT'][0]['DEP_DATE'] + " " + booking1['Data']['FLIGHT'][0]['DEP_TIME'] + ":00", "%Y-%m-%d %H:%M:%S").strftime("%d-%m-%Y %H:%M:%S")
                            departure_time.append(arrival_time2)
                            flight_to.append(booking1['Data']['FLIGHT'][0]['DES_NAME'])
                            flight_from.append(booking1['Data']['FLIGHT'][0]['ORG_NAME'])
                            is_return_flight.append('0')
                            print("INNNNNNNNNNNN ELSEEEEEEEEEEEEEEEE")

                    sub_total = 118+int(ticket_price)

                    url_assign = settings.API_BASE_URL + "assign_flight_booking"
                    payload11 = {'ticket_no': ticket_number, 'pnr_no': pnr_no, 'portal_used': "",
                               'booking_id': last_booking_id, 'user_id': user_id, 'user_type': login_type, 'flight_no': flight_no,
                               'flight_name': flight_name, 'arrival_time': arrival_time,
                               'departure_time': departure_time, 'flight_to': flight_to, 'flight_from': flight_from,
                               'no_of_stops': no_of_stops, 'seat_type': seat_type, 'flight_type': flight_type,
                               'trip_type': trip_type, 'fare_type': fare_type, 'meal_is_include': meal_is_include,
                               'no_of_passanger': no_of_passanger, 'employee_booking_id': employee_booking_id,
                               'ticket_price': ticket_price, 'management_fee': '100',
                               'tax_mng_amt': '18', 'tax_on_management_fee': '18',
                               'tax_on_management_fee_percentage': '18',
                               'sub_total': sub_total,
                               'management_fee_igst': 18, 'management_fee_cgst': 0,
                               'management_fee_sgst': 0,
                               'management_fee_igst_rate': int(ticket_price)*0.18,
                               'management_fee_cgst_rate': 0,
                               'management_fee_sgst_rate': 0, 'cgst': 0, 'sgst': 0,
                               'igst': int(sub_total)*0.18,
                               'oper_ticket_price': ticket_price, 'oper_commission': "",
                               'oper_commission_type': "",
                               'oper_cotrav_billing_entity': "1",
                               'cotrav_billing_entity': '1',
                               'oper_cgst': 0, 'oper_sgst': 0, 'oper_igst': 18,
                               'client_ticket_path': '', 'client_ticket': '1',
                               'vender_ticket': '',
                               'vender_ticket_path': '', 'is_client_sms': '1',
                               'is_client_email': '1',
                               'igst_amount': int(ticket_price)*0.18, 'cgst_amount': 0, 'sgst_amount': 0,
                               'operator_id': '1','vendor_booking_id':vendor_booking,'is_return_flight':is_return_flight,
                               }
                    print("payrol  .....")
                    print(payload11)
                    company11 = getDataFromAPI(login_type, access_token, url_assign, payload11)
                    print(company11)
                    messages.success(request, 'Flight Booking Added Successfully..!')
                    return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Flight Booking Not Added..!')
                return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/2", {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Flight Booking Not Added..!')
            return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/2", {'message': "Operation Successfully"})


    else:
        return HttpResponseRedirect("/login")



####################### Download MIS ##################################

def dateonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError:
        return ''


def download_taxi_bookings(request):
    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

        # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [

        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',

        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'Passengers',

        'Booking Date',
        'Booking Time',

        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        'Approver1 Action',

        'Approver1 Name',

        'Approver1 Date',

        'Approver1 Time',

        'Approver2 Action',

        'Approver2 Name',

        'Approver2 Date',

        'Approver2 Time',

        'Approved Date',

        'Approved Time',

        'Approved By',

        'Canceled Date',

        'Canceled Time',

        'Canceled By',

        'Assigned Date',

        'Assigned Time',

        'Assigned By',

        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',

        'Package Name',
        'Tour Type',
        'Vehicle Type',

        'Driver Name',
        'Driver Contact	',
        'Taxi Reg No.',
        'No. Of Seats',

        'Current Booking Status',

        'Hours Done',

        'Allowed Hours',

        'Extra Hours',

        'Kms Done',

        'Allowed Kms',

        'Extra Kms',

        'Extra Hours Charges',

        'Base Price',

        'Management Fee',
        'Tax on management fee',
        'Sub Total',
        'Cotrav Billing Entity',
        'IGST',
        'CGST',
        'SGST',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',

        'Estimated Amount',
        'Is Auto Approved',
        'Bill ID',
        'Bill Date',
        'Billing Entity',

        'Is Auto Approved',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        spoc_status =''
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by= ''
        spoc_canceled_date =''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

            if bk['spoc_status'] == 1:
                spoc_status = "In-Active"
            else:
                spoc_status = "Active"

            if (act['action'] == 3 and act['user_type'] == 6):
                print('canceled')
                spoc_canceled_by = act['employee_name']
                spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [
            bk['reference_no'],
            bk['assessment_city_id'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],

            bk['zone_name'],
            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),
            dateonly(bk['pickup_datetime']),
            timeonly(bk['pickup_datetime']),

            bk['package_name'],
            bk['tour_type'],
            bk['taxi_type_name'],

            bk['driver_name'],
            bk['driver_contact'],
            'Taxi Reg No.',
            bk['no_of_seats'],

            bk['cotrav_status'],

            bk['ci_hours_done'],

            bk['ci_allowed_hours'],

            bk['ci_extra_hours'],

            bk['ci_kms_done'],

            bk['ci_allowed_kms'],

            bk['ci_extra_kms'],

            bk['ci_extra_hr_charges'],

            bk['base_rate'],

            bk['ci_management_fee'],
            bk['ci_tax_on_management_fee'],
            bk['ci_sub_total'],
            bk['ci_cotrav_billing_entity'],
            bk['ci_igst'],
            bk['ci_cgst'],
            bk['ci_sgst'],
            bk['ci_management_fee_igst'],
            bk['ci_management_fee_cgst'],
            bk['ci_management_fee_sgst'],
            bk['ci_management_fee_igst_rate'],
            bk['ci_management_fee_cgst_rate'],
            bk['ci_management_fee_sgst_rate'],

            '',
            '',
            '',
            '',
            '',
            '',

            bk['client_status'],

            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Bus Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Reason For Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",
        'Approver1 Action',
        'Approver1 Name',
        'Approver1 Date',
        'Approver1 Time',
        'Approver2 Action',
        'Approver2 Name',
        'Approver2 Date',
        'Approver2 Time',
        'Approved Date',
        'Approved Time',
        'Approved By',
        'Canceled Date',
        'Canceled Time',
        'Canceled By',
        'Assigned Date',
        'Assigned Time',
        'Assigned By',
        'Passanger Name',
        'Pickup City',
        'Drop City',
        'Journey Date',
        'Journey Time',
        'Current Booking Status',
        'Bus Type Allocated',
        'PNR Number',
        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',
        'Is Auto Approved',
        'Bill ID',
        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by=''
        spoc_canceled_date=''
        spoc_status=''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']


                # Define the data for each cell in the row
        row = [

            bk['reference_no'],

            bk['assessment_code'],
            bk['assessment_city_id'],

            bk['reason_booking'],

            bk['zone_name'],

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            dateonly(bk['booking_datetime']),

            timeonly(bk['booking_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            passanger_list,

            bk['pickup_location'],

            bk['drop_location'],

            dateonly(bk['pickup_from_datetime']),

            timeonly(bk['pickup_from_datetime']),

            bk['cotrav_status'],

            bk['assign_bus_type_id'],

            bk['pnr_no'],

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            '',

            '',

            bk['client_status'],
            bk['cotrav_status'],

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "Pickup City",
        "Drop City",
        "Booking Date",
        "Booking Time",
        "Journey Date",
        "Journey Time",
        "SPOC Status",
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        "Zone",
        "Coach Type Allocated",
        "Quota Used",

        'No of seats',
        'Operator name',
        'Operator contact',
        'Train name',
        'Ticket no',
        'pnr no',
        'Assign bus type id',
        'Seat no',
        'Portal used',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status=''
        spoc_canceled_by=''
        spoc_canceled_date=''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            bk['zone_name'],
            bk['train_type_priority_1'],
            bk['seat_no'],


            bk['no_of_seats'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['train_name'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['seat_no'],
            bk['portal_used'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "From City",
        "To City",
        "Booking Date",
        "Booking Time",
        "Departure Date",
        "Departure Time",

        "Return Date",
        "Booking Status",

        "SPOC Status",
        "SPOC Cancel By",
        "SPOC Cancel Date",

        "Usage Type",
        "Trip Type",
        "Flight Type",
        "Seat Type",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        'First Flight Name',
        'First Flight No',
        'First Flight PNR No.',
        'First Flight From',
        'First Flight To',
        'First Flight Departure Date',
        'First Flight Departure time',
        'First Flight Arrival Date',
        'First Flight Arrival time',

        'Second Flight Name',
        'Second Flight No',
        'Second Flight PNR No.',
        'Second Flight From',
        'Second Flight To',
        'Second Flight Departure Date',
        'Second Flight Departure Time',
        'Second Flight Arrival Datetime',
        'Second Flight Arrival Time',

        'Third Flight Name',
        'Third Flight No',
        'Third Flight PNR No.',
        'Third Flight From',
        'Third Flight To',
        'Third Flight Departure Date',
        'Third Flight Departure Time',
        'Third Flight Arrival Date',
        'Third Flight Arrival Time',

        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        spoc_status= ''
        spoc_canceled_by=''
        spoc_canceled_date=''


        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            "Booking Remarks",
            bk['from_location'],
            bk['to_location'],

            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),

            "",
            bk['cotrav_status'],

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            flight_name1,
            flight_no1,
            pnr_no1,
            from_city1,
            to_city1,
            dateonly(departure_datetime1),
            timeonly(departure_datetime1),
            dateonly(arrival_datetime1),
            timeonly(arrival_datetime1),

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            dateonly(departure_datetime2),
            timeonly(departure_datetime2),
            dateonly(arrival_datetime2),
            timeonly(arrival_datetime2),

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            dateonly(departure_datetime3),
            timeonly(departure_datetime3),
            dateonly(arrival_datetime3),
            timeonly(arrival_datetime3),

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']



    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Bucket Request Date',
        'Bucket Request Time',
        'Bucket Approved Date',
        'Bucket Approved Time',
        'Preferred Hotel',
        'Assigned Hotel',
        'Assigned Hotel Address',
        'Hotel Contact',
        'Assign Date',
        'Assign Time',
        'TaxiVaxi Status',
        'Employees Name',
        'No. of Persons',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Higher Bucket Requested',
        'Reason for Higher Bucket',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',
        'Current Booking Status',
        'No. of Nights',
        'Room Type',
        'Room Occupancy',
        'Per Night Price',
        'Total Room Price',
        'Tax On Room Cancellation',
        'Ticket Price',
        'Management Fee',
        'Tax on management fee',
        'Sub total',
        'Cotrav billing entity',
        'igst',
        'cgst',
        'sgst',
        'Management fee igst',
        'Management fee cgst',
        'Management fee sgst',
        'Management fee igst rate',
        'Management fee cgst rate',
        'Management fee sgst rate',
        'Is Auto Approved',
        'Bill ID',
        'Is TBA Booking',
        'Is Offline Booking',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            'Billing Entity',
            'Travel request Code',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            '',
            '',
            '',
            '',

            bk['preferred_hotel'],
            bk['assign_hotel_id'],
            bk['operator_name'],
            bk['operator_contact'],
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['status_cotrav'],
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),
            bk['status_cotrav'],
            '',
            bk['room_type_name'],
            bk['hotel_type_name'],
            bk['bucket_price_1'],
            bk['bucket_price_1'],
            '',
            bk['ticket_price'],

            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['sub_total'],
            bk['billing_entity_id'],
            bk['igst'],
            bk['cgst'],
            bk['sgst'],
            bk['management_fee_igst'],
            bk['management_fee_cgst'],
            bk['management_fee_sgst'],
            bk['management_fee_igst_rate'],
            bk['management_fee_cgst_rate'],
            bk['management_fee_sgst_rate'],

            '',
            '',
            '',
            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],


        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_employees(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Employees']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyEmployees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Employees'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Spoc Name",
        "Core Employee ID",
        "Employee Company ID",
        "Employee Name",
        "Employee Email",
        "Employee Phone No",
        "Login UserName",
        "Age",
        "Gender",
        "ID Proof Type",
        "ID No",
        "Is CXO",
        "Designation",
        "Home City",
        "Home Address",
        "Assistant ID",
        "Date Of Birth",
        "Billing Entity",
        "Last Login",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_cxo'] == 0:
            is_cxo = "No"
        else:
            is_cxo = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['user_name'],
            bk['core_employee_id'],
            bk['core_employee_id'],
            bk['employee_cid'],
            bk['employee_name'],
            bk['employee_email'],
            bk['employee_contact'],
            bk['username'],
            bk['age'],
            bk['gender'],
            bk['id_proof_type'],
            bk['id_proof_no'],
            is_cxo,
            bk['designation'],
            bk['home_city'],
            bk['home_address'],
            bk['assistant_id'],
            bk['date_of_birth'],
            bk['billing_entity_id'],
            bk['last_login'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_billing_entities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Entitys']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=BillingEntitys.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Billing Entitys'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Billing City",
        "Entity Name",
        "Contact Person Name",
        "Contact Person Email",
        "Contact Person Phone No",
        "Address Line 1",
        "Address Line 2",
        "Address Line 3",
        "GST NO",
        "PAN No",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['billing_city'],
            bk['entity_name'],
            bk['contact_person_name'],
            bk['contact_person_email'],
            bk['contact_person_no'],
            bk['address_line_1'],
            bk['address_line_2'],
            bk['address_line_3'],
            bk['gst_id'],
            bk['pan_no'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response

def download_spocs(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Spocs']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySpocs.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Spocs'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgroup Name",
        "Spoc Company ID",
        "Spoc Name",
        "Spoc Email",
        "Spoc Contact No ",
        "Login UserName ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['user_cid'],
            bk['user_name'],
            bk['email'],
            bk['user_contact'],
            bk['username'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def taxi_billing(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        url = settings.API_BASE_URL + "spoc_taxi_bookings"
        payload = {'booking_type': id, 'spoc_id':request.user.id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/spoc_taxi_billing.html",{'bookings': booking, 'billing_type': id })
        else:
            return render(request, "Company/Spoc/spoc_taxi_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def bus_billing(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        url = settings.API_BASE_URL + "spoc_bus_bookings"
        payload = {'booking_type': id, 'spoc_id':request.user.id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/spoc_bus_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Spoc/spoc_bus_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def train_billing(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        url = settings.API_BASE_URL + "spoc_train_bookings"
        payload = {'booking_type': id, 'spoc_id':request.user.id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/spoc_train_billing.html", {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Company/Spoc/spoc_train_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def flight_billing(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        url = settings.API_BASE_URL + "spoc_flight_bookings"
        payload = {'booking_type': id, 'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/spoc_flight_billing.html",{'bookings': booking,'billing_type': id})
        else:
            return render(request, "Company/Spoc/spoc_flight_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def hotel_billing(request,id):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        url = settings.API_BASE_URL + "spoc_hotel_bookings"
        payload = {'booking_type': id, 'spoc_id':request.user.id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Spoc/spoc_hotel_billing.html",{'bookings': booking,'billing_type': id})
        else:
            return render(request, "Company/Spoc/spoc_hotel_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/login")


def taxi_billing_verify(request):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            #print(payload)
            vry_url = settings.API_BASE_URL + "spoc_verify_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_revise_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def bus_billing_verify(request):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_verify_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_revise_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def train_billing_verify(request):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_verify_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_revise_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def hotel_billing_verify(request):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_verify_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_revise_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def flight_billing_verify(request):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id':verify_id, 'user_id':request.user.id, 'corporate_id':corporate_id, 'invoice_id':invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_verify_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id,'invoice_comments':invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "spoc_revise_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                return HttpResponseRedirect(current_url, {})
            else:
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/login")


def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]

def dateonly(dt=''):
    try:
        if(dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''

def timeonly(dt=''):
    try:
        if(dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError :
        return ''


def razor_charge(request):
    if 'spoc_login_type' in request.session:
        login_type = request.session['spoc_login_type']
        access_token = request.session['spoc_access_token']

        pnr = request.POST.get('pnr', '')
        payload ={'pnr':pnr}

        url_save = settings.API_BASE_URL + "get_flight_pnr_details"
        booking1 = getDataFromAPI(login_type, access_token, url_save, payload)
        print(booking1['Data']['FLIGHT'])
        messages.success(request, 'Payment Successfully..!'+str(booking1['Data']['FLIGHT']))
        return HttpResponseRedirect("/Corporate/Spoc/flight-bookings/2", {'message': "Operation Successfully"})

