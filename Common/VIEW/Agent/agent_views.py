import os

from django.conf import settings
from django.core.files.storage import default_storage
from django.shortcuts import render, redirect
import requests
import json
import string
import random
from datetime import date, datetime
from django.views.decorators.csrf import csrf_exempt
from django_global_request.middleware import get_request
# Create your views here.
from django.contrib.auth import authenticate, login as auth_login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.hashers import make_password
from Common.models import Corporate_Agent_Login_Access_Token
from django.http import HttpResponseRedirect, FileResponse
from django.contrib import messages
from landing.cotrav_messeging import Excelexport, Render
from django.core.mail import EmailMultiAlternatives, get_connection
from openpyxl import Workbook
from django.http import HttpResponse
from django.utils.encoding import smart_str

def upload_file_getpath(request):
    if request.method == 'POST':
        file_up = request.FILES['email_attachment']
        print("befour Upload")
        get_path = file_upload_get_path(file_up)
        print("After Upload")
        print(get_path)
        return render(request, 'Agent/upload.html', {})
    else:
        return render(request, 'Agent/upload.html', {})


def generate_pdf_file(request):
    render = Render()
    name = {'abc':'abc','xyz':'xyz'}
    path = render.render_to_file("Agent/xhtmltopdftest.html",name)
    print(path)
    connection = get_connection()  # uses SMTP server specified in settings.py
    connection.open()  # If you don't open the connection manually, Django will automatically open, then tear down the connection in msg.send()
    mail_to = "balwant@taxivaxi.in"
    email_body = "Test"
    email_subject = "Test"
    html_content = path['1']
    msg = EmailMultiAlternatives(email_subject, email_body, 'balwant@taxivaxi.in', [mail_to])
    msg.attach_alternative(html_content, "text/html")
    msg.attach_file(html_content)
    res = msg.send(fail_silently=True)
    print("email send")
    connection.close()  # Cleanup
    return 1


def agent_homepage(request):
    if 'agent_login_type' in request.session:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {}

        rm = request.user.is_relationship_manager
        om = request.user.is_operation_manager
        sa = request.user.is_super_admin

        if (sa):
            agent_user_type = 101
        elif (rm):
            agent_user_type = 102
        elif (om):
            agent_user_type = 103
        else:
            agent_user_type = 101

        print("agents User type")
        print(request.user)
        print(user_type)
        print(agent_user_type)

        url = settings.API_BASE_URL + "agent_dashboard"
        data = getDataFromAPI(user_type, access_token, url, payload)
        dataDashboard = data['Dashboard']

        url = settings.API_BASE_URL + "spocs"
        data = getDataFromAPI(user_type, access_token, url, payload)
        dataSpocs = data['Spocs']

        return render(request, 'Agent/agent_home_page.html',
                      {'user': request.user, 'dataDashboard': dataDashboard, 'agent_user_type': agent_user_type, 'dataSpocs':dataSpocs})
    else:
        return HttpResponseRedirect("/agents/login")


def cotrav_communication(request):
    if 'agent_login_type' in request.session:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            msg_type = request.POST.get('msg_type', '')
            if msg_type == 'notification':
                msg_head = request.POST.get('msg_head', '')
                msg_title = request.POST.get('msg_title', '')
                msg_text = request.POST.get('msg_text', '')
                url = settings.API_BASE_URL + "send_broadcast_notification"
                payload = {'msg_head': msg_head, 'msg_title':msg_title, 'msg_text':msg_text}
                taxi = getDataFromAPI(user_type, access_token, url, payload)
                if taxi['success'] == 1:
                    messages.success(request, taxi['message'])
                    return render(request, "Agent/cotrav_communication.html", {'user': request.user})
                else:
                    return render(request, "Agent/cotrav_communication.html", {'user': request.user})
            elif msg_type == 'message':
                mobile_nos = request.POST.get('mobile_nos', '')
                msg_text = request.POST.get('msg_text', '')
                url = settings.API_BASE_URL + "send_message_to_moblies"
                payload = {'mobile_nos': mobile_nos, 'msg_text': msg_text}
                taxi = getDataFromAPI(user_type, access_token, url, payload)
                if taxi['success'] == 1:
                    messages.success(request, taxi['message'])
                    return render(request, "Agent/cotrav_communication.html", {'user': request.user})
                else:
                    return render(request, "Agent/cotrav_communication.html", {'user': request.user})
            elif msg_type == 'mail':
                email_subject = request.POST.get('email_subject', '')
                email_body = request.POST.get('email_body', '')
                email_to = request.POST.get('email_to', '')
                url = settings.API_BASE_URL + "send_mail_to_user"
                payload = {'email_subject': email_subject, 'email_body': email_body, 'email_to': email_to}
                taxi = getDataFromAPI(user_type, access_token, url, payload)
                print(taxi)
                if taxi['success'] == 1:
                    messages.success(request, taxi['message'])
                    return render(request, "Agent/cotrav_communication.html", {'user': request.user})
                else:
                    messages.error(request, taxi['message'])
                    return render(request, "Agent/cotrav_communication.html", {'user': request.user})

        else:
            return render(request, 'Agent/cotrav_communication.html', {'user': request.user})
    else:
        return HttpResponseRedirect("/agents/login")

def user_profile(request):
    return render(request, 'Agent/user_profile.html', {'user': request.user})


def agent_login_action(request):
    context = {}
    if request.method == 'POST':
        username = request.POST.get('email', '')
        password = request.POST.get('password', '')
        user = authenticate(username=username, post_password=password, login_type="10")
        print(user)
        if user is not None:
            if user:
                request.session.set_expiry(7200)  # sets the exp. value of the session
                user_type_login = request.session['agent_login_type']
                access_token_login = request.session['agent_access_token']
                auth_login(request, user, backend='Common.backends.CustomCompanyUserAuth')  # the user is now logged in
                request.session['agent_access_token'] = access_token_login
                request.session['agent_login_type'] = user_type_login
                request.session['login_type'] = user_type_login
                request.session.set_expiry(7200)  # sets the exp. value of the session
                messages.success(request, 'Login Successful..!')
                return HttpResponseRedirect("/agents/agent_home")
        else:
            messages.error(request, 'Invalid Email Or Password..!')
            return render(request,'Agent/corporate_agent_login.html',context)
    else:
        return render(request,'Agent/corporate_agent_login.html',{})


def agent_logout_action(request):
    if 'agent_login_type' in request.session:
        request = get_request()
        access_token = request.session['agent_access_token']
        user = Corporate_Agent_Login_Access_Token.objects.get(access_token=access_token)
        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        del request.session['agent_login_type']
        del request.session['agent_access_token']
        return HttpResponseRedirect("/agents/login")
    else:
        return HttpResponseRedirect("/agents/login")


def taxi_types(request):
    if request.method == "POST":
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            type_name = request.POST.get('type_name', '')

            taxitype_id = request.POST.get('taxitype_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')
            operation_message = ""
            url = ""
            payload = {}
            if taxitype_id:
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_taxi_type"
                    payload = {'taxitype_id':taxitype_id,'user_id':user_id,'user_type':user_type}
                    operation_message = "Taxi Type Deleted Successfully..!"
                else:
                    url = settings.API_BASE_URL + "update_taxi_type"
                    payload = {'type_name':type_name,'taxitype_id':taxitype_id,'user_id':user_id,'user_type':user_type}
                    operation_message = "Taxi Type Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_taxi_type"
                payload = {'type_name':type_name,'user_id':user_id,'user_type':user_type}
                operation_message = "Taxi Type Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/taxi-types", {'message': "Operation Successfully"})
            else:
                messages.error(request, taxi['message'])
                return HttpResponseRedirect("/agents/taxi-types", {'message': "Operation Failed"})

        else:
            return HttpResponseRedirect("/agents/login")

    else:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "taxi_types"
        payload = {'': ''}
        taxi = getDataFromAPI(user_type, access_token, url, payload)

        if taxi['success'] == 1:
            taxi_data = taxi['taxi_types']
            return render(request, "Agent/taxitypes.html", {'types': taxi_data})
        else:
            return render(request, "Agent/taxitypes.html", {'Types': {}})


def taxi_models(request):
    if request.method == "POST":
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            brand_name = request.POST.get('brand_name', '')
            model_name = request.POST.get('model_name', '')
            taxitype_id = request.POST.get('taxitype_id', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')
            model_id = request.POST.get('model_id', '')

            url = ""
            payload = {}
            if model_id:
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_taxi_model"
                    payload = {'model_id':model_id,'user_id':user_id,'user_type':user_type}
                    operation_message = "Taxi Model Deleted Successfully..!"
                else:
                    url = settings.API_BASE_URL + "update_taxi_model"
                    payload = {'brand_name':brand_name,'model_name':model_name,'taxitype_id':taxitype_id,'no_of_seats':no_of_seats,'model_id':model_id,'user_id':user_id,'user_type':user_type}
                    operation_message = "Taxi Model Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_taxi_model"
                payload = {'brand_name':brand_name,'model_name':model_name,'taxitype_id':taxitype_id,'no_of_seats':no_of_seats,'user_id':user_id,'user_type':user_type}
                operation_message = "Taxi Model Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/taxi-models", {'message': "Operation Successfully"})
            else:
                messages.error(request, taxi['message'])
                return HttpResponseRedirect("/agents/taxi-models", {'message': "Operation Failed"})

        else:
            return HttpResponseRedirect("/agents/login")

    else:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "taxi_types"
        payload = {'': ''}
        taxi = getDataFromAPI(user_type, access_token, url, payload)

        url_m = settings.API_BASE_URL + "taxi_models"
        payload = {'': ''}
        taxi_model = getDataFromAPI(user_type, access_token, url_m, payload)

        if taxi['success'] == 1:
            taxi_data = taxi['taxi_types']
            taxi_model = taxi_model['Models']
            return render(request, "Agent/taxi_models.html", {'types': taxi_data, 'taxi_models': taxi_model})
        else:
            return render(request, "Agent/taxi_models.html", {'Types': {}})


def taxis(request):
    if request.method == "POST":
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            model_id = request.POST.get('model_id', '')
            taxi_reg_no = request.POST.get('taxi_reg_no', '')
            make_year = request.POST.get('make_year', '')
            garage_location = request.POST.get('garage_location', '')
            garage_distance = request.POST.get('garage_distance', '')

            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')
            taxi_id = request.POST.get('taxi_id', '')

            print(taxi_id)
            print(delete_id)

            url = ""
            payload = {}
            if taxi_id:
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_taxi"
                    payload = {'taxi_id':taxi_id,'user_id':user_id,'user_type':user_type}
                    operation_message ="Taxi Deleted Successfully..!"
                else:
                    url = settings.API_BASE_URL + "update_taxi"
                    payload = {'model_id':model_id,'taxi_reg_no':taxi_reg_no,'make_year':make_year,'garage_location':garage_location,'garage_distance':garage_distance,'taxi_id':taxi_id,'user_id':user_id,'user_type':user_type}
                    operation_message="Taxi Updated Successfull..!"
            else:
                url = settings.API_BASE_URL + "add_taxi"
                payload = {'model_id':model_id,'taxi_reg_no':taxi_reg_no,'make_year':make_year,'garage_location':garage_location,'garage_distance':garage_distance,'user_id':user_id,'user_type':user_type}
                operation_message="Taxi Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/taxis", {'message': "Operation Successfully"})
            else:
                messages.error(request, taxi['message'])
                return HttpResponseRedirect("/agents/taxis", {'message': "Operation Failed"})

        else:
            return HttpResponseRedirect("/agents/login")

    else:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "taxis"
        payload = {'': ''}
        taxi = getDataFromAPI(user_type, access_token, url, payload)

        url_m = settings.API_BASE_URL + "taxi_models"
        payload = {'': ''}
        taxi_model = getDataFromAPI(user_type, access_token, url_m, payload)

        if taxi['success'] == 1:
            taxi = taxi['Taxis']
            taxi_model = taxi_model['Models']
            return render(request, "Agent/taxis.html", {'taxis': taxi, 'taxi_models': taxi_model})
        else:
            return render(request, "Agent/taxis.html", {'Types': {}})



def add_company(request):
    if request.method == "POST":
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_name = request.POST.get('corporate_name', '')
            corporate_code = request.POST.get('corporate_code', '')
            contact_person_name = request.POST.get('contact_person_name', '')
            contact_person_no = request.POST.get('contact_person_no', '')
            contact_person_email = request.POST.get('contact_person_email', '')
            bill_corporate_name = request.POST.get('bill_corporate_name', '')
            address_line_1 = request.POST.get('address_line_1', '')
            address_line_2 = request.POST.get('address_line_2', '')
            address_line_3 = request.POST.get('address_line_3', '')
            billing_city_id = request.POST.get('billing_city_id', '')
            gst_id = request.POST.get('gst_id', '')

            has_billing_spoc_level = request.POST.get('has_billing_spoc_level', '')
            has_billing_admin_level = request.POST.get('has_billing_admin_level', '')
            has_auth_level = request.POST.get('has_auth_level', '')
            no_of_auth_level = request.POST.get('no_of_auth_level', '')
            has_assessment_codes = request.POST.get('has_assessment_codes', '')
            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_spoc = request.POST.get('is_spoc', '')
            has_self_booking_access = request.POST.get('has_self_booking_access', '')
            will_do_realtime_payment = request.POST.get('will_do_realtime_payment', '')
            tds_on_management_fee_only = request.POST.get('tds_on_management_fee_only', '')

            user_id = request.POST.get('cotrav_agent_id', '')

            corporate_id = request.POST.get('corporate_id')
            delete_id = request.POST.get('delete_id')

            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')

            if corporate_id:
                password = ''
            else:
                password = "taxi123"
                corporate_id = 0

            payload = {'corporate_name':corporate_name,'corporate_code':corporate_code,'contact_person_name':contact_person_name,'contact_person_no':contact_person_no,
                     'contact_person_email': contact_person_email,'bill_corporate_name':bill_corporate_name,'address_line_1':address_line_1,
                      'address_line_2': address_line_2,'address_line_3': address_line_3, 'gst_id': gst_id,'has_billing_spoc_level':has_billing_spoc_level,
                      'has_auth_level': has_auth_level,'no_of_auth_level':no_of_auth_level,'has_assessment_codes':
                      has_assessment_codes,'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,'has_billing_admin_level': has_billing_admin_level,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,'is_spoc':is_spoc,'password':password,'cotrav_agent_id':user_id,
                       'user_type':user_type,'billing_city_id':billing_city_id,'will_do_realtime_payment':will_do_realtime_payment,'has_self_booking_access':has_self_booking_access,
                       'is_send_email':is_send_email,'is_send_sms':is_send_sms,'tds_on_management_fee_only':tds_on_management_fee_only}

            url = settings.API_BASE_URL + "add_company"
            company = getDataFromAPI(user_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Company Added Successfully..!')
                return HttpResponseRedirect("/agents/companies", {'message': "Added Successfully"})
            else:
                messages.error(request, 'Failed to Add Company..!')
                return HttpResponseRedirect("/agents/companies", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")
    else:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {'some': 'data'}

        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        cities = cities["Cities"]
        return render(request, 'Agent/add_company.html', {'cities':cities})



def companies(request):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        url = settings.API_BASE_URL+"companies"
        payload = {'some': 'data'}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            corporates_data = company['Corporates']
            return render(request,"Agent/companies.html",{'companies':corporates_data})
        else:
            return render(request,"Agent/companies.html",{'companies':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def edit_company(request, id):
    if request.method == "POST":
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_name = request.POST.get('corporate_name', '')
            corporate_code = request.POST.get('corporate_code', '')
            contact_person_name = request.POST.get('contact_person_name', '')
            contact_person_no = request.POST.get('contact_person_no', '')
            contact_person_email = request.POST.get('contact_person_email', '')
            has_auth_level = request.POST.get('has_auth_level', '')
            no_of_auth_level = request.POST.get('no_of_auth_level', '')
            has_assessment_codes = request.POST.get('has_assessment_codes', '')
            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            has_self_booking_access = request.POST.get('has_self_booking_access', '')
            will_do_realtime_payment = request.POST.get('will_do_realtime_payment', '')
            has_billing_spoc_level = request.POST.get('has_billing_spoc_level', '')
            has_billing_admin_level = request.POST.get('has_billing_admin_level', '')
            tds_on_management_fee_only = request.POST.get('tds_on_management_fee_only', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')

            user_id = request.POST.get('user_id', '')

            corporate_id = request.POST.get('corporate_id', '')

            url = settings.API_BASE_URL + "update_company"
            payload = {'corporate_name':corporate_name,'corporate_code':corporate_code,'contact_person_name':contact_person_name,'contact_person_no':contact_person_no,
                     'contact_person_email': contact_person_email,'has_billing_spoc_level':has_billing_spoc_level,
                      'has_auth_level': has_auth_level,'no_of_auth_level':no_of_auth_level,'has_assessment_codes':
                      has_assessment_codes,'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,'tds_on_management_fee_only':tds_on_management_fee_only,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,'has_billing_admin_level':has_billing_admin_level,
                       'will_do_realtime_payment':will_do_realtime_payment,'has_self_booking_access':has_self_booking_access,
            'corporate_id': corporate_id,'user_id':user_id,'user_type':user_type,'is_send_email':is_send_email,'is_send_sms':is_send_sms}
            print(payload)
            company = getDataFromAPI(user_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Company Updated Successfully..!')
                return HttpResponseRedirect("/agents/companies", {'message': "Updated Successfully"})
            else:
                messages.error(request, 'Failed to Update Company..!')
                return HttpResponseRedirect("/agents/companies", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")
    else:
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "view_company"
            payload = {'corporate_id': id}
            company = getDataFromAPI(user_type, access_token, url, payload)
            companys = company['Corporates']
            return render(request, 'Agent/edit_company.html', {'companys': companys})
        else:
            return HttpResponseRedirect("/agents/login")


def delete_company(request,id):
    request = get_request()
    if 'agent_login_type' in request.session:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        user_id = request.POST.get('user_id', '')

        url = settings.API_BASE_URL+"delete_company"
        payload = {'corporate_id': corporate_id,'user_id':user_id}
        company = getDataFromAPI(user_type, access_token, url, payload)
        if company['success'] == 1:
            messages.success(request, 'company Deleted Successfully..!')
            return HttpResponseRedirect("/agents/companies", {'message': "Deleted Successfully"})
        else:
            messages.error(request, 'Failed to Delete company..!')
            return HttpResponseRedirect("/agents/companies", {'message': "Fail"})
    else:
        return HttpResponseRedirect("/agents/login")


def add_company_document(request):
    request = get_request()
    if 'agent_login_type' in request.session:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        user_id = request.POST.get('user_id', '')
        current_url = request.POST.get('current_url', '')
        document_name = request.POST.get('document_name', '')
        document_desc = request.POST.get('document_desc', '')

        global booking_email
        booking_email = ''
        if request.FILES:
            file_up = request.FILES.get('document', False)
            if file_up:
                file_up = request.FILES['document']
                booking_email = file_company_doc_upload(file_up)
            else:
                booking_email = None
        else:
            booking_email = None

        document = booking_email

        url = settings.API_BASE_URL+"add_company_document"
        payload = {'corporate_id': corporate_id,'user_id':user_id, 'document_name':document_name, 'document_desc':document_desc, 'document':document}
        company = getDataFromAPI(user_type, access_token, url, payload)
        if company['success'] == 1:
            messages.success(request, 'Document Added Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Deleted Successfully"})
        else:
            messages.error(request, 'Failed to Add Document..!')
            return HttpResponseRedirect(current_url, {'message': "Fail"})
    else:
        return HttpResponseRedirect("/agents/login")


def delete_company_document(request):
    request = get_request()
    if 'agent_login_type' in request.session:
        user_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        document_id = request.POST.get('document_id', '')
        user_id = request.POST.get('user_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL+"delete_company_document"
        payload = {'document_id': document_id,'user_id':user_id}
        company = getDataFromAPI(user_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            messages.success(request, 'Document Deleted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Deleted Successfully"})
        else:
            messages.error(request, 'Failed to Delete Document..!')
            return HttpResponseRedirect(current_url, {'message': "Fail"})
    else:
        return HttpResponseRedirect("/agents/login")


def company_admins(request, id):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': id}
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        r = requests.post(url, data=payload, headers=headers)
        company = json.loads(r.text)

        url_companies = settings.API_BASE_URL + "companies"
        companies = getDataFromAPI(login_type, access_token, url_companies, payload)
        companies = companies['Corporates']

        if company['success'] == 1:
            admins = company['Admins']
            return render(request, "Agent/company_admins.html", {'admins': admins,'companies':companies})
        else:
            return render(request, "Agent/company_admins.html", {'admins': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def company_billing_entities(request, id):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': id}

        company = getDataFromAPI(login_type, access_token, url, payload)
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)

        url_companies = settings.API_BASE_URL + "companies"
        companies = getDataFromAPI(login_type, access_token, url_companies, payload)
        companies = companies['Corporates']

        if company['success'] == 1:
            entities = company['Entitys']
            print(entities)
            cities = cities["Cities"]
            return render(request, "Agent/billing_entities.html",{'billing_entities': entities,"cities": cities,'companies': companies})
        else:
            return render(request, "Agent/billing_entities.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_rates(request, id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': id}
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        r = requests.post(url, data=payload, headers=headers)
        company = json.loads(r.text)

        url_companies = settings.API_BASE_URL + "companies"
        companies = getDataFromAPI(login_type, access_token, url_companies, payload)
        companies = companies['Corporates']
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        url_taxi_type = settings.API_BASE_URL + "taxi_types"
        taxi_type = getDataFromAPI(login_type, access_token, url_taxi_type, payload)

## sanket added
        payload = {}
        url_taxi_packages = settings.API_BASE_URL + "taxi_packages"
        taxi_packages = getDataFromAPI(login_type, access_token, url_taxi_packages, payload)
        taxi_packages = taxi_packages['packages']

        if company['success'] == 1:
            company_rates = company['Corporate_Retes']
            cities = cities["Cities"]
            taxi_type = taxi_type['taxi_types']
            return render(request, "Agent/company_rates.html", {'corporate_rates': company_rates,'companies':companies,'cities':cities,'taxi_types':taxi_type,'taxi_packages':taxi_packages})
        else:
            return render(request, "Agent/company_rates.html", {'entities': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_groups(request, id):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': id}
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        r = requests.post(url, data=payload, headers=headers)
        company = json.loads(r.text)

        url_companies = settings.API_BASE_URL + "companies"
        companies = getDataFromAPI(login_type, access_token, url_companies, payload)
        companies = companies['Corporates']

        if company['success'] == 1:
            groups = company['Groups']
            return render(request, "Agent/groups.html", {'groups': groups,'companies':companies})
        else:
            return render(request, "Agent/groups.html", {'groups': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_subgroups(request, id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': id}
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        r = requests.post(url, data=payload, headers=headers)
        company = json.loads(r.text)

        url_companies = settings.API_BASE_URL + "companies"
        companies = getDataFromAPI(login_type, access_token, url_companies, payload)
        companies = companies['Corporates']

        if company['success'] == 1:
            url2 = settings.API_BASE_URL + "groups"
            subgroups = company['Subgroups']
            r = requests.post(url2, data=payload, headers=headers)
            gr = json.loads(r.text)
            groups = gr['Groups']
            return render(request, "Agent/subgroups.html", {'subgroups': subgroups, 'groups': groups,'companies':companies})
        else:
            return render(request, "Agent/subgroups.html", {'subgroups': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_spocs(request, id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': id}
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        r = requests.post(url, data=payload, headers=headers)
        company = json.loads(r.text)
        if company['success'] == 1:
            spocs = company['Spocs']
            return render(request, "Agent/spocs.html", {'spocs': spocs})
        else:
            return render(request, "Agent/spocs.html", {'spocs': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_employees(request, id):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': id}
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        r = requests.post(url, data=payload, headers=headers)
        company = json.loads(r.text)
        if company['success'] == 1:
            employees = company['Employees']
            return render(request, "Agent/employees.html", {'employees': employees})
        else:
            return render(request, "Agent/employees.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_management_fees(request):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        if request.method == 'POST':
            corporate_id = request.POST.get('corporate_id', '')
            user_id = request.POST.get('user_id', '')
            service_fees_type_id = request.POST.get('service_fees_type_id', '')
            service_fees_type_value = request.POST.get('service_fees_type_value', '')
            service_fees_type = request.POST.get('service_fees_type', '')
            fees_id = request.POST.get('fees_id', '')
            delete_id = request.POST.get('delete_id', '')
            print(delete_id)
            print("idsisi")
            payload = {'corporate_id': corporate_id, 'service_fees_type_id': service_fees_type_id, 'service_fees_type_value': service_fees_type_value,
                      'service_fees_type':service_fees_type,'fees_id':fees_id ,'login_type': login_type, 'user_id': user_id}

            if fees_id:
                if delete_id:
                    print("i m here")
                    url = settings.API_BASE_URL + "delete_corporate_management_fee"
                    operation_msg = "Corporate Management Fee Deleted Successfully..!"
                else:
                    url = settings.API_BASE_URL + "update_corporate_management_fee"
                    operation_msg = "Corporate Management Fee Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_corporate_management_fee"
                operation_msg = "Corporate Management Fee Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, operation_msg)
                return HttpResponseRedirect("/agents/company-management-fees", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/company-management-fees", {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "corporate_management_fee"
            payload = {'corporate_id': id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            fees = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            url_types = settings.API_BASE_URL + "service_fee_types"
            fee_types = getDataFromAPI(login_type, access_token, url_types, payload)

            if fees['success'] == 1:
                fees = fees['Fees']
                companies = company1['Corporates']
                fee_types= fee_types['Types']
                return render(request, "Agent/corporate_management_fee.html", {'fees': fees,'companies':companies,'fee_types':fee_types})
            else:
                return render(request, "Agent/corporate_management_fee.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_operation_management(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        user_id = request.user.id;

        if request.method == 'POST':
            corporate_id = request.POST.get('corporate_id', '')
            service_type_id = request.POST.get('service_type_id', '')
            agent_id = request.POST.get('agent_id', '')
            is_active = request.POST.get('is_active', '')

            oms_id = request.POST.get('oms_id', '')

            payload = {'corporate_id':corporate_id,'service_type_id':service_type_id,'agent_id':agent_id,'oms_id':oms_id,'is_active':is_active,'login_type':login_type,'user_id':user_id}
            print(payload)
            if oms_id:
                url = settings.API_BASE_URL + "update_operation_managements"
                message = "Operation Manager Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operation_managements"
                message = "Operation Manager Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, message)
                return HttpResponseRedirect("/agents/company-operation-management", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/company-operation-management", {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "operation_managements"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            url_service_type = settings.API_BASE_URL + "service_types"
            service_type = getDataFromAPI(login_type, access_token, url_service_type, payload)

            url_agets = settings.API_BASE_URL + "agents"
            agents = getDataFromAPI(login_type, access_token, url_agets, payload)

            if company['success'] == 1:
                access = company['Access']
                companies = company1['Corporates']
                service_types = service_type['Types']
                agents = agents['Agents']
                return render(request, "Agent/operation_managements.html", {'access': access,'companies':companies,'service_types':service_types,'agents':agents})
            else:
                return render(request, "Agent/operation_managements.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def company_relationship_manager(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        user_id = request.user.id;

        if request.method == 'POST':
            corporate_id = request.POST.get('corporate_id', '')
            rm_level_1_id = request.POST.get('rm_level_1_id', '')
            rm_level_2_id = request.POST.get('rm_level_2_id', '')
            is_active = request.POST.get('is_active', '')

            rms_id = request.POST.get('rms_id', '')

            payload = {'corporate_id':corporate_id,'rm_level_1_id':rm_level_1_id,'rm_level_2_id':rm_level_2_id,'rms_id':rms_id,'is_active':is_active,'login_type':login_type,'user_id':user_id}

            if rms_id:
                url = settings.API_BASE_URL + "update_relationship_managements"
                message = "Relationship Manager Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_relationship_managements"
                message = "Relationship Manager Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, message)
                return HttpResponseRedirect("/agents/company-relationship-manager", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/company-relationship-manager", {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "relationship_managements"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            url_agets = settings.API_BASE_URL + "agents"
            agents = getDataFromAPI(login_type, access_token, url_agets, payload)

            if company['success'] == 1:
                access = company['Access']
                companies = company1['Corporates']
                agents = agents['Agents']
                return render(request, "Agent/relationship_manager.html", {'access': access,'companies':companies,'agents':agents})
            else:
                return render(request, "Agent/relationship_manager.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def assessment_cities(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        user_id = request.user.id;

        if request.method == 'POST':
            corporate_id = request.POST.get('corporate_id', '')
            city_name = request.POST.get('city_name', '')
            print(city_name)

            city_id = request.POST.get('city_id', '')

            payload = {'corporate_id':corporate_id,'city_name':city_name,'city_id':city_id,'login_type':login_type,'user_id':user_id}

            if city_id:
                url = settings.API_BASE_URL + "update_assessment_cities"
                opr_msg = "Assessment City Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_assessment_cities"
                opr_msg = "Assessment City Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect("/agents/assessment_cities/0", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/assessment_cities/0", {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "assessment_cities"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)


            if company['success'] == 1:
                cities = company['Cities']
                companies = company1['Corporates']
                return render(request, "Agent/assessment_cities.html", {'cities': cities,'companies':companies})
            else:
                return render(request, "Agent/assessment_cities.html", {'employees': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def assessment_codes(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        user_id = request.user.id;
        if request.method == 'POST':

            corporate_id = request.POST.get('corporate_id', '')
            assessment_code = request.POST.get('assessment_code', '')
            code_desc = request.POST.get('code_desc', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            service_from = request.POST.get('service_from', '')
            service_to = request.POST.get('service_to', '')

            code_id = request.POST.get('code_id', '')

            payload = {'corporate_id': corporate_id, 'assessment_code': assessment_code, 'code_desc': code_desc,'from_date':from_date,'to_date':to_date,
                       'login_type': login_type, 'user_id': user_id,'service_from':service_from,'service_to':service_to,'code_id':code_id}

            if code_id:
                url = settings.API_BASE_URL + "update_assessment_codes"
                opr_msg = "Assessment Code Updated Successfully...!"
            else:
                url = settings.API_BASE_URL + "add_assessment_codes"
                opr_msg = "Assessment Code Added Successfully...!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect("/agents/assessment_codes/0", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/assessment_codes/0", {'message': "Record Not Added"})

        else:
            url = settings.API_BASE_URL + "assessment_codes"
            payload = {'corporate_id':id}
            headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)

            url_comp = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url_comp, payload)

            if company['success'] == 1:
                codes = company['Codes']
                companies = company1['Corporates']
                return render(request, "Agent/assessment_codes.html", {'codes': codes,'companies':companies})
            else:
                return render(request, "Agent/assessment_codes.html", {'codes': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def delete_assessment_codes(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        code_id = request.POST.get('code_id')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            payload = {'code_id': code_id, 'user_id': user_id, 'login_type': login_type}
            url = settings.API_BASE_URL + "delete_assessment_codes"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Assessment Code Deleted Successfully..!')
                return HttpResponseRedirect("/agents/assessment_codes/0", {'message': "Deleted Successfully"})
            else:
                messages.error(request, 'Failed to Deleted Assessment Code .!')
                return HttpResponseRedirect("/agents/assessment_codes/0", {'message': "Failed"})
        else:
            return HttpResponseRedirect("/agents/login")


def delete_assessment_cities(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        city_id = request.POST.get('city_id')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            payload = {'city_id': city_id, 'user_id': user_id, 'login_type': login_type}
            url = settings.API_BASE_URL + "delete_assessment_cities"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Assessment City Deleted Successfully..!')
                return HttpResponseRedirect("/agents/assessment_cities/0", {'message': "Deleted Successfully"})
            else:
                messages.error(request, 'Failed to Delete Assessment City..!')
                return HttpResponseRedirect("/agents/assessment_cities/0", {'message': "Failed"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_company_rate(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')

        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            package_name = request.POST.get('package_name', '')
            package_name_txt = request.POST.get('package_name_txt', '')
            city_id = request.POST.get('city_id')
            taxi_type = request.POST.get('taxi_type', '')
            tour_type = request.POST.get('tour_type', '')
            kms = request.POST.get('kms', '')
            hours = request.POST.get('hours', '')
            km_rate = request.POST.get('km_rate', '')
            hour_rate = request.POST.get('hour_rate', '')
            base_rate = request.POST.get('base_rate', '')
            night_rate = request.POST.get('night_rate', '')
            night_start = request.POST.get('night_start', '')
            night_end = request.POST.get('night_end', '')

            delete_id = request.POST.get('delete_id')
            rate_id = request.POST.get('rate_id')

            if not package_name:
                package_name = package_name_txt

            payload = {'corporate_id': corporate_id,'package_name':package_name,'city_id':city_id,'taxi_type':taxi_type,
            'tour_type':tour_type,'kms':kms,'hours':hours,'km_rate':km_rate,'hour_rate':hour_rate,'base_rate':base_rate,'night_rate':night_rate,
            'user_id': user_id, 'user_type': user_type,'rate_id':rate_id,'is_delete': delete_id,'night_start':night_start,'night_end':night_end }

            print(payload)

            if rate_id:
                url = settings.API_BASE_URL + "update_company_rates"
                opr_msg = "Company Rate Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_company_rates"
                    opr_msg = "Company Rate Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_company_rates"
                opr_msg = "Company Rate Added Successfully..!"

            company = getDataFromAPI(user_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect("/agents/rates/0", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/rates/0", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_company_entity(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            entity_name = request.POST.get('entity_name', '')
            billing_city_id = request.POST.get('billing_city_id')
            contact_person_name = request.POST.get('contact_person_name', '')
            contact_person_email = request.POST.get('contact_person_email', '')
            contact_person_no = request.POST.get('contact_person_no', '')
            address_line_1 = request.POST.get('address_line_1', '')
            address_line_2 = request.POST.get('address_line_2', '')
            address_line_3 = request.POST.get('address_line_3', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            entity_id = request.POST.get('entity_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token,
                       'entity_name': entity_name, 'billing_city_id': billing_city_id,
                       'contact_person_name': contact_person_name, 'contact_person_email': contact_person_email,
                       'contact_person_no': contact_person_no, 'address_line_1': address_line_1,
                       'address_line_2': address_line_2,
                       'address_line_3': address_line_3, 'gst_id': gst_id, 'pan_no': pan_no, 'entity_id': entity_id,
                       'is_delete': delete_id, }
            print(payload)

            url = ""
            if entity_id:
                url = settings.API_BASE_URL + "update_billing_entity"
                operation_message = "Company Entity Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_billing_entity"
                    operation_message = "Company Entity Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_billing_entity"
                operation_message = "Company Entity Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/billing_entities/0", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/billing_entities//0", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_company_group(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            group_name = request.POST.get('group_name', '')
            zone_name = request.POST.get('zone_name')

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))
            password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'group_name': group_name, 'zone_name': zone_name,'access_token_auth':access_token_auth,'name':name,
                       'email':email,'cid':cid,'contact_no':contact_no,'is_radio':is_radio,'is_local':is_local,'is_outstation':is_outstation,'is_bus':is_bus,
                       'is_train':is_train,'is_hotel':is_hotel,'is_meal':is_meal,'is_flight':is_flight,'is_water_bottles':is_water_bottles,'is_reverse_logistics':is_reverse_logistics,
                       'password':password,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = settings.API_BASE_URL + "add_group"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Group Added Successfully..!')
                return HttpResponseRedirect("/agents/groups/" + str(id), {'message': "Added Successfully"})
            else:
                messages.error(request, 'Failed to Add Company Group..!')
                return HttpResponseRedirect("/agents/groups/" + str(id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            subgroup_name = request.POST.get('group_name', '')
            group_id = request.POST.get('group_id', '')

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))
            password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'subgroup_name': subgroup_name, 'group_id': group_id,'name':name,
                       'email':email,'cid':cid,'contact_no':contact_no,'is_radio':is_radio,'is_local':is_local,'is_outstation':is_outstation,'is_bus':is_bus,
                       'is_train':is_train,'is_hotel':is_hotel,'is_meal':is_meal,'is_flight':is_flight,'is_water_bottles':is_water_bottles,'is_reverse_logistics':is_reverse_logistics,
                       'password':password,'access_token_auth':access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = settings.API_BASE_URL + "add_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Company Sub-Group Added Successfully..!')

                return HttpResponseRedirect("/agents/subgroups/" + str(id), {'message': "Added Successfully"})
            else:
                messages.error(request, 'Failed To Add company Sub-Group..!')
                return HttpResponseRedirect("/agents/subgroups/" + str(id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def update_company_group(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        group_id = request.POST.get('group_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            group_name = request.POST.get('group_name', '')
            zone_name = request.POST.get('zone_name')

            payload = {'group_id': group_id, 'access_token': access_token, 'group_name': group_name, 'zone_name': zone_name,
                       'user_id': user_id, 'login_type': login_type}

            url = settings.API_BASE_URL + "update_group"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Group Update Successfully..!')
                return HttpResponseRedirect("/agents/view-company-group/" + group_id, {'message': "Update Successfully"})
            else:
                messages.error(request, 'Failed To Update Company Group..!')
                return HttpResponseRedirect("/agents/view-company-group/" + group_id, {'message': "Record Not Updated"})
        else:
            return HttpResponseRedirect("/agents/login")



def update_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        subgroup_id = request.POST.get('subgroup_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            group_name = request.POST.get('group_name', '')

            payload = {'subgroup_id': subgroup_id, 'access_token': access_token, 'group_name': group_name,
                       'user_id': user_id, 'login_type': login_type}

            print(payload)
            url = settings.API_BASE_URL + "update_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Company Sub-Group Update Successfully..!')
                return HttpResponseRedirect("/agents/view-company-subgroup/" + str(id),{'message': "Update Successfully"})
            else:
                messages.error(request, 'Failed To Update Company Sub-Group..!')
                return HttpResponseRedirect("/agents/view-company-subgroup/" + str(id),{'message': "Record Not Updated"})
        else:
            return HttpResponseRedirect("/agents/login")


def delete_company_group(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        group_id = request.POST.get('group_id')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            access_token_auth = request.session['agent_access_token']

            payload = {'group_id': group_id, 'user_id': user_id, 'login_type': login_type, 'access_token': access_token,
                       'access_token_auth': access_token_auth}
            url = settings.API_BASE_URL + "delete_group"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Company Group Deleted Successfully..!')
                return HttpResponseRedirect("/agents/groups/" + str(id), {'message': "Delete Successfully"})
            else:
                messages.error(request, 'Failed To Delete Company Group..!')
                return HttpResponseRedirect("/agents/groups/" + str(id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")



def delete_company_subgroup(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            subgroup_id = request.POST.get('subgroup_id')
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            access_token_auth = request.session['agent_access_token']

            payload = {'subgroup_id': subgroup_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'access_token_auth': access_token_auth}
            url = settings.API_BASE_URL + "delete_subgroup"
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(payload)
            if company['success'] == 1:
                messages.success(request, 'Company Sub-Group Deleted Successfully..!')
                return HttpResponseRedirect("/agents/subgroups/" + str(id), {'message': "Delete Successfully"})
            else:
                messages.error(request, 'Failed to Delete Company Sub-Group..!')
                return HttpResponseRedirect("/agents/subgroups/" + str(id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")



def add_company_group_auth(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            access_token_auth = request.session['agent_access_token']

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            group_id = request.POST.get('group_id')
            group_auth_id = request.POST.get('group_auth_id')
            delete_id = request.POST.get('delete_id')

            if group_id:
                group_auth_id = group_auth_id
                password = "taxi123"

            if group_auth_id:
                group_auth_id = group_auth_id
            else:
                group_auth_id = 0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'group_id': group_id, 'delete_id': delete_id, 'password': password, 'group_auth_id': group_auth_id,
                       'access_token_auth': access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = ""
            if group_auth_id:
                url = settings.API_BASE_URL + "update_group_auth"
                opr_msg = "Group Authenticator Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_group_auth"
                    opr_msg = "Group Authenticator Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_group_auth"
                opr_msg = "Group Authenticator Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect("/agents/view-company-group/" + group_id, {'message': "Added Successfully"})
            else:
                messages.error(request,  company['message'])
                return HttpResponseRedirect("/agents/view-company-group/" + group_id, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_company_subgroup_auth(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            access_token_auth = request.session['agent_access_token']

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            subgroup_id = request.POST.get('subgroup_id')
            subgroup_auth_id = request.POST.get('subgroup_auth_id')
            delete_id = request.POST.get('delete_id')

            if subgroup_id:
                subgroup_auth_id = subgroup_auth_id
                password = "taxi123"

            if subgroup_auth_id:
                subgroup_auth_id = subgroup_auth_id
            else:
                subgroup_auth_id = 0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'subgroup_id': subgroup_id, 'delete_id': delete_id, 'password': password,
                       'subgroup_auth_id': subgroup_auth_id, 'access_token_auth': access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = ""
            print(payload)
            if subgroup_auth_id:
                url = settings.API_BASE_URL + "update_subgroup_auth"
                opr_msg = "Subgroup Authenticator Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_subgroup_auth"
                    opr_msg = "Subgroup Authenticator Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_subgroup_auth"
                opr_msg = "Subgroup Authenticator Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(url)
            print(company)
            if company['success'] == 1:
                messages.success(request, opr_msg)
                return HttpResponseRedirect("/agents/view-company-subgroup/" + subgroup_id,{'message': "Added Successfully"})
            else:
                messages.error(request,  company['success'])
                return HttpResponseRedirect("/agents/view-company-subgroup/" + subgroup_id, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")



def add_company_admins(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            access_token_auth = ''.join(random.choice(string.ascii_uppercase + string.digits) for _ in range(60))

            name = request.POST.get('name', '')
            email = request.POST.get('email', '')
            cid = request.POST.get('cid', '')
            contact_no = request.POST.get('contact_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            has_billing_access = request.POST.get('has_billing_access', '')
            admin_id = request.POST.get('admin_id')

            delete_id = request.POST.get('delete_id')

            if admin_id:
                password = ''
            else:
                password = "taxi123"

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'name': name, 'email': email, 'cid': cid, 'contact_no': contact_no,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'admin_id': admin_id, 'delete_id': delete_id, 'password': password,
                       'access_token_auth': access_token_auth,'is_send_email':is_send_email,'is_send_sms':is_send_sms,'has_billing_access':has_billing_access}

            url = ""
            if admin_id:
                url = settings.API_BASE_URL + "update_admin"
                operation_message = "Company Admin Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_admin"
                    operation_message = "Company Admin Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_admin"
                operation_message = "Company Admin Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(url)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/admins/" + str(id), {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/admins/" + str(id), {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")



def add_spocs(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            group_id = request.POST.get('group_id', '')
            subgroup_id = request.POST.get('subgroup_id', '')
            user_cid = request.POST.get('user_cid', '')

            user_name = request.POST.get('user_name', '')
            user_contact = request.POST.get('user_contact', '')
            email = request.POST.get('email', '')
            username = request.POST.get('email', '')
            budget = request.POST.get('budget', '')
            expense = request.POST.get('budget', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')
            is_send_email = request.POST.get('is_send_email', '')
            is_send_sms = request.POST.get('is_send_sms', '')
            spoc_id = request.POST.get('spoc_id')

            delete_id = request.POST.get('delete_id')

            if spoc_id:
                password = ''
            else:
                password = "taxi123"
                spoc_id =0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'group_id': group_id, 'subgroup_id': subgroup_id, 'user_cid': user_cid, 'user_name': user_name,
                       'user_contact':user_contact,'email':email,'username':username,'budget':budget,'expense':expense,
                       'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics,
                       'spoc_id': spoc_id, 'delete_id': delete_id, 'password': password,'is_send_email':is_send_email,'is_send_sms':is_send_sms}

            url = ""
            print(payload)
            if spoc_id:
                url = settings.API_BASE_URL + "update_spoc"
                operation_message= "Spoc Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_spoc"
                    operation_message= "Spoc Deactivated Successfully..!"
                if delete_id == '0':
                    url = settings.API_BASE_URL + "delete_spoc"
                    operation_message= "Spoc Activated Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_spoc"
                operation_message = "Spoc Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(url)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/spocs/0", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/spocs/0", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")
    else:
        request = get_request()
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
        if 'agent_login_type' in request.session:
            url_spoc = settings.API_BASE_URL + "view_spoc"
            payload = {'spoc_id': id}
            r = requests.post(url_spoc, data=payload, headers=headers)
            company_spoc = json.loads(r.text)
            spocs = company_spoc['Spoc']

            for spoc in spocs:
                corporate_id = spoc['corporate_id']

            url = settings.API_BASE_URL + "groups"
            payload = {'corporate_id': 0}

            r = requests.post(url, data=payload, headers=headers)
            company = json.loads(r.text)
            groups = company['Groups']
            url_subgroup = settings.API_BASE_URL + "subgroups"
            r = requests.post(url_subgroup, data=payload, headers=headers)
            company_sub = json.loads(r.text)
            subgroups = company_sub['Subgroups']

            url_companies = settings.API_BASE_URL + "companies"
            companies = getDataFromAPI(login_type, access_token, url_companies, payload)
            companies = companies['Corporates']

            if id:
                return render(request, 'Agent/add_spoc.html', {'groups': groups, 'subgroups': subgroups, 'spoc':spocs, 'companies':companies})
            else:
                return render(request, 'Agent/add_spoc.html', {'groups': groups, 'subgroups': subgroups,'companies':companies})
        else:
            return HttpResponseRedirect("/agents/login")


def add_employee(request, id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            spoc_id = request.POST.get('spoc_id', '')
            billing_entity_id = request.POST.get('billing_entity_id', '')
            core_employee_id = request.POST.get('core_employee_id', '')
            employee_cid = request.POST.get('employee_cid', '')

            employee_name = request.POST.get('employee_name', '')
            employee_email = request.POST.get('employee_email', '')
            username = request.POST.get('employee_email', '')
            employee_contact = request.POST.get('employee_contact', '')
            age = request.POST.get('age')
            gender = request.POST.get('gender')
            id_proof_type = request.POST.get('id_proof_type')

            id_proof_no = request.POST.get('id_proof_no', '')
            is_active = request.POST.get('is_active','')
            has_dummy_email = request.POST.get('has_dummy_email')
            fcm_regid = request.POST.get('fcm_regid')
            is_cxo = request.POST.get('is_cxo')
            designation = request.POST.get('designation', '')
            home_city = request.POST.get('home_city', '')
            home_address = request.POST.get('home_address', '')
            reporting_manager = request.POST.get('reporting_manager', '')
            employee_band = request.POST.get('employee_band', '')
            modal_add_employee = request.POST.get('modal_add_employee', '')

            date_of_birth = request.POST.get('date_of_birth', '')
            if date_of_birth and date_of_birth != 'None':
                age = calculate_age(date_of_birth)
            else:
                age = 0

            if is_cxo == '1':
                assistant_id = request.POST.get('assistant_id', '')
                if assistant_id == ' ':
                    assistant_id =0
            else:
                assistant_id = 0

            employee_id = request.POST.get('employee_id')

            delete_id = request.POST.get('delete_id')

            if employee_id:
                password = ''
            else:
                password = "taxi123"
                employee_id =0

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,'spoc_id':spoc_id,'core_employee_id':core_employee_id,
                       'access_token': access_token,'employee_cid':employee_cid,'employee_name':employee_name,'employee_email':employee_email,
                       'employee_contact':employee_contact,'age':age,'gender':gender,'id_proof_type':id_proof_type,'id_proof_no':id_proof_no,
                       'is_active':is_active,'has_dummy_email':has_dummy_email,'fcm_regid':fcm_regid,'is_cxo':is_cxo,'employee_id': employee_id,
                       'designation':designation,'home_city':home_city,'home_address':home_address,'assistant_id':assistant_id,'date_of_birth':date_of_birth,
                       'delete_id': delete_id, 'password': password,'billing_entity_id':billing_entity_id,'username':username,'reporting_manager':reporting_manager,
                       'employee_band':employee_band}

            url = ""
            print(payload)
            if employee_id:
                url = settings.API_BASE_URL + "update_employee"
                operation_message = "Employee Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_employee"
                    operation_message = "Employee Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_employee"
                operation_message = "Employee Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(url)
            print(company)
            if company['success'] == 1:
                if modal_add_employee:
                    return HttpResponse(1)
                else:
                    messages.success(request, operation_message)
                    return HttpResponseRedirect("/agents/employees/0", {'message': "Added Successfully"})
            else:
                if modal_add_employee:
                    return HttpResponse(0)
                else:
                    messages.error(request, company['message'])
                    return HttpResponseRedirect("/agents/employees/0", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")
    else:
        request = get_request()

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url_emp = settings.API_BASE_URL + "view_employee"
            payload = {'employee_id': id,'corporate_id': 0}
            company_emp = getDataFromAPI(login_type, access_token, url_emp, payload)
            employees = company_emp['Employee']

            url_companies = settings.API_BASE_URL + "companies"
            companies = getDataFromAPI(login_type, access_token, url_companies, payload)
            companies = companies['Corporates']

            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']


            url_spoc = settings.API_BASE_URL + "spocs"
            company_spoc = getDataFromAPI(login_type, access_token, url_spoc, payload)
            spocs = company_spoc['Spocs']

            url_entity = settings.API_BASE_URL + "billing_entities"
            company_entity = getDataFromAPI(login_type, access_token, url_entity, payload)
            entitys = company_entity['Entitys']

            url_emp = settings.API_BASE_URL + "employees"
            employees1 = getDataFromAPI(login_type, access_token, url_emp, payload)
            employeess = employees1['Employees']


            if id:
                return render(request, 'Agent/add_employee.html', {'employee':employees,'companies':companies,'cities':cities,'spocs':spocs,'entitys':entitys,'employeess':employeess})
            else:
                return render(request, 'Agent/add_employee.html', {'companies':companies,'cities':cities,'spocs':spocs,'entitys':entitys,'employeess':employeess})
        else:
            return HttpResponseRedirect("/agents/login")



def add_agent(request,id):
    if request.method == 'POST':
        request = get_request()

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            user_id = request.POST.get('user_id', '')

            emp_id = request.POST.get('emp_id', '')
            username = request.POST.get('username', '')
            contact_no = request.POST.get('contact_no', '')
            email = request.POST.get('email', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')

            has_billing_access = request.POST.get('has_billing_access', '')
            has_voucher_payment_access = request.POST.get('has_voucher_payment_access', '')
            has_voucher_approval_access = request.POST.get('has_voucher_approval_access', '')
            is_super_admin = request.POST.get('is_super_admin', '')
            is_relationship_manager = request.POST.get('is_relationship_manager', '')
            is_operation_manager = request.POST.get('is_operation_manager', '')

            agent_id = request.POST.get('agents_id','')

            delete_id = request.POST.get('delete_id')

            if agent_id:
                password = ''
            else:
                password = "taxi123"
                agent_id =0

            payload = {'emp_id': emp_id,'username': username,'contact_no': contact_no,'email': email,'is_radio': is_radio,'is_local': is_local,
            'is_outstation': is_outstation,'is_bus': is_bus,'is_train': is_train,'is_hotel': is_hotel,'is_meal':is_meal,'is_flight':is_flight,
            'is_water_bottles':  is_water_bottles,'is_reverse_logistics':
            is_reverse_logistics,'has_billing_access':has_billing_access,'has_voucher_payment_access':has_voucher_payment_access,
            'has_voucher_approval_access': has_voucher_approval_access,'is_super_admin':is_super_admin,'password':password,'is_operation_manager':is_operation_manager,
            'user_id':user_id,'user_type':login_type,'agent_id':agent_id,'delete_id':delete_id,'is_relationship_manager':is_relationship_manager}

            url = ""
            print(payload)
            if agent_id:
                url = settings.API_BASE_URL + "update_agent"
                operation_message = "Agent Update Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_agent"
                    operation_message = "Agent Deactivated Successfully..!"
                if delete_id == '2':
                    url = settings.API_BASE_URL + "activate_agent"
                    operation_message = "Agent Activated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_agent"
                operation_message = "Agent Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/agents", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/agents", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if id:

            if 'agent_login_type' in request.session:
                login_type = request.session['agent_login_type']
                access_token = request.session['agent_access_token']

                url_agent = settings.API_BASE_URL + "view_agent"
                payload = {'agent_id': id}
                agent = getDataFromAPI(login_type, access_token, url_agent, payload)
                agents = agent['Agent']
                return render(request, 'Agent/add_agent.html', {'agent':agents})
            else:
                return HttpResponseRedirect("/agents/login")
        else:
            return render(request, 'Agent/add_agent.html', {})



def view_company_group(request, id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_group"
        payload = {'group_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_group_auth"
        grp_auths = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            groups = company['Groups']
            grp_auths = grp_auths['Groups']
            return render(request, "Agent/view_groups.html", {'group': groups, 'grp_auths': grp_auths})
        else:
            return render(request, "Agent/view_groups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_company_subgroup(request, id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_subgroup"
        payload = {'subgroup_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_subgroup_auth"
        subgrp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(subgrp_auths)
        if company['success'] == 1:
            subgroups = company['SubGroups']
            subgrp_auths = subgrp_auths['SubGroups']
            return render(request, "Agent/view_subgroups.html",{'subgroup': subgroups, 'subgrp_auths': subgrp_auths})
        else:
            return render(request, "Agent/view_subgroups.html", {'group': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_agents(request):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL+"agents"
        payload = {'some': 'data'}
        agents = getDataFromAPI(login_type, access_token, url, payload)
        if agents['success'] == 1:
            agents = agents['Agents']
            return render(request,"Agent/agents.html",{'agents':agents})
        else:
            return render(request,"Agent/agents.html",{'agents':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def hotels(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL+"hotels"
        payload = {}
        hotels = getDataFromAPI(login_type, access_token, url, payload)
        if hotels['success'] == 1:
            hotels = hotels['Hotels']
            return render(request,"Agent/hotels.html",{'hotels':hotels})
        else:
            return render(request,"Agent/hotels.html",{'hotels':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def hotel_booking_portals(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL+"hotel_booking_portals"
        payload = {}
        hotels = getDataFromAPI(login_type, access_token, url, payload)
        if hotels['success'] == 1:
            hotels = hotels['Portals']
            return render(request,"Agent/hotel_booking_portals.html",{'portals':hotels})
        else:
            return render(request,"Agent/hotel_booking_portals.html",{'hotels':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def add_hotel(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            cotrav_agent_id = request.POST.get('cotrav_agent_id', '')

            hotel_name = request.POST.get('hotel_name', '')
            hotel_email = request.POST.get('hotel_email', '')
            hotel_contact = request.POST.get('hotel_contact', '')
            website = request.POST.get('website', '')

            hotel_address = request.POST.get('hotel_address', '')
            contact_name = request.POST.get('contact_name', '')
            contact_email = request.POST.get('contact_email', '')
            contact_no = request.POST.get('contact_no', '')

            beneficiary_name = request.POST.get('beneficiary_name', '')
            beneficiary_account_no = request.POST.get('beneficiary_account_no', '')
            bank_name = request.POST.get('bank_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')

            is_service_tax_applicable = request.POST.get('is_service_tax_applicable', '')
            tds_rate = request.POST.get('tds_rate', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            hotel_id = request.POST.get('hotel_id', '')
            delete_id = request.POST.get('delete_id')

            payload = {'type':type,'hotel_name':hotel_name,'hotel_email':hotel_email,'hotel_contact':hotel_contact,'hotel_address':hotel_address,'contact_name':contact_name,
                       'website':website,'beneficiary_account_no':beneficiary_account_no,'ifsc_code':ifsc_code,'bank_name':bank_name,
                       'is_service_tax_applicable':is_service_tax_applicable,'contact_email':contact_email,'contact_no':contact_no,
                       'beneficiary_name':beneficiary_name,'tds_rate':tds_rate,'gst_id':gst_id,'pan_no':pan_no,'hotel_id':hotel_id,'user_id':cotrav_agent_id,'user_type':login_type}

            url = ""

            if hotel_id:
                url = settings.API_BASE_URL + "update_hotel"
                operation_message = "Operator Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_hotel"
                    operation_message = "Operator Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_hotel"
                operation_message = "Operator Added Successfully..!"

            operator = getDataFromAPI(login_type, access_token, url, payload)
            print(operator)
            if operator['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/hotels", {'message': "Added Successfully"})
            else:
                messages.error(request, operator['message'])
                return HttpResponseRedirect("/agents/hotels", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            if id:
                login_type = request.session['agent_login_type']
                access_token = request.session['agent_access_token']
                payload = {'hotel_id':id}
                url = settings.API_BASE_URL + "view_hotel"
                operator = getDataFromAPI(login_type, access_token, url, payload)
                operator = operator['Hotels']
                return render(request, 'Agent/add_hotel.html', {'operators':operator})
            else:
                return render(request, 'Agent/add_hotel.html', {})
        else:
            return HttpResponseRedirect("/agents/login")


def add_hotel_portals(request, id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            cotrav_agent_id = request.POST.get('cotrav_agent_id', '')

            hotel_name = request.POST.get('hotel_name', '')
            hotel_email = request.POST.get('hotel_email', '')
            hotel_contact = request.POST.get('hotel_contact', '')
            website = request.POST.get('website', '')

            hotel_address = request.POST.get('hotel_address', '')
            contact_name = request.POST.get('contact_name', '')
            contact_email = request.POST.get('contact_email', '')
            contact_no = request.POST.get('contact_no', '')

            beneficiary_name = request.POST.get('beneficiary_name', '')
            beneficiary_account_no = request.POST.get('beneficiary_account_no', '')
            bank_name = request.POST.get('bank_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')

            is_service_tax_applicable = request.POST.get('is_service_tax_applicable', '')
            tds_rate = request.POST.get('tds_rate', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            hotel_id = request.POST.get('hotel_id', '')
            delete_id = request.POST.get('delete_id')
            portal_vendor_id = request.POST.get('portal_vendor_id')

            payload = {'type':6,'hotel_name':hotel_name,'hotel_email':hotel_email,'hotel_contact':hotel_contact,'hotel_address':hotel_address,'contact_name':contact_name,
                       'website':website,'beneficiary_account_no':beneficiary_account_no,'ifsc_code':ifsc_code,'bank_name':bank_name,
                       'is_service_tax_applicable':is_service_tax_applicable,'contact_email':contact_email,'contact_no':contact_no,
                       'beneficiary_name':beneficiary_name,'tds_rate':tds_rate,'gst_id':gst_id,'pan_no':pan_no,'hotel_id':hotel_id,
                       'user_id':cotrav_agent_id,'user_type':login_type,'portal_vendor_id':portal_vendor_id}

            url = ""
            print(payload)

            if hotel_id:
                url = settings.API_BASE_URL + "update_hotel_portal"
                operation_message = "Hotel Portal Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_hotel_portal"
                    operation_message = "Hotel Portal Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_hotel_portal"
                operation_message = "Hotel Portal Added Successfully..!"

            operator = getDataFromAPI(login_type, access_token, url, payload)
            print(operator)
            if operator['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/hotel_booking_portals", {'message': "Added Successfully"})
            else:
                messages.error(request, operator['message'])
                return HttpResponseRedirect("/agents/hotel_booking_portals", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            if id:
                login_type = request.session['agent_login_type']
                access_token = request.session['agent_access_token']
                payload = {'hotel_id':id}
                url = settings.API_BASE_URL + "view_hotel_portal"
                operator = getDataFromAPI(login_type, access_token, url, payload)
                operator = operator['Portals']
                return render(request, 'Agent/add_hotel_portal.html', {'operators':operator})
            else:
                return render(request, 'Agent/add_hotel_portal.html', {})
        else:
            return HttpResponseRedirect("/agents/login")


def hotel_contacts(request, id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            hotel_id = request.POST.get('hotel_id', '')
            operator_address = request.POST.get('operator_address', '')
            contact_name = request.POST.get('contact_name', '')
            contact_email = request.POST.get('contact_email', '')
            contact_no = request.POST.get('contact_no', '')

            contact_id = request.POST.get('contact_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')
            url = ""
            payload = {}

            if contact_id:
                if delete_id == '1':
                    print("in delete")
                    url = settings.API_BASE_URL + "delete_hotel_contact"
                    payload = {'contact_id':contact_id,'user_id':user_id,'user_type':user_type}
                    operation_message= "Hotel Contact Deleted Successfully..!"
                else:
                    print("in edit")
                    url = settings.API_BASE_URL + "update_hotel_contact"
                    payload = {'contact_id':contact_id,'hotel_id':hotel_id,'operator_address':operator_address,'contact_name':contact_name,'contact_email':contact_email,
                               'contact_no':contact_no,'user_id':user_id,'user_type':user_type}
                    operation_message = "Hotel Contact Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_hotel_contact"
                payload = {'hotel_id':hotel_id,'operator_address':operator_address,'contact_name':contact_name,'contact_email':contact_email,
                               'contact_no':contact_no,'user_id':user_id,'user_type':user_type}
                operation_message = "Hotel Contact Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/hotel_contacts/"+str(id), {'message': "Added Successfully"})
            else:
                messages.error(request, taxi['message'])
                return HttpResponseRedirect("/agents/hotel_contacts/"+str(id), {'message': "Added Successfully"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL+"hotel_contacts"
            payload = {'hotel_id':id}
            operator_contacts = getDataFromAPI(login_type, access_token, url, payload)

            if operator_contacts['success'] == 1:
                operator_contacts = operator_contacts['HotelContacts']
                return render(request,"Agent/hotel_contacts.html",{'hotel_contacts':operator_contacts,'hotel_id':id})
            else:
                return render(request,"Agent/hotel_contacts.html",{'hotel_contacts':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def hotel_banks(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            hotel_id = request.POST.get('hotel_id', '')
            beneficiary_name = request.POST.get('beneficiary_name', '')
            beneficiary_account_no = request.POST.get('beneficiary_account_no', '')
            bank_name = request.POST.get('bank_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')

            bank_id = request.POST.get('bank_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')

            url = ""
            payload = {}

            if bank_id:
                if delete_id == '1':
                    print("in delete")
                    url = settings.API_BASE_URL + "delete_hotel_bank"
                    payload = {'bank_id': bank_id, 'user_id': user_id, 'user_type': user_type}
                    operation_message = "Hotel Bank Deleted Successfully..!"
                else:
                    print("in edit")
                    url = settings.API_BASE_URL + "update_hotel_bank"
                    payload = {'bank_id': bank_id, 'hotel_id': hotel_id,'beneficiary_name':beneficiary_name,
                               'beneficiary_account_no':beneficiary_account_no,'bank_name':bank_name,'ifsc_code':ifsc_code,
                               'user_id': user_id, 'user_type': user_type}
                    operation_message = "Hotel Bank Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_hotel_bank"
                payload = {'hotel_id': hotel_id,'beneficiary_name':beneficiary_name,'beneficiary_account_no':beneficiary_account_no,'bank_name':bank_name,'ifsc_code':ifsc_code,
                           'user_id': user_id, 'user_type': user_type}
                operation_message = "Hotel Bank Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)
            print(taxi)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/hotel_banks/"+str(id), {'message': "Added Successfully"})
            else :
                messages.error(request,taxi['message'])
                return HttpResponseRedirect("/agents/hotel_banks/"+str(id), {'message': "Operation Failed"})

        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "hotel_banks"
            payload = {'hotel_id': id}
            operator_banks = getDataFromAPI(login_type, access_token, url, payload)
            if operator_banks['success'] == 1:
                operator_banks = operator_banks['HotelBanks']
                return render(request, "Agent/hotel_banks.html", {'operator_banks': operator_banks,'hotel_id':id})
            else:
                return render(request, "Agent/hotel_banks.html", {'operator_contacts': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def operators(request):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL+"operators"
        payload = {}
        operators = getDataFromAPI(login_type, access_token, url, payload)
        if operators['success'] == 1:
            operators = operators['Operators']
            return render(request,"Agent/operators.html",{'operators':operators})
        else:
            return render(request,"Agent/operators.html",{'operators':{}})
    else:
        return HttpResponseRedirect("/agents/login")



def operator_contacts(request, id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            operator_id = request.POST.get('operator_id', '')
            operator_address = request.POST.get('operator_address', '')
            contact_name = request.POST.get('contact_name', '')
            contact_email = request.POST.get('contact_email', '')
            contact_no = request.POST.get('contact_no', '')

            contact_id = request.POST.get('contact_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')
            url = ""
            payload = {}

            if contact_id:
                if delete_id == '1':
                    print("in delete")
                    url = settings.API_BASE_URL + "delete_operator_contact"
                    payload = {'contact_id':contact_id,'user_id':user_id,'user_type':user_type}
                    operation_message= "Operator Contact Deleted Successfully..!"
                else:
                    print("in edit")
                    url = settings.API_BASE_URL + "update_operator_contact"
                    payload = {'contact_id':contact_id,'operator_id':operator_id,'operator_address':operator_address,'contact_name':contact_name,'contact_email':contact_email,
                               'contact_no':contact_no,'user_id':user_id,'user_type':user_type}
                    operation_message = "Operator Contact Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operator_contact"
                payload = {'operator_id':operator_id,'operator_address':operator_address,'contact_name':contact_name,'contact_email':contact_email,
                               'contact_no':contact_no,'user_id':user_id,'user_type':user_type}
                operation_message = "Operator Contact Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/operator_contacts/"+str(id), {'message': "Added Successfully"})
            else:
                messages.error(request, taxi['message'])
                return HttpResponseRedirect("/agents/operator_contacts/"+str(id), {'message': "Added Successfully"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL+"operator_contacts"
            payload = {'operator_id':id}
            operator_contacts = getDataFromAPI(login_type, access_token, url, payload)

            if operator_contacts['success'] == 1:
                operator_contacts = operator_contacts['OperatorContacts']
                return render(request,"Agent/operator_contacts.html",{'operator_contacts':operator_contacts,'operator_id':id})
            else:
                return render(request,"Agent/operator_contacts.html",{'operator_contacts':{}})
    else:
        return HttpResponseRedirect("/agents/login")



def operator_banks(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            operator_id = request.POST.get('operator_id', '')
            beneficiary_name = request.POST.get('beneficiary_name', '')
            beneficiary_account_no = request.POST.get('beneficiary_account_no', '')
            bank_name = request.POST.get('bank_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')

            bank_id = request.POST.get('bank_id', '')
            user_id = request.POST.get('user_id', '')
            delete_id = request.POST.get('delete_id', '')

            url = ""
            payload = {}

            if bank_id:
                if delete_id == '1':
                    print("in delete")
                    url = settings.API_BASE_URL + "delete_operator_bank"
                    payload = {'bank_id': bank_id, 'user_id': user_id, 'user_type': user_type}
                    operation_message = "Operator Bank Deleted Successfully..!"
                else:
                    print("in edit")
                    url = settings.API_BASE_URL + "update_operator_bank"
                    payload = {'bank_id': bank_id, 'operator_id': operator_id,'beneficiary_name':beneficiary_name,
                               'beneficiary_account_no':beneficiary_account_no,'bank_name':bank_name,'ifsc_code':ifsc_code,
                               'user_id': user_id, 'user_type': user_type}
                    operation_message = "Operator Bank Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operator_bank"
                payload = {'operator_id': operator_id,'beneficiary_name':beneficiary_name,'beneficiary_account_no':beneficiary_account_no,'bank_name':bank_name,'ifsc_code':ifsc_code,
                           'user_id': user_id, 'user_type': user_type}
                operation_message = "Operator Bank Added Successfully..!"

            taxi = getDataFromAPI(user_type, access_token, url, payload)
            print(taxi)

            if taxi['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/operator_banks/"+str(id), {'message': "Added Successfully"})
            else :
                messages.error(request,taxi['message'])
                return HttpResponseRedirect("/agents/operator_banks/"+str(id), {'message': "Operation Failed"})

        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "operator_banks"
            payload = {'operator_id': id}
            operator_banks = getDataFromAPI(login_type, access_token, url, payload)
            if operator_banks['success'] == 1:
                operator_banks = operator_banks['OperatorBanks']
                return render(request, "Agent/operator_banks.html", {'operator_banks': operator_banks,'operator_id':id})
            else:
                return render(request, "Agent/operator_banks.html", {'operator_contacts': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def delete_operator(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        operator_id = request.POST.get('operator_id', '')
        user_id = request.POST.get('user_id', '')

        url = settings.API_BASE_URL+"delete_operator"
        payload = {'operator_id': operator_id,'user_id':user_id,'user_type':login_type}
        operators = getDataFromAPI(login_type, access_token, url, payload)
        if operators['success'] == 1:
            messages.success(request, 'Operator Operation Successful..!')
            return HttpResponseRedirect("/agents/operators", {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Delete Operator..!')
            return HttpResponseRedirect("/agents/operators", {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")



def add_operator(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            cotrav_agent_id = request.POST.get('cotrav_agent_id', '')

            type = request.POST.get('type', '')
            username = request.POST.get('username', '')
            operator_name = request.POST.get('operator_name', '')
            operator_email = request.POST.get('operator_email', '')

            operator_contact = request.POST.get('operator_contact', '')
            website = request.POST.get('website', '')
            operator_address = request.POST.get('operator_address', '')

            is_service_tax_applicable = request.POST.get('is_service_tax_applicable', '')
            service_tax_number = request.POST.get('service_tax_number', '')
            night_start_time = request.POST.get('night_start_time', '')
            night_end_time = request.POST.get('night_end_time', '')
            tds_rate = request.POST.get('tds_rate', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            is_radio = request.POST.get('is_radio', '')
            is_local = request.POST.get('is_local', '')
            is_outstation = request.POST.get('is_outstation', '')
            is_bus = request.POST.get('is_bus', '')
            is_train = request.POST.get('is_train', '')
            is_hotel = request.POST.get('is_hotel', '')
            is_meal = request.POST.get('is_meal', '')
            is_flight = request.POST.get('is_flight', '')
            is_water_bottles = request.POST.get('is_water_bottles', '')
            is_reverse_logistics = request.POST.get('is_reverse_logistics', '')

            operator_id = request.POST.get('operator_id', '')
            delete_id = request.POST.get('delete_id')

            if operator_id:
                password = ''
            else:
                password = "taxi123"
                operator_id = 0

            payload = {'type':type,'username':username,'password':password,'operator_name':operator_name,'operator_email':operator_email,'operator_contact':operator_contact,
                       'website':website,'operator_address':operator_address,
                       'is_service_tax_applicable':is_service_tax_applicable,'service_tax_number':service_tax_number,'night_start_time':night_start_time,
                       'night_end_time':night_end_time,'tds_rate':tds_rate,'gst_id':gst_id,'pan_no':pan_no,'operator_id':operator_id,
                       'user_id':cotrav_agent_id,'user_type':login_type,'is_radio': is_radio, 'is_local': is_local, 'is_outstation': is_outstation, 'is_bus': is_bus,
                       'is_train': is_train, 'is_hotel': is_hotel, 'is_meal': is_meal, 'is_flight': is_flight,
                       'is_water_bottles': is_water_bottles, 'is_reverse_logistics': is_reverse_logistics}

            url = ""

            if operator_id:
                url = settings.API_BASE_URL + "update_operator"
                operation_message = "Operator Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_operator"
                    operation_message = "Operator Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operator"
                operation_message = "Operator Added Successfully..!"

            operator = getDataFromAPI(login_type, access_token, url, payload)
            print(operator)
            if operator['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/operators", {'message': "Added Successfully"})
            else:
                messages.error(request, operator['message'])
                return HttpResponseRedirect("/agents/operators", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            if id:
                login_type = request.session['agent_login_type']
                access_token = request.session['agent_access_token']
                payload = {'operator_id':id}
                url = settings.API_BASE_URL + "view_operator"
                operator = getDataFromAPI(login_type, access_token, url, payload)
                operator = operator['Operator']
                return render(request, 'Agent/add_operator.html', {'operators':operator})
            else:
                return render(request, 'Agent/add_operator.html', {})
        else:
            return HttpResponseRedirect("/agents/login")



def operator_rates(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "operator_rates"
        payload = {}
        op_rates = getDataFromAPI(login_type, access_token, url, payload)
        if op_rates['success'] == 1:
            op_rates = op_rates['Rates']
            return render(request, "Agent/operator_rates.html", {'op_rates': op_rates})
        else:
            return render(request, "Agent/operator_rates.html", {'op_rates': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def add_operator_rate(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            user_id = request.POST.get('cotrav_agent_id', '')

            operator_id = request.POST.get('operator_id', '')
            city_id = request.POST.get('city_id', '')
            taxi_type_id = request.POST.get('taxi_type_id', '')
            package_name = request.POST.get('package_name', '')
            package_name_txt = request.POST.get('package_name_txt', '')
            if not package_name:
                package_name = package_name_txt
            tour_type = request.POST.get('tour_type', '')
            kms = request.POST.get('kms', '')
            hours = request.POST.get('hours', '')
            km_rate = request.POST.get('km_rate', '')
            hour_rate = request.POST.get('hour_rate', '')
            base_rate = request.POST.get('base_rate', '')
            night_rate = request.POST.get('night_rate', '')
            fuel_rate = request.POST.get('fuel_rate', '')

            rate_id = request.POST.get('rate_id', '')



            payload = {'operator_id':operator_id,'city_id':city_id,'taxi_type_id':taxi_type_id,'package_name':package_name,'tour_type':tour_type,
                       'kms':kms,'hours':hours,'km_rate':km_rate,'hour_rate':hour_rate,'base_rate':base_rate,'night_rate':night_rate,'fuel_rate':fuel_rate,
                       'rate_id':rate_id,'user_id':user_id,'login_type':login_type}

            url = ""

            if rate_id:
                url = settings.API_BASE_URL + "update_operator_rate"
                operation_message = "Operator Rate Updated Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_operator_rate"
                operation_message = "Operator Rate Added Successfully..!"

            operator = getDataFromAPI(login_type, access_token, url, payload)

            if operator['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/operator-rates", {'message': "Operation Successfully"})
            else:
                payload = {'operator_id': id}
                url = settings.API_BASE_URL + "operator_rate"
                operators = getDataFromAPI(login_type, access_token, url, payload)
                operators = operators['Rates']

                # url_cities = settings.API_BASE_URL + "cities"
                # cities = getDataFromAPI(login_type, access_token, url_cities, payload)
                # cities = cities['Cities']
                cities = ""

                url_taxi_type = settings.API_BASE_URL + "taxi_types"
                taxi_types = getDataFromAPI(login_type, access_token, url_taxi_type, payload)
                taxi_types = taxi_types['taxi_types']

                url_taxi_packages = settings.API_BASE_URL + "taxi_packages"
                taxi_packages = getDataFromAPI(login_type, access_token, url_taxi_packages, payload)
                taxi_packages = taxi_packages['packages']

                return render(request, 'Agent/add_operator_rate.html', {'operators':operators,'cities':cities,'taxi_types':taxi_types,'taxi_packages':taxi_packages})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            if id:
                if 'agent_login_type' in request.session:
                    login_type = request.session['agent_login_type']
                    access_token = request.session['agent_access_token']
                    payload = {'rate_id': id}
                    url = settings.API_BASE_URL + "view_operator_rate"
                    op_rates = getDataFromAPI(login_type, access_token, url, payload)
                    op_rates = op_rates['Rate']

                    url = settings.API_BASE_URL + "operators"
                    operators = getDataFromAPI(login_type, access_token, url, payload)
                    operators = operators['Operators']

                    url_cities = settings.API_BASE_URL + "cities"
                    cities = getDataFromAPI(login_type, access_token, url_cities, payload)
                    cities = cities['Cities']

                    url_taxi_type = settings.API_BASE_URL + "taxi_types"
                    taxi_types = getDataFromAPI(login_type, access_token, url_taxi_type, payload)
                    taxi_types = taxi_types['taxi_types']

                    url_taxi_packages = settings.API_BASE_URL + "taxi_packages"
                    taxi_packages = getDataFromAPI(login_type, access_token, url_taxi_packages, payload)
                    taxi_packages = taxi_packages['packages']

                    return render(request, 'Agent/add_operator_rate.html', {'operator_rates': op_rates,'operators':operators,'cities':cities,'taxi_types':taxi_types,'taxi_packages':taxi_packages})
                else:
                    return HttpResponseRedirect("/agents/login")
            else:
                if 'agent_login_type' in request.session:
                    login_type = request.session['agent_login_type']
                    access_token = request.session['agent_access_token']

                    payload = {'operator_id': id}
                    url = settings.API_BASE_URL + "operators"
                    operators = getDataFromAPI(login_type, access_token, url, payload)
                    operators = operators['Operators']

                    url_cities = settings.API_BASE_URL + "cities"
                    cities = getDataFromAPI(login_type, access_token, url_cities, payload)
                    cities = cities['Cities']

                    url_taxi_type = settings.API_BASE_URL + "taxi_types"
                    taxi_types = getDataFromAPI(login_type, access_token, url_taxi_type, payload)
                    taxi_types = taxi_types['taxi_types']

                    url_taxi_packages = settings.API_BASE_URL + "taxi_packages"
                    taxi_packages = getDataFromAPI(login_type, access_token, url_taxi_packages, payload)
                    taxi_packages = taxi_packages['packages']

                    return render(request, 'Agent/add_operator_rate.html', {'operators':operators,'cities':cities,'taxi_types':taxi_types,'taxi_packages':taxi_packages})
                else:
                    return HttpResponseRedirect("/agents/login")
        else:
            return HttpResponseRedirect("/agents/login")



def delete_operator_rate(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        rate_id = request.POST.get('rate_id', '')
        user_id = request.POST.get('user_id', '')

        url = settings.API_BASE_URL+"delete_operator_rate"
        payload = {'rate_id': rate_id,'user_id':user_id,'user_type':login_type}
        print(payload)
        operators = getDataFromAPI(login_type, access_token, url, payload)
        if operators['success'] == 1:
            messages.success(request, 'Operator Rate Deleted Successfully..!')
            return HttpResponseRedirect("/agents/operator-rates", {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Delete Operator..!')
            return HttpResponseRedirect("/agents/operator-rates", {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")


def operator_drivers(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "operator_drivers"
        payload = {}
        op_drivers = getDataFromAPI(login_type, access_token, url, payload)
        if op_drivers['success'] == 1:
            drivers = op_drivers['Drivers']
            return render(request, "Agent/operator_drivers.html", {'op_drivers': drivers})
        else:
            return render(request, "Agent/operator_drivers.html", {'op_drivers': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def add_operator_driver(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            user_id = request.POST.get('cotrav_agent_id', '')

            operator_id = request.POST.get('operator_id', '')
            driver_name = request.POST.get('driver_name', '')
            driver_contact = request.POST.get('driver_contact', '')
            driver_email = request.POST.get('driver_email', '')
            licence_no = request.POST.get('licence_no', '')
            fcm_regid = request.POST.get('fcm_regid', '')
            taxi_id = request.POST.get('taxi_id', '')

            driver_id = request.POST.get('driver_id', '')
            if driver_id:
                password = ''
            else:
                password = "taxi123"
                driver_id = 0

            payload = {'operator_id':operator_id,'driver_name':driver_name,'driver_contact':driver_contact,'driver_email':driver_email,'licence_no':licence_no,
                          'fcm_regid':fcm_regid,'taxi_id':taxi_id,'password':password,'driver_id':driver_id,'user_id':user_id,'login_type':login_type}

            url = ""

            if driver_id:
                url = settings.API_BASE_URL + "update_operator_driver"
                operation_message = 'Operator Driver Updated Successfully..!'
            else:
                url = settings.API_BASE_URL + "add_operator_driver"
                operation_message = 'Operator Driver Added Successfully..!'

            op_drivers = getDataFromAPI(login_type, access_token, url, payload)

            if op_drivers['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/operator-drivers", {'message': "Operation Successfully"})
            else:
                messages.error(request, op_drivers['message'])
                return HttpResponseRedirect("/agents/operator-drivers", {'message': "Operation Failed"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            if id:
                if 'agent_login_type' in request.session:
                    login_type = request.session['agent_login_type']
                    access_token = request.session['agent_access_token']
                    payload = {'driver_id': id}
                    url = settings.API_BASE_URL + "view_operator_driver"
                    op_drivers = getDataFromAPI(login_type, access_token, url, payload)
                    op_drivers = op_drivers['Drivers']

                    url = settings.API_BASE_URL + "operators"
                    operators = getDataFromAPI(login_type, access_token, url, payload)
                    operators = operators['Operators']

                    url_taxi = settings.API_BASE_URL + "taxis"
                    taxi = getDataFromAPI(login_type, access_token, url_taxi, payload)
                    taxi = taxi['Taxis']


                    return render(request, 'Agent/add_operator_driver.html', {'operator_drivers': op_drivers,'operators':operators,'taxies':taxi})
                else:
                    return HttpResponseRedirect("/agents/login")
            else:
                if 'agent_login_type' in request.session:
                    login_type = request.session['agent_login_type']
                    access_token = request.session['agent_access_token']

                    payload = {'operator_id': id}
                    url = settings.API_BASE_URL + "operators"
                    operators = getDataFromAPI(login_type, access_token, url, payload)
                    operators = operators['Operators']

                    url_taxi = settings.API_BASE_URL + "taxis"
                    taxi = getDataFromAPI(login_type, access_token, url_taxi, payload)
                    taxi = taxi['Taxis']

                    return render(request, 'Agent/add_operator_driver.html', {'operators':operators,'taxies':taxi})
                else:
                    return HttpResponseRedirect("/agents/login")
        else:
            return HttpResponseRedirect("/agents/login")


def delete_operator_driver(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        driver_id = request.POST.get('driver_id', '')
        user_id = request.POST.get('user_id', '')

        url = settings.API_BASE_URL+"delete_operator_driver"
        payload = {'driver_id': driver_id,'user_id':user_id,'user_type':login_type}
        operators = getDataFromAPI(login_type, access_token, url, payload)

        if operators['success'] == 1:
            messages.success(request, 'Operator Driver Deleted Successfully..!')
            return HttpResponseRedirect("/agents/operator-drivers", {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Delete Operator Driver..!')
            return HttpResponseRedirect("/agents/operator-drivers", {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")



def taxi_bookings(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_taxi_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "companies"
        payload = {'some': 'data'}
        company1 = getDataFromAPI(login_type, access_token, url, payload)
        companies = company1['Corporates']

        opr_url = settings.API_BASE_URL + "operators"
        operators = getDataFromAPI(login_type, access_token, opr_url, payload)
        operators = operators['Operators']

        opr_url = settings.API_BASE_URL + "agents"
        get_agents = getDataFromAPI(login_type, access_token, opr_url, payload)
        agents = get_agents['Agents']

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/taxi_bookings.html",{'bookings': booking,'booking_type':id,'corporates':companies,'operators':operators, 'agents':agents})
        else:
            return render(request, "Agent/taxi_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_taxi_booking(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url_taxi = settings.API_BASE_URL + "taxi_types"
        taxies = getDataFromAPI(login_type, access_token, url_taxi, payload)
        taxies = taxies['taxi_types']

        url_cities = settings.API_BASE_URL + "cities"
        taxies1 = getDataFromAPI(login_type, access_token, url_cities, payload)
        cities = taxies1['Cities']

        if company['success'] == 1:
            booking = company['Bookings']
            print(booking[0]['no_of_seats'])
            no_emp = int(booking[0]['no_of_seats'])
            no_of_emp_rabge = {range(no_emp+1 , 6)}
            return render(request, "Agent/view_taxi_booking.html",{'bookings': booking, 'taxies':taxies, 'cities':cities,'no_of_emp_rabge':range(no_emp , 6)})
        else:
            return render(request, "Agent/view_taxi_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def assign_operator_taxi_boooking(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        current_url = request.POST.get('current_url', '')
        booking_id = request.POST.get('booking_id', '')
        operator_id = request.POST.get('operator_id', '')
        user_id = request.POST.get('user_id', '')
        operator_contact = request.POST.get('operator_contact', '')
        operator_email = request.POST.get('operator_email', '')
        operator_package_id = request.POST.get('operator_package_id', '')

        url = settings.API_BASE_URL + "assign_operator_taxi_boooking"
        payload = {'operator_id':operator_id,'operator_email':operator_email,'operator_contact':operator_contact,'booking_id': booking_id,
        'user_id':user_id,'login_type':login_type, 'operator_package_id':operator_package_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, "Operator Assign Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.success(request, "Failed To Assign Operator")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def add_taxi_booking(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')

            global booking_email
            booking_email = ''
            if request.FILES:
                    file_up = request.FILES.get('email_attachment',False)
                    if file_up:
                        file_up = request.FILES['email_attachment']
                        booking_email = file_upload_get_path(file_up)
                    else:
                        booking_email = None
            else:
                booking_email = None

            bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
            corporate_id = request.POST.get('corporate_id', '')
            entity_id = request.POST.get('entity_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            tour_type = request.POST.get('tour_type', '')
            pickup_city = request.POST.get('pickup_city', '')
            pickup_location = request.POST.get('pickup_location', '')
            drop_location = request.POST.get('drop_location', '')
            pickup_datetime = request.POST.get('pickup_datetime', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            taxi_type = request.POST.get('taxi_type', '')
            package_id = request.POST.get('package_id', '')
            no_of_days = request.POST.get('no_of_days', '')
            actual_city_id = request.POST.get('current_city_id', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')
            is_email = request.POST.get('is_email', '')
            is_sms = request.POST.get('is_sms', '')

            if tour_type == 1 or tour_type == '1':
                url_add_city = settings.API_BASE_URL + "add_city_name"
                pickup_details = [x.strip() for x in pickup_city.split(',')]
                city_data = {'login_type': login_type, 'access_token': access_token, 'city_name': pickup_details[0], 'state_id': '1'}
                city_id = getDataFromAPI(login_type, access_token, url_add_city, city_data)
                for conty_id in city_id['id']:
                    actual_city_id = conty_id['id']

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'entity_id':entity_id,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'tour_type':tour_type,'pickup_city':actual_city_id,
                       'pickup_location':pickup_location,'drop_location':drop_location,'pickup_datetime':pickup_datetime+':00','booking_datetime':booking_datetime+':00','taxi_type':taxi_type,
                       'package_id':package_id,'no_of_days':no_of_days,'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,
                       'is_sms':is_sms,'is_email':is_email,'assessment_code': assessment_code, 'assessment_city_id': assessment_city_id,
                       'booking_email':booking_email,'bta_code_travel_req_no':bta_code_travel_req_no}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_taxi_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
            print(booking)
            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/agents/taxi-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Taxi Booking Not Added Successfully..!')
                return HttpResponseRedirect("/agents/taxi-bookings/2", {'message': "Operation Successfully"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "companies"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            companies = company['Corporates']

            url_city = settings.API_BASE_URL + "city_by_package"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities['Cities']

            url_taxi = settings.API_BASE_URL + "taxi_types"
            taxies = getDataFromAPI(login_type, access_token, url_taxi, payload)
            taxies = taxies['taxi_types']

            if id:
                return render(request, 'Agent/add_taxi_booking.html', {'companies':companies,'cities':cities,'taxies':taxies})
            else:
                return render(request, 'Agent/add_taxi_booking.html', {'companies':companies,'cities':cities,'taxies':taxies})
        else:
            return HttpResponseRedirect("/agents/login")


def accept_taxi_booking(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        booking_id = request.POST.get('booking_id', '')
        user_id = request.POST.get('user_id', '')
        accept_id = request.POST.get('accept_id', '')
        reject_id = request.POST.get('reject_id', '')
        current_url = request.POST.get('current_url', '')
        user_comment = request.POST.get('user_comment', '')

        url = ""
        if accept_id:
            url = settings.API_BASE_URL + "accept_taxi_booking"
            operation_message ="Taxi Booking Accepted Successfully..!"

        if reject_id:
            url = settings.API_BASE_URL + "reject_taxi_booking"
            operation_message = "Taxi Booking Rejected Successfully..!"

        payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment,'accept_id':accept_id,'reject_id':reject_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, operation_message)
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Failed to Accept Taxi Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")


def assign_taxi_booking(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')

            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')

            vendor_booking_id = request.POST.get('vendor_booking_id', '')
            operator_package_id = request.POST.get('operator_package_id', '')
            operator_id = request.POST.get('operator_id', '')
            driver_contact = request.POST.get('driver_contact', '')
            driver_id = request.POST.get('driver_id', '')
            taxi_id = request.POST.get('taxi_id', '')
            taxi_types = request.POST.get('taxi_types', '')
            taxi_model = request.POST.get('taxi_model', '')
            tour_type = request.POST.get('tour_typ_save', '')

            taxi_type_id =0
            taxi_model_id= 0
            taxi_act_id =0
            oper_id= 0
            driver_id_id=0

            is_client_sms = request.POST.get('is_client_sms', '')
            is_client_email = request.POST.get('is_client_email', '')
            is_driver_sms = request.POST.get('is_driver_sms', '')

            url = settings.API_BASE_URL + "assign_taxi_booking"
            payload = {'vendor_booking_id':vendor_booking_id,'is_client_sms':is_client_sms,
                       'is_client_email':is_client_email,'is_driver_sms':is_driver_sms,'operator_package_id':operator_package_id,
                       'booking_id': booking_id,'user_id':user_id,'user_type':login_type , 'operator_id': operator_id ,
                       'driver_contact': driver_contact , 'driver_id': driver_id , 'taxi_id': taxi_id , 'taxi_types': taxi_types , 'taxi_model': taxi_model , 'tour_type': tour_type , 'taxi_type_id': taxi_type_id , 'taxi_model_id': taxi_model_id }
            print(payload)
            print("aSSIGN TAXI")
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Taxi Booking Assigned..!')
                return HttpResponseRedirect("/agents/taxi-bookings/3", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Assign Taxi Booking..!')
                return HttpResponseRedirect("/agents/taxi-bookings/3", {'message': "Operation Failed"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'booking_id': id, 'service_type':'1'}
            opr_url = settings.API_BASE_URL + "get_operators_by_service_type"
            operators = getDataFromAPI(login_type, access_token, opr_url, payload)
            operators = operators['Operators']

            drivers_url = settings.API_BASE_URL + "operator_drivers"
            operator_drivers = getDataFromAPI(login_type, access_token, drivers_url, payload)
            operator_drivers = operator_drivers['Drivers']

            url_taxi = settings.API_BASE_URL + "taxi_models"
            taxis = getDataFromAPI(login_type, access_token, url_taxi, payload)
            models = taxis['Models']

            url_taxi_types = settings.API_BASE_URL + "taxi_types"
            url_taxi_types = getDataFromAPI(login_type, access_token, url_taxi_types, payload)
            taxi_types = url_taxi_types['taxi_types']

            url_taxis = settings.API_BASE_URL + "taxis"
            url_taxis = getDataFromAPI(login_type, access_token, url_taxis, payload)
            taxis = url_taxis['Taxis']

            url = settings.API_BASE_URL + "view_taxi_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)
            booking = booking['Bookings']

            url1 = settings.API_BASE_URL + "get_operator_package"
            payload1 = {'operator_id': 0}
            booking1 = getDataFromAPI(login_type, access_token, url1, payload1)
            opr_rates = booking1['Rate']

            return render(request, 'Agent/assign_taxi_booking.html',
                          {'bookings': booking, 'operators': operators, 'operator_drivers': operator_drivers,
                           'models': models,'taxi_types':taxi_types,'taxis':taxis,'opr_rates':opr_rates})

    else:
        return HttpResponseRedirect("/agents/login")


def add_taxi_invoice(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')

            global payin_slip
            email_attachment = request.POST.get('payin_slip', '')
            print(email_attachment)
            if email_attachment:
                payin_slip = email_attachment
                pass
            else:
                if request.FILES:
                    print("in file")
                    file_up = request.FILES.get('payin_slip', False)
                    if file_up:
                        file_up = request.FILES['payin_slip']
                        payin_slip = file_upload_get_path(file_up)
                    else:
                        payin_slip = None
                else:
                    payin_slip = None

            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')

            tax_on_management_fee = request.POST.get('tax_on_management_fee', '')
            tax_on_management_fee_percentage = request.POST.get('tax_on_management_fee_percentage', '')
            management_fee_igst = request.POST.get('management_fee_igst', '')
            management_fee_cgst= request.POST.get('management_fee_cgst', '')
            management_fee_sgst = request.POST.get('management_fee_sgst', '')
            management_fee_igst_rate = request.POST.get('management_fee_igst_rate', '')
            management_fee_cgst_rate = request.POST.get('management_fee_cgst_rate', '')
            management_fee_sgst_rate = request.POST.get('management_fee_sgst_rate', '')
            igst_amount = request.POST.get('igst_amount', '')
            cgst_amount = request.POST.get('cgst_amount', '')
            sgst_amount = request.POST.get('sgst_amount', '')

            hours_done = request.POST.get('hours_done')
            allowed_hours = request.POST.get('allowed_hours')
            extra_hours = request.POST.get('extra_hours')
            charge_hour = request.POST.get('charge_hour')
            days = request.POST.get('days')
            start_km = request.POST.get('start_km')
            end_km = request.POST.get('end_km')
            kms_done = request.POST.get('kms_done')
            allowed_kms = request.POST.get('allowed_kms')
            extra_kms = request.POST.get('extra_kms')
            extra_km_rate = request.POST.get('extra_km_rate')

            cotrav_billing_entity = request.POST.get('cotrav_billing_entity')
            bb_entity = request.POST.get('bb_entity')
            radio_rate = request.POST.get('radio_rate')

            base_rate = request.POST.get('base_rate')
            extra_hr_charges = request.POST.get('extra_hr_charges')
            extra_km_charges = request.POST.get('extra_km_charges')
            driver_allowance = request.POST.get('driver_allowance')
            total_excluding_tax = request.POST.get('total_excluding_tax')
            other_charges = request.POST.get('other_charges')
            total = request.POST.get('total')
            sub_total = request.POST.get('sub_total')

            url = settings.API_BASE_URL + "add_taxi_invoice"
            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'tax_on_management_fee':tax_on_management_fee,'tax_on_management_fee_percentage':tax_on_management_fee_percentage,
                      'management_fee_igst':management_fee_igst,'management_fee_cgst':management_fee_cgst,'management_fee_sgst':management_fee_sgst,'management_fee_igst_rate':management_fee_igst_rate,
                       'management_fee_sgst_rate':management_fee_sgst_rate,'igst_amount':igst_amount,'cgst_amount':cgst_amount,'sgst_amount':sgst_amount,'hours_done':hours_done,
                       'extra_hours':extra_hours,'charge_hour':charge_hour,'days':days,'start_km':start_km,'end_km':end_km,'kms_done':kms_done,'allowed_kms':allowed_kms,
                       'cotrav_billing_entity':cotrav_billing_entity,'bb_entity':bb_entity,'radio_rate':radio_rate,'base_rate':base_rate,'extra_hr_charges':extra_hr_charges,
                       'extra_km_charges':extra_km_charges,'driver_allowance':driver_allowance,'total_excluding_tax':total_excluding_tax,'other_charges':other_charges,
                       'total':total,'sub_total':sub_total,'management_fee_cgst_rate':management_fee_cgst_rate,'extra_kms':extra_kms,'extra_km_rate':extra_km_rate,'allowed_hours':allowed_hours,
                       'payin_slip':payin_slip}
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Taxi Invoice Added Successfully..!')
                return HttpResponseRedirect("/agents/taxi-bookings/3", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Added Taxi Invoice..!')
                return HttpResponseRedirect("/agents/taxi-bookings/3", {'message': "Operation Failed"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "view_taxi_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)
            booking = booking['Bookings']

            billing_entity = booking[0]['billing_entity_id']
            if billing_entity:
                url_hotels = settings.API_BASE_URL + "view_billing_entitie"
                payload = {'entity_id': billing_entity}
                bus_entity = getDataFromAPI(login_type, access_token, url_hotels, payload)
                bus_operators = bus_entity['Entitys']
                be_name = bus_operators[0]['entity_name']
                be_gst = bus_operators[0]['gst_id']
            else:
                be_name = ""
                be_gst = ""

            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            c_entity = getDataFromAPI(login_type, access_token, url, payload)
            c_entity = c_entity['Enitity']

            return render(request, 'Agent/add_taxi_invoice.html', {'bookings': booking,'be_name':be_name,'be_gst':be_gst,'c_entitys':c_entity })

    else:
        return HttpResponseRedirect("/agents/login")


def add_new_dutyslip(request):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')

            global payin_slip
            email_attachment = request.POST.get('payin_slip', '')
            print(email_attachment)
            if email_attachment:
                payin_slip = email_attachment
                pass
            else:
                if request.FILES:
                    print("in file")
                    file_up = request.FILES.get('payin_slip', False)
                    if file_up:
                        file_up = request.FILES['payin_slip']
                        payin_slip = file_upload_get_path(file_up)
                    else:
                        payin_slip = None
                else:
                    payin_slip = None

            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')

            url = settings.API_BASE_URL + "add_new_dutyslip"
            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'payin_slip':payin_slip}
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Duty Slip Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Added Duty Slip..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")
    
    
def add_new_employee(request):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')

            billing_entity_id = request.POST.get('billing_entity_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            user_id = request.POST.get('user_id', '')
            core_employee_id = request.POST.get('core_employee_id', '')
            employee_cid = request.POST.get('employee_cid', '')

            employee_name = request.POST.get('employee_name', '')
            employee_email = request.POST.get('employee_email', '')
            username = request.POST.get('employee_email', '')
            employee_contact = request.POST.get('employee_contact', '')

            date_of_birth = request.POST.get('date_of_birth', '')
            if date_of_birth and date_of_birth != 'None':
                age = calculate_age(date_of_birth)
            else:
                age = 0

            gender = request.POST.get('gender')

            url = settings.API_BASE_URL + "add_employee"
            payload = {'corporate_id':corporate_id,'spoc_id': spoc_id, 'billing_entity_id':billing_entity_id, 'user_id':user_id, 'user_type':login_type,'core_employee_id':core_employee_id,'employee_cid':employee_cid,
            'employee_name':employee_name,'employee_email':employee_email,'username':username, 'employee_contact':employee_contact, 'date_of_birth':date_of_birth, 'age':age, 'gender':gender}
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Employee Added Successfully..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Added Employee..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")

############################## BUS  ######################################


def bus_bookings(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_bus_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "companies"
        payload = {'some': 'data'}
        company1 = getDataFromAPI(login_type, access_token, url, payload)
        companies = company1['Corporates']

        opr_url = settings.API_BASE_URL + "agents"
        get_agents = getDataFromAPI(login_type, access_token, opr_url, payload)
        agents = get_agents['Agents']

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/bus_bookings.html",{'bookings': booking,'booking_type':id,'corporates':companies,'agents':agents})
        else:
            return render(request, "Agent/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_bus_booking(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Agent/view_bus_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def add_bus_booking(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')

            global booking_email
            booking_email = ''
            if request.FILES:
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None
            bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            bus_type = request.POST.get('bus_type', '')
            bus_type2 = request.POST.get('bus_type2', '')
            bus_type3 = request.POST.get('bus_type3', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_bus = request.POST.get('preferred_bus', '')
            preferred_board_point = request.POST.get('preferred_board_point', '')
            preferred_drop_point = request.POST.get('preferred_drop_point', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')
            is_email = request.POST.get('is_email', '')
            is_sms = request.POST.get('is_sms', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'bus_type':bus_type,'booking_datetime':booking_datetime+':00',
            'journey_datetime':journey_datetime+':00','journey_datetime_to':journey_datetime_to+':00','entity_id':entity_id,'reason_booking':reason_booking,'no_of_seats':no_of_seats,
                       'preferred_bus':preferred_bus,'employees':employees,'is_email':is_email,'is_sms':is_sms,'bta_code_travel_req_no':bta_code_travel_req_no,
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city_id,'booking_email':booking_email,
                       'preferred_board_point':preferred_board_point, 'preferred_drop_point':preferred_drop_point,'bus_type2':bus_type2,'bus_type3':bus_type3}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_bus_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
            print(booking)
            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/agents/bus-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Bus Booking..!')
                return HttpResponseRedirect("/agents/bus-bookings/2", {'message': "Operation Failed"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "companies"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            companies = company['Corporates']

            if id:
                return render(request, 'Agent/add_bus_booking.html', {'companies':companies})
            else:
                return render(request, 'Agent/add_bus_booking.html', {'companies':companies})
        else:
            return HttpResponseRedirect("/agents/login")



def accept_bus_booking(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            user_comment = request.POST.get('user_comment', '')

            current_url = request.POST.get('current_url', '')

            url = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "accept_bus_booking"
                operation_message="Bus Booking Accepted Successfully..!"

            if reject_id == '1':
                url = settings.API_BASE_URL + "reject_bus_booking"
                operation_message="Bus Booking Rejected Successfully..!"

            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Accept Bus Booking..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
        else:
            return render(request, "Agent/bus_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def assign_bus_booking(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            client_ticket_path=''
            vender_ticket_path=''
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')

            ticket_no = request.POST.get('ticket_no', '')
            pnr_no = request.POST.get('pnr_no', '')
            assign_bus_type_id = request.POST.get('assign_bus_type_id', '')
            seat_no= request.POST.getlist('seat_no', '')
            employee_id= request.POST.getlist('employee_id', '')
            portal_used = request.POST.get('portal_used', '')
            operator_name = request.POST.get('operator_name', '')
            operator_contact = request.POST.get('operator_contact', '')
            boarding_point = request.POST.get('boarding_point', '')
            boarding_datetime = request.POST.get('boarding_datetime', '')
            droping_point = request.POST.get('droping_point', '')

            is_client_sms = request.POST.get('is_client_sms', '')
            is_client_email = request.POST.get('is_client_email', '')
            is_driver_sms = request.POST.get('is_driver_sms', '')

            ticket_price = request.POST.get('ticket_price', '')
            management_fee = request.POST.get('management_fee', '')
            tax_mng_amt = request.POST.get('tax_mng_amt', '')
            tax_on_management_fee = request.POST.get('tax_on_management_fee', '')
            tax_on_management_fee_percentage = request.POST.get('tax_on_management_fee_percentage', '')
            sub_total = request.POST.get('sub_total', '')
            management_fee_igst = request.POST.get('management_fee_igst', '')
            management_fee_cgst = request.POST.get('management_fee_cgst', '')
            management_fee_sgst = request.POST.get('management_fee_sgst', '')
            management_fee_igst_rate = request.POST.get('management_fee_igst_rate', '')
            management_fee_cgst_rate = request.POST.get('management_fee_cgst_rate', '')
            management_fee_sgst_rate = request.POST.get('management_fee_sgst_rate', '')
            cgst = request.POST.get('cgst', '')
            sgst = request.POST.get('sgst', '')
            igst = request.POST.get('igst', '')

            oper_ticket_price = request.POST.get('oper_ticket_price', '')

            oper_commission = request.POST.get('operator_commission', '')

            oper_commission_type = request.POST.get('oper_commission_type', '')

            oper_cotrav_billing_entity = request.POST.get('oper_cotrav_billing_entity', '')
            cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')

            oper_cgst = request.POST.get('oper_cgst', '')
            oper_sgst = request.POST.get('oper_sgst', '')
            oper_igst = request.POST.get('oper_igst', '')

            igst_amount = request.POST.get('igst_amount', '')
            cgst_amount = request.POST.get('cgst_amount', '')
            sgst_amount = request.POST.get('sgst_amount', '')

            if request.FILES:
                    file_up = request.FILES.get('busticketToUpload',False)
                    if file_up:
                        save_path = os.path.join(settings.MEDIA_ROOT, 'bus/client_ticket', str(file_up))
                        path1 = default_storage.save(save_path, request.FILES['busticketToUpload'])
                        client_ticket_path = path1
                    else:
                        client_ticket_path = None

                    file_up2 = request.FILES.get('vendorticketToUpload',False)
                    if file_up2:
                        save_path = os.path.join(settings.MEDIA_ROOT, 'bus/vender_ticket', str(file_up2))
                        path2 = default_storage.save(save_path, request.FILES['vendorticketToUpload'])
                        vender_ticket_path = path2
                    else:
                        vender_ticket_path = None

            client_ticket = request.POST.get('client_ticket')

            url = settings.API_BASE_URL + "assign_bus_booking"
            payload = {'ticket_no': ticket_no, 'pnr_no': pnr_no, 'assign_bus_type_id': assign_bus_type_id,
                       'seat_no': seat_no, 'portal_used': portal_used,'employee_id':employee_id
                , 'operator_name': operator_name, 'operator_contact': operator_contact,
                       'boarding_point': boarding_point, 'boarding_datetime': boarding_datetime,
                       'booking_id': booking_id, 'user_id': user_id, 'user_type': login_type,
                       'ticket_price': ticket_price, 'management_fee': management_fee, 'tax_mng_amt': tax_mng_amt
                , 'tax_on_management_fee': tax_on_management_fee,
                       'tax_on_management_fee_percentage': tax_on_management_fee_percentage, 'sub_total': sub_total,
                       'management_fee_igst': management_fee_igst, 'management_fee_cgst': management_fee_cgst,
                       'management_fee_sgst': management_fee_sgst, 'management_fee_igst_rate': management_fee_igst_rate,
                       'management_fee_cgst_rate': management_fee_cgst_rate,
                       'management_fee_sgst_rate': management_fee_sgst_rate,'cgst':cgst,'sgst':sgst,'igst':igst,
                       'oper_ticket_price': oper_ticket_price, 'oper_commission':oper_commission ,
                       'oper_commission_type':oper_commission_type ,'oper_cotrav_billing_entity':oper_cotrav_billing_entity ,
                       'oper_cgst':oper_cgst,'oper_sgst':oper_sgst,'oper_igst':oper_igst,'client_ticket_path':client_ticket_path,
                       'vender_ticket_path':vender_ticket_path,'cotrav_billing_entity':cotrav_billing_entity,'is_client_sms':is_client_sms,
                       'is_client_email':is_client_email,'is_driver_sms':is_driver_sms,'igst_amount': igst_amount,'cgst_amount': cgst_amount,
                       'sgst_amount': sgst_amount,'client_ticket':client_ticket,'droping_point':droping_point}

            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Bus Booking Assign successfully..!')
                return HttpResponseRedirect("/agents/bus-bookings/3", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Assign Bus Booking..1')
                return HttpResponseRedirect("/agents/bus-bookings/3", {'message': "Operation Failed"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'booking_id': id}

            url = settings.API_BASE_URL + "view_bus_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)
            booking = booking['Bookings']

            billing_entity = booking[0]['billing_entity_id']
            if billing_entity:
                url_hotels = settings.API_BASE_URL + "view_billing_entitie"
                payload = {'entity_id': billing_entity}
                bus_entity = getDataFromAPI(login_type, access_token, url_hotels, payload)
                bus_operators = bus_entity['Entitys']
                be_name = bus_operators[0]['entity_name']
                be_gst = bus_operators[0]['gst_id']
            else:
                be_name = ""
                be_gst = ""

            url_hotels = settings.API_BASE_URL + "get_operators_by_service_type"
            payload = {'service_type': 4}
            hotels = getDataFromAPI(login_type, access_token, url_hotels, payload)
            bus_operators = hotels['Operators']

            url_hotels_port = settings.API_BASE_URL + "bus_booking_portals"
            payload = {}
            hotels1 = getDataFromAPI(login_type, access_token, url_hotels_port, payload)
            bus_portal = hotels1['Portals']

            ######################
            payload = {}
            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            c_entity = getDataFromAPI(login_type, access_token, url, payload)
            c_entity = c_entity['Enitity']

            #########################

            return render(request, 'Agent/assign_bus_booking.html', {'bookings': booking,'bus_operators':bus_operators,'c_entitys':c_entity,'be_name':be_name,'be_gst':be_gst,'bus_portal':bus_portal})

    else:
        return HttpResponseRedirect("/agents/login")


############################## TRAIN  ######################################


def train_bookings(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_train_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "companies"
        payload = {'some': 'data'}
        company1 = getDataFromAPI(login_type, access_token, url, payload)
        companies = company1['Corporates']

        opr_url = settings.API_BASE_URL + "agents"
        get_agents = getDataFromAPI(login_type, access_token, opr_url, payload)
        agents = get_agents['Agents']

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/train_bookings.html",{'bookings': booking,'booking_type':id,'corporates':companies,'agents':agents})
        else:
            return render(request, "Agent/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_train_booking(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Agent/view_train_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def add_train_booking(request,id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')

            global booking_email
            booking_email = ''
            if request.FILES:
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None
            bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
            corporate_id = request.POST.get('corporate_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            from_location = request.POST.get('from', '')
            to_location = request.POST.get('to', '')
            train_type = request.POST.get('train_type', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            journey_datetime = request.POST.get('journey_datetime', '')
            journey_datetime_to = request.POST.get('journey_datetime_to', '')
            entity_id = request.POST.get('entity_id', '')
            preferred_train = request.POST.get('preferred_train', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')

            is_email = request.POST.get('is_email', '')
            is_sms = request.POST.get('is_sms', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'login_type':login_type,'user_id':user_id,'access_token':access_token,'corporate_id': corporate_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'from':from_location,'to':to_location,'train_type':train_type,'booking_datetime':booking_datetime+':00',
            'journey_datetime':journey_datetime+':00','journey_datetime_to':journey_datetime_to+':00','entity_id':entity_id,'reason_booking':reason_booking,'no_of_seats':no_of_seats,
                       'preferred_train':preferred_train,'employees':employees,'is_email':is_email,'is_sms':is_sms,'assessment_code': assessment_code,
                       'assessment_city_id': assessment_city_id,'booking_email':booking_email,'bta_code_travel_req_no':bta_code_travel_req_no}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_train_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
            print(booking)
            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/agents/train-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Add Train Booking..!')
                return HttpResponseRedirect("/agents/train-bookings/2", {'message': "Operation Failed"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "companies"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            companies = company['Corporates']

            url_train = settings.API_BASE_URL + "train_types"
            trains = getDataFromAPI(login_type, access_token, url_train, payload)
            types = trains['Types']

            # url_railway_stations = settings.API_BASE_URL + "railway_stations"
            # trains1 = getDataFromAPI(login_type, access_token, url_railway_stations, payload)
            # railway_stations = trains1['Stations']
            railway_stations = ""

            if id:

                return render(request, 'Agent/add_train_booking.html', {'companies':companies,'types':types,'railway_stations':railway_stations})
            else:

                return render(request, 'Agent/add_train_booking.html', {'companies':companies,'types':types,'railway_stations':railway_stations})
        else:
            return HttpResponseRedirect("/agents/login")


def accept_train_booking(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            current_url = request.POST.get('current_url', '')
            user_comment = request.POST.get('user_comment', '')

            url = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "accept_train_booking"
                operation_message="Train Booking Accepted successfully..!"

            if reject_id == '1':
                url = settings.API_BASE_URL + "reject_train_booking"
                operation_message="Train Booking Rejected successfully..!"

            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Accept Train Booking..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
        else:
            return render(request, "Agent/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def assign_train_booking(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')

            train_name = request.POST.get('train_name', '')
            ticket_no = request.POST.get('ticket_no', '')
            pnr_no = request.POST.get('pnr_no', '')
            assign_bus_type_id = request.POST.get('assign_bus_type_id', '')
            seat_no= request.POST.getlist('seat_no', '')
            coach= request.POST.getlist('coach', '')
            employee_id= request.POST.getlist('employee_id', '')
            portal_used = request.POST.get('portal_used', '')
            quota_used = request.POST.get('quota_used', '')
            operator_name = request.POST.get('operator_name', '')
            operator_contact = request.POST.get('operator_contact', '')
            boarding_point = request.POST.get('boarding_point', '')
            boarding_datetime = request.POST.get('boarding_datetime', '')
            #boarding_datetime = datetime.strptime(boarding_datetime, '%d/%m/%Y %H:%M:%S')

            is_client_sms = request.POST.get('is_client_sms', '')
            is_client_email = request.POST.get('is_client_email', '')
            is_driver_sms = request.POST.get('is_driver_sms', '')

            client_ticket_path = ''

            ticket_price = request.POST.get('ticket_price', '')
            management_fee = request.POST.get('management_fee', '')
            tax_mng_amt = request.POST.get('tax_mng_amt', '')
            tax_on_management_fee = request.POST.get('tax_on_management_fee', '')
            tax_on_management_fee_percentage = request.POST.get('tax_on_management_fee_percentage', '')
            sub_total = request.POST.get('sub_total', '')
            management_fee_igst = request.POST.get('management_fee_igst', '')
            management_fee_cgst = request.POST.get('management_fee_cgst', '')
            management_fee_sgst = request.POST.get('management_fee_sgst', '')
            management_fee_igst_rate = request.POST.get('management_fee_igst_rate', '')
            management_fee_cgst_rate = request.POST.get('management_fee_cgst_rate', '')
            management_fee_sgst_rate = request.POST.get('management_fee_sgst_rate', '')
            cgst = request.POST.get('cgst', '')
            sgst = request.POST.get('sgst', '')
            igst = request.POST.get('igst', '')
            cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')


            if request.FILES:
                file_up = request.FILES.get('busticketToUpload', False)
                if file_up:
                    save_path = os.path.join(settings.MEDIA_ROOT, 'train/client_ticket', str(file_up))
                    path1 = default_storage.save(save_path, request.FILES['busticketToUpload'])
                    client_ticket_path = path1
                else:
                    client_ticket_path = None

            url = settings.API_BASE_URL + "assign_train_booking"


            payload = {'ticket_no':ticket_no,'pnr_no':pnr_no,'assign_bus_type_id':assign_bus_type_id,'seat_no':seat_no,'coach':coach,'portal_used':portal_used
                ,'operator_name':operator_name,'operator_contact':operator_contact,'boarding_point':boarding_point,'boarding_datetime':boarding_datetime,
                       'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'train_name':train_name , 'ticket_price': ticket_price, 'management_fee': management_fee, 'tax_mng_amt': tax_mng_amt
                , 'tax_on_management_fee': tax_on_management_fee,'quota_used':quota_used,
                       'tax_on_management_fee_percentage': tax_on_management_fee_percentage, 'sub_total': sub_total,
                       'management_fee_igst': management_fee_igst, 'management_fee_cgst': management_fee_cgst,
                       'management_fee_sgst': management_fee_sgst, 'management_fee_igst_rate': management_fee_igst_rate,
                       'management_fee_cgst_rate': management_fee_cgst_rate,'cotrav_billing_entity':cotrav_billing_entity,
                       'management_fee_sgst_rate': management_fee_sgst_rate,'cgst':cgst,'sgst':sgst,'igst':igst,'client_ticket_path':client_ticket_path,
                       'is_client_sms':is_client_sms,'is_client_email':is_client_email,'employee_id':employee_id}
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Train Booking assigned Successfully..!')
                return HttpResponseRedirect("/agents/train-bookings/3", {'message': "Operation Successfully"})

            else:
                messages.error(request, 'Failed To assign Train Booking..!')
                return HttpResponseRedirect("/agents/train-bookings/3", {'message': "Operation Failed"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'booking_id': id}

            url = settings.API_BASE_URL + "view_train_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)
            booking = booking['Bookings']
            print(booking)

            for pasg_id in booking[0]['Passangers']:

                print("passenger id is ")
                print(pasg_id['id'])

            url_train = settings.API_BASE_URL + "train_types"
            trains = getDataFromAPI(login_type, access_token, url_train, payload)
            types = trains['Types']

            url_train1 = settings.API_BASE_URL + "irctc_accounts"
            trains1 = getDataFromAPI(login_type, access_token, url_train1, payload)
            accounts = trains1['Accounts']

            payload = {}
            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            c_entity = getDataFromAPI(login_type, access_token, url, payload)
            c_entity = c_entity['Enitity']

            billing_entity = booking[0]['billing_entity_id']
            if billing_entity:
                url_hotels = settings.API_BASE_URL + "view_billing_entitie"
                payload = {'entity_id': billing_entity}
                bus_entity = getDataFromAPI(login_type, access_token, url_hotels, payload)
                bus_operators = bus_entity['Entitys']
                be_name = bus_operators[0]['entity_name']
                be_gst = bus_operators[0]['gst_id']
            else:
                be_name = ""
                be_gst = ""

            return render(request, 'Agent/assign_train_booking.html',{'bookings': booking, 'accounts':accounts,'types':types,'c_entitys':c_entity,'be_gst':be_gst,'be_name':be_name})

    else:
        return HttpResponseRedirect("/agents/login")

############################## HOTELS  ######################################


def add_hotel_booking(request, id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')

            global booking_email
            booking_email = ''
            if request.FILES:
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None
            # spoc_id = request.POST.get('spoc_id', '')
            current_url = request.POST.get('current_url', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no')
            corporate_id = request.POST.get('corporate_id')
            booking_email_copy = request.POST.get('booking_email_copy')
            from_city = request.POST.get('from_city')
            city_area = request.POST.get('city_area')
            preferred_hotel_area = request.POST.get('preferred_hotel_area')
            check_in_date = request.POST.get('check_in_date')
            check_out_date = request.POST.get('check_out_date')
            room_type_priority1 = request.POST.get('room_type_priority1')
            room_type_priority2 = request.POST.get('room_type_priority2')
            room_occupancy = request.POST.get('room_occupancy')
            preferred_hotel = request.POST.get('preferred_hotel')
            booking_date = request.POST.get('booking_datetime')
            no_of_nights = request.POST.get('no_of_nights', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            billing_entity = request.POST.get('billing_entity')
            reason_for_booking = request.POST.get('reason_for_booking')
            is_email = request.POST.get('is_email', '')
            is_sms = request.POST.get('is_sms', '')
            no_of_seats = request.POST.get('no_of_seats')

            # post variables end

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1, no_of_emp):
                employees.append(request.POST.get('employee_id_' + str(i), ''))
                print(employees)

            payload = {'login_type': login_type, 'user_id': user_id, 'access_token': access_token,
                       'corporate_id': corporate_id, 'spoc_id': spoc_id, 'group_id': group_id,
                       'subgroup_id': subgroup_id, 'from_city_id': from_city, 'from_area_id': city_area,
                       'preferred_area': preferred_hotel_area, 'checkin_datetime': check_in_date+':00',
                       'checkout_datetime': check_out_date+':00', 'bucket_priority_1': room_type_priority1,
                       'bucket_priority_2': room_type_priority2, 'room_type_id': room_occupancy,
                       'preferred_hotel': preferred_hotel, 'booking_datetime': booking_date+':00',
                       'assessment_code': assessment_code, 'assessment_city_id': assessment_city_id,'bta_code_travel_req_no':bta_code_travel_req_no,
                       'billing_entity_id': billing_entity, 'employees': employees,'reason_booking':reason_for_booking,'no_of_seats':no_of_seats,
                       'is_email':is_email,'is_sms':is_sms,'booking_email':booking_email,'no_of_nights':no_of_nights}

            print(payload)
            url_taxi_booking = settings.API_BASE_URL + "add_hotel_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)
            if booking['success'] == 1:
                messages.success(request, str(booking['message']))

                return HttpResponseRedirect("/agents/hotel-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed to Accept Hotel Booking..!')

                return HttpResponseRedirect("/agents/hotel-bookings/2", {'message': "Operation Failed"})
        else:
                return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "companies"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            companies = company['Corporates']

            # url_city = settings.API_BASE_URL + "cities"
            # cities = getDataFromAPI(login_type, access_token, url_city, payload)
            # cities = cities['Cities']
            cities = ""

            if id:
                return render(request, 'Agent/add_hotel_booking.html', {'companies': companies, 'cities': cities, })
            else:
                return render(request, 'Agent/add_hotel_booking.html', {'companies': companies, 'cities': cities, })
        else:
            return HttpResponseRedirect("/agents/login")



def hotel_bookings(request, id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_hotel_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "companies"
        payload = {'some': 'data'}
        company1 = getDataFromAPI(login_type, access_token, url, payload)
        companies = company1['Corporates']
        
        opr_url = settings.API_BASE_URL + "agents"
        get_agents = getDataFromAPI(login_type, access_token, opr_url, payload)
        agents = get_agents['Agents']

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/hotel_bookings.html", {'bookings': booking, 'booking_type':id,'corporates':companies, 'agents':agents})
        else:
            return render(request, "Agent/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_hotel_booking(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Agent/view_hotel_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def accept_hotel_booking(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            current_url = request.POST.get('current_url', '')
            user_comment = request.POST.get('user_comment', '')

            url = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "accept_hotel_booking"
                operation_message = "Hotel Booking Accepted successfully..!"

            if reject_id == '1':
                url = settings.API_BASE_URL + "reject_hotel_booking"
                operation_message = "Hotel Booking Rejected successfully..!"

            payload = {'booking_id': booking_id, 'user_id': user_id, 'user_type': login_type, 'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed To Accept Hotel Booking..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
        else:
            return render(request, "Agent/hotel_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def assign_hotel_booking(request, id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            assign_hotel_id = request.POST.get('assign_hotel_id', '')
            no_of_days = request.POST.get('no_of_days', '')

            assign_room_type = request.POST.get('assign_room_type', '')
            is_ac_room = request.POST.get('is_ac_room', '')
            daily_brakefast = request.POST.get('daily_brakefast', '')
            is_prepaid = request.POST.get('is_prepaid', '')
            agent_booking_id = request.POST.get('agent_booking_id', '')
            comment = request.POST.get('comment', '')

            vender_ticket_path = ''

            client_ticket_path = ''

            total_room_price = request.POST.get('total_room_price', '')
            voucher_number = request.POST.get('voucher_number', '')
            portal_used = request.POST.get('portal_used', '')
            commission_earned = request.POST.get('commission_earned', '')

            user_id = request.POST.get('user_id')
            booking_id = request.POST.get('booking_id')

            ticket_price = request.POST.get('ticket_price', '')
            management_fee = request.POST.get('management_fee', '')
            tax_mng_amt = request.POST.get('tax_mng_amt', '')
            tax_on_management_fee = request.POST.get('tax_on_management_fee', '')
            tax_on_management_fee_percentage = request.POST.get('tax_on_management_fee_percentage', '')
            sub_total = request.POST.get('sub_total', '')
            management_fee_igst = request.POST.get('management_fee_igst', '')
            management_fee_cgst = request.POST.get('management_fee_cgst', '')
            management_fee_sgst = request.POST.get('management_fee_sgst', '')
            management_fee_igst_rate = request.POST.get('management_fee_igst_rate', '')
            management_fee_cgst_rate = request.POST.get('management_fee_cgst_rate', '')
            management_fee_sgst_rate = request.POST.get('management_fee_sgst_rate', '')
            cgst = request.POST.get('cgst', '')
            sgst = request.POST.get('sgst', '')
            igst = request.POST.get('igst', '')

            oper_ticket_price = request.POST.get('oper_ticket_price', '')

            oper_commission = request.POST.get('operator_commission', '')

            oper_commission_type = request.POST.get('oper_commission_type', '')

            oper_cotrav_billing_entity = request.POST.get('oper_cotrav_billing_entity', '')
            cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')

            oper_cgst = request.POST.get('oper_cgst', '')
            oper_sgst = request.POST.get('oper_sgst', '')
            oper_igst = request.POST.get('oper_igst', '')
            igst_amount = request.POST.get('igst_amount', '')
            cgst_amount = request.POST.get('cgst_amount', '')
            sgst_amount = request.POST.get('sgst_amount', '')

            is_client_sms = request.POST.get('is_client_sms', '')
            is_client_email = request.POST.get('is_client_email', '')

            if request.FILES:
                file_up = request.FILES.get('busticketToUpload', False)
                if file_up:
                    save_path = os.path.join(settings.MEDIA_ROOT, 'bus/client_ticket', str(file_up))
                    path1 = default_storage.save(save_path, request.FILES['busticketToUpload'])
                    client_ticket_path = path1
                else:
                    client_ticket_path = None

                file_up2 = request.FILES.get('vendorticketToUpload', False)
                if file_up2:
                    save_path = os.path.join(settings.MEDIA_ROOT, 'bus/vender_ticket', str(file_up2))
                    path2 = default_storage.save(save_path, request.FILES['vendorticketToUpload'])
                    vender_ticket_path = path2
                else:
                    vender_ticket_path = None

            client_ticket = request.POST.get('client_ticket')
            print(client_ticket)

            vender_ticket = request.POST.get('vender_ticket')
            print(vender_ticket)

            url = settings.API_BASE_URL + "assign_hotel_booking"

            payload = {'assign_hotel_id': assign_hotel_id,'assign_room_type':assign_room_type,'is_ac_room':is_ac_room,'daily_brakefast':daily_brakefast,
            'is_prepaid':is_prepaid,'agent_booking_id':agent_booking_id,'comment':comment,'user_id':user_id,'user_type': login_type,'total_room_price':total_room_price,
            'voucher_number':voucher_number,'portal_used':portal_used,'commission_earned':commission_earned,'booking_id':booking_id,
                       'ticket_price': ticket_price, 'management_fee': management_fee, 'tax_mng_amt': tax_mng_amt
                , 'tax_on_management_fee': tax_on_management_fee,'cotrav_billing_entity':cotrav_billing_entity,
                       'tax_on_management_fee_percentage': tax_on_management_fee_percentage, 'sub_total': sub_total,
                       'management_fee_igst': management_fee_igst, 'management_fee_cgst': management_fee_cgst,
                       'management_fee_sgst': management_fee_sgst, 'management_fee_igst_rate': management_fee_igst_rate,
                       'management_fee_cgst_rate': management_fee_cgst_rate,
                       'management_fee_sgst_rate': management_fee_sgst_rate, 'cgst': cgst, 'sgst': sgst, 'igst': igst,
                       'oper_ticket_price': oper_ticket_price, 'oper_commission': oper_commission,
                       'oper_commission_type': oper_commission_type,'client_ticket':client_ticket,'vender_ticket':vender_ticket,
                       'oper_cotrav_billing_entity': oper_cotrav_billing_entity,'no_of_days':no_of_days,
                       'oper_cgst': oper_cgst, 'oper_sgst': oper_sgst, 'oper_igst': oper_igst,
                       'client_ticket_path': client_ticket_path,'is_client_sms':is_client_sms,'is_client_email':is_client_email,
                       'vender_ticket_path': vender_ticket_path,'igst_amount': igst_amount,'cgst_amount': cgst_amount,'sgst_amount': sgst_amount
                       }
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)

            print(company)

            if company['success'] == 1:
                messages.success(request, 'Hotel Booking Assigned Successfully..!')
                return HttpResponseRedirect("/agents/hotel-bookings/3", {'message': "Operation Successfully"})

            else:
                messages.error(request, 'Failed To Assign hotel Booking..!')
                return HttpResponseRedirect("/agents/hotel-bookings/3", {'message': "Operation Failed"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'booking_id': id}

            url = settings.API_BASE_URL + "view_hotel_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)

            booking = booking['Bookings']

            billing_entity = booking[0]['billing_entity_id']
            if billing_entity:
                url_hotels = settings.API_BASE_URL + "view_billing_entitie"
                payload = {'entity_id': billing_entity}
                bus_entity = getDataFromAPI(login_type, access_token, url_hotels, payload)
                bus_operators = bus_entity['Entitys']
                be_name = bus_operators[0]['entity_name']
                be_gst = bus_operators[0]['gst_id']
            else:
                be_name = ""
                be_gst = ""

            url_room_types = settings.API_BASE_URL + "hotel_types"
            payload = {'booking_id': id}
            room_types = getDataFromAPI(login_type, access_token, url_room_types, payload)
            room_types = room_types['Types']

            url_room_types = settings.API_BASE_URL + "hotel_booking_portals"
            payload = {'booking_id': id}
            portals = getDataFromAPI(login_type, access_token, url_room_types, payload)
            portals = portals['Portals']

            url_hotels = settings.API_BASE_URL + "get_operators_by_service_type"
            payload = {'service_type': 6}
            hotels = getDataFromAPI(login_type, access_token, url_hotels, payload)
            hotel_operators = hotels['Operators']

            ######################
            payload = {}
            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            c_entity = getDataFromAPI(login_type, access_token, url, payload)
            c_entity = c_entity['Enitity']
            print(c_entity)

            #########################

            return render(request, 'Agent/assign_hotel_booking.html',{'bookings': booking,'portals':portals,'room_types':room_types,
             'c_entitys':c_entity,'hotel_operators':hotel_operators,'be_name':be_name,'be_gst':be_gst})

    else:
        return HttpResponseRedirect("/agents/login")


############################## FLIGHT  ######################################

def flight_bookings(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_flight_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "companies"
        payload = {'some': 'data'}
        company1 = getDataFromAPI(login_type, access_token, url, payload)
        companies = company1['Corporates']

        opr_url = settings.API_BASE_URL + "agents"
        get_agents = getDataFromAPI(login_type, access_token, opr_url, payload)
        agents = get_agents['Agents']

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/flight_bookings.html",{'bookings': booking,'booking_type':id,'corporates':companies,'agents':agents})
        else:
            return render(request, "Agent/flight_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")



def view_flight_booking(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url_access = settings.API_BASE_URL + "get_airports"
        data = getDataFromAPI(login_type, access_token, url_access, payload)
        airports = data['Airports']

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/view_flight_booking.html",{'bookings': booking, 'airports':airports})
        else:
            return render(request, "Agent/view_flight_booking.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def add_flight_booking(request, id):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            user_id = request.POST.get('user_id', '')

            global booking_email
            booking_email =''
            if request.FILES:
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None

            bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
            corporate_id = request.POST.get('corporate_id', '')
            entity_id = request.POST.get('entity_id', '')
            spoc_id = request.POST.get('spoc_id', '')
            spoc_details = [x.strip() for x in spoc_id.split(',')]

            spoc_id = spoc_details[0]
            group_id = spoc_details[1]
            subgroup_id = spoc_details[2]

            usage_type = request.POST.get('usage_type', '')
            trip_type = request.POST.get('trip_type', '')
            seat_type = request.POST.get('seat_type', '')
            from_city = request.POST.get('from_city', '')
            to_city = request.POST.get('to_city', '')
            booking_datetime = request.POST.get('booking_datetime', '')
            departure_datetime = request.POST.get('departure_datetime', '')
            return_datetime = request.POST.get('return_datetime', '')
            preferred_flight = request.POST.get('preferred_flight', '')
            assessment_code = request.POST.get('assessment_code', '')
            billing_entity_id = request.POST.get('billing_entity_id', '')
            reason_booking = request.POST.get('reason_booking', '')
            no_of_seats = request.POST.get('no_of_seats', '')
            assessment_code = request.POST.get('assessment_code', '')
            assessment_city_id = request.POST.get('assessment_city_id', '')

            is_email = request.POST.get('is_email', '')
            is_sms = request.POST.get('is_sms', '')

            employees = []
            no_of_emp = int(no_of_seats) + 1
            for i in range(1,no_of_emp):
                employees.append(request.POST.get('employee_id_'+str(i), ''))
                print(employees)

            payload = {'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'entity_id':entity_id,'spoc_id':spoc_id,'group_id':group_id,
                       'subgroup_id':subgroup_id,'usage_type':usage_type,'trip_type':trip_type,'seat_type':seat_type,'from_city':from_city,'to_city':to_city,
                       'booking_datetime':booking_datetime+':00','departure_datetime':departure_datetime,'preferred_flight':preferred_flight,
                       'reason_booking':reason_booking,'no_of_seats':no_of_seats,'employees':employees,'billing_entity_id':billing_entity_id,
                       'is_email':is_email,'is_sms':is_sms,'assessment_code': assessment_code, 'assessment_city_id': assessment_city_id,
                       'booking_email':booking_email,'bta_code_travel_req_no':bta_code_travel_req_no,'return_datetime':return_datetime}
            print(payload)

            url_taxi_booking = settings.API_BASE_URL + "add_flight_booking"
            booking = getDataFromAPI(login_type, access_token, url_taxi_booking, payload)

            if booking['success'] == 1:
                messages.success(request, str(booking['message']))
                return HttpResponseRedirect("/agents/flight-bookings/2", {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed To Add Flight Booking..!')
                return HttpResponseRedirect("/agents/flight-bookings/2", {'message': "Operation Failed"})
        else:
            return HttpResponseRedirect("/agents/login")

    else:
        request = get_request()
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            url = settings.API_BASE_URL + "companies"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            companies = company['Corporates']

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            if id:
                return render(request, 'Agent/add_flight_booking.html', {'companies':companies, 'airports':airports})
            else:
                return render(request, 'Agent/add_flight_booking.html', {'companies':companies, 'airports':airports})
        else:
            return HttpResponseRedirect("/agents/login")


def assign_flight_booking(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            no_of_passanger = request.POST.get('no_of_passanger', '')

            client_ticket_path = ''
            vender_ticket_path = ''

            operator_id = request.POST.get('operator_name', '')
            meal_is_include = request.POST.get('meal_is_include', '')
            fare_type = request.POST.get('fare_type', '')
            trip_type = request.POST.get('trip_type', '')
            flight_type = request.POST.get('flight_type', '')
            seat_type = request.POST.get('seat_type', '')
            no_of_stops = request.POST.get('no_of_stops', '')
            no_of_stop_rt = request.POST.get('no_of_stop_rt', '')

            ticket_number = request.POST.getlist('ticket_number', '')
            employee_booking_id = request.POST.getlist('employee_booking_id', '')

            flight_from = request.POST.getlist('from', '')
            flight_to = request.POST.getlist('to', '')
            departure_time = request.POST.getlist('departure_time', '')
            arrival_time = request.POST.getlist('arival_time', '')
            flight_name = request.POST.getlist('flight_name', '')
            flight_no = request.POST.getlist('flight_no', '')
            pnr_no = request.POST.getlist('pnr_number', '')

            portal_used = request.POST.get('portal_used', '')

            is_client_sms = request.POST.get('is_client_sms', '')
            is_client_email = request.POST.get('is_client_email', '')
            is_driver_sms = request.POST.get('is_driver_sms', '')

            ticket_price = request.POST.get('ticket_price', '')
            management_fee = request.POST.get('management_fee', '')
            tax_mng_amt = request.POST.get('tax_mng_amt', '')
            tax_on_management_fee = request.POST.get('tax_on_management_fee', '')
            tax_on_management_fee_percentage = request.POST.get('tax_on_management_fee_percentage', '')
            sub_total = request.POST.get('sub_total', '')
            management_fee_igst = request.POST.get('management_fee_igst', '')
            management_fee_cgst = request.POST.get('management_fee_cgst', '')
            management_fee_sgst = request.POST.get('management_fee_sgst', '')
            management_fee_igst_rate = request.POST.get('management_fee_igst_rate', '')
            management_fee_cgst_rate = request.POST.get('management_fee_cgst_rate', '')
            management_fee_sgst_rate = request.POST.get('management_fee_sgst_rate', '')
            cgst = request.POST.get('cgst', '')
            sgst = request.POST.get('sgst', '')
            igst = request.POST.get('igst', '')

            oper_ticket_price = request.POST.get('oper_ticket_price', '')

            oper_commission = request.POST.get('operator_commission', '')

            oper_commission_type = request.POST.get('oper_commission_type', '')

            oper_cotrav_billing_entity = request.POST.get('oper_cotrav_billing_entity', '')
            cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')

            oper_cgst = request.POST.get('oper_cgst', '')
            oper_sgst = request.POST.get('oper_sgst', '')
            oper_igst = request.POST.get('oper_igst', '')

            igst_amount = request.POST.get('igst_amount', '')
            cgst_amount = request.POST.get('cgst_amount', '')
            sgst_amount = request.POST.get('sgst_amount', '')

            if not igst_amount.isdigit():
                igst_amount =0

            if not cgst_amount.isdigit():
                cgst_amount =0

            if not sgst_amount.isdigit():
                sgst_amount =0

            if request.FILES:
                file_up = request.FILES.get('busticketToUpload', False)
                if file_up:
                    save_path = os.path.join(settings.MEDIA_ROOT, 'flight/client_ticket', str(file_up))
                    path1 = default_storage.save(save_path, request.FILES['busticketToUpload'])
                    client_ticket_path = path1
                else:
                    client_ticket_path = ""

                file_up2 = request.FILES.get('vendorticketToUpload', False)
                if file_up2:
                    save_path = os.path.join(settings.MEDIA_ROOT, 'flight/vender_ticket', str(file_up2))
                    path2 = default_storage.save(save_path, request.FILES['vendorticketToUpload'])
                    vender_ticket_path = path2
                else:
                    vender_ticket_path = ""

            client_ticket = request.POST.get('client_ticket')
            print(client_ticket)

            vender_ticket = request.POST.get('vender_ticket')
            print(vender_ticket)

            url = settings.API_BASE_URL + "assign_flight_booking"
            payload = {'ticket_no':ticket_number,'pnr_no':pnr_no,'portal_used':portal_used
                ,'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'flight_no':flight_no,'flight_name':flight_name,'arrival_time':arrival_time,
                       'departure_time':departure_time,'flight_to':flight_to,'flight_from':flight_from,'no_of_stops':no_of_stops,'seat_type':seat_type,'flight_type':flight_type,
                       'trip_type':trip_type,'fare_type':fare_type,'meal_is_include':meal_is_include,'no_of_passanger':no_of_passanger,'employee_booking_id':employee_booking_id,
                       'ticket_price': ticket_price, 'management_fee': management_fee, 'tax_mng_amt': tax_mng_amt
                , 'tax_on_management_fee': tax_on_management_fee,'no_of_stop_rt':no_of_stop_rt,
                       'tax_on_management_fee_percentage': tax_on_management_fee_percentage, 'sub_total': sub_total,
                       'management_fee_igst': management_fee_igst, 'management_fee_cgst': management_fee_cgst,
                       'management_fee_sgst': management_fee_sgst, 'management_fee_igst_rate': management_fee_igst_rate,
                       'management_fee_cgst_rate': management_fee_cgst_rate,
                       'management_fee_sgst_rate': management_fee_sgst_rate, 'cgst': cgst, 'sgst': sgst, 'igst': igst,
                       'oper_ticket_price': oper_ticket_price, 'oper_commission': oper_commission,
                       'oper_commission_type': oper_commission_type,
                       'oper_cotrav_billing_entity': oper_cotrav_billing_entity,'cotrav_billing_entity':cotrav_billing_entity,
                       'oper_cgst': oper_cgst, 'oper_sgst': oper_sgst, 'oper_igst': oper_igst,
                       'client_ticket_path': client_ticket_path,'client_ticket':client_ticket,'vender_ticket':vender_ticket,
                       'vender_ticket_path': vender_ticket_path,'is_client_sms':is_client_sms,'is_client_email':is_client_email,
                       'igst_amount': igst_amount,'cgst_amount': cgst_amount,'sgst_amount': sgst_amount,'operator_id':operator_id
                       }
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, 'Operation Successfully')
                return HttpResponseRedirect("/agents/flight-bookings/3", {'message': "Operation Successfully"})

            else:
                messages.error(request, 'Operation Failed')
                return HttpResponseRedirect("/agents/flight-bookings/3", {'message': "Operation Failed"})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'booking_id': id}

            url = settings.API_BASE_URL + "view_flight_booking"
            payload = {'booking_id': id}
            booking = getDataFromAPI(login_type, access_token, url, payload)
            booking = booking['Bookings']

            billing_entity = booking[0]['billing_entity_id']
            if billing_entity:
                url_hotels = settings.API_BASE_URL + "view_billing_entitie"
                payload = {'entity_id': billing_entity}
                bus_entity = getDataFromAPI(login_type, access_token, url_hotels, payload)
                bus_operators = bus_entity['Entitys']
                be_name = bus_operators[0]['entity_name']
                be_gst = bus_operators[0]['gst_id']
            else:
                be_name = ""
                be_gst = ""

            payload = {}
            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            c_entity = getDataFromAPI(login_type, access_token, url, payload)
            c_entity = c_entity['Enitity']

            url_hotels = settings.API_BASE_URL + "get_operators_by_service_type"
            payload = {'service_type': 7}
            hotels = getDataFromAPI(login_type, access_token, url_hotels, payload)
            flight_operators = hotels['Operators']

            url_access = settings.API_BASE_URL + "get_airports"
            data = getDataFromAPI(login_type, access_token, url_access, payload)
            airports = data['Airports']

            return render(request, 'Agent/assign_flight_booking.html',{'bookings': booking,
            'c_entitys':c_entity,'be_name':be_name,'be_gst':be_gst,'flight_operators':flight_operators,'airports':airports})

    else:
        return HttpResponseRedirect("/agents/login")


def cancel_flight_booking_passengers(request,id):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            tax_on_management_fee = 0
            refund_amount = 0
            cancel_comment=0
            management_fee_igst = 0
            management_fee_cgst = 0
            management_fee_sgst = 0
            management_fee_igst_rate = 0
            management_fee_cgst_rate = 0
            management_fee_sgst_rate = 0
            cgst = 0
            sgst = 0
            igst = 0
            igst_amount = 0
            cgst_amount = 0
            sgst_amount = 0
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            current_url = request.POST.get('current_url', '')
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            no_of_passenger = request.POST.get('no_of_passanger', '')
            employees = request.POST.getlist('cancel_employee_id', '')
            refund_amount = request.POST.get('refund_amount', '')
            cancel_comment = request.POST.get('cancel_comment', '')
            igst_rate = request.POST.get('igst_rate', '')
            if igst_rate:
                igst_rate = int(igst_rate)
            else:
                igst_rate = 0

            ticket_price = request.POST.get('ticket_price', '')
            if ticket_price:
                pass
            else:
                ticket_price = 0
            old_ticket_price = int(ticket_price)
            ticket_price = int(ticket_price) - int(refund_amount)
            management_fee = request.POST.get('management_fee', '')
            tax_mng_amt = ticket_price*0.18
            tax_on_management_fee = int(management_fee)*0.18
            tax_on_management_fee_percentage = 18
            sub_total = ticket_price+int(management_fee)+tax_mng_amt+tax_on_management_fee+tax_on_management_fee_percentage

            if int(igst_rate) > 0:
                management_fee_igst = 18
                management_fee_cgst = 0
                management_fee_sgst = 0
                management_fee_igst_rate = 18
                management_fee_cgst_rate = 0
                management_fee_sgst_rate = 0
                cgst = 18
                sgst = 0
                igst = 0
                igst_amount = sub_total * 0.18
                cgst_amount = 0
                sgst_amount = 0
            else:
                management_fee_igst = 0
                management_fee_cgst = 9
                management_fee_sgst = 9
                management_fee_igst_rate = 0
                management_fee_cgst_rate = 9
                management_fee_sgst_rate = 9
                cgst = 0
                sgst = 9
                igst = 9
                igst_amount = 0
                cgst_amount = sub_total * 0.9
                sgst_amount = sub_total * 0.9

            url = settings.API_BASE_URL + "cancel_flight_booking_passengers"
            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'no_of_passenger':no_of_passenger,
                       'ticket_price': ticket_price, 'management_fee': management_fee, 'tax_mng_amt': tax_mng_amt,
                       'tax_on_management_fee': tax_on_management_fee,'refund_amount':refund_amount,'cancel_comment':cancel_comment,
                       'tax_on_management_fee_percentage': tax_on_management_fee_percentage, 'sub_total': sub_total,
                       'management_fee_igst': management_fee_igst, 'management_fee_cgst': management_fee_cgst,
                       'management_fee_sgst': management_fee_sgst, 'management_fee_igst_rate': management_fee_igst_rate,
                       'management_fee_cgst_rate': management_fee_cgst_rate,'old_ticket_price':old_ticket_price,
                       'management_fee_sgst_rate': management_fee_sgst_rate, 'cgst': cgst, 'sgst': sgst, 'igst': igst,
                       'igst_amount': igst_amount,'cgst_amount': cgst_amount,'sgst_amount': sgst_amount,'employees':employees
                       }
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, 'Employee Canceled Successfully')
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})

            else:
                messages.error(request, 'Employee Cancel Failed')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
    else:
        return HttpResponseRedirect("/agents/login")


def accept_flight_booking(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            current_url = request.POST.get('current_url', '')
            user_comment = request.POST.get('user_comment', '')
            url = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "accept_flight_booking"
                operation_message = 'Flight Booking Accepted Successfully..!'

            if reject_id == '1':
                url = settings.API_BASE_URL + "reject_flight_booking"
                operation_message = "Flight Booking Rejected Successfully..!"

            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
            else:
                messages.error(request, 'Failed To Accept Flight Booking..!')
                return HttpResponseRedirect(current_url, {'message': "Operation Failed"})
        else:
            return render(request, "Agent/train_bookings.html", {'': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def dateonly(dt=''):
    try:
        if(dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "-" + str(datetime_object.month) + "-" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if(dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError :
        return ''


def download_taxi_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Passengers',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',
        'Package Name',
        'Tour Type',
        'Vehicle Type',
        'Driver Name',
        'Driver Contact',
        'Taxi Reg No.',
        'No. Of Seats',
        'Assign Date',
        'Assign Time',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',
        'Current Booking Status',
        'Hours Done',
        'Allowed Hours',
        'Extra Hours',
        'Rate per Hour',
        'Extra Hour Charge',
        'Rate per Hour',
        'Kms Done',
        'Allowed Kms',
        'Extra Kms',
        'Rate Per Km',
        'Driver/Night Charge',
        'Base Price',
        'Reimbursement Tax',
        'Extras',
        'Usage Charge',
        'Client Management Fee',
        'Client Tax On Management Fee',
        'Client Sub Total',
        'Client Cotrav Billing Entity',
        'Client IGST',
        'Client CGST',
        'Client SGST',
        'Client Management Fee IGST',
        'Client Management Fee CGST',
        'Client Management Fee SGST',
        'Client Management Fee IGST Rate',
        'Client Management Fee CGST Rate',
        'Client Management Fee SGST Rate',

        'Is Auto Approved',
        'Bill ID',
        'Bill Date',
        'Billing Entity',
        'Cotrav Status',
        'Client Status',
    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies

    for bk in booking:
        row_num += 1
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        spoc_status=''
        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"


        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    print("approver date")
                    print(approved_date)
                    approved_status = 'Accepted'

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']

        row = [bk['reference_no'],
               bk['city_name'],
               bk['assessment_code'],
               bk['assessment_city_id'],
               bk['reason_booking'],
               bk['zone_name'],
               bk['group_name'],
               bk['subgroup_name'],
               bk['spoc_name'],
               passanger_list,
               dateonly(bk['booking_date']),
               timeonly(bk['booking_date']),
               spoc_status,
               approved_by,
               dateonly(approved_date),
               timeonly(approved_date),
               approved_status,
               bk['pickup_location'],
               bk['drop_location'],
               dateonly(bk['pickup_datetime']),
               timeonly(bk['pickup_datetime']),
               '',
               '',
               bk['package_name'],

               bk['tour_type'],

               bk['taxi_type_name'],
               bk['driver_name'],
               bk['driver_contact'],
               bk['licence_no'],
               bk['no_of_seats'],
               dateonly(assigned_date),
               timeonly(assigned_date),
               canceled_by,
               '',
               dateonly(canceled_date),
               timeonly(canceled_date),
               bk['last_action_by'],
               bk['ci_hours_done'],
               bk['ci_allowed_hours'],
               bk['ci_extra_hours'],
               bk['ci_charge_hour'],
               bk['ci_extra_hr_charges'],
               bk['ci_charge_hour'],
               bk['ci_kms_done'],
               bk['ci_allowed_kms'],
               bk['ci_extra_kms'],
               bk['ci_extra_kms'],
               bk['ci_driver_allowance'],
               bk['ci_base_rate'],
               '',
               bk['ci_extra_kms'],
               bk['ci_extra_km_rate'],
               bk['ci_management_fee'],
               bk['ci_tax_on_management_fee'],
               bk['ci_sub_total'],
               bk['ci_cotrav_billing_entity'],
               bk['ci_igst'],
               bk['ci_cgst'],
               bk['ci_sgst'],
               bk['ci_management_fee_igst'],
               bk['ci_management_fee_cgst'],
               bk['ci_management_fee_sgst'],
               bk['ci_management_fee_igst_rate'],
               bk['ci_management_fee_cgst_rate'],
               bk['ci_management_fee_sgst_rate'],
               '',
               '',
               '',
               bk['billing_entity_id'],
               bk['cotrav_status'],
               bk['client_status'],

               ]
        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    if 'login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            #print(booking)

            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            )
            response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
                date=datetime.now().strftime('%Y-%m-%d'),
            )
            workbook = Workbook()

            # Get active worksheet/tab
            worksheet = workbook.active
            worksheet.title = 'Bus Bookings'

            # Define the titles for columns

            columns = [
                "Booking ID",
                "Assessment Code",
                "Assessment City",
                "Zone",
                "Group Name",
                "Subgroup Name",
                "SPOC Name",
                "Passenger Name",
                "Booking Date",
                "Booking Time",
                "SPOC Status",
                "Approver Name",
                "Approved Date",
                "Approved Time",
                "Approver Status",
                "Pickup City",
                "Drop City",
                "Journey Date",
                "Journey Time",
                "Cotrav Status",
                "Assign Date",
                "Assign Time",
                "Rejected By",
                "Reject Reason",
                "Reject Date",
                "Reject Time",
                "Current Booking Status",
                "Bus Type Allocated",
                "PNR Number",

                'Ticket No',
                'Vendor Booking Id',
                'Voucher No',
                'Booking Eail',
                'Pickup To Datetime',
                'Preferred Bus',
                'Boarding Point',
                'Boarding Datetime',
                'Bus Type Priority 1',
                'Bus Type Priority 2',
                'Bus Type Priority 3',
                'No Of Seats',
                'Last Action By',
                'Operator Name',
                'Operator Contact',
                'Assign Bus Type ID',
                'Seat No',
                'Portal Used',
                'Status Client',
                'Status Cotrav',
                'Invoice Generated',
                'Invoice ID',
                'Group ID',
                'Subgroup ID',
                'Spoc ID',
                'Corporate ID',
                'Billing Entity ID',
                'Reason Booking',
                'Cotrav Agent Name',


                'Client Ticket Price',
                'Client Management Fee',
                'Client Tax On Mng. fee',
                'Client Sub Total',
                'Client Cotrav Billing Entity',
                'Client IGST',
                'Client CGST',
                'Client SGST',
                'Client IGST Mng. Fee',
                'Client CGST Mng. Fee',
                'Client SGST Mng. Fee',
                'Client IGST Mng. Fee Rate',
                'Client CGST Mng. Fee Rate',
                'Client SGST Mng. Fee Rate',


                "Cancellation Charge",
                "Service Tax on Cancellation",
                "Is Auto Approved",
                "Bill ID",

                'Cotrav Status',
                'Client Status',

            ]
            row_num = 1

            # Assign the titles for each cell of the header
            for col_num, column_title in enumerate(columns, 1):
                cell = worksheet.cell(row=row_num, column=col_num)
                cell.value = column_title

            # Iterate through all movies
            for bk in booking:
                row_num += 1
                passanger_list = ''
                created_by = ''
                created_date = ''
                accepted_by = ''
                accepted_date = ''
                approved_by = ''
                approved_date = ''
                approved_status = ''
                assigned_by = ''
                assigned_date = ''
                canceled_by = ''
                canceled_date = ''
                b_date = ''
                b_time = ''
                spoc_status = ''
                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"


                if (bk['Passangers']):
                    if len(bk['Passangers']) > 1:
                        print(bk['Passangers'])
                        for psg in bk['Passangers']:
                            passanger_list = passanger_list + str(psg['employee_name']) + ","
                    else:
                        passanger_list = bk['Passangers'][0]['employee_name']
                if len(bk['Actions']) >= 0:
                    for act in bk['Actions']:
                        if (act['action'] == 1):
                            print('created')
                            created_by = act['employee_name']
                            created_date = act['action_date']
                        if (act['action'] == 2):
                            print('accepted')
                            accepted_by = act['employee_name']
                            accepted_date = act['action_date']
                        if (act['action'] == 2):
                            print('approved')
                            approved_by = act['employee_name']
                            approved_date = act['action_date']
                            approved_status = 'accepted'

                        if (act['action'] == 3):
                            print('canceled')
                            canceled_by = act['employee_name']
                            canceled_date = act['action_date']

                        if (act['action'] == 4):
                            print('assigned')
                            assigned_by = act['employee_name']
                            assigned_date = act['action_date']


                row = [
                    bk['reference_no'],
                    bk['assessment_code'],
                    bk['assessment_city_id'],
                    bk['zone_name'],
                    bk['group_name'],
                    bk['subgroup_name'],
                    bk['spoc_name'],
                    passanger_list,
                    b_date,
                    b_time,
                    spoc_status,
                    approved_by,
                    dateonly(approved_date),
                    timeonly(approved_date),
                    approved_status,
                    bk['pickup_location'],
                    bk['drop_location'],
                    bk['pickup_from_datetime'],
                    dateonly(bk['boarding_datetime']),
                    bk['cotrav_status'],

                    dateonly(assigned_date),
                    timeonly(assigned_date),
                    canceled_by,
                    '',
                    dateonly(canceled_date),
                    timeonly(canceled_date),
                    bk['status_cotrav'],
                    bk['assign_bus_type_id'],
                    bk['pnr_no'],

                    bk['ticket_no'],
                    bk['vendor_booking_id'],
                    bk['voucher_no'],
                    bk['booking_email'],
                    bk['pickup_to_datetime'],
                    bk['preferred_bus'],
                    bk['boarding_point'],
                    bk['boarding_datetime'],
                    bk['bus_type_priority_1'],
                    bk['bus_type_priority_2'],
                    bk['bus_type_priority_3'],
                    bk['no_of_seats'],
                    bk['last_action_by'],
                    bk['operator_name'],
                    bk['operator_contact'],
                    bk['assign_bus_type_id'],
                    bk['seat_no'],
                    bk['portal_used'],
                    bk['status_client'],
                    bk['status_cotrav'],
                    bk['is_invoice'],
                    bk['invoice_id'],
                    bk['group_id'],
                    bk['subgroup_id'],
                    bk['spoc_id'],
                    bk['corporate_id'],
                    bk['billing_entity_id'],
                    bk['reason_booking'],
                    bk['cotrav_agent_name'],


                    bk['client_ticket_price'],
                    bk['client_mang_fee'],
                    bk['client_tax_on_mang_fee'],
                    bk['client_sub_total'],
                    bk['client_cotrav_billing_entity'],
                    bk['client_igst'],
                    bk['client_cgst'],
                    bk['client_sgst'],
                    bk['client_mng_fee_igst'],
                    bk['client_mng_fee_cgst'],
                    bk['client_mng_fee_sgst'],
                    bk['client_mng_fee_igst_rate'],
                    bk['client_mng_fee_cgst_rate'],
                    bk['client_mng_fee_sgst_rate'],


                    "",
                    "",
                    "",
                    "",

                    bk['client_status'],
                    bk['cotrav_status'],

                ]

                # Assign the data for each cell of the row
                for col_num, cell_value in enumerate(row, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    cell.value = cell_value

        workbook.save(response)

        return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            #print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [
        "Booking ID",
        "Assessment Code",
        "Travel Request Code",
        "From City",
        "To City",
        "SPOC Name",
        "Passengers",
        "Passenger IDs",
        "Boarding Date",
        "Boarding Time",
        "Booking Date",
        "Booking Time",
        "Boarding Point",
        "Train Name",
        "Train No.",
        "PNR No.",
        "Coach Type",
        "Quota Used",

        "Ticket Price",
        "Surcharge",
        'Client Ticket Price',
        'Client Management Fee',
        'Client Tax On Mng. fee',
        'Client Sub Total',
        'Client Cotrav Billing Entity',
        'Client IGST',
        'Client CGST',
        'Client SGST',
        'Client IGST Mng. Fee',
        'Client CGST Mng. Fee',
        'Client SGST Mng. Fee',
        'Client IGST Mng. Fee Rate',
        'Client CGST Mng. Fee Rate',
        'Client SGST Mng. Fee Rate',

        "Vendor Taxable Amount",
        "Vendor GST Amount",
        "Vendor Bill Date",
        "Vendor Bill No.",
        "Vendor GSTIN",
        "HSN/SAC Code",
        "Invoice Created",

        "Client Status",
        "Cotrav Status",
    ]


    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''



        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'accepted'

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']



        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            "",
            bk['pickup_location'],
            bk['drop_location'],
            bk['spoc_name'],
            passanger_list,
            "",
            dateonly(bk['boarding_datetime']),
            timeonly(bk['boarding_datetime']),
            bk['booking_datetime'],
            bk['boarding_datetime'],
            bk['boarding_point'],
            bk['train_name'],
            bk['train_name'],
            bk['pnr_no'],
            bk['seat_no'],
            "",

            bk['client_ticket_price'],
            "",
            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],

            "",
            "",
            "",
            "",
            "",
            "",
            "",

            bk['client_status'],
            bk['cotrav_status'],


        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            #print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Booking Date',
        'Usage Type',
        'Trip Type',
        'Flight Type',
        'Seat Type',
        'From City',
        'To City',
        'Departure Date',
        'Departure Time',
        'Passengers',
        'Approved Date',
        'Approved Time',
        'Reject Date',
        'Reject Time',
        'Assign Date',
        'Assign Time',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'First Flight Name',
        'First Flight No',
        'First PNR No',
        'First Flight From City',
        'First Flight To City',
        'First Flight Departure Datetime',
        'First Flight Arrival Datetime',

        'Second Flight Name',
        'Second Flight No',
        'Second PNR No',
        'Second Flight From City',
        'Second Flight To City',
        'Second Flight Departure Datetime',
        'Second Flight Arrival Datetime',

        'Third Flight Name',
        'Third Flight No',
        'Third PNR No',
        'Third Flight From City',
        'Third Flight To City',
        'Third Flight Departure Datetime',
        'Third Flight Arrival Datetime',

        'Client Ticket Price',
        'Client Management Fee',
        'Client Tax On Mng. fee',
        'Client Sub Total',
        'Client Cotrav Billing Entity',
        'Client IGST',
        'Client CGST',
        'Client SGST',
        'Client IGST Mng. Fee',
        'Client CGST Mng. Fee',
        'Client SGST Mng. Fee',
        'Client IGST Mng. Fee Rate',
        'Client CGST Mng. Fee Rate',
        'Client SGST Mng. Fee Rate',

        'SPOC Comment',
        'Meal Charges',
        'Cancellation Charge',
        'Tax On Cancellation',
        'Is Auto Approved',
        'Bill ID',
        'Cotrav Status',
        'Client Status',

    ]


    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['booking_datetime']):
            datetime_str = bk['booking_datetime']
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) +"/"+ str(datetime_object.month) + "/" + str(datetime_object.year)
            booking_time = str(datetime_object.hour) +":" + str(datetime_object.hour)


        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'accepted'

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            dateonly(bk['booking_datetime']),
            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],
            bk['from_location'],
            bk['to_location'],
            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),
            passanger_list,
            dateonly(approved_date),
            timeonly(approved_date),
            dateonly(canceled_date),
            timeonly(canceled_date),
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['spoc_name'],

            flight_name1 ,
            flight_no1 ,
            pnr_no1 ,
            from_city1 ,
            to_city1 ,
            departure_datetime1 ,
            arrival_datetime1 ,

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            departure_datetime2,
            arrival_datetime2,

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            departure_datetime3,
            arrival_datetime3,

            bk['client_ticket_price'],
            bk['client_mang_fee'],
            bk['client_tax_on_mang_fee'],
            bk['client_sub_total'],
            bk['client_cotrav_billing_entity'],
            bk['client_igst'],
            bk['client_cgst'],
            bk['client_sgst'],
            bk['client_mng_fee_igst'],
            bk['client_mng_fee_cgst'],
            bk['client_mng_fee_sgst'],
            bk['client_mng_fee_igst_rate'],
            bk['client_mng_fee_cgst_rate'],
            bk['client_mng_fee_sgst_rate'],


            '',
            '',
            '',
            '',
            '',
            '',
            bk['cotrav_status'],
            bk['client_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    if 'login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)


        if company['success'] == 1:
            booking = company['Bookings']

            #print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Bucket Request Date',
        'Bucket Request Time',
        'Bucket Approved Date',
        'Bucket Approved Time',
        'Preferred Hotel',
        'Assigned Hotel',
        'Assigned Hotel Address',
        'Hotel Contact',
        'Assign Date',
        'Assign Time',
        'TaxiVaxi Status',
        'Employees Name',
        'No. of Persons',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Higher Bucket Requested',
        'Reason for Higher Bucket',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',
        'Current Booking Status',
        'No. of Nights',
        'Room Type',
        'Room Occupancy',
        'Per Night Price',
        'Total Room Price',
        'Tax On Room Cancellation Charge',
        'Client Management Fee',
        'Client Tax On Mng. fee',
        'Client Sub Total',
        'Client Cotrav Billing Entity',
        'Client IGST',
        'Client CGST',
        'Client SGST',
        'Client IGST Mng. Fee',
        'Client CGST Mng. Fee',
        'Client SGST Mng. Fee',
        'Client IGST Mng. Fee Rate',
        'Client CGST Mng. Fee Rate',
        'Client SGST Mng. Fee Rate',
        'Is Auto Approved',
        'Bill ID',
        'Is TBA Booking',
        'Is Offline Booking',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = "Accept"

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = "Reject"

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'accepted'

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            '',
            '',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            '',
            '',
            '',
            '',

            bk['preferred_hotel'],
            bk['assign_hotel_id'],
            bk['operator_name'],
            bk['operator_contact'],
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['status_cotrav'],
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),
            bk['status_cotrav'],
            '',
            bk['room_type_name'],
            bk['hotel_type_name'],
            bk['bucket_price_1'],
            bk['bucket_price_1'],
            '',
            bk['ticket_price'],

            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['sub_total'],
            bk['billing_entity_id'],
            bk['igst'],
            bk['cgst'],
            bk['sgst'],
            bk['management_fee_igst'],
            bk['management_fee_cgst'],
            bk['management_fee_sgst'],
            bk['management_fee_igst_rate'],
            bk['management_fee_cgst_rate'],
            bk['management_fee_sgst_rate'],

            '',
            '',

            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_billing_entities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': 0}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Entitys']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=BillingEntitys.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Billing Entitys'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Billing City",
        "Entity Name",
        "Contact Person Name",
        "Contact Person Email",
        "Contact Person Phone No",
        "Address Line 1",
        "Address Line 2",
        "Address Line 3",
        "GST NO",
        "PAN No",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['billing_city'],
            bk['entity_name'],
            bk['contact_person_name'],
            bk['contact_person_email'],
            bk['contact_person_no'],
            bk['address_line_1'],
            bk['address_line_2'],
            bk['address_line_3'],
            bk['gst_id'],
            bk['pan_no'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_rates(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Corporate_Retes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyRates.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Rates'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City",
        "Taxi Type",
        "Tour Type",
        "KMS",
        "HOURS",
        "KM Rate",
        "HR Rate",
        "Base Rate",
        "Night Rate",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['tour_type'] == 2:
            tour_type = "Local"
        else:
            tour_type = "Outstation"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            bk['taxi_type'],
            tour_type,
            bk['kms'],
            bk['hours'],
            bk['km_rate'],
            bk['hour_rate'],
            bk['base_rate'],
            bk['night_rate'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_assessment_cities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "assessment_cities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Cities']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=AssessmentCities.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Assessment Cities'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_assessment_codes(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "assessment_codes"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Codes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=AssessmentCodes.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Assessment Codes'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Assessment Code",
        "Code Description",
        "From Date",
        "To Date",
        "Service From Date",
        "Service To Date",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_active'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['assessment_code'],
            bk['code_desc'],
            bk['from_date'],
            bk['to_date'],
            bk['service_from'],
            bk['service_to'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_groups(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Groups']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyGroups.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Groups'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Zone Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['zone_name'],

            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_subgroups(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Subgroups']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySubGroups.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company SubGroups'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgoup Name",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],

            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_admins(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Admins']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyAdmins.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Admins'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Admin Name",
        "Admin Email",
        "Admin Contact No ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['name'],
            bk['email'],
            bk['contact_no'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_spocs(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Spocs']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySpocs.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Spocs'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgroup Name",
        "Spoc Company ID",
        "Spoc Name",
        "Spoc Email",
        "Spoc Contact No ",
        "Login UserName ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['user_cid'],
            bk['user_name'],
            bk['email'],
            bk['user_contact'],
            bk['username'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_employees(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Employees']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyEmployees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Employees'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Spoc Name",
        "Core Employee ID",
        "Employee Company ID",
        "Employee Name",
        "Employee Email",
        "Employee Phone No",
        "Login UserName",
        "Age",
        "Gender",
        "ID Proof Type",
        "ID No",
        "Is CXO",
        "Designation",
        "Home City",
        "Home Address",
        "Assistant ID",
        "Date Of Birth",
        "Billing Entity",
        "Last Login",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_cxo'] == 0:
            is_cxo = "No"
        else:
            is_cxo = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['user_name'],
            bk['core_employee_id'],
            bk['core_employee_id'],
            bk['employee_cid'],
            bk['employee_name'],
            bk['employee_email'],
            bk['employee_contact'],
            bk['username'],
            bk['age'],
            bk['gender'],
            bk['id_proof_type'],
            bk['id_proof_no'],
            is_cxo,
            bk['designation'],
            bk['home_city'],
            bk['home_address'],
            bk['assistant_id'],
            bk['date_of_birth'],
            bk['billing_entity_id'],
            bk['last_login'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def taxi_billing(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_taxi_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/agent_taxi_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Agent/agent_taxi_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/agents/login")


def bus_billing(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_bus_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        #print(company)
        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/agent_bus_billing.html", {'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Agent/agent_bus_billing.html", {'billing_type': id})

    else:
        return HttpResponseRedirect("/agents/login")


def train_billing(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_train_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/agent_train_billing.html",{'bookings': booking, 'billing_type': id})
        else:
            return render(request, "Agent/agent_train_billing.html", {'billing_type': id})
    else:
        return HttpResponseRedirect("/agents/login")


def flight_billing(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_flight_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/agent_flight_billing.html",
                          {'bookings': booking,'billing_type': id, 'corporates': companies})
        else:
            return render(request, "Agent/agent_flight_billing.html", {})
    else:
        return HttpResponseRedirect("/agents/login")


def hotel_billing(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "agent_hotel_bookings"
        payload = {'booking_type': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Agent/agent_hotel_billing.html",{'bookings': booking,'billing_type': id, 'corporates': companies})
        else:
            return render(request, "Agent/agent_hotel_billing.html", {})
    else:
        return HttpResponseRedirect("/agents/login")


def taxi_billing_verify(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id, 'invoice_id': invoice_id}
            #print(payload)
            vry_url = settings.API_BASE_URL + "agent_verify_taxi_bookings"
            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Verify Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Verify..!")
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_revise_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Revise Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Revise..!")
                return HttpResponseRedirect(current_url, {})
        elif 'update_invoice' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_update_taxi_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Updated Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Updated..!")
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def bus_billing_verify(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_verify_bus_bookings"
            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Verify Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Verify..!")
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_revise_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Revise Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Revise..!")
                return HttpResponseRedirect(current_url, {})
        elif 'update_invoice' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            print("updateetetet")
            vry_url = settings.API_BASE_URL + "agent_update_bus_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Updated Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Updated..!")
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def train_billing_verify(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_verify_train_bookings"
            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Verify Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Verify..!")
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_revise_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Revise Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Revise..!")
                return HttpResponseRedirect(current_url, {})
        elif 'update_invoice' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_update_train_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Updated Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Updated..!")
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def hotel_billing_verify(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_verify_hotel_bookings"
            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Verify Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Verify..!")
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_revise_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Revise Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Revise..!")
                return HttpResponseRedirect(current_url, {})
        elif 'update_invoice' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_update_hotel_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Updated Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Updated..!")
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def flight_billing_verify(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        if 'verify' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            current_url = request.POST.get('current_url')
            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_verify_flight_bookings"
            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Verify Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Verify..!")
                return HttpResponseRedirect(current_url, {})
        elif 'revise' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_revise_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Revise Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Revise..!")
                return HttpResponseRedirect(current_url, {})
        elif 'update_invoice' in request.POST:
            verify_id = request.POST.getlist('booking_ids')
            corporate_id = request.POST.getlist('corporate_ids')
            invoice_id = request.POST.getlist('invoice_ids')
            invoice_comments = request.POST.getlist('invoice_comments')
            current_url = request.POST.get('current_url')

            payload = {'verify_id': verify_id, 'user_id': request.user.id, 'corporate_id': corporate_id,
                       'invoice_id': invoice_id, 'invoice_comments': invoice_comments}
            print(payload)
            vry_url = settings.API_BASE_URL + "agent_update_flight_bookings"

            verify = getDataFromAPI(login_type, access_token, vry_url, payload)
            if verify['success'] == 1:
                messages.success(request, "Invoice Updated Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, "Invoice Not Updated..!")
                return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def cotrav_billing_entities(request):
    if 'agent_login_type' in request.session:
        if id:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'': id}
            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            operator = operator['Enitity']

            company = getDataFromAPI(login_type, access_token, url, payload)
            url_city = settings.API_BASE_URL + "cities"
            cities = getDataFromAPI(login_type, access_token, url_city, payload)
            cities = cities["Cities"]
            
            return render(request, 'Agent/cotrav_billing_entities.html', {'billing_entities': operator, 'cities':cities})
        else:
            return render(request, 'Agent/cotrav_billing_entities.html', {})
    else:
        return HttpResponseRedirect("/agents/login")


def add_cotrav_billing_entities(request,id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            entity_name = request.POST.get('entity_name', '')
            billing_city_id = request.POST.get('billing_city_id')
            contact_person_name = request.POST.get('contact_person_name', '')
            contact_person_email = request.POST.get('contact_person_email', '')
            contact_person_no = request.POST.get('contact_person_no', '')
            address_line_1 = request.POST.get('address_line_1', '')
            address_line_2 = request.POST.get('address_line_2', '')
            address_line_3 = request.POST.get('address_line_3', '')
            gst_id = request.POST.get('gst_id', '')
            pan_no = request.POST.get('pan_no', '')

            entity_id = request.POST.get('entity_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token,
                       'entity_name': entity_name, 'billing_city_id': billing_city_id,
                       'contact_person_name': contact_person_name, 'contact_person_email': contact_person_email,
                       'contact_person_no': contact_person_no, 'address_line_1': address_line_1,
                       'address_line_2': address_line_2,
                       'address_line_3': address_line_3, 'gst_id': gst_id, 'pan_no': pan_no, 'entity_id': entity_id,
                       'is_delete': delete_id, }

            url = ""
            if entity_id:
                url = settings.API_BASE_URL + "update_cotrav_billing_entities"
                operation_message = "Company Entity Updaed Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_cotrav_billing_entities"
                    operation_message = "Company Entity Deleted Successfully..!"

            else:
                url = settings.API_BASE_URL + "add_cotrav_billing_entities"
                operation_message = "Company Entity Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/cotrav-billing-entities", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/cotrav-billing-entities", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def bill_create(request):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            
            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0
            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            date_type = request.POST.get('date_type', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            payload = {'corporate_id': int(corporate_id),'service_type':service_type,'invoice_type':invoice_type,'date_type':date_type,'from_date':from_date,'to_date':to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}

            return render(request, 'Agent/bills_create.html', {'bills': operator,'companies':companies,'data':payload})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Agent/bills_create.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/agents/login")


def get_all_generated_bills(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_all_generated_bills"
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = ''
        if company['success'] == 1:
            companies = company['Bill']
        else:
            companies = {}
        return render(request, 'Agent/bills_geterated.html', {'bills': companies, 'bill_type':id})
    else:
        return HttpResponseRedirect("/agents/login")


def get_all_bill_payment_status(request,id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_all_bill_payment_status"
        company = getDataFromAPI(login_type, access_token, url, payload)

        companies = ''
        comp_accounts = ''
        cotrav_accounts = ''

        if company['success'] == 1:
            companies = company['Bill']

            url = settings.API_BASE_URL + "get_cotrav_accounts"
            cotrav_accounts = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                cotrav_accounts = cotrav_accounts['Accounts']
            else:
                cotrav_accounts = ""

            url = settings.API_BASE_URL + "get_corporate_accounts"
            comp_accounts = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                comp_accounts = comp_accounts['Accounts']
            else:
                comp_accounts = ""

        else:
            companies = ''
        return render(request, 'Agent/bill_payment_status.html', {'bills': companies, 'bill_type':id, 'cotrav_accounts':cotrav_accounts,'comp_accounts':comp_accounts})
    else:
        return HttpResponseRedirect("/agents/login")


def bill_create_nontax_invoice(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        current_url = request.POST.get('current_url')

        invoice_id = request.POST.getlist('invoice_ids')
        service_types = request.POST.getlist('service_types')
        management_fees = request.POST.getlist('management_fees')
        management_fee_igsts = request.POST.getlist('management_fee_igsts')
        management_fee_cgsts = request.POST.getlist('management_fee_cgsts')
        management_fee_sgsts = request.POST.getlist('management_fee_sgsts')
        sub_totals = request.POST.getlist('sub_totals')
        cotrav_billing_ids = request.POST.get('cotrav_entity_id')
        billing_ids = request.POST.get('client_entity_id')
        po_number = request.POST.get('po_number')
        bill_date = request.POST.get('bill_date')
        tds_yes = request.POST.getlist('tds_yes')
        corporate_ids = request.POST.getlist('corporate_ids')

        payload = {'invoice_ids': invoice_id,'service_types':service_types,'management_fees':management_fees,'management_fee_igsts':management_fee_igsts,
                   'management_fee_cgsts':management_fee_cgsts,'management_fee_sgsts':management_fee_sgsts,'sub_totals':sub_totals,'billing_ids':billing_ids,
                   'cotrav_billing_ids':cotrav_billing_ids,'tds_yes':tds_yes,'corporate_ids':corporate_ids,'po_number':po_number,'bill_date':bill_date+' 00:00:00'}
        print(payload)

        url = settings.API_BASE_URL + "bill_create_nontax_invoice"
        operator = getDataFromAPI(login_type, access_token, url, payload)
        print(operator)
        
        if operator['success'] == 1:
            messages.success(request, "Bill Generated Success")
            #filename = 'D:/Taxivaxi_Python_Projects/CoTrav/media/Bill_PDF/Voucher_32939.pdf'
            filename = operator['voucher_path']
            file_path = os.path.join(filename)
            if os.path.exists(file_path):
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                    return response
            else:
                messages.error(request, "Bill Generated Failed")
                return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Bill Generated Failed")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def generate_final_bill(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        current_url = request.POST.get('current_url')
        bill_id = request.POST.getlist('bill_id')
        is_tds = request.POST.getlist('is_tds')

        payload = {'bill_id': bill_id, 'is_tds':is_tds}
        print(payload)

        url = settings.API_BASE_URL + "bill_create_final_invoice"
        operator = getDataFromAPI(login_type, access_token, url, payload)
        print(operator)

        if operator['success'] == 1:
            messages.success(request, "Bill Generated Success")
            #filename = 'D:/Taxivaxi_Python_Projects/CoTrav/media/Bill_PDF/Voucher_32939.pdf'
            filename = operator['voucher_path']
            file_path = os.path.join(filename)
            if os.path.exists(file_path):
                with open(file_path, 'rb') as fh:
                    response = HttpResponse(fh.read(), content_type="application/vnd.ms-excel")
                    response['Content-Disposition'] = 'inline; filename=' + os.path.basename(file_path)
                    return response
            else:
                messages.error(request, "Bill Generated Failed")
                return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Bill Generated Failed")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")

def add_booking_tracking_status(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        booking_id = request.POST.get('booking_id', '')
        user_id = request.POST.get('user_id', '')
        booking_type = request.POST.get('booking_type', '')
        user_comment = request.POST.get('user_comment', '')
        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id, 'user_id':user_id, 'booking_type':booking_type,'user_comment':user_comment}
        url = settings.API_BASE_URL + "add_booking_tracking_status"
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Tracking Status Added Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To Add tracking Status..!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def add_booking_assign_to_agent(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        booking_id = request.POST.get('booking_id', '')
        user_id = request.POST.get('user_id', '')
        to_user_id = request.POST.get('to_user_id', '')
        booking_type = request.POST.get('booking_type', '')
        user_comment = request.POST.get('user_comment', '')
        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id, 'user_id':user_id, 'booking_type':booking_type,'user_comment':user_comment,'to_user_id':to_user_id}
        url = settings.API_BASE_URL + "add_booking_assign_to_agent"
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Assign Agent Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To AAssign Agent.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def change_booking_status(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        booking_id = request.POST.get('booking_id', '')
        user_id = request.POST.get('user_id', '')
        to_user_id = request.POST.get('to_user_id', '')
        status_id = request.POST.get('status_id', '')
        user_comment = request.POST.get('user_comment', '')
        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id, 'user_id':user_id, 'status_id':status_id,'user_comment':user_comment,'to_user_id':to_user_id}
        url = settings.API_BASE_URL + "change_booking_status"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Assign Agent Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To AAssign Agent.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def edit_taxi_booking(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        global booking_email
        email_attachment = request.POST.get('email_attachment', '')
        print(email_attachment)
        if email_attachment:
            booking_email = email_attachment
            pass
        else:
            if request.FILES:
                print("in file")
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None

        user_id = request.POST.get('user_id', '')
        booking_id = request.POST.get('booking_id', '')
        tour_type = request.POST.get('tour_type', '')
        current_city_id = request.POST.get('current_city_id', '')
        taxi_type = request.POST.get('taxi_type', '')
        no_of_days = request.POST.get('no_of_days', '')
        pickup_location = request.POST.get('pickup_location', '')
        drop_location = request.POST.get('drop_location', '')
        pickup_datetime = request.POST.get('pickup_datetime', '')
        billing_entity_id = request.POST.get('billing_entity_id', '')
        cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')
        rate_id = request.POST.get('rate_id', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        reason_booking = request.POST.get('reason_booking', '')
        assessment_code = request.POST.get('assessment_code', '')
        assessment_city_id = request.POST.get('assessment_city_id', '')

        spoc_id = request.POST.get('spoc_id', '')
        spoc_details = [x.strip() for x in spoc_id.split(',')]

        spoc_id = spoc_details[0]
        group_id = spoc_details[1]
        subgroup_id = spoc_details[2]

        employees = []
        no_of_emp = int(no_of_seats) + 1
        for i in range(1, no_of_emp):
            employees.append(request.POST.get('employee_id_' + str(i), ''))

        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id, 'tour_type':tour_type, 'current_city_id':current_city_id,'taxi_type':taxi_type,'no_of_days':no_of_days,
                   'pickup_location':pickup_location, 'drop_location':drop_location, 'pickup_datetime':pickup_datetime+':00','cotrav_billing_entity':cotrav_billing_entity,
                   'billing_entity_id':billing_entity_id,'rate_id':rate_id,'no_of_seats':no_of_seats,'employees':employees,'booking_email':booking_email,
                   'reason_booking':reason_booking,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,'spoc_id':spoc_id,'group_id':group_id,
                   'subgroup_id':subgroup_id,'user_id':user_id}
        url = settings.API_BASE_URL + "edit_taxi_booking"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Edit Taxi Booking Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To Edit Taxi Booking.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def edit_bus_booking(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        global booking_email
        email_attachment = request.POST.get('email_attachment', '')
        print(email_attachment)
        if email_attachment:
            booking_email = email_attachment
            pass
        else:
            if request.FILES:
                print("in file")
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None

        user_id = request.POST.get('user_id', '')
        booking_id = request.POST.get('booking_id', '')
        bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
        pickup_location = request.POST.get('pickup_location', '')
        drop_location = request.POST.get('drop_location', '')
        pickup_from_datetime = request.POST.get('pickup_from_datetime', '')
        pickup_to_datetime = request.POST.get('pickup_to_datetime', '')
        bus_type = request.POST.get('bus_type', '')
        bus_type2 = request.POST.get('bus_type2', '')

        billing_entity_id = request.POST.get('billing_entity_id', '')
        cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        reason_booking = request.POST.get('reason_booking', '')
        assessment_code = request.POST.get('assessment_code', '')
        assessment_city_id = request.POST.get('assessment_city_id', '')

        spoc_id = request.POST.get('spoc_id', '')
        spoc_details = [x.strip() for x in spoc_id.split(',')]

        spoc_id = spoc_details[0]
        group_id = spoc_details[1]
        subgroup_id = spoc_details[2]

        employees = []
        no_of_emp = int(no_of_seats) + 1
        for i in range(1, no_of_emp):
            employees.append(request.POST.get('employee_id_' + str(i), ''))

        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id, 'journey_datetime':pickup_from_datetime+':00', 'journey_datetime_to':pickup_to_datetime+':00',
                   'from':pickup_location, 'to':drop_location, 'cotrav_billing_entity':cotrav_billing_entity,
                   'billing_entity_id':billing_entity_id,'no_of_seats':no_of_seats,'employees':employees,'booking_email':booking_email,
                   'reason_booking':reason_booking,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,'spoc_id':spoc_id,'group_id':group_id,
                   'subgroup_id':subgroup_id,'user_id':user_id,'bus_type':bus_type,'bus_type2':bus_type2,'bta_code_travel_req_no':bta_code_travel_req_no}
        url = settings.API_BASE_URL + "edit_bus_booking"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Edit Taxi Booking Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To Edit Taxi Booking.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def edit_train_booking(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        global booking_email
        email_attachment = request.POST.get('email_attachment', '')
        print(email_attachment)
        if email_attachment:
            booking_email = email_attachment
            pass
        else:
            if request.FILES:
                print("in file")
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None

        user_id = request.POST.get('user_id', '')
        booking_id = request.POST.get('booking_id', '')
        bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
        pickup_location = request.POST.get('pickup_location', '')
        drop_location = request.POST.get('drop_location', '')
        pickup_from_datetime = request.POST.get('pickup_from_datetime', '')
        pickup_to_datetime = request.POST.get('pickup_to_datetime', '')
        train_type_priority_1 = request.POST.get('train_type_priority_1', '')
        train_type_priority_2 = request.POST.get('train_type_priority_2', '')

        billing_entity_id = request.POST.get('billing_entity_id', '')
        cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        reason_booking = request.POST.get('reason_booking', '')
        assessment_code = request.POST.get('assessment_code', '')
        assessment_city_id = request.POST.get('assessment_city_id', '')

        spoc_id = request.POST.get('spoc_id', '')
        spoc_details = [x.strip() for x in spoc_id.split(',')]

        spoc_id = spoc_details[0]
        group_id = spoc_details[1]
        subgroup_id = spoc_details[2]

        employees = []
        no_of_emp = int(no_of_seats) + 1
        for i in range(1, no_of_emp):
            employees.append(request.POST.get('employee_id_' + str(i), ''))

        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id,'journey_datetime_to':pickup_to_datetime+':00',
                   'from':pickup_location, 'to':drop_location, 'journey_datetime':pickup_from_datetime+':00','cotrav_billing_entity':cotrav_billing_entity,
                   'billing_entity_id':billing_entity_id,'no_of_seats':no_of_seats,'employees':employees,'booking_email':booking_email,
                   'reason_booking':reason_booking,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,'spoc_id':spoc_id,'group_id':group_id,
                   'subgroup_id':subgroup_id,'user_id':user_id,'train_type_priority_1':train_type_priority_1,'train_type_priority_2':train_type_priority_2,
                   'bta_code_travel_req_no':bta_code_travel_req_no}
        url = settings.API_BASE_URL + "edit_train_booking"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Edit Train Booking Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To Edit Train Booking.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def edit_hotel_booking(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        global booking_email
        email_attachment = request.POST.get('email_attachment', '')
        print(email_attachment)
        if email_attachment:
            booking_email = email_attachment
            pass
        else:
            if request.FILES:
                print("in file")
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None

        user_id = request.POST.get('user_id', '')
        booking_id = request.POST.get('booking_id', '')
        bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
        bucket_priority_1 = request.POST.get('bucket_priority_1', '')
        bucket_priority_2 = request.POST.get('bucket_priority_2', '')
        hotel_type = request.POST.get('hotel_type', '')
        from_city_id = request.POST.get('from_city_id', '')
        from_area_id = request.POST.get('from_area_id', '')
        preferred_area_name = request.POST.get('preferred_area_name', '')
        checkin_datetime = request.POST.get('checkin_datetime', '')
        checkout_datetime = request.POST.get('checkout_datetime', '')

        billing_entity_id = request.POST.get('billing_entity_id', '')
        cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        reason_booking = request.POST.get('reason_booking', '')
        assessment_code = request.POST.get('assessment_code', '')
        assessment_city_id = request.POST.get('assessment_city_id', '')

        spoc_id = request.POST.get('spoc_id', '')
        spoc_details = [x.strip() for x in spoc_id.split(',')]

        spoc_id = spoc_details[0]
        group_id = spoc_details[1]
        subgroup_id = spoc_details[2]

        employees = []
        no_of_emp = int(no_of_seats) + 1
        for i in range(1, no_of_emp):
            employees.append(request.POST.get('employee_id_' + str(i), ''))

        current_url = request.POST.get('current_url', '')
        payload = {'booking_id': booking_id,'checkout_datetime':checkout_datetime+":00",'preferred_area_name':preferred_area_name,
                   'from_city_id':from_city_id, 'from_area_id':from_area_id, 'checkin_datetime':checkin_datetime+":00",'cotrav_billing_entity':cotrav_billing_entity,
                   'billing_entity_id':billing_entity_id, 'no_of_seats':no_of_seats,'employees':employees,'booking_email':booking_email,
                   'reason_booking':reason_booking,'assessment_code':assessment_code,'assessment_city_id':assessment_city_id,'spoc_id':spoc_id,
                   'group_id':group_id,'bta_code_travel_req_no':bta_code_travel_req_no,'bucket_priority_1':bucket_priority_1,'bucket_priority_2':bucket_priority_2,
                   'subgroup_id':subgroup_id,'user_id':user_id,'hotel_type':hotel_type}
        url = settings.API_BASE_URL + "edit_hotel_booking"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Edit Hotel Booking Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To Edit Hotel Booking.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def edit_flight_booking(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        global booking_email
        email_attachment = request.POST.get('email_attachment', '')
        print(email_attachment)
        if email_attachment:
            booking_email = email_attachment
            pass
        else:
            if request.FILES:
                print("in file")
                file_up = request.FILES.get('email_attachment', False)
                if file_up:
                    file_up = request.FILES['email_attachment']
                    booking_email = file_upload_get_path(file_up)
                else:
                    booking_email = None
            else:
                booking_email = None

        booking_id = request.POST.get('booking_id', '')
        bta_code_travel_req_no = request.POST.get('bta_code_travel_req_no', '')
        corporate_id = request.POST.get('corporate_id', '')
        user_id = request.POST.get('user_id', '')
        spoc_id = request.POST.get('spoc_id', '')

        usage_type = request.POST.get('usage_type', '')
        trip_type = request.POST.get('trip_type', '')
        seat_type = request.POST.get('seat_type', '')
        from_city = request.POST.get('from_city', '')
        to_city = request.POST.get('to_city', '')
        departure_datetime = request.POST.get('departure_datetime', '')
        return_datetime = request.POST.get('return_datetime', '')
        preferred_flight = request.POST.get('preferred_flight', '')
        billing_entity_id = request.POST.get('billing_entity_id', '')
        reason_booking = request.POST.get('reason_booking', '')
        no_of_seats = request.POST.get('no_of_seats', '')
        assessment_code = request.POST.get('assessment_code', '')
        assessment_city_id = request.POST.get('assessment_city_id', '')
        cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')

        spoc_details = [x.strip() for x in spoc_id.split(',')]

        spoc_id = spoc_details[0]
        group_id = spoc_details[1]
        subgroup_id = spoc_details[2]

        employees = []
        no_of_emp = int(no_of_seats) + 1
        for i in range(1, no_of_emp):
            employees.append(request.POST.get('employee_id_' + str(i), ''))

        current_url = request.POST.get('current_url', '')
        payload = {'user_id': user_id, 'user_type': login_type, 'corporate_id': corporate_id,'spoc_id': spoc_id, 'group_id': group_id,
                   'subgroup_id': subgroup_id, 'usage_type': usage_type, 'trip_type': trip_type, 'seat_type': seat_type,'from_city': from_city, 'to_city': to_city,
                    'departure_datetime': departure_datetime,'preferred_flight': preferred_flight,'booking_id':booking_id,
                   'reason_booking': reason_booking, 'no_of_seats': no_of_seats, 'employees': employees, 'billing_entity_id': billing_entity_id,'assessment_code': assessment_code,
                   'assessment_city_id': assessment_city_id, 'booking_email': booking_email, 'bta_code_travel_req_no': bta_code_travel_req_no,'return_datetime':return_datetime}
        url = settings.API_BASE_URL + "edit_flight_booking"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Edit Flight Booking Successfully..!")
            return HttpResponseRedirect(current_url, {})
        else:
            messages.error(request, "Failed To Edit Taxi Booking.!")
            return HttpResponseRedirect(current_url, {})
    else:
        return HttpResponseRedirect("/agents/login")


def pay_bill(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            bill_id = request.POST.get('bill_id', '')
            user_id = request.POST.get('user_id', '')
            payment_mode = request.POST.get('payment_mode', '')
            paid_amount = request.POST.get('paid_amount', '')
            payment_ref_no = request.POST.get('payment_ref_no', '')
            paid_by = request.POST.get('paid_by', '')
            paid_to = request.POST.get('paid_to', '')
            payment_dateTime = request.POST.get('payment_dateTime', '')
            attachments = request.POST.get('attachments', '')
            user_comment = request.POST.get('user_comment', '')
            total_paid = request.POST.get('total_paid', '')
            total_balance = request.POST.get('total_balance', '')
            total = request.POST.get('total', '')
            paybale = request.POST.get('paybale', '')

            global booking_email
            booking_email = ''
            if request.FILES:
                    file_up = request.FILES.get('attachments',False)
                    if file_up:
                        file_up = request.FILES['attachments']
                        booking_email = file_upload_get_path(file_up)
                    else:
                        booking_email = None
            else:
                booking_email = None

            current_url = request.POST.get('current_url', '')


            url = settings.API_BASE_URL + "pay_bill"
            payload = {'bill_id': bill_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment,'payment_mode':payment_mode,'paid_amount':paid_amount,
                       'payment_ref_no':payment_ref_no, 'paid_by':paid_by, 'payment_dateTime':payment_dateTime, 'attachments':booking_email,
                       'total_balance':total_balance,'total_paid':total_paid,'total':total,'paid_to':paid_to,'paybale':paybale}
            print(payload)
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, "Payment Successfully..!")
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, 'Failed to Accept Bill..!')
                return HttpResponseRedirect(current_url, {})
        else:
            return HttpResponseRedirect("/agents/bill-nontax-invoice/1")
    else:
        return HttpResponseRedirect("/agents/login")


def accept_bill(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            booking_id = request.POST.get('booking_id', '')
            user_id = request.POST.get('user_id', '')
            accept_id = request.POST.get('accept_id', '')
            reject_id = request.POST.get('reject_id', '')
            current_url = request.POST.get('current_url', '')
            user_comment = request.POST.get('user_comment', '')

            url = ""
            operation_message = ""
            if accept_id == '1':
                url = settings.API_BASE_URL + "accept_bill"
                operation_message="Bill Accepted successfully..!"

            if reject_id == '1':
                url = settings.API_BASE_URL + "reject_bill"
                operation_message="Bill Rejected successfully..!"

            payload = {'booking_id': booking_id,'user_id':user_id,'user_type':login_type,'user_comment':user_comment}

            company = getDataFromAPI(login_type, access_token, url, payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect(current_url, {})
            else:
                messages.error(request, 'Failed to Accept Bill..!')
                return HttpResponseRedirect(current_url, {})
        else:
            return HttpResponseRedirect("/agents/bill-nontax-invoice/1")
    else:
        return HttpResponseRedirect("/agents/login")


def edit_bill_detail(request, id):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            bill_id = request.POST.get('bill_id', '')
            bill_number = request.POST.get('bill_number', '')
            user_id = request.POST.get('user_id', '')
            corporate_id = request.POST.get('corporate_id', '')
            client_billing_entity = request.POST.get('client_billing_entity', '')
            cotrav_billing_entity = request.POST.get('cotrav_billing_entity', '')
            current_url = request.POST.get('current_url', '')
            po_id = request.POST.get('po_id', '')
            management_fee = request.POST.get('management_fee', '')
            tds_deducted_by_client = request.POST.get('tds_deducted_by_client', '')
            system_calculated_tds = request.POST.get('system_calculated_tds', '')
            igst = request.POST.get('igst', '')
            cgst = request.POST.get('cgst', '')
            sgst = request.POST.get('sgst', '')
            gst_paid = request.POST.get('gst_paid', '')
            total_amount = request.POST.get('total_amount', '')
            bill_created_date = request.POST.get('bill_created_date', '')

            url = settings.API_BASE_URL + "update_bill"
            operation_message="Bill Updated successfully..!"
            payload = {'bill_id':bill_id,'bill_number': bill_number,'user_id':user_id,'user_type':login_type,'corporate_id':corporate_id,'client_billing_entity':client_billing_entity,
                       'cotrav_billing_entity':cotrav_billing_entity,'po_id':po_id,'management_fee':management_fee,'tds_deducted_by_client':tds_deducted_by_client,
                       'system_calculated_tds':system_calculated_tds,'igst':igst,'cgst':cgst,'sgst':sgst,'gst_paid':gst_paid,'total_amount':total_amount,'bill_created_date':bill_created_date+' 00:00'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            print(payload)
            print(company)
            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/bill-nontax-invoice/1")
            else:
                messages.error(request, 'Failed to Updated Bill..!')
                return HttpResponseRedirect("/agents/bill-nontax-invoice/1")
        else:
            url = settings.API_BASE_URL + "view_bill"
            payload = {'bill_id':id}
            company = getDataFromAPI(login_type, access_token, url, payload)
            bills = company['Bill']
            
            url = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url, payload)
            companies = company1['Corporates']

            url2 = settings.API_BASE_URL + "billing_entities"
            co_entity = getDataFromAPI(login_type, access_token, url2, payload)
            print(co_entity)
            co_entity = co_entity['Entitys']

            url = settings.API_BASE_URL + "get_cotrav_billing_entities"
            c_entity = getDataFromAPI(login_type, access_token, url, payload)
            c_entity = c_entity['Enitity']

            url = settings.API_BASE_URL + "get_po_number_by_corporate"
            c_entity1 = getDataFromAPI(login_type, access_token, url, payload)
            po_nos = c_entity1['PO_NUMBERS']

            
            return render(request, 'Agent/edit_bill.html', {'bills': bills, 'companies':companies, 'c_entitys':c_entity, 'comp_enitys':co_entity,'po_nos':po_nos})
    else:
        return HttpResponseRedirect("/agents/login")


def corporate_podetails(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_corporate_podetails"
        company = getDataFromAPI(login_type, access_token, url, payload)
        pos = ''
        companies = ''
        if company['success'] == 1:
            pos = company['PO']
            url = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url, payload)
            companies = company1['Corporates']
        else:
            pos = ''
            companies = ''

        return render(request, 'Agent/corporate_podetails.html', {'pos': pos, 'companies':companies})
    else:
        return HttpResponseRedirect("/agents/login")


def corporate_accounts(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_corporate_accounts"
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        pos = ''
        companies = ''
        if company['success'] == 1:
            pos = company['Accounts']
            url = settings.API_BASE_URL + "companies"
            company1 = getDataFromAPI(login_type, access_token, url, payload)
            companies = company1['Corporates']
        else:
            pos = ''
            companies = ''

        return render(request, 'Agent/corporate_accounts.html', {'accounts': pos, 'companies':companies})
    else:
        return HttpResponseRedirect("/agents/login")


def cotrav_accounts(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        payload = {'bill_type': id}
        url = settings.API_BASE_URL + "get_cotrav_accounts"
        company = getDataFromAPI(login_type, access_token, url, payload)
        pos = ''
        companies = ''
        if company['success'] == 1:
            pos = company['Accounts']
        else:
            pos = ''
            companies = ''

        return render(request, 'Agent/cotrav_accounts.html', {'accounts': pos, 'companies':companies})
    else:
        return HttpResponseRedirect("/agents/login")


def add_company_podetail(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            po_number = request.POST.get('po_number')
            po_date = request.POST.get('po_date', '')
            po_amount = request.POST.get('po_amount', '')
            po_copy = request.POST.get('po_copy', '')
            po_balance = request.POST.get('po_balance', '')

            bill_id = request.POST.get('bill_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'po_number': po_number, 'po_date': po_date+' 00:00', 'po_amount': po_amount, 'po_copy': po_copy,
                       'po_balance': po_balance,  'bill_id': bill_id, 'is_delete': delete_id, }

            url = ""
            if bill_id:
                url = settings.API_BASE_URL + "update_podetails"
                operation_message = "Company PO Detail Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_podetails"
                    operation_message = "Company PO Detail Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_podetails"
                operation_message = "Company PO Detail Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/corporate-podetails", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/corporate-podetails", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_company_accounts(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            bank_name = request.POST.get('bank_name')
            bank_branch = request.POST.get('bank_branch', '')
            acoount_no = request.POST.get('acoount_no', '')
            acoount_holder_name = request.POST.get('acoount_holder_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')
            micr_code = request.POST.get('micr_code', '')

            account_id = request.POST.get('account_id')

            delete_id = request.POST.get('delete_id')

            payload = {'corporate_id': corporate_id, 'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'bank_name': bank_name, 'bank_branch': bank_branch, 'acoount_no': acoount_no, 'acoount_holder_name': acoount_holder_name,
                       'ifsc_code': ifsc_code,'micr_code':micr_code,  'account_id': account_id, 'is_delete': delete_id, }

            url = ""
            if account_id:
                url = settings.API_BASE_URL + "update_corporate_account"
                operation_message = "Company Account Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_corporate_account"
                    operation_message = "Company Account Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_corporate_account"
                operation_message = "Company Account Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/corporate-accounts", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/corporate-accounts", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def add_cotrav_accounts(request, id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            bank_name = request.POST.get('bank_name')
            bank_branch = request.POST.get('bank_branch', '')
            acoount_no = request.POST.get('acoount_no', '')
            acoount_holder_name = request.POST.get('acoount_holder_name', '')
            ifsc_code = request.POST.get('ifsc_code', '')
            micr_code = request.POST.get('micr_code', '')

            account_id = request.POST.get('account_id')

            delete_id = request.POST.get('delete_id')

            payload = {'user_id': user_id, 'login_type': login_type,
                       'access_token': access_token, 'bank_name': bank_name, 'bank_branch': bank_branch, 'acoount_no': acoount_no, 'acoount_holder_name': acoount_holder_name,
                       'ifsc_code': ifsc_code,'micr_code':micr_code,  'account_id': account_id, 'is_delete': delete_id, }

            url = ""
            if account_id:
                url = settings.API_BASE_URL + "update_cotrav_account"
                operation_message = "Cotrav Account Updated Successfully..!"
                if delete_id == '1':
                    url = settings.API_BASE_URL + "delete_cotrav_account"
                    operation_message = "Cotrav Account Deleted Successfully..!"
            else:
                url = settings.API_BASE_URL + "add_cotrav_account"
                operation_message = "Cotrav Account Added Successfully..!"

            company = getDataFromAPI(login_type, access_token, url, payload)

            if company['success'] == 1:
                messages.success(request, operation_message)
                return HttpResponseRedirect("/agents/cotrav-accounts", {'message': "Added Successfully"})
            else:
                messages.error(request, company['message'])
                return HttpResponseRedirect("/agents/cotrav-accounts", {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def dashboard_search_api_call(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        serveType = int(request.POST.get('serveType', ''))

        booking_type = int(request.POST.get('booking_type', ''))

        bookings_from_date = request.POST.get('bookings_from_date', '')

        bookings_to_date = request.POST.get('bookings_to_date', '')


        booking_id = ""
        pickup_location = ""
        pickup_date1 = ""
        pickup_date2 = ""
        pickup_date3 = ""
        pickup_date4 = ""
        spoc_id = ""
        operator_name = ""

        city = ""
        pnr_no = ""
        ass_code = ""

        checkin_date = ""
        voucher_no = ""
        hotel_name = ""

        search_serve_url = ""

        whereClause = "1 "

        if serveType == 1 :
            booking_id = request.POST.get('booking_id', '')
            pickup_location = request.POST.get('pickup_location', '')
            pickup_date1 = request.POST.get('pickup_date1', '')
            spoc_id = request.POST.get('spoc_id', '')
            operator_name = request.POST.get('operator_name', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/taxi_bookings.html'
        if serveType == 2 :
            booking_id = request.POST.get('booking_id', '')
            pickup_date2 = request.POST.get('pickup_date2', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/bus_bookings.html'
        if serveType == 3 :
            booking_id = request.POST.get('booking_id', '')
            pickup_date3 = request.POST.get('pickup_date3', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/train_bookings.html'
        if serveType == 4:
            booking_id = request.POST.get('booking_id', '')
            pickup_date4 = request.POST.get('pickup_date4', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            pnr_no = request.POST.get('pnr_no', '')
            ass_code = request.POST.get('ass_code', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/flight_bookings.html'
        if serveType == 5 :
            booking_id = request.POST.get('booking_id', '')
            checkin_date = request.POST.get('checkin_date', '')
            city = request.POST.get('city', '')
            spoc_id = request.POST.get('spoc_id', '')
            ass_code = request.POST.get('ass_code', '')
            voucher_no = request.POST.get('voucher_no', '')
            hotel_name = request.POST.get('hotel_name', '')
            current_url = request.POST.get('current_url', '')
            search_serve_url = 'Agent/hotel_bookings.html'


        if (bookings_from_date and bookings_to_date) :
            bookings_from_date = bookings_from_date + ' 00:00:00'
            bookings_to_date = bookings_to_date + ' 00:00:00'

            bookings_from_date_object = datetime.strptime(bookings_from_date, '%d-%m-%Y %H:%M:%S')
            bookings_to_date_object = datetime.strptime(bookings_to_date, '%d-%m-%Y %H:%M:%S')

            bookings_from_date = bookings_from_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            bookings_to_date = bookings_to_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            print(bookings_from_date)

            print(bookings_to_date)

            whereClause = whereClause + "AND " + "b.booking_date BETWEEN CAST('" + bookings_from_date + "' AS DATE) AND CAST('" + bookings_to_date + "' AS DATE) "

        if pickup_location:
            whereClause = whereClause + "AND " + "b.pickup_location LIKE '%" + pickup_location + "%' "

        if pickup_date1:
            pickup_date1 = pickup_date1 + ' 00:00:00'
            pickup_date1_object = datetime.strptime(pickup_date1, '%d-%m-%Y %H:%M:%S')
            pickup_date1 = pickup_date1_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_datetime) = CAST('" + pickup_date1 + "' AS DATE) "

        if pickup_date2:
            pickup_date2 = pickup_date2 + ' 00:00:00'
            pickup_date2_object = datetime.strptime(pickup_date2, '%d-%m-%Y %H:%M:%S')
            pickup_date2 = pickup_date2_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_from_datetime) = CAST('" + pickup_date2 + "' AS DATE) "

        if pickup_date3:
            pickup_date3 = pickup_date3 + ' 00:00:00'
            pickup_date3_object = datetime.strptime(pickup_date3, '%d-%m-%Y %H:%M:%S')
            pickup_date3 = pickup_date3_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.pickup_from_datetime) = CAST('" + pickup_date3 + "' AS DATE) "

        if pickup_date4:
            pickup_date4 = pickup_date4 + ' 00:00:00'
            pickup_date4_object = datetime.strptime(pickup_date4, '%d-%m-%Y %H:%M:%S')
            pickup_date4 = pickup_date4_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.departure_datetime) = CAST('" + pickup_date4 + "' AS DATE) "

        if spoc_id:
            whereClause = whereClause + "AND "  + "b.spoc_id = '" + spoc_id + "' "

        if operator_name:
            whereClause = whereClause

        if city:
            whereClause = whereClause + "AND " + "b.pickup_location LIKE '%" + city + "%' "

        if pnr_no:
            whereClause = whereClause + "AND " + "b.pnr_no = '" + pnr_no + "' "

        if ass_code:
            whereClause = whereClause + "AND " + "b.assessment_code = '" + ass_code + "' "

        if checkin_date:
            checkin_date = checkin_date + ' 00:00:00'
            checkin_date_object = datetime.strptime(checkin_date, '%d-%m-%Y %H:%M:%S')
            checkin_date = checkin_date_object.strftime("%Y-%m-%d (%H:%M:%S.%f)")

            whereClause = whereClause + "AND " + "DATE(b.checkin_datetime) = CAST('" + checkin_date + "' AS DATE) "

        if voucher_no:
            whereClause = whereClause + "AND " + "b.voucher_no = '" + voucher_no + "' "

        if hotel_name:
            whereClause = whereClause + "AND " + "ht.name LIKE '%" + hotel_name + "%' "

        if booking_id:
            whereClause = "b.reference_no = '" + booking_id + "' "

        current_url = "/agents/agent_home"

        search_serve_url = "Agent/dashboard_search_result.html"

        payload = {'whereClause': whereClause , 'serveType': serveType , 'booking_type': booking_type }
        url = settings.API_BASE_URL + "dashboard_search_bookings"
        print(payload)
        verify = getDataFromAPI(login_type, access_token, url, payload)
        print(verify)
        if verify['success'] == 1:
            messages.success(request, "Search Result..!")
            return render(request, search_serve_url , {'bookings':verify['Result'] , 'serveType':serveType } )

        else:
            messages.error(request, "Sory for error...!")
            return HttpResponse("error")
    else:
        return HttpResponseRedirect("/agents/login")


def softdeleated_operators(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        oper_tbl_type = id
        url = settings.API_BASE_URL+"softdeleated_operators"
        payload = {'type': oper_tbl_type }
        operators = getDataFromAPI(login_type, access_token, url, payload)

        if operators['success'] == 1:
            operators = operators['Operators']

            if oper_tbl_type == 1:
                html_page = "Agent/softdeleated_operators.html"
                oper_arg = {'operators': operators}
            elif oper_tbl_type == 2:
                html_page = "Agent/softdeleated_operator_rates.html"
                oper_arg = {'op_rates': operators}
            elif oper_tbl_type == 3:
                html_page = "Agent/softdeleated_operator_drivers.html"
                oper_arg = {'op_drivers': operators}
            else:
                html_page = "Agent/softdeleated_operators.html"
                oper_arg = {'operators': operators}

            return render(request,html_page,oper_arg)
        else:
            return render(request,"Agent/softdeleated_operators.html",{'operators':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def softdeleated_companies(request,id):
    request = get_request()

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        oper_tbl_type = id
        print('typeeeee')
        print(oper_tbl_type)
        url = settings.API_BASE_URL+"softdeleated_companies"
        payload = {'type': oper_tbl_type }
        operators = getDataFromAPI(login_type, access_token, url, payload)

        if operators['success'] == 1:
            operators = operators['Companies']

            if oper_tbl_type == 1:
                html_page = "Agent/softdeleated_companies.html"
                oper_arg = {'companies': operators}
            elif oper_tbl_type == 2:
                html_page = "Agent/softdeleated_cotrav_billing_entities.html"
                oper_arg = {'billing_entities': operators}
            elif oper_tbl_type == 3:
                html_page = "Agent/softdeleated_company_rates.html"
                oper_arg = {'corporate_rates': operators}
            elif oper_tbl_type == 4:
                html_page = "Agent/softdeleated_groups.html"
                oper_arg = {'groups': operators}
            elif oper_tbl_type == 5:
                html_page = "Agent/softdeleated_subgroups.html"
                oper_arg = {'subgroups': operators}
            elif oper_tbl_type == 6:
                html_page = "Agent/softdeleated_company_admins.html"
                oper_arg = {'admins': operators}
            elif oper_tbl_type == 7:
                html_page = "Agent/softdeleated_employees.html"
                oper_arg = {'employees': operators}
            elif oper_tbl_type == 8:
                html_page = "Agent/softdeleated_assessment_cities.html"
                oper_arg = {'cities': operators}
            elif oper_tbl_type == 9:
                html_page = "Agent/softdeleated_assessment_codes.html"
                oper_arg = {'codes': operators}
            elif oper_tbl_type == 10:
                html_page = "Agent/softdeleated_corporate_management_fee.html"
                oper_arg = {'fees': operators}
            else:
                html_page = "Agent/softdeleated_corporate_management_fee.html"
                oper_arg = {'fees': operators}

            return render(request,html_page,oper_arg)
        else:
            return render(request,"Agent/softdeleated_operators.html",{'operators':{}})
    else:
        return HttpResponseRedirect("/agents/login")




def softdeleated_taxies(request,id):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']
        taxi_tbl_type = id
        url = settings.API_BASE_URL + "softdeleated_taxi"
        payload = {'type': taxi_tbl_type}
        Taxi = getDataFromAPI(login_type, access_token, url, payload)

        if Taxi['success'] == 1:
            Taxi = Taxi['Taxi']

            if taxi_tbl_type == 1:
                html_page = "Agent/softdeleated_taxis.html"
                taxi_arg = {'taxis': Taxi}
            elif taxi_tbl_type == 2:
                html_page = "Agent/softdeleated_taxitypes.html"
                taxi_arg = {'types': Taxi}
            elif taxi_tbl_type == 3:
                html_page = "Agent/softdeleated_taxi_models.html"
                taxi_arg = {'taxi_models': Taxi}
            else:
                html_page = "Agent/softdeleated_taxi_models.html"
                taxi_arg = {'taxi_models': Taxi}

            return render(request, html_page, taxi_arg)
        else:
            return render(request, "Agent/softdeleated_operators.html", {'operators': {}})
    else:
        return HttpResponseRedirect("/agents/login")


def update_softdeleated(request):
    if request.method == 'POST':

        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            record_id = int(request.POST.get('record_id', ''))
            html_page = request.POST.get('current_url', '')
            table_name = request.POST.get('table_name', '')
            current_url = request.POST.get('current_url', '')


            # payload = {'record_id': record_id , 'table_name': table_name }
            payload = {'table_name': table_name , 'record_id': record_id }
            url = settings.API_BASE_URL + "softdeleated_update"

            Taxi = getDataFromAPI(login_type, access_token, url, payload)

            print(Taxi)

            if Taxi['success'] == 1:
                return HttpResponseRedirect(html_page, {'message': "Operator is now avaiable to use"})
            else:
                return HttpResponseRedirect(current_url)
        else:
            return HttpResponseRedirect("/agents/login")
    else:

        return HttpResponseRedirect("/agents/login")


def upload_client_ticket(request):
    if request.method == 'POST':
        if 'agent_login_type' in request.session:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            user_id = int(request.POST.get('user_id', ''))
            booking_id = request.POST.get('booking_id', '')
            booking_type = request.POST.get('booking_type', '')
            is_client = request.POST.get('is_client', '')
            current_url = request.POST.get('current_url', '')

            client_ticket = request.POST.get('client_ticket')
            client_ticket_path = ""

            if not client_ticket == 1:
                if request.FILES:
                    file_up = request.FILES.get('ticketToUpload', False)
                    if file_up:
                        file_up = request.FILES['ticketToUpload']
                        if is_client == 1:
                            client_ticket_path = client_ticket_upload_get_path(file_up)
                        else:
                            client_ticket_path = vendor_ticket_upload_get_path(file_up)
                    else:
                        client_ticket_path = None

            payload = {'booking_id': booking_id , 'user_id': user_id, 'booking_type':booking_type, 'is_client':is_client, 'client_ticket_path':client_ticket_path,
                       'client_ticket':client_ticket}
            print(payload)
            url = settings.API_BASE_URL + "upload_new_ticket"

            Taxi = getDataFromAPI(login_type, access_token, url, payload)

            print(Taxi)

            if Taxi['success'] == 1:
                messages.success(request, "Ticket Uploaded Successfully...!")
                return HttpResponseRedirect(current_url, {'message': "Operator is now avaiable to use"})
            else:
                messages.error(request, "Ticket Not Upload...!"+Taxi['message'])
                return HttpResponseRedirect(current_url)
        else:
            return HttpResponseRedirect("/agents/login")
    else:

        return HttpResponseRedirect("/agents/login")


def master_select(request,id):
    request = get_request()
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        html_page = request.POST.get('current_url', '')
        table_name = request.POST.get('table_name', '')
        current_url = request.POST.get('current_url', '')

        table_flag = id
        table_title = ""

        if(table_flag == 1 ):
            table_name = "flight_airports"
            table_title = "Flight Airports"
        elif(table_flag == 2):
            table_name = "invoice_status"
            table_title = "Invoice Status"
        elif (table_flag == 3):
            table_name = "irctc_accounts"
            table_title = "Irctc Accounts"
        elif (table_flag == 4):
            table_name = "nationality"
            table_title = "Nationality"
        elif (table_flag == 5):
            table_name = "status_client"
            table_title = "Status Client"
        elif (table_flag == 6):
            table_name = "status_cotrav"
            table_title = "Status Cotrav"
        elif (table_flag == 7):
            table_name = "train_types"
            table_title = "Train Types"
        elif (table_flag == 8):
            table_name = "user_type"
            table_title = "User Type"
        else:
            table_name = "flight_airports"
            table_flag == 1
            table_title = "Flight Airports"


        html_page = "Agent/master_table.html"


        # payload = {'record_id': record_id , 'table_name': table_name }
        payload = {'table_name': table_name}

        url = settings.API_BASE_URL + "master-select"

        masters = getDataFromAPI(login_type, access_token, url, payload)

        #print(masters)

        if masters['success'] == 1:

            masters = masters['Records']

            return render(request, html_page, {'masters': masters,'table_flag':table_flag , 'table_title':table_title} )

        else:
            return HttpResponseRedirect(current_url)
    else:
        return HttpResponseRedirect("/agents/login")

def add_master(request):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')
        payload = {}
        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            html_page = request.POST.get('current_url', '')
            table_name = request.POST.get('table_name', '')
            current_url = request.POST.get('current_url', '')

            table_flag = int(request.POST.get('table_flag', ''))

            col_name = ""

            col_value = ""

            if (table_flag == 1):
                city_name = request.POST.get('city_name', '')

                country_name = request.POST.get('country_name', '')

                code = request.POST.get('code', '')

                table_name = "flight_airports"

                col_name = 'city_name' + ',' + 'country_name' + ',' + 'code'

                col_value = "'" + str(city_name) + "','" + str(country_name) + "','" + str(code) + "'"

            elif (table_flag == 2):

                status_name = request.POST.get('status_name', '')

                table_name = "invoice_status"

                col_name = 'status_name'

                col_value = "'" + str(status_name) + "'"

            elif (table_flag == 3):

                username = request.POST.get('username', '')
                password = request.POST.get('password', '')
                usage_limit = request.POST.get('usage_limit', '')
                is_used = request.POST.get('is_used', '')
                used_by = request.POST.get('used_by', '')
                usage_started_at = request.POST.get('usage_started_at', '')
                monthly_usage_count = request.POST.get('monthly_usage_count', '')
                booking_ids = request.POST.get('booking_ids', '')

                table_name = "irctc_accounts"

                col_name = 'username' + ',' + 'password' + ',' + 'usage_limit' + ',' + 'is_used' + ',' + 'used_by' + ',' + 'usage_started_at' + ',' + 'monthly_usage_count' + ',' + 'booking_ids'

                col_value = "'" + str(username) + "','" + str(password) + "','" + str(usage_limit) + "'" + "','" + str(is_used) + "'" + "','" + str(used_by) + "'" + "','" + str(usage_started_at) + "'" + "','" + str(monthly_usage_count) + "'" + "','" + str(booking_ids) + "'"

            elif (table_flag == 4):
                name = request.POST.get('name', '')

                code = request.POST.get('code', '')

                table_name = "nationality"

                col_name = 'name' + ',' + 'code'

                col_value = "'" + str(name) + "','" + str(code) + "'"

            elif (table_flag == 5):
                status = request.POST.get('status', '')

                description = request.POST.get('description', '')

                table_name = "status_client"

                col_name = 'status' + ',' + 'description'

                col_value = "'" + str(status) + "','" + str(description) + "'"

            elif (table_flag == 6):
                status = request.POST.get('status', '')

                description = request.POST.get('description', '')

                table_name = "status_cotrav"

                col_name = 'status' + ',' + 'description'

                col_value = "'" + str(status) + "','" + str(description) + "'"

            elif (table_flag == 7):
                name = request.POST.get('name', '')

                table_name = "train_types"

                col_name = 'name'

                col_value = "'" + str(name)  + "'"

            elif (table_flag == 8):
                name = request.POST.get('name', '')

                table_name = "user_type"

                col_name = 'name'

                col_value = "'" + str(name) + "'"

            payload = {'table_name':table_name,'col_name':col_name,'col_value':col_value}

            url = settings.API_BASE_URL + "master-add"

            company = getDataFromAPI(user_type, access_token, url, payload)

            if company['success'] == 1:

                return HttpResponseRedirect(current_url, {'message': "Added Successfully"})
            else:

                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def update_master(request,id):
    if request.method == 'POST':
        request = get_request()
        user_id = request.POST.get('user_id', '')
        corporate_id = request.POST.get('corporate_id', '')
        payload = {}

        if 'agent_login_type' in request.session:
            user_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            html_page = request.POST.get('current_url', '')
            table_name = request.POST.get('table_name', '')

            current_url = request.POST.get('current_url', '')

            table_flag = int(request.POST.get('table_flag', ''))

            set_clause = ""

            record_id = ""

            idd = int(request.POST.get('idd', ''))

            if (table_flag == 1):
                city_name = request.POST.get('city_name', '')

                country_name = request.POST.get('country_name', '')

                code = request.POST.get('code', '')

                table_name = "flight_airports"

                record_id = idd

                set_clause = "`city_name`='" + str(city_name) + "',`country_name`='" + str(country_name) + "',`code`='" + str(code) + "'"

            elif (table_flag == 2):

                status_name = request.POST.get('status_name', '')

                table_name = "invoice_status"

                record_id = idd

                set_clause = "`status_name`='" + str(status_name) + "'"

            elif (table_flag == 3):

                username = request.POST.get('username', '')
                password = request.POST.get('password', '')
                usage_limit = request.POST.get('usage_limit', '')
                is_used = request.POST.get('is_used', '')
                used_by = request.POST.get('used_by', '')
                usage_started_at = request.POST.get('usage_started_at', '')
                monthly_usage_count = request.POST.get('monthly_usage_count', '')
                booking_ids = request.POST.get('booking_ids', '')

                record_id = idd

                table_name = "irctc_accounts"

                set_clause = "`username`='" + str(username) + "',`password`='" + str(password) + "',`usage_limit`='" + str(usage_limit) + "',`is_used`='" + str(is_used) + "',`used_by`='" + str(used_by) + "',`usage_started_at`='" + str(usage_started_at) + "',`monthly_usage_count`='" + str(monthly_usage_count) + "',`booking_ids`=" + str(booking_ids) + "'"

            elif (table_flag == 4):
                name = request.POST.get('name', '')

                code = request.POST.get('code', '')

                table_name = "nationality"

                record_id = idd

                set_clause = "`name`='" + str(name) + "',`code`='" + str(code) + "'"

            elif (table_flag == 5):
                status = request.POST.get('status', '')

                description = request.POST.get('description', '')

                table_name = "status_client"

                record_id = idd

                set_clause = "`status`='" + str(status) + "',`description`='" + str(description) + "'"

            elif (table_flag == 6):
                status = request.POST.get('status', '')

                description = request.POST.get('description', '')

                table_name = "status_cotrav"

                record_id = idd

                set_clause = "`status`='" + str(status) + "',`description`='" + str(description) + "'"

            elif (table_flag == 7):
                name = request.POST.get('name', '')

                table_name = "train_types"

                record_id = idd

                set_clause = "`name`='" + str(name)  + "'"

            elif (table_flag == 8):
                name = request.POST.get('name', '')

                table_name = "user_type"

                record_id = idd

                set_clause = "`name`='" + str(name)  + "'"

            payload = {'table_name':table_name,'set_clause':set_clause,'record_id':record_id}

            url = settings.API_BASE_URL + "master-update"

            company = getDataFromAPI(user_type, access_token, url, payload)

            if company['success'] == 1:

                return HttpResponseRedirect(current_url, {'message': "Updated Successfully"})
            else:

                return HttpResponseRedirect(current_url, {'message': "Record Not Added"})
        else:
            return HttpResponseRedirect("/agents/login")


def get_all_cotrav_visa(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            pass
        else:
            url = settings.API_BASE_URL+"get_visa_services"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                corporates_data = company['Visa']
                return render(request,"Agent/visa_services.html",{'visa_services':corporates_data})
            else:
                return render(request,"Agent/visa_services.html",{'visa_services':{}})
    else:
        return HttpResponseRedirect("/agents/login")


def get_all_cotrav_visa_requests(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        if request.method == 'POST':
            pass
        else:
            url = settings.API_BASE_URL+"get_all_cotrav_visa_requests"
            payload = {'some': 'data'}
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                corporates_data = company['Visa']
                return render(request,"Agent/visa_services.html",{'visa_services':corporates_data})
            else:
                return render(request,"Agent/visa_services.html",{'visa_services':{}})
    else:
        return HttpResponseRedirect("/agents/login")

def add_new_visa_request(request):
    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        url = settings.API_BASE_URL + "companies"
        payload = {'some': 'data'}
        company = getDataFromAPI(login_type, access_token, url, payload)
        companies = company['Corporates']

        url_cities = settings.API_BASE_URL + "get_nationality"
        taxies1 = getDataFromAPI(login_type, access_token, url_cities, payload)
        nationalities = taxies1['Nationality']

        url_cities1 = settings.API_BASE_URL + "get_countries"
        taxies11 = getDataFromAPI(login_type, access_token, url_cities1, payload)
        Country = taxies11['Country']

        url_cities111 = settings.API_BASE_URL + "get_states"
        taxies1ds1 = getDataFromAPI(login_type, access_token, url_cities111, payload)
        states = taxies1ds1['State']

        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        cities = cities['Cities']





        return render(request,"Agent/add_visa_services.html",{'companies':companies, 'nationalities':nationalities, 'countrys':Country, 'states':states, 'cities':cities})

    else:
        return HttpResponseRedirect("/agents/login")



def reports_invoice(request):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0

            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            date_type = request.POST.get('date_type', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')

            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            payload = {'corporate_id': int(corporate_id), 'service_type': service_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}

            print("payload")

            print(payload)

            url = settings.API_BASE_URL + "report_invoice"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            #print(operator)
            if operator['success'] == 1:
                operator = operator['Reports']
            else:
                operator = {}

            return render(request, 'Agent/reports_invoice.html',
                          {'Reports': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Agent/reports_invoice.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/agents/login")



def reports_client_billing(request):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0
            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            bill_status = request.POST.get('bill_status', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'bill_status': bill_status, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "report_client_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}

            return render(request, 'Agent/reports_client_billing.html',
                          {'bills': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Agent/reports_client_billing.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/agents/login")



def reports_operator_billing(request):
    if 'agent_login_type' in request.session:
        if request.method == 'POST':
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']

            corporate_id = request.POST.get('corporate_id', '')
            if not corporate_id:
                corporate_id = 0
            service_type = request.POST.get('service_type', '')
            invoice_type = request.POST.get('invoice_type', '')
            date_type = request.POST.get('date_type', '')
            from_date = request.POST.get('from_date', '')
            to_date = request.POST.get('to_date', '')
            payload = {'corporate_id': int(corporate_id), 'service_type': service_type, 'invoice_type': invoice_type,
                       'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
            print(payload)
            url = settings.API_BASE_URL + "companies"

            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}

            return render(request, 'Agent/bills_create.html',
                          {'bills': operator, 'companies': companies, 'data': payload})
        else:
            login_type = request.session['agent_login_type']
            access_token = request.session['agent_access_token']
            payload = {'': id}

            url = settings.API_BASE_URL + "companies"
            company = getDataFromAPI(login_type, access_token, url, payload)
            if company['success'] == 1:
                companies = company['Corporates']
            else:
                companies = {}

            url = settings.API_BASE_URL + "get_all_bills"
            operator = getDataFromAPI(login_type, access_token, url, payload)
            print("Billl")
            print(operator)
            if operator['success'] == 1:
                operator = operator['Bill']
            else:
                operator = {}
            return render(request, 'Agent/reports_operator_billing.html', {'bills': operator, 'companies': companies})
    else:
        return HttpResponseRedirect("/agents/login")



def download_invoice_reports(request):

    request = get_request()

    booking = ''
    service_text = ""
    from_date = ""
    to_date = ""

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        if not corporate_id:
            corporate_id = 0
        service_type = int(request.POST.get('service_type', ''))
        date_type = request.POST.get('date_type', '')
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')


        if service_type == 1:

            service_text = "taxi"

        elif service_type == 2:

            service_text = "bus"

        elif service_type == 3:

            service_text = "train"

        elif service_type == 4:

            service_text = "flight"

        elif service_type == 5:

            service_text = "hotel"

        else:

            service_text = "all"



        payload = {'corporate_id': int(corporate_id), 'service_type': service_type,'date_type': date_type, 'from_date': from_date, 'to_date': to_date}
        print(payload)
        url = settings.API_BASE_URL + "report_invoice"
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Reports']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report_invoice_'+ service_text +'_'+ from_date +'_'+ to_date + '.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Invoice Reports'

    # Define the titles for columns

    columns = [
            "Sr.No",
            "reference_no",
            "assessment_code",
             "assessment_city_id",
            "pickup_location",
            "drop_location",
            "spoc_id",
            "user_name",
            "user_contact",
            "booking_date",
            "booking_time",
            "boarding_date",
            "boarding_time",
            "boarding_point",
            "portal_used",
            "operator_name",
            "operator_contact",
            "ticket_no",
            "pnr_no",
            "assign_bus_type_id",
            "ticket_price",
            "management_fee",
            "tax_on_management_fee",
            "tax_on_management_fee_percentage",
            "sub_total",
            "vi_ticket_price",
            "vender_commission",

            "vender_commission",
            "invoice_status",
    ]



    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1

        if service_type == 1:
            bk['booking_datetime'] = bk['booking_date']
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['portal_used'] = ''
            bk['ticket_no'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''
            bk['ticket_price'] = ''

        if service_type == 3:
            bk['vi_ticket_price'] = ''
            bk['vender_commission'] = ''
            bk['invoice_status'] = ''

        if service_type == 4 :
            bk['assessment_city_id'] = ''
            bk['pickup_location'] = bk['from_location']
            bk['drop_location'] = bk['to_location']
            bk['boarding_datetime'] = ''
            bk['boarding_point'] = ''
            bk['portal_used'] = ''
            bk['operator_name'] = ''
            bk['operator_contact'] = ''
            bk['pnr_no'] = ''
            bk['assign_bus_type_id'] = ''

        if service_type == 5 :
             bk['pickup_location'] = ''
             bk['drop_location'] = ''
             bk['boarding_datetime'] = ''
             bk['boarding_point'] = ''
             bk['ticket_no'] = ''
             bk['pnr_no'] = ''
             bk['assign_bus_type_id'] = ''


        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['pickup_location'],
            bk['drop_location'],
            bk['spoc_id'],
            bk['user_name'],
            bk['user_contact'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['boarding_datetime']),
            timeonly(bk['boarding_datetime']),
            bk['boarding_point'],
            bk['portal_used'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['ticket_price'],
            bk['management_fee'],
            bk['tax_on_management_fee'],
            bk['tax_on_management_fee_percentage'],
            bk['sub_total'],
            bk['vi_ticket_price'],


            bk['vender_commission'],

            bk['vender_commission'],
            bk['invoice_status'],

        ]





        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_client_bill_reports(request):

    request = get_request()

    booking = ''

    if 'agent_login_type' in request.session:
        login_type = request.session['agent_login_type']
        access_token = request.session['agent_access_token']

        corporate_id = request.POST.get('corporate_id', '')
        if not corporate_id:
            corporate_id = 0
        service_type = int(request.POST.get('service_type', ''))
        bill_status = int(request.POST.get('bill_status', ''))
        from_date = request.POST.get('from_date', '')
        to_date = request.POST.get('to_date', '')

        service_text = ""

        if service_type == 1:

            service_text = "taxi"

        elif service_type == 2:

            service_text = "bus"

        elif service_type == 3:

            service_text = "train"

        elif service_type == 4:

            service_text = "flight"

        elif service_type == 5:

            service_text = "hotel"

        else:

            service_text = "all"


        if bill_status == 1 :

            bill_text = "Unpaid"

        elif bill_status == 2 :

            bill_text = "partial"

        elif bill_status == 3:

            bill_text = "paid"

        else:
            bill_text = "All"


        payload = {'corporate_id': int(corporate_id), 'service_type': service_type,'bill_status': bill_status, 'from_date': from_date, 'to_date': to_date}
        print(payload)
        url = settings.API_BASE_URL + "report_client_bills"
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bill']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=report_client_bills_'+ service_text +'_'+ from_date +'_'+ to_date + '_' + bill_text +'.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Client Bill Reports'

    # Define the titles for columns

    columns = [
        "Sr.No",
        "CorporatemName ",
        "Bill Number",
        "No Of Invoices",
        "Cotrav Billing Entity",
        "Client  Billing Entity ",
        "Billing Type",
        "TDS  Deducted  By Client",
        "System  Calculated TDS",
        "IGST",
        "CGST",
        "SGST",
        "Total Amount",
        "Is Paid",
        "Payment Status",
        "Total GST Paid",
        "Management Fee",
        "Outstanding Pending Payment",
        "Paid Total Amount",
        "Balance Total Amount",
        "Advance Payment",
        "Is Offline",
        "Reimbursement Voucher",
        "ID Taxable Amount",
        "Nontaxable Amount",
        "PO Id",
        "Bill Created Date",
        "Bill Final Date",
        "User Comment",
        "Cotrav Status",
        "Client Status",
    ]



    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_id'],
            bk['bill_number'],
            bk['no_of_invoices'],
            bk['cotrav_billing_entity'],
            bk['client_billing_entity'],
            bk['billing_type'],
            bk['tds_deducted_by_client'],
            bk['system_calculated_tds'],
            bk['igst'],
            bk['cgst'],

            bk['sgst'],
            bk['total_amount'],
            bk['is_paid'],

            bk['payment_status'],
            bk['total_gst_paid'],
            bk['management_fee'],
            bk['outstanding_pending_payment'],

            bk['paid_total_amount'],
            bk['balance_total_amount'],
            bk['advance_payment'],

            bk['is_offline'],
            bk['reimbursement_voucher_id'],
            bk['taxable_amount'],
            bk['nontaxable_amount'],
            bk['po_id'],
            bk['bill_created_date'],
            bk['bill_final_date'],
            bk['user_comment'],
            bk['cotrav_status'],
            bk['client_status'],

        ]



        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response	



def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
            dict(zip([col[0] for col in desc], row))
            for row in cursor.fetchall()
    ]


def calculate_age(born):
    print(born)
    dt_str = datetime.strptime(born, '%d-%m-%Y')
    today = date.today()
    return today.year - dt_str.year - ((today.month, today.day) < (dt_str.month, dt_str.day))


def pdftest(request):
    return None


def file_upload_get_path(file_name):
    save_path = os.path.join(settings.MEDIA_ROOT, 'booking_email', str(file_name))
    path = default_storage.save(save_path, file_name)
    download_path = os.path.join(settings.MEDIA_URL,'booking_email/', str(file_name))
    return download_path


def client_ticket_upload_get_path(file_name):
    save_path = os.path.join(settings.MEDIA_ROOT, 'client_ticket', str(file_name))
    path = default_storage.save(save_path, file_name)
    download_path = os.path.join(settings.MEDIA_URL,'client_ticket/', str(file_name))
    return download_path


def vendor_ticket_upload_get_path(file_name):
    save_path = os.path.join(settings.MEDIA_ROOT, 'vendor_ticket', str(file_name))
    path = default_storage.save(save_path, file_name)
    download_path = os.path.join(settings.MEDIA_URL,'vendor_ticket/', str(file_name))
    return download_path

def file_company_doc_upload(file_name):
    save_path = os.path.join(settings.MEDIA_ROOT, 'company_doc', str(file_name))
    path = default_storage.save(save_path, file_name)
    download_path = os.path.join(settings.MEDIA_URL,'company_doc/', str(file_name))
    return download_path


def upload_visa_doc_get_path(file_name):
    a = datetime.now()
    save_path = os.path.join(settings.MEDIA_ROOT, 'visa_doc/'+str(int(a.strftime('%d%m%Y%H%M')))+"/", str(file_name))
    path = default_storage.save(save_path, file_name)
    download_path = os.path.join(settings.MEDIA_URL,'visa_doc/'+str(int(a.strftime('%d%m%Y%H%M')))+"/", str(file_name))
    return download_path