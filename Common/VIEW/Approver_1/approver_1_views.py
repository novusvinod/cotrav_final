from datetime import date, datetime
from django.http import HttpResponseRedirect
from openpyxl import Workbook
from datetime import date, datetime
from django.http import HttpResponseRedirect, HttpResponse
from django.conf import settings
from django.shortcuts import render, redirect
import requests
import json
from django_global_request.middleware import get_request
from django.contrib.auth.decorators import login_required
from django.http import HttpResponseRedirect
from openpyxl import Workbook
from django.contrib import messages

from Common.models import Corporate_Approves_1_Login_Access_Token
from landing.cotrav_messeging import Excelexport

def logout_action(request):
    if 'approves_1_login_type' in request.session:
        request = get_request()
        access_token = request.session['approves_1_access_token']
        user = Corporate_Approves_1_Login_Access_Token.objects.get(access_token=access_token)
        del request.session['approves_1_login_type']
        del request.session['approves_1_access_token']

        user.expiry_date = datetime.now()  # change field
        user.save()  # this will update only
        #logout(request)  # the user is now LogOut
        return redirect("/login")
    else:
        return redirect("/login")

def homepage(request):
    print("i ma session Data Home Page")
    for key, value in request.session.items():
        print('{} => {}'.format(key, value))
    print("i ma session Data")
    if 'approves_1_login_type' in request.session:
        user_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        print("print from home")
        print(user_type)
        print(access_token)
        print("print from home")

        payload = {'approver_1_id': request.user.id}
        url = settings.API_BASE_URL + "approver_1_dashboard"
        data = getDataFromAPI(user_type, access_token, url, payload)
        dataDashboard = data['Dashboard']
        return render(request, 'Company/Approver_1/home_page.html', {'user': request.user,'dataDashboard':dataDashboard})
    else:
        return HttpResponseRedirect("/login")


def user_profile(request):
    return render(request, 'Company/Approver_1/user_profile.html', {'user': request.user})


def company_admins(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        url = settings.API_BASE_URL + "admins"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            admins = company['Admins']
            return render(request, "Company/Approver_1/company_admins.html", {'admins': admins})
        else:
            return render(request, "Company/Approver_1/company_admins.html", {'admins': {}})
    else:
        return redirect('/login')


def company_billing_entities(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        url_city = settings.API_BASE_URL + "cities"
        cities = getDataFromAPI(login_type, access_token, url_city, payload)
        if company['success'] == 1:
            entities = company['Entitys']
            cities = cities["Cities"]
            return render(request, "Company/Approver_1/billing_entities.html",
                          {'billing_entities': entities, "cities": cities, })
        else:
            return render(request, "Company/Approver_1/billing_entities.html", {'entities': {}})
    else:
        return redirect('/login')


def company_rates(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            company_rates = company['Corporate_Retes']
            return render(request, "Company/Approver_1/company_rates.html", {'corporate_rates': company_rates})
        else:
            return render(request, "Company/Approver_1/company_rates.html", {'entities': {}})
    else:
        return redirect('/login')


def company_groups(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "groups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            groups = company['Groups']
            return render(request, "Company/Approver_1/groups.html", {'groups': groups})
        else:
            return render(request, "Company/Approver_1/groups.html", {'groups': {}})
    else:
        return redirect('/login')


def company_subgroups(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "subgroups"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            url2 = settings.API_BASE_URL + "groups"
            subgroups = company['Subgroups']
            gr = getDataFromAPI(login_type, access_token, url2, payload)
            groups = gr['Groups']
            return render(request, "Company/Approver_1/subgroups.html", {'subgroups': subgroups, 'groups': groups})
        else:
            return render(request, "Company/Approver_1/subgroups.html", {'subgroups': {}})
    else:
        return redirect('/login')


def company_spocs(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            spocs = company['Spocs']
            return render(request, "Company/Approver_1/spocs.html", {'spocs': spocs})
        else:
            return render(request, "Company/Approver_1/spocs.html", {'spocs': {}})
    else:
        return redirect('/login')


def company_employees(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        if company['success'] == 1:
            employees = company['Employees']
            return render(request, "Company/Approver_1/employees.html", {'employees': employees})
        else:
            return render(request, "Company/Approver_1/employees.html", {'employees': {}})
    else:
        return redirect('/login')


def view_company_group(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_group"
        payload = {'group_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        url = settings.API_BASE_URL + "view_group_auth"
        grp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(grp_auths)
        if company['success'] == 1:
            groups = company['Groups']
            grp_auths = grp_auths['Groups']
            return render(request, "Company/Approver_1/view_groups.html", {'group': groups, 'grp_auths': grp_auths})
        else:
            return render(request, "Company/Approver_1/view_groups.html", {'group': {}})
    else:
        return redirect('/login')


def view_company_subgroup(request, id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_subgroup"
        payload = {'subgroup_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        subgrp_auths = getDataFromAPI(login_type, access_token, url, payload)
        print(subgrp_auths)
        if company['success'] == 1:
            subgroups = company['SubGroups']
            subgrp_auths = subgrp_auths['SubGroups']
            return render(request, "Company/Approver_1/view_subgroups.html",
                          {'subgroup': subgroups, 'subgrp_auths': subgrp_auths})
        else:
            return render(request, "Company/Approver_1/view_subgroups.html", {'group': {}})
    else:
        return redirect('/login')

#################################### TAXI #############################################

def taxi_bookings(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "approver_1_taxi_bookings"
        payload = {'approver_1_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/taxi_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Approver_1/taxi_bookings.html", {'': {}})
    else:
        return redirect('/login')


def view_taxi_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_taxi_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/view_taxi_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Approver_1/view_taxi_booking.html", {'': {}})
    else:
        return redirect('/login')


def accept_taxi_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_accept_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Accept Taxi Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Fails"})
    else:
        return redirect('/login')


def reject_taxi_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_reject_taxi_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Taxi Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Reject Taxi Booking..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Fails"})
    else:
        return redirect('/login')

############################################## BUS ############################################


def bus_bookings(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "approver_1_bus_bookings"
        payload = {'approver_1_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/bus_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Approver_1/bus_bookings.html", {'': {}})
    else:
        return redirect('/login')


def view_bus_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_bus_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/view_bus_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Approver_1/view_bus_booking.html", {'': {}})
    else:
        return redirect('/login')


def accept_bus_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_accept_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Accept Bus Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def reject_bus_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_reject_bus_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Bus Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Reject Bus Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')

############################################## TRAIN ############################################


def train_bookings(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "approver_1_train_bookings"
        payload = {'approver_1_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/train_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Approver_1/train_bookings.html", {'': {}})
    else:
        return redirect('/login')


def view_train_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_train_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/view_train_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Approver_1/view_train_booking.html", {'': {}})
    else:
        return redirect('/login')


def accept_train_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_accept_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Accept Train Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def reject_train_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_reject_train_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Train Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Reject Train Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')

############################################## HOTELS ############################################


def hotel_bookings(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "approver_1_hotel_bookings"
        payload = {'approver_1_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/hotel_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Approver_1/hotel_bookings.html", {'': {}})
    else:
        return redirect('/login')


def view_hotel_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_hotel_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/view_hotel_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Approver_1/view_hotel_booking.html", {'': {}})
    else:
        return redirect('/login')


def accept_hotel_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_accept_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Accept Hotel Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def reject_hotel_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_reject_hotel_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Hotel Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url , {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail To Reject Hotel Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')

############################################## FLIGHT ############################################


def flight_bookings(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id

        url = settings.API_BASE_URL + "approver_1_flight_bookings"
        payload = {'approver_1_id': user_id,'booking_type':id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/flight_bookings.html",{'bookings': booking,'booking_type':id})
        else:
            return render(request, "Company/Approver_1/flight_bookings.html", {'': {}})
    else:
        return redirect('/login')


def view_flight_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "view_flight_booking"
        payload = {'booking_id': id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']
            return render(request, "Company/Approver_1/view_flight_booking.html",{'bookings': booking})
        else:
            return render(request, "Company/Approver_1/view_flight_booking.html", {'': {}})
    else:
        return redirect('/login')


def accept_flight_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_accept_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        print(payload)
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Accepted Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Accept Flight Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')


def reject_flight_booking(request,id):
    request = get_request()

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']
        user_id = request.user.id
        booking_id = request.POST.get('booking_id', '')
        current_url = request.POST.get('current_url', '')

        url = settings.API_BASE_URL + "approver_1_reject_flight_booking"
        payload = {'booking_id': booking_id,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            messages.success(request, 'Flight Booking Rejected Successfully..!')
            return HttpResponseRedirect(current_url, {'message': "Operation Successfully"})
        else:
            messages.error(request, 'Fail to Reject Flight Booking..!')
            return HttpResponseRedirect(current_url,{'message': "Operation Fails"})
    else:
        return redirect('/login')

####################### Download MIS ##################################

def dateonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_date = str(datetime_object.day) + "/" + str(datetime_object.month) + "/" + str(datetime_object.year)
            return booking_date
        else:
            return ''
    except ValueError:
        return ''


def timeonly(dt=''):
    try:
        if (dt):
            datetime_str = dt
            datetime_object = datetime.strptime(datetime_str, '%d-%m-%Y %H:%M')
            booking_time = str(datetime_object.hour) + ":" + str(datetime_object.hour)
            return booking_time
        else:
            return ''
    except ValueError:
        return ''


def download_taxi_bookings(request):
    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_taxi_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

        # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-taxi-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Taxi Bookings'

    # Define the titles for columns

    columns = [

        'Booking ID',
        'City',
        'Assessment Code',
        'Assessment City',
        'Reason for Booking',
        'Zone',

        'Group Name',
        'Subgroup Name',
        'SPOC Name',

        'Passengers',

        'Booking Date',
        'Booking Time',

        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        'Approver1 Action',

        'Approver1 Name',

        'Approver1 Date',

        'Approver1 Time',

        'Approver2 Action',

        'Approver2 Name',

        'Approver2 Date',

        'Approver2 Time',

        'Approved Date',

        'Approved Time',

        'Approved By',

        'Canceled Date',

        'Canceled Time',

        'Canceled By',

        'Assigned Date',

        'Assigned Time',

        'Assigned By',

        'Pickup Location',
        'Drop Location',
        'Pickup Date',
        'Pickup Time',
        'Drop Date',
        'Drop Time',

        'Package Name',
        'Tour Type',
        'Vehicle Type',

        'Driver Name',
        'Driver Contact	',
        'Taxi Reg No.',
        'No. Of Seats',



        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1
        spoc_status =''
        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by= ''
        spoc_canceled_date =''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

            if bk['spoc_status'] == 1:
                spoc_status = "In-Active"
            else:
                spoc_status = "Active"

            if (act['action'] == 3 and act['user_type'] == 6):
                print('canceled')
                spoc_canceled_by = act['employee_name']
                spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [
            bk['reference_no'],
            bk['assessment_city_id'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],

            bk['zone_name'],
            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_date']),
            timeonly(bk['booking_date']),
            dateonly(bk['pickup_datetime']),
            timeonly(bk['pickup_datetime']),

            bk['package_name'],
            bk['tour_type'],
            bk['taxi_type_name'],

            bk['driver_name'],
            bk['driver_contact'],
            '',
            bk['no_of_seats'],

            bk['client_status'],

            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_bus_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_bus_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-bus-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Bus Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Assessment Code',
        'Assessment City',
        'Reason For Booking',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        "SPOC Cancel Date",
        "SPOC Cancel Time",
        'Approver1 Action',
        'Approver1 Name',
        'Approver1 Date',
        'Approver1 Time',
        'Approver2 Action',
        'Approver2 Name',
        'Approver2 Date',
        'Approver2 Time',
        'Approved Date',
        'Approved Time',
        'Approved By',
        'Canceled Date',
        'Canceled Time',
        'Canceled By',
        'Assigned Date',
        'Assigned Time',
        'Assigned By',
        'Passanger Name',
        'Pickup City',
        'Drop City',
        'Journey Date',
        'Journey Time',
        'Current Booking Status',
        'Bus Type Allocated',
        'PNR Number',

        "Client Status",
        "Cotrav Status",

    ]
    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_canceled_by=''
        spoc_canceled_date=''
        spoc_status=''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']


                # Define the data for each cell in the row
        row = [

            bk['reference_no'],

            bk['assessment_code'],
            bk['assessment_city_id'],

            bk['reason_booking'],

            bk['zone_name'],

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            dateonly(bk['booking_datetime']),

            timeonly(bk['booking_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            passanger_list,

            bk['pickup_location'],

            bk['drop_location'],

            dateonly(bk['pickup_from_datetime']),

            timeonly(bk['pickup_from_datetime']),

            bk['cotrav_status'],

            bk['assign_bus_type_id'],

            bk['pnr_no'],

            bk['client_status'],
            bk['cotrav_status'],

        ]
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_train_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_train_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-train-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Train Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "Pickup City",
        "Drop City",
        "Booking Date",
        "Booking Time",
        "Journey Date",
        "Journey Time",
        "SPOC Status",
        "SPOC Cancel Date",
        "SPOC Cancel Time",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        "Zone",
        "Coach Type Allocated",
        "Quota Used",

        'No of seats',
        'Operator name',
        'Operator contact',
        'Train name',
        'Ticket no',
        'pnr no',
        'Assign bus type id',
        'Seat no',
        'Portal used',

        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''
        spoc_status=''
        spoc_canceled_by=''
        spoc_canceled_date=''

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['reason_booking'],
            bk['pickup_location'],
            bk['drop_location'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),
            dateonly(bk['pickup_from_datetime']),
            timeonly(bk['pickup_from_datetime']),

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            bk['zone_name'],
            bk['train_type_priority_1'],
            bk['seat_no'],


            bk['no_of_seats'],
            bk['operator_name'],
            bk['operator_contact'],
            bk['train_name'],
            bk['ticket_no'],
            bk['pnr_no'],
            bk['assign_bus_type_id'],
            bk['seat_no'],
            bk['portal_used'],

            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_flight_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_flight_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Bookings']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-flight-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Flight Bookings'

    # Define the titles for columns

    columns = [

        "Booking ID",
        "Assessment Code",
        "Assessment City",
        "Booking Remarks",
        "From City",
        "To City",
        "Booking Date",
        "Booking Time",
        "Departure Date",
        "Departure Time",

        "Return Date",
        "Booking Status",

        "SPOC Status",
        "SPOC Cancel By",
        "SPOC Cancel Date",

        "Usage Type",
        "Trip Type",
        "Flight Type",
        "Seat Type",

        "Approver1 Action",
        "Approver1 Name",
        "Approver1 Action  Date",
        "Approver1 Action  Time",

        "Approver2 Action",
        "Approver2 Name",
        "Approver2 Action  Date",
        "Approver2 Action  Time",

        "Approved Date",
        "Approved Time",
        "Approved By",

        "Reject Date",
        "Reject Time",
        "Reject By",

        "Assign Date",
        "Assign Time",
        'Assigned By',

        "Group Name",
        "Subgroup Name",
        "SPOC Name",

        "Passengers",

        'First Flight Name',
        'First Flight No',
        'First Flight PNR No.',
        'First Flight From',
        'First Flight To',
        'First Flight Departure Date',
        'First Flight Departure time',
        'First Flight Arrival Date',
        'First Flight Arrival time',

        'Second Flight Name',
        'Second Flight No',
        'Second Flight PNR No.',
        'Second Flight From',
        'Second Flight To',
        'Second Flight Departure Date',
        'Second Flight Departure Time',
        'Second Flight Arrival Datetime',
        'Second Flight Arrival Time',

        'Third Flight Name',
        'Third Flight No',
        'Third Flight PNR No.',
        'Third Flight From',
        'Third Flight To',
        'Third Flight Departure Date',
        'Third Flight Departure Time',
        'Third Flight Arrival Date',
        'Third Flight Arrival Time',



        "Client Status",
        "Cotrav Status",

    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        flight_name1 = ''
        flight_no1 = ''
        pnr_no1 = ''
        from_city1 = ''
        to_city1 = ''
        departure_datetime1 = ''
        arrival_datetime1 = ''

        flight_name2 = ''
        flight_no2 = ''
        pnr_no2 = ''
        from_city2 = ''
        to_city2 = ''
        departure_datetime2 = ''
        arrival_datetime2 = ''

        flight_name3 = ''
        flight_no3 = ''
        pnr_no3 = ''
        from_city3 = ''
        to_city3 = ''
        departure_datetime3 = ''
        arrival_datetime3 = ''

        spoc_status= ''
        spoc_canceled_by=''
        spoc_canceled_date=''


        if len(bk['Flights']) == 1:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

        if len(bk['Flights']) == 2:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

        if len(bk['Flights']) == 3:
            flight_name1 = bk['Flights'][0]['flight_name']
            flight_no1 = bk['Flights'][0]['flight_no']
            pnr_no1 = bk['Flights'][0]['pnr_no']
            from_city1 = bk['Flights'][0]['from_city']
            to_city1 = bk['Flights'][0]['to_city']
            departure_datetime1 = bk['Flights'][0]['departure_datetime']
            arrival_datetime1 = bk['Flights'][0]['arrival_datetime']

            flight_name2 = bk['Flights'][1]['flight_name']
            flight_no2 = bk['Flights'][1]['flight_no']
            pnr_no2 = bk['Flights'][1]['pnr_no']
            from_city2 = bk['Flights'][1]['from_city']
            to_city2 = bk['Flights'][1]['to_city']
            departure_datetime2 = bk['Flights'][1]['departure_datetime']
            arrival_datetime2 = bk['Flights'][1]['arrival_datetime']

            flight_name3 = bk['Flights'][2]['flight_name']
            flight_no3 = bk['Flights'][2]['flight_no']
            pnr_no3 = bk['Flights'][2]['pnr_no']
            from_city3 = bk['Flights'][2]['from_city']
            to_city3 = bk['Flights'][2]['to_city']
            departure_datetime3 = bk['Flights'][2]['departure_datetime']
            arrival_datetime3 = bk['Flights'][2]['arrival_datetime']

        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"

                if (act['action'] == 3 and act['user_type'] == 6):
                    print('canceled')
                    spoc_canceled_by = act['employee_name']
                    spoc_canceled_date = act['action_date']

        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            bk['assessment_code'],
            bk['assessment_city'],
            "Booking Remarks",
            bk['from_location'],
            bk['to_location'],

            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            dateonly(bk['departure_datetime']),
            timeonly(bk['departure_datetime']),

            "",
            bk['cotrav_status'],

            spoc_status,
            spoc_canceled_by,
            spoc_canceled_date,

            bk['usage_type'],
            bk['trip_type'],
            bk['flight_type'],
            bk['seat_type'],

            approver1_action,

            approver1,

            dateonly(approver1_date),

            timeonly(approver1_date),

            approver2_action,

            approver2,

            dateonly(approver2_date),

            timeonly(approver2_date),

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            dateonly(canceled_date),

            timeonly(canceled_date),

            canceled_by,

            dateonly(assigned_date),

            timeonly(assigned_date),

            assigned_by,

            bk['group_name'],

            bk['subgroup_name'],

            bk['spoc_name'],

            passanger_list,

            flight_name1,
            flight_no1,
            pnr_no1,
            from_city1,
            to_city1,
            dateonly(departure_datetime1),
            timeonly(departure_datetime1),
            dateonly(arrival_datetime1),
            timeonly(arrival_datetime1),

            flight_name2,
            flight_no2,
            pnr_no2,
            from_city2,
            to_city2,
            dateonly(departure_datetime2),
            timeonly(departure_datetime2),
            dateonly(arrival_datetime2),
            timeonly(arrival_datetime2),

            flight_name3,
            flight_no3,
            pnr_no3,
            from_city3,
            to_city3,
            dateonly(departure_datetime3),
            timeonly(departure_datetime3),
            dateonly(arrival_datetime3),
            timeonly(arrival_datetime3),



            bk['client_status'],
            bk['cotrav_status'],

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_hotel_bookings(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    user_type = {
        1: 'Admin',
        2: 'Approver 2',
        3: 'Approver 1',
        4: 'SPOC',
        5: 'Flight Approver',
        10: 'Agent',
        6: 'Employee',
    }

    cotrav_sts = {
        1: 'Created',
        2: 'Accepted',
        3: 'Canceled',
        4: 'Assigned',

    }

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        user_id = request.POST.get('user_id', '')
        filter_by = request.POST.get('filter_by', '')
        booking_from_datetime = request.POST.get('booking_from_datetime', '')
        booking_to_datetime = request.POST.get('booking_to_datetime', '')
        booking_status = request.POST.get('booking_status', '')
        company_name = request.POST.get('company_name', '')

        url = settings.API_BASE_URL + "report_hotel_booking"
        payload = {'filter_by': filter_by, 'booking_from_datetime': booking_from_datetime,
                   'booking_to_datetime': booking_to_datetime, 'booking_status': booking_status,
                   'company_name': company_name,'user_id':user_id}
        company = getDataFromAPI(login_type, access_token, url, payload)
        print(company)
        if company['success'] == 1:
            booking = company['Bookings']



    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename={date}-hotel-bookings.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Hotel Bookings'

    # Define the titles for columns

    columns = [
        'Booking ID',
        'Billing Entity',
        'Travel request Code',
        'Assessment Code',
        'Assessment City',
        'Zone',
        'Group Name',
        'Subgroup Name',
        'From City',
        'To City',
        'SPOC Name',
        'Booking Date',
        'Booking Time',
        'SPOC Status',
        'Approver Name',
        'Approved Date',
        'Approved Time',
        'Approver Status',
        'Bucket Request Date',
        'Bucket Request Time',
        'Bucket Approved Date',
        'Bucket Approved Time',
        'Preferred Hotel',
        'Assigned Hotel',
        'Assigned Hotel Address',
        'Hotel Contact',
        'Assign Date',
        'Assign Time',
        'TaxiVaxi Status',
        'Employees Name',
        'No. of Persons',
        'Check IN Date',
        'Check IN Time',
        'Check OUT Date',
        'Check OUT Time',
        'Booking Reason',
        'Higher Bucket Requested',
        'Reason for Higher Bucket',
        'Rejected By',
        'Reject Reason',
        'Reject Date',
        'Reject Time',

        'Is Prepaid',
        'Daily Breakfast',
        'Is Room AC',

        "Client Status",
        "Cotrav Status",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    for bk in booking:
        row_num += 1

        passanger_list = ''
        created_by = ''
        created_date = ''
        accepted_by = ''
        accepted_date = ''
        approved_by = ''
        approved_date = ''
        approved_status = ''
        assigned_by = ''
        assigned_date = ''
        canceled_by = ''
        canceled_date = ''
        created_user_type = ''
        accepted_user_type = ''
        approved_user_type = ''
        canceled_user_type = ''
        assigned_user_type = ''
        approver1 = ''
        approver1_date = ''
        approver1_action = ''

        approver2 = ''
        approver2_date = ''
        approver2_action = ''

        spoc_status = ''
        is_prepaid = ''
        daily_brakefast = ''
        is_ac_room = ''

        if bk['spoc_status'] == 1:
            spoc_status = "In-Active"
        else:
            spoc_status = "Active"
        if bk['is_prepaid'] == 1:
            is_prepaid = "Yes"
        else:
            is_prepaid = "No"
        if bk['daily_brakefast'] == 1:
            daily_brakefast = "Yes"
        else:
            daily_brakefast = "No"
        if bk['is_ac_room'] == 1:
            is_ac_room = "Yes"
        else:
            is_ac_room = "No"


        if (bk['Passangers']):
            if len(bk['Passangers']) > 1:
                print(bk['Passangers'])
                for psg in bk['Passangers']:
                    passanger_list = passanger_list + str(psg['employee_name']) + ","
            else:
                passanger_list = bk['Passangers'][0]['employee_name']
        if len(bk['Actions']) >= 0:
            for act in bk['Actions']:
                if (act['user_type'] == 2):
                    approver1 = act['employee_name']
                    approver1_date = act['action_date']
                    approver1_action = cotrav_sts.get(act['action'], " ")

                if (act['user_type'] == 3):
                    approver2 = act['employee_name']
                    approver2_date = act['action_date']
                    approver2_action = cotrav_sts.get(act['action'], " ")

                if (act['action'] == 1):
                    print('created')
                    created_by = act['employee_name']
                    created_date = act['action_date']
                    created_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('accepted')
                    accepted_by = act['employee_name']
                    accepted_date = act['action_date']
                    accepted_user_type = user_type.get(act['user_type'], " ")
                if (act['action'] == 2):
                    print('approved')
                    approved_by = act['employee_name']
                    approved_date = act['action_date']
                    approved_status = 'Accepted'
                    approved_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 3):
                    print('canceled')
                    canceled_by = act['employee_name']
                    canceled_date = act['action_date']
                    canceled_user_type = user_type.get(act['user_type'], " ")

                if (act['action'] == 4):
                    print('assigned')
                    assigned_by = act['employee_name']
                    assigned_date = act['action_date']
                    assigned_user_type = user_type.get(act['user_type'], " ")

                if bk['spoc_status'] == 1:
                    spoc_status = "In-Active"
                else:
                    spoc_status = "Active"
        # Define the data for each cell in the row
        row = [

            bk['reference_no'],
            'Billing Entity',
            'Travel request Code',
            bk['assessment_code'],
            bk['assessment_city_id'],
            bk['zone_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['from_city_name'],

            bk['from_area_id_name'],
            bk['spoc_name'],
            dateonly(bk['booking_datetime']),
            timeonly(bk['booking_datetime']),

            spoc_status,

            approver1,

            dateonly(approved_date),

            timeonly(approved_date),

            approved_by,

            '',
            '',
            '',
            '',

            bk['preferred_hotel'],
            bk['assign_hotel_id'],
            bk['operator_name'],
            bk['operator_contact'],
            dateonly(assigned_date),
            timeonly(assigned_date),
            bk['status_cotrav'],
            passanger_list,
            bk['no_of_seats'],
            dateonly(bk['checkin_datetime']),
            timeonly(bk['checkin_datetime']),
            dateonly(bk['checkout_datetime']),
            timeonly(bk['checkout_datetime']),
            bk['reason_booking'],
            bk['bucket_priority_1'],
            bk['bucket_priority_2'],
            canceled_by,
            '',
            dateonly(canceled_date),
            timeonly(canceled_date),

            is_prepaid,
            daily_brakefast,
            is_ac_room,

            bk['client_status'],
            bk['cotrav_status'],


        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_billing_entities(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "billing_entities"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Entitys']

            # print(booking)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=BillingEntitys.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Billing Entitys'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Billing City",
        "Entity Name",
        "Contact Person Name",
        "Contact Person Email",
        "Contact Person Phone No",
        "Address Line 1",
        "Address Line 2",
        "Address Line 3",
        "GST NO",
        "PAN No",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['billing_city'],
            bk['entity_name'],
            bk['contact_person_name'],
            bk['contact_person_email'],
            bk['contact_person_no'],
            bk['address_line_1'],
            bk['address_line_2'],
            bk['address_line_3'],
            bk['gst_id'],
            bk['pan_no'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_rates(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "company_rates"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Corporate_Retes']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyRates.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Rates'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "City",
        "Taxi Type",
        "Tour Type",
        "KMS",
        "HOURS",
        "KM Rate",
        "HR Rate",
        "Base Rate",
        "Night Rate",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['tour_type'] == 2:
            tour_type = "Local"
        else:
            tour_type = "Outstation"
        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['city_name'],
            bk['taxi_type'],
            tour_type,
            bk['kms'],
            bk['hours'],
            bk['km_rate'],
            bk['hour_rate'],
            bk['base_rate'],
            bk['night_rate'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def download_spocs(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "spocs"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Spocs']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanySpocs.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Spocs'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Group Name",
        "Subgroup Name",
        "Spoc Company ID",
        "Spoc Name",
        "Spoc Email",
        "Spoc Contact No ",
        "Login UserName ",
        "Last Login ",
        "Is Radio booking",
        "Is Local booking",
        "Is Outstation booking",
        "Is Bus booking",
        "Is Train booking",
        "Is Hotel booking",
        "Is Flight booking",
        "Is Water Bottles booking",
        "Is Reverse Logistics booking",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_radio'] == 0:
            is_radio = "No"
        else:
            is_radio = "Yes"

        if bk['is_local'] == 0:
            is_local = "No"
        else:
            is_local = "Yes"

        if bk['is_outstation'] == 0:
            is_outstation = "No"
        else:
            is_outstation = "Yes"
        if bk['is_bus'] == 0:
            is_bus = "No"
        else:
            is_bus = "Yes"
        if bk['is_train'] == 0:
            is_train = "No"
        else:
            is_train = "Yes"
        if bk['is_hotel'] == 0:
            is_hotel = "No"
        else:
            is_hotel = "Yes"
        if bk['is_flight'] == 0:
            is_flight = "No"
        else:
            is_flight = "Yes"
        if bk['is_water_bottles'] == 0:
            is_water_bottles = "No"
        else:
            is_water_bottles = "Yes"
        if bk['is_reverse_logistics'] == 0:
            is_reverse_logistics = "No"
        else:
            is_reverse_logistics = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['group_name'],
            bk['subgroup_name'],
            bk['user_cid'],
            bk['user_name'],
            bk['email'],
            bk['user_contact'],
            bk['username'],
            bk['last_login'],
            is_radio,
            is_local,
            is_outstation,
            is_bus,
            is_train,
            is_hotel,
            is_flight,
            is_water_bottles,
            is_reverse_logistics,
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response



def download_employees(request):
    """
    Downloads all movies as Excel file with a single worksheet
    """
    request = get_request()

    booking = ''

    if 'approves_1_login_type' in request.session:
        login_type = request.session['approves_1_login_type']
        access_token = request.session['approves_1_access_token']

        url = settings.API_BASE_URL + "employees"
        payload = {'corporate_id': request.user.corporate_id}
        company = getDataFromAPI(login_type, access_token, url, payload)

        if company['success'] == 1:
            booking = company['Employees']

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = 'attachment; filename=CompanyEmployees.xlsx'.format(
        date=datetime.now().strftime('%Y-%m-%d'),
    )
    workbook = Workbook()

    # Get active worksheet/tab
    worksheet = workbook.active
    worksheet.title = 'Company Employees'

    # Define the titles for columns

    columns = [
        "Sr No",
        "Corporate Name",
        "Spoc Name",
        "Core Employee ID",
        "Employee Company ID",
        "Employee Name",
        "Employee Email",
        "Employee Phone No",
        "Login UserName",
        "Age",
        "Gender",
        "ID Proof Type",
        "ID No",
        "Is CXO",
        "Designation",
        "Home City",
        "Home Address",
        "Assistant ID",
        "Date Of Birth",
        "Billing Entity",
        "Last Login",
        "Is Active",
    ]

    row_num = 1

    # Assign the titles for each cell of the header
    for col_num, column_title in enumerate(columns, 1):
        cell = worksheet.cell(row=row_num, column=col_num)
        cell.value = column_title

    # Iterate through all movies
    is_deleted = ""
    tour_type = ''
    for bk in booking:
        row_num += 1
        if bk['is_deleted'] == 0:
            is_deleted = "Active"
        else:
            is_deleted = "In-Active"

        if bk['is_cxo'] == 0:
            is_cxo = "No"
        else:
            is_cxo = "Yes"

        # Define the data for each cell in the row
        row = [

            row_num - 1,
            bk['corporate_name'],
            bk['user_name'],
            bk['core_employee_id'],
            bk['employee_cid'],
            bk['employee_name'],
            bk['employee_email'],
            bk['employee_contact'],
            bk['username'],
            bk['age'],
            bk['gender'],
            bk['id_proof_type'],
            bk['id_proof_no'],
            is_cxo,
            bk['designation'],
            bk['home_city'],
            bk['home_address'],
            bk['assistant_id'],
            bk['date_of_birth'],
            bk['billing_entity_id'],
            bk['last_login'],
            is_deleted,

        ]

        # Assign the data for each cell of the row
        for col_num, cell_value in enumerate(row, 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            cell.value = cell_value

    workbook.save(response)

    return response


def getDataFromAPI(login_type, access_token, url, payload):
    headers = {'Authorization': 'Token ' + access_token, 'usertype': login_type}
    r = requests.post(url, data=payload, headers=headers)
    api_response = json.loads(r.text)
    r.close()
    return api_response


def dictfetchall(cursor):
    "Returns all rows from a cursor as a dict"
    desc = cursor.description
    return [
        dict(zip([col[0] for col in desc], row))
        for row in cursor.fetchall()
    ]
