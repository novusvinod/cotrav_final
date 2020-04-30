from django.shortcuts import render,redirect

from django.http import HttpResponse

from django.core.mail import EmailMultiAlternatives, get_connection

from django.core.mail import send_mail

from django.core import mail

from django.utils.html import strip_tags

from django.template.loader import render_to_string

from datetime import datetime
from datetime import timedelta
from openpyxl import Workbook

from io import BytesIO

from django.template.loader import get_template

from xhtml2pdf import pisa

import os
from random import randint

from django.core.files import File

from threading import Thread, activeCount
from landing.utils import render_to_pdf
from landing.models import ManagementFee

TAX_PERC = 0.18

OPER_TAX_PERC = 0.18


class BookingEmail:

    def __init__(self,booking_id,pickup_city,drop_city,preferred_train,pickup_time,drop_time):

        self.booking_id = booking_id
        self.pickup_city = pickup_city
        self.drop_city = drop_city
        self.preferred_train = preferred_train
        self.pickup_time = pickup_time
        self.drop_time = drop_time
        self.Passangers = [{'employee_name':'sannn','age':'12'}]
        self.send_to = "sanketongmel@gmail.com"

    def send_email (self,file: list):

        subject, from_email, to = 'hello', 'contact@cotrav.co', self.send_to
        text_content = 'This email message contain train booking details.'
        html_content = render_to_string("email_voucher_template/booking_email_template.html",{'booking_id':self.booking_id,'pickup_city':self.pickup_city,'Passangers':self.Passangers})

        msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
        msg.content_subtype = "html"
        msg.attach_alternative(html_content, "text/html")
        msg.attach_file(file[1])
        # msg.attach(attach.name, attach.read(), attach.content_type)
        #msg.send()
        res = msg.send()

        return (res)


    def send_sms (self):

        self.res = 1

        return self.res


class Bill():

    def __init__(self, params):
        self.params = params


    def get_file_path(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/bill_template/bill_single_invoice.html', params)
        file = Render.render_to_bill_pdf('pdf_voucher_template/bill_template/bill_single_invoice.html', params)
        return file

    def get_file_path_multiple_invoice(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/bill_template/bill_multiple_invoice.html', params)
        file = Render.render_to_bill_pdf('pdf_voucher_template/bill_template/bill_multiple_invoice.html', params)
        return file

    def get_file_path_rembusment(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/bill_template/bill_single_invoice.html', params)
        file = Render.render_to_bill_pdf_bill('pdf_voucher_template/bill_template/bill_single_invoice.html', params)
        return file

    def get_file_path_tax(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/bill_template/bill_single_invoice.html', params)
        file = Render.render_to_bill_pdf_bill('pdf_voucher_template/bill_template/bill_single_invoice.html', params)
        return file

class Bus():

    def __init__(self, params):
        self.params = params


    def get(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/bus_voucher.html', params)
        file = Render.render_to_file('pdf_voucher_template/bus_voucher.html', params)

        return file


class Flight():

    def __init__(self, params):
        self.params = params

    def get(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/flight_voucher.html', params)
        file = Render.render_to_file('pdf_voucher_template/flight_voucher.html', params)
        return file


class Hotel():
    def __init__(self, params):
        self.params = params

    def get(self, request):
        params = self.params
        pdf = render_to_pdf('pdf_voucher_template/hotel_voucher.html', params)
        file = Render.render_to_file('pdf_voucher_template/hotel_voucher.html', params)
        return file


class SignupEmail():

    def __init__(self,company,company_location,cp_name,cp_no,cp_email,message):
         self.company = company
         self.company_location = company_location
         self.cp_name = cp_name
         self.cp_no = cp_no
         self.cp_email = "vinod@taxivaxi.com "
         self.message = message
         self.send_to = cp_email

    def send_email(self):
         email_subject = "Thanks for joining us"
         tempalte = "signup_welcome.html"
         email_body = 'Thank You For Showing Interest. We will Contact You Soon'
         connection = get_connection()  # uses SMTP server specified in settings.py
         connection.open()  # If you don't open the connection manually, Django will automatically open, then tear down the connection in msg.send()
         html_content = render_to_string(tempalte, {'company':self.company,'location':self.company_location,'cp_name':self.cp_name,'cp_no':self.cp_no,'cp_email':self.cp_email,'message':self.message})
         msg = EmailMultiAlternatives(email_subject, email_body, 'balwant@taxivaxi.in', [self.cp_email])
         msg.attach_alternative(html_content, "text/html")

         res = msg.send(fail_silently=True)
         connection.close()  # Cleanup
         # res = 1
         
         return (res)

    def reminder_email(self):
        email_subject = "Lead Generation Reminder"
        tempalte = "signup_welcome.html"
        email_body = 'Lead With Following Company Details Contacted US Again..!'
        connection = get_connection()  # uses SMTP server specified in settings.py
        connection.open()  # If you don't open the connection manually, Django will automatically open, then tear down the connection in msg.send()
        html_content = render_to_string(tempalte, {'company': self.company, 'location': self.company_location,
                                                   'cp_name': self.cp_name, 'cp_no': self.cp_no,
                                                   'cp_email': self.cp_email, 'message': self.message})
        msg = EmailMultiAlternatives(email_subject, email_body, 'balwant@taxivaxi.in', [self.cp_email])
        msg.attach_alternative(html_content, "text/html")

        res = msg.send(fail_silently=True)
        connection.close()  # Cleanup
        # res = 1

        return (res)

    

class LeadGenerationEmail():

    def __init__(self,company,company_location,cp_name,cp_no,cp_email,message,status):
         self.company = company
         self.company_location = company_location
         self.cp_name = cp_name
         self.cp_no = cp_no
         self.cp_email = cp_email
         self.message = message
         self.status = status
         self.send_to = "sanketongmel@gmail.com,vinod@taxivaxi.com,balwant@taxivaxi.in"


    def lead_create_send_email(self):
         subject, from_email, to = 'New Lead Is Generated', self.cp_email , self.send_to
         text_content = 'New Lead is generated..!!'
         html_content = render_to_string("landing/lead_generated_email_template.html", {'company':self.company,'location':self.company_location,'cp_name':self.cp_name,'cp_no':self.cp_no,'cp_email':self.cp_email,'message':self.message,'status':self.status})

         msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
         msg.content_subtype = "html"
         msg.attach_alternative(html_content, "text/html")
         
         #msg.attach_file(file[1])
         res = msg.send(fail_silently=True)
         
         return 1 

    def lead_updated_send_email(self,agent):
         self.send_to = self.send_to + " , " + agent
         subject, from_email, to = 'Lead updated intimation', self.cp_email , self.send_to
         text_content = 'lead is updated and agent is assigned..!!'
         html_content = render_to_string("landing/lead_updated_email_template.html", {'company':self.company,'location':self.company_location,'cp_name':self.cp_name,'cp_no':self.cp_no,'cp_email':self.cp_email,'message':self.message,'status':self.status,'agent_assigned':agent})

         msg = EmailMultiAlternatives(subject, text_content, from_email, [to])
         msg.content_subtype = "html"
         msg.attach_alternative(html_content, "text/html")
         
         #msg.attach_file(file[1])
         res = msg.send(fail_silently=True)
         
         return 1      



class Excelexport():

    def __init__(self,ar):
        self.row = []
        self.keys = []
        self.value = []
        self.row_num = 1

        self.jsonar = ar

        self.columns = []


    def myexport(self):

        workbook = Workbook()
    
        # Get active worksheet/tab
        worksheet = workbook.active
        worksheet.title = 'Bookings'

        # get headings of coloumn from key

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
        #response = content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',

        response['Content-Disposition'] = 'attachment; filename={date}-movies.xlsx'.format(date=datetime.now().strftime('%Y-%m-%d-%H:%M:%S'),)


        for keyval in self.jsonar:

            for key in keyval:

                self.keys.append(key)


        self.keys = list(dict.fromkeys(self.keys)) 

        self.columns = self.keys 

         # Assign the titles for each cell of the header
        for col_num, column_title in enumerate(self.columns, 1):
            cell = worksheet.cell(row=self.row_num, column=col_num)
            cell.value = column_title
      

        # get each row values

        for book in self.jsonar:

            self.row_num += 1

            for key in book:

                if key == 'Passangers':

                    print("Passenger found")

                else:    
                    #print(key)
                    #print('##')
                    #keys.append(key)
                    val = book[key]
                    #print(val)
                    #print('####')
                    self.value.append(val)

            self.row = self.value

            for col_num, cell_value in enumerate(self.row, 1):
                cell = worksheet.cell(row=self.row_num, column=col_num)
                cell.value = cell_value

            self.value.clear()    

        workbook.save(response)  

        return response



class Render:

    @staticmethod
    def render(path: str, params: dict):
        template = get_template(path)
        html = template.render(params)
        response = BytesIO()
        file = open("my.file.pdf", "wb")
        pdf = pisa.pisaDocument(BytesIO(html.encode("UTF-8")), file)
        file.close()
        if not pdf.err:
            return HttpResponse(response.getvalue(), content_type='application/pdf')
        else:
            return HttpResponse("Error Rendering PDF", status=400)

    @staticmethod
    def render_to_file(path: str, params: dict):
        template = get_template(path)
        html = template.render(params)
        file_name = "{0}{1}.pdf".format('Voucher_', randint(1, 1000000))
        file_path = os.path.join(os.path.abspath(os.path.dirname("__file__")), "media//Email_Voucher_PDF", file_name)
        with open(file_path, 'wb') as pdf:
            pisa.pisaDocument(BytesIO(html.encode("UTF-8")), pdf)
        return [file_name, file_path]

    @staticmethod
    def render_to_bill_pdf(path: str, params: dict):
        template = get_template(path)
        html = template.render(params)
        file_name = "{0}{1}.pdf".format('Voucher_', randint(1, 1000000))
        file_path = os.path.join(os.path.abspath(os.path.dirname("__file__")), "media//Bill_PDF", file_name)
        with open(file_path, 'wb') as pdf:
            pisa.pisaDocument(BytesIO(html.encode("UTF-8")), pdf)
        return [file_name, file_path]

    def render_to_bill_pdf_bill(path: str, params: dict):
        template = get_template(path)
        html = template.render(params)
        file_name = "{0}{1}.pdf".format('Bill_', randint(1, 1000000))
        file_path = os.path.join(os.path.abspath(os.path.dirname("__file__")), "media//Bill_PDF", file_name)
        with open(file_path, 'wb') as pdf:
            pisa.pisaDocument(BytesIO(html.encode("UTF-8")), pdf)
        return [file_name, file_path]


class TaxCalc():

    def __init__(self):
        self.base_rate = 100
        self.ticket_price = 0
        self.management_fee = 0
        self.tax_on_management_fee = 0
        self.tax_on_management_fee_percentage = 0
        self.sub_total = 0
        self.cotrav_billing_entity = 0
        self.igst = 0
        self.cgst = 0
        self.sgst = 0
        self.management_fee_igst = 0
        self.management_fee_cgst = 0
        self.management_fee_sgst = 0
        self.management_fee_igst_rate = 0
        self.management_fee_cgst_rate = 0
        self.management_fee_sgst_rate = 0

    def taxOnManagement(self):
        pass


    def taxIgst(self):
        pass

class TaxCalc():

    def __init__( self , corporateid , service_type , cotrav_billing_entity , corporate_billing_entity , ticket_price , no_of_passanger = 1 , oper_ticket_price = 0 , oper_cotrav_billing_entity = 0 , oper_billing_entity = 0 ):
        self.ticket_price = int(ticket_price)
        self.management_fee = 0
        self.tax_on_management_fee = 0
        self.tax_on_management_fee_percentage = TAX_PERC * 100
        self.tax_amount_on_management_fee = 0
        self.tax_perscentage = TAX_PERC;
        self.sub_total = 0
        self.booking_billing_entity = corporate_billing_entity
        self.cotrav_billing_entity = cotrav_billing_entity
        self.igst = 0
        self.cgst = 0
        self.sgst = 0
        self.management_fee_igst = 0
        self.management_fee_cgst = 0
        self.management_fee_sgst = 0
        self.management_fee_igst_rate = 0
        self.management_fee_cgst_rate = 0
        self.management_fee_sgst_rate = 0

        self.service_fees_type = int(service_type)
        self.no_of_passanger = int(no_of_passanger)

        self.oper_tax_perscentage = OPER_TAX_PERC;
        self.oper_ticket_price = oper_ticket_price
        self.oper_cotrav_billing_entity = oper_cotrav_billing_entity
        self.oper_billing_entity = oper_billing_entity

        self.gst_perc = 0.0
        self.gst_paid = 0.0

        self.oper_cgst_amount = 0
        self.oper_sgst_amount = 0
        self.oper_igst_amount = 0

        self.oper_cgst = 0.0
        self.oper_sgst = 0.0
        self.oper_igst = 0.0

        try:
            abc = ManagementFee.objects.get(corporate_id = int(corporateid) , service_fees_type_id = int(service_type) )

            self.service_fees_type = abc.service_fees_type

            print('service fee type')
            print(self.service_fees_type)

            if (self.service_fees_type == 2):

                print(self.ticket_price)

                if (self.ticket_price == 0):

                    self.management_fee = (abc.service_fees_type_value / 100) * self.no_of_passanger

                    self.management_fee = round(self.management_fee,2)

                else:

                    self.management_fee = ( self.ticket_price * (abc.service_fees_type_value /100 ) ) * self.no_of_passanger

                    self.management_fee = round(self.management_fee, 2)

            else:
                self.management_fee = abc.service_fees_type_value * self.no_of_passanger

                print(abc.service_fees_type_value)

                print(self.service_fees_type)


        except ManagementFee.DoesNotExist:

            self.management_fee = 100

            print('data not found')


    def taxOnManagement(self):
        self.tax_amount_on_management_fee =  self.management_fee * self.tax_perscentage
        self.tax_amount_on_management_fee = round(self.tax_amount_on_management_fee,2)

        return self.tax_amount_on_management_fee


    def gst(self):
        val1 = self.booking_billing_entity
        val2 = self.cotrav_billing_entity

        self.taxOnManagement()

        v1 = val1[:2]
        v2 = val2[:2]

        if ( v1 == v2 ):
            self.cgst = self.tax_perscentage / 2
            self.sgst = self.tax_perscentage / 2

            self.management_fee_cgst = self.cgst
            self.management_fee_sgst = self.sgst

            self.management_fee_cgst_rate = self.tax_amount_on_management_fee / 2
            self.management_fee_sgst_rate = self.tax_amount_on_management_fee / 2

        else:
            self.igst = self.tax_perscentage

            self.management_fee_igst = self.igst

            self.management_fee_igst_rate = self.tax_amount_on_management_fee

        return {'cgst': self.cgst , 'sgst': self.sgst , 'igst': self.igst }


    def hotel_gst(self):

        if ( self.ticket_price > 0 and self.ticket_price <1999 ) :

            self.hotel_gst = 0

        elif ( self.ticket_price > 1001 and self.ticket_price < 2500 ):

            self.hotel_gst = 12

        elif (self.ticket_price > 2501 and self.ticket_price < 7500):

            self.hotel_gst = 18

        else:

            self.hotel_gst = 28

        val1 = self.booking_billing_entity
        val2 = self.cotrav_billing_entity

        self.taxOnManagement()

        v1 = val1[:2]
        v2 = val2[:2]

        if (v1 == v2):

            if( self.hotel_gst == 0 ):
                self.cgst = 0
                self.sgst = 0
                self.management_fee_cgst_rate = self.tax_amount_on_management_fee / 2
                self.management_fee_sgst_rate = self.tax_amount_on_management_fee / 2

            else:

                self.cgst = ( self.hotel_gst / 2)

                self.sgst = ( self.hotel_gst / 2)

                self.management_fee_cgst = self.cgst
                self.management_fee_sgst = self.sgst

                self.management_fee_cgst_rate = self.tax_amount_on_management_fee / 2
                self.management_fee_sgst_rate = self.tax_amount_on_management_fee / 2

        else:

            self.igst = self.hotel_gst

            self.management_fee_igst = self.igst

            self.management_fee_igst_rate = self.tax_amount_on_management_fee


        return {'cgst': self.cgst, 'sgst': self.sgst, 'igst': self.igst}




    def cgst(self):

        return self.cgst


    def sgst(self):

        return self.sgst


    def igst(self):

        return self.igst



    def total_billing_amount(self):

        mang_fee_tax_amt = self.taxOnManagement()

        self.sub_total = int(self.ticket_price) + int(self.management_fee) + int(mang_fee_tax_amt)

        return self.sub_total



    def oper_tax_calc(self):

        #val1 = self.booking_billing_entity
        #val2 = self.cotrav_billing_entity

        val1 = self.oper_billing_entity
        val2 = self.cotrav_billing_entity

        print('oper_billing_entity')
        print(val1)
        print('cotrav_billing_entity')
        print(val2)


        self.amount = float(self.oper_ticket_price) * self.tax_perscentage

        v1 = val1[:2]
        v2 = val2[:2]

        if (v1 == v2):

            self.oper_cgst = self.oper_tax_perscentage / 2
            self.oper_sgst = self.oper_tax_perscentage / 2

            self.oper_cgst_amount = self.amount / 2
            self.oper_sgst_amount = self.amount / 2

            self.gst_perc = self.oper_tax_perscentage / 2

        else:

            self.oper_igst = self.oper_tax_perscentage
            self.oper_igst_amount = self.amount
            self.gst_perc = self.oper_tax_perscentage

        self.gst_paid = float(self.oper_ticket_price) * self.gst_perc

        return {'cgst': self.oper_cgst, 'sgst': self.oper_sgst, 'igst': self.oper_igst, 'gst_paid': self.gst_paid,
                'gst_perc': self.gst_perc, 'oper_cgst_amount': self.oper_cgst_amount,
                'oper_sgst_amount': self.oper_sgst_amount, 'oper_igst_amount': self.oper_igst_amount}



    def detailTax(self):

        tax = {

            "ticket_price": self.ticket_price,
            'management_fee': self.management_fee,
            'tax_on_management_fee': self.tax_on_management_fee,
            'tax_on_management_fee_percentage': self.tax_on_management_fee_percentage,
            'tax_amount_on_management_fee': self.tax_amount_on_management_fee,
            'tax_perscentage': self.tax_perscentage,
            'sub_total': self.sub_total,
            'booking_billing_entity': self.booking_billing_entity,
            'cotrav_billing_entity': self.cotrav_billing_entity,
            'igst': self.igst,
            'cgst': self.cgst,
            'sgst': self.sgst,

            'management_fee_igst': self.management_fee_igst,
            'management_fee_cgst': self.management_fee_cgst,
            'management_fee_sgst': self.management_fee_sgst,

            'management_fee_igst_rate': self.management_fee_igst_rate,
            'management_fee_cgst_rate': self.management_fee_cgst_rate,
            'management_fee_sgst_rate': self.management_fee_sgst_rate,

            'oper_cgst': self.oper_cgst,
            'oper_sgst': self.oper_sgst,
            'oper_igst': self.oper_igst,

            'gst_paid': self.gst_paid,
            'gst_perc': self.gst_perc,

            'oper_cgst_amount': self.oper_cgst_amount,
            'oper_sgst_amount': self.oper_sgst_amount,
            'oper_igst_amount': self.oper_igst_amount

        }

        return tax



